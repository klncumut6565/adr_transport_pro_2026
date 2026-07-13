"""DatabaseManager (monolitten satırı satırına)."""

from __future__ import annotations

import base64
import json
import logging
import os
import re
import shutil
import sqlite3
import uuid
from dataclasses import dataclass, field
from datetime import datetime, date, timedelta
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    from rapidfuzz import fuzz, process
    RAPIDFUZZ_AVAILABLE = True
except ImportError:
    RAPIDFUZZ_AVAILABLE = False

from .constants import *  # noqa: F403
from .models import *  # noqa: F403


def _adr_engine():
    """Geç import: db<->engines döngü riskine karşı (yalnız içe aktarıcı kullanır)."""
    from .engines import ADREngine
    return ADREngine

class DatabaseManager:
    """SQLite veritabani yonetimi - her islemde yeni baglanti."""

    def __init__(self, db_path: str = None):
        if db_path is None:
            app_dir = Path.home() / ".adr_transport_pro"
            app_dir.mkdir(exist_ok=True)
            db_path = str(app_dir / "adr_database.db")
        self.db_path = db_path
        self.connection = None
        self.init_database()
        self._setup_backup_system()

    @property
    def conn(self) -> sqlite3.Connection:
        """SecurityManager uyumluluğu icin kisayol."""
        return self._get_conn()

    def _get_conn(self) -> sqlite3.Connection:
        if self.connection is None:
            self.connection = sqlite3.connect(
                self.db_path, timeout=30, check_same_thread=False
            )
            self.connection.row_factory = sqlite3.Row
            self.connection.execute("PRAGMA journal_mode=WAL")
            self.connection.execute("PRAGMA synchronous=NORMAL")
            self.connection.execute("PRAGMA busy_timeout=30000")
            self.connection.execute("PRAGMA cache_size=-8000")
        return self.connection

    def init_database(self):
        conn = self._get_conn()
        cursor = conn.cursor()

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                type TEXT NOT NULL CHECK(type IN ('sender', 'receiver', 'carrier')),
                name TEXT NOT NULL,
                tax_number TEXT,
                tax_office TEXT,
                mersis_no TEXT,
                address TEXT,
                city TEXT,
                district TEXT,
                phone TEXT,
                email TEXT,
                contact_person TEXT,
                is_favorite INTEGER DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS drivers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                full_name TEXT NOT NULL,
                tc_no TEXT UNIQUE,
                phone TEXT,
                src5_no TEXT,
                src5_expiry TEXT,
                license_class TEXT,
                license_expiry TEXT,
                is_active INTEGER DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS vehicles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plate TEXT NOT NULL UNIQUE,
                trailer_plate TEXT,
                adr_compliance_cert_no TEXT,
                adr_compliance_expiry TEXT,
                inspection_date TEXT,
                inspection_expiry TEXT,
                tank_info TEXT,
                vehicle_type TEXT,
                max_capacity REAL DEFAULT 0,
                is_active INTEGER DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS chemicals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                un_number TEXT NOT NULL,
                classification_code TEXT DEFAULT '',
                proper_shipping_name_tr TEXT,
                proper_shipping_name_en TEXT,
                class_code TEXT,
                packing_group TEXT,
                tunnel_code TEXT,
                transport_category TEXT,
                segregation_group TEXT,
                special_provisions TEXT,
                lq_allowed INTEGER DEFAULT 0,
                eq_allowed INTEGER DEFAULT 0,
                limited_quantity TEXT DEFAULT '',
                excepted_quantity TEXT DEFAULT '',
                hazard_labels TEXT
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS shipments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                document_no TEXT NOT NULL UNIQUE,
                document_date TEXT,
                status TEXT DEFAULT 'Taslak',
                sender_id INTEGER,
                receiver_id INTEGER,
                carrier_id INTEGER,
                driver_id INTEGER,
                vehicle_id INTEGER,
                total_points REAL DEFAULT 0,
                orange_plate_required INTEGER DEFAULT 0,
                written_instructions_required INTEGER DEFAULT 0,
                driver_adr_required INTEGER DEFAULT 0,
                tunnel_restriction_code TEXT,
                exemption_type TEXT DEFAULT 'Yok',
                is_validated INTEGER DEFAULT 0,
                validation_errors TEXT,
                notes TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                created_by TEXT,
                FOREIGN KEY (sender_id)   REFERENCES companies(id),
                FOREIGN KEY (receiver_id) REFERENCES companies(id),
                FOREIGN KEY (carrier_id)  REFERENCES companies(id),
                FOREIGN KEY (driver_id)   REFERENCES drivers(id),
                FOREIGN KEY (vehicle_id)  REFERENCES vehicles(id)
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS shipment_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                shipment_id INTEGER NOT NULL,
                chemical_id INTEGER,
                un_number TEXT,
                proper_name TEXT,
                class_code TEXT,
                packing_group TEXT,
                packaging_type TEXT,
                packaging_count INTEGER DEFAULT 1,
                net_quantity REAL DEFAULT 0,
                gross_quantity REAL DEFAULT 0,
                unit TEXT DEFAULT 'kg',
                is_lq INTEGER DEFAULT 0,
                is_eq INTEGER DEFAULT 0,
                lq_max_per_package REAL DEFAULT 0,
                eq_max_per_package REAL DEFAULT 0,
                notes TEXT,
                FOREIGN KEY (shipment_id) REFERENCES shipments(id) ON DELETE CASCADE,
                FOREIGN KEY (chemical_id) REFERENCES chemicals(id)
            )
        """)

        # Migration: eski DB'lere tunnel_code / transport_category / segregation_group / classification_code ekle
        for col, default in [
            ("tunnel_code",         "''"),
            ("transport_category",  "''"),   # [DÜZELTİLDİ] '2' varsayılanı kaldırıldı
            ("segregation_group",   "''"),
            ("classification_code", "''"),
        ]:
            try:
                cursor.execute(
                    f"ALTER TABLE shipment_items ADD COLUMN {col} TEXT DEFAULT {default}"
                )
            except Exception:
                pass  # Kolon zaten varsa hata görmezden gel

        # Mevcut DB'de transport_category DEFAULT '2' ile eklenmiş BOŞLUK sorununu
        # çözmek için: artık ALTER TABLE DEFAULT '' kullanılıyor (yukarıda düzeltildi).
        # NOT: TC=2 geçerli bir ADR taşıma kategorisidir (UN 1203 Benzin gibi).
        #      Bu değerleri silmek YANLIŞ olur — UPDATE ifadeleri KALDIRILDI.

        # ------------------------------------------------------------------
        # [v4.2 MIGRASYON] chemicals: limit alanlari + bilesik anahtar
        # ------------------------------------------------------------------
        chem_cols = {row[1] for row in cursor.execute(
            "PRAGMA table_info(chemicals)").fetchall()}
        for col in ("classification_code", "limited_quantity", "excepted_quantity"):
            if col not in chem_cols:
                cursor.execute(
                    f"ALTER TABLE chemicals ADD COLUMN {col} TEXT DEFAULT ''")

        # Eski semalarda chemicals uzerinde hatali UNIQUE kisitlari vardi:
        # once sadece un_number (UN1950 gibi 12 varyantli maddeleri kirardi),
        # sonra un_number+classification_code+packing_group (bu da yanlis --
        # resmi ADR Tablo A'da bu UCLU AYNI olup yalnizca ozel hukum (6. sutun)
        # ile ayrilan gercekten farkli satirlar var, ornegin UN1133 F1 PG II
        # 640C ve 640D varyantlari; bu kisit boyle satirlari sessizce
        # birbirinin uzerine yaziyordu: 2939 gecerli satirdan yalnizca 2873'u
        # kaliyordu, 66 satir kayboluyordu). Artik chemicals uzerinde is
        # mantigi acisindan hicbir UNIQUE kisit yok; her Tablo A satiri kendi
        # kaydi olarak saklanir.
        table_sql = cursor.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='chemicals'"
        ).fetchone()
        _sql = (table_sql[0] or "") if table_sql else ""
        if "un_number TEXT NOT NULL UNIQUE" in _sql or "UNIQUE(un_number" in _sql:
            cursor.execute("ALTER TABLE chemicals RENAME TO chemicals_eski")
            cursor.execute("""
                CREATE TABLE chemicals (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    un_number TEXT NOT NULL,
                    classification_code TEXT DEFAULT '',
                    proper_shipping_name_tr TEXT,
                    proper_shipping_name_en TEXT,
                    class_code TEXT,
                    packing_group TEXT,
                    tunnel_code TEXT,
                    transport_category TEXT,
                    segregation_group TEXT,
                    special_provisions TEXT,
                    lq_allowed INTEGER DEFAULT 0,
                    eq_allowed INTEGER DEFAULT 0,
                    limited_quantity TEXT DEFAULT '',
                    excepted_quantity TEXT DEFAULT '',
                    hazard_labels TEXT
                )
            """)
            cursor.execute("""
                INSERT INTO chemicals
                    (id, un_number, classification_code, proper_shipping_name_tr,
                     proper_shipping_name_en, class_code, packing_group, tunnel_code,
                     transport_category, segregation_group, special_provisions,
                     lq_allowed, eq_allowed, limited_quantity, excepted_quantity,
                     hazard_labels)
                SELECT id, un_number, classification_code, proper_shipping_name_tr,
                       proper_shipping_name_en, class_code, packing_group, tunnel_code,
                       transport_category, segregation_group, special_provisions,
                       lq_allowed, eq_allowed, limited_quantity, excepted_quantity,
                       hazard_labels
                FROM chemicals_eski
            """)
            cursor.execute("DROP TABLE chemicals_eski")

        # companies.logo_path (PDF antet)
        comp_cols = {row[1] for row in cursor.execute(
            "PRAGMA table_info(companies)").fetchall()}
        if "logo_path" not in comp_cols:
            cursor.execute(
                "ALTER TABLE companies ADD COLUMN logo_path TEXT DEFAULT ''")

        # Firma urun katalogu: ayni ADR karsiligini paylasan ticari urunler
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS company_products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_name TEXT NOT NULL DEFAULT '',
                trade_name TEXT NOT NULL,
                un_number TEXT NOT NULL,
                classification_code TEXT DEFAULT '',
                packing_group TEXT DEFAULT '',
                limited_quantity TEXT DEFAULT '',
                excepted_quantity TEXT DEFAULT '',
                transport_category TEXT DEFAULT '',
                tunnel_code TEXT DEFAULT '',
                special_provisions TEXT DEFAULT '',
                source_file TEXT DEFAULT '',
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(trade_name, un_number, classification_code)
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS packaging_types (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name_tr TEXT,
                name_en TEXT,
                description TEXT,
                max_gross_mass REAL,
                un_tested INTEGER DEFAULT 0
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT DEFAULT CURRENT_TIMESTAMP,
                user_name TEXT,
                action TEXT,
                table_name TEXT,
                record_id INTEGER,
                old_values TEXT,
                new_values TEXT,
                ip_address TEXT
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT,
                description TEXT
            )
        """)

        cursor.execute('CREATE INDEX IF NOT EXISTS idx_companies_name   ON companies(name)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_companies_type   ON companies(type)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_chemicals_un     ON chemicals(un_number)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_chemicals_name   ON chemicals(proper_shipping_name_tr)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_shipments_docno  ON shipments(document_no)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_shipments_date   ON shipments(document_date)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_shipment_items_shipment ON shipment_items(shipment_id)')

        conn.commit()

    def _setup_backup_system(self):
        backup_dir = Path(self.db_path).parent / "backups"
        backup_dir.mkdir(exist_ok=True)

        today = datetime.now().strftime("%Y%m%d")
        backup_file = backup_dir / f"adr_backup_{today}.db"

        if not backup_file.exists():
            try:
                if self.connection:
                    self.connection.close()
                    self.connection = None
                shutil.copy2(self.db_path, backup_file)
            except Exception:
                pass
            finally:
                self._get_conn()

        cutoff = datetime.now() - timedelta(days=30)
        for f in backup_dir.glob("adr_backup_*.db"):
            try:
                file_date = datetime.strptime(f.stem.split("_")[-1], "%Y%m%d")
                if file_date < cutoff:
                    f.unlink()
            except:
                pass

    def execute(self, query: str, params: tuple = ()) -> List[sqlite3.Row]:
        cursor = self._get_conn().cursor()
        cursor.execute(query, params)
        return cursor.fetchall()

    def execute_one(self, query: str, params: tuple = ()) -> Optional[sqlite3.Row]:
        cursor = self._get_conn().cursor()
        cursor.execute(query, params)
        return cursor.fetchone()

    def execute_insert(self, query: str, params: tuple = ()) -> int:
        conn = self._get_conn()
        cursor = conn.cursor()
        cursor.execute(query, params)
        conn.commit()
        return cursor.lastrowid

    def execute_update(self, query: str, params: tuple = ()) -> int:
        conn = self._get_conn()
        cursor = conn.cursor()
        cursor.execute(query, params)
        conn.commit()
        return cursor.rowcount

    def execute_delete(self, query: str, params: tuple = ()) -> int:
        return self.execute_update(query, params)

    def close(self):
        if self.connection:
            try:
                self.connection.commit()
                self.connection.close()
            except Exception:
                pass
            finally:
                self.connection = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    # --- COMPANY CRUD ---

    def add_company(self, company: Company) -> int:
        existing = self.execute_one(
            "SELECT id FROM companies WHERE LOWER(name)=LOWER(?) AND type=?",
            (company.name.strip(), company.type)
        )
        if existing:
            return existing["id"]

        query = """
            INSERT INTO companies (type, name, tax_number, tax_office, mersis_no,
                address, city, district, phone, email, contact_person, is_favorite)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        return self.execute_insert(query, (
            company.type, company.name, company.tax_number, company.tax_office,
            company.mersis_no, company.address, company.city, company.district,
            company.phone, company.email, company.contact_person, int(company.is_favorite)
        ))

    def update_company(self, company: Company) -> int:
        query = """
            UPDATE companies SET type=?, name=?, tax_number=?, tax_office=?,
                mersis_no=?, address=?, city=?, district=?, phone=?, email=?,
                contact_person=?, is_favorite=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """
        return self.execute_update(query, (
            company.type, company.name, company.tax_number, company.tax_office,
            company.mersis_no, company.address, company.city, company.district,
            company.phone, company.email, company.contact_person, int(company.is_favorite),
            company.id
        ))

    def delete_company(self, company_id: int) -> int:
        return self.execute_delete("DELETE FROM companies WHERE id=?", (company_id,))

    def get_company(self, company_id: int) -> Optional[Company]:
        row = self.execute_one("SELECT * FROM companies WHERE id=?", (company_id,))
        if row:
            return self._row_to_company(row)
        return None

    def get_companies(self, company_type: str = None, search: str = None,
                      favorite_only: bool = False) -> List[Company]:
        query = "SELECT * FROM companies WHERE 1=1"
        params = []
        if company_type:
            query += " AND type=?"
            params.append(company_type)
        if favorite_only:
            query += " AND is_favorite=1"
        if search:
            query += " AND (name LIKE ? OR tax_number LIKE ? OR city LIKE ?)"
            params.extend([f"%{search}%"] * 3)
        query += " ORDER BY name"
        rows = self.execute(query, tuple(params))
        return [self._row_to_company(r) for r in rows]

    def _row_to_company(self, row: sqlite3.Row) -> Company:
        return Company(
            id=row["id"], type=row["type"], name=row["name"],
            tax_number=row["tax_number"] or "", tax_office=row["tax_office"] or "",
            mersis_no=row["mersis_no"] or "", address=row["address"] or "",
            city=row["city"] or "", district=row["district"] or "",
            phone=row["phone"] or "", email=row["email"] or "",
            contact_person=row["contact_person"] or "",
            is_favorite=bool(row["is_favorite"]),
            created_at=row["created_at"] or "", updated_at=row["updated_at"] or ""
        )

    # --- DRIVER CRUD ---

    def add_driver(self, driver: Driver) -> int:
        query = """
            INSERT INTO drivers (full_name, tc_no, phone,
                src5_no, src5_expiry, license_class, license_expiry)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """
        return self.execute_insert(query, (
            driver.full_name, driver.tc_no, driver.phone,
            driver.src5_no, driver.src5_expiry,
            driver.license_class, driver.license_expiry
        ))

    def update_driver(self, driver: Driver) -> int:
        query = """
            UPDATE drivers SET full_name=?, tc_no=?, phone=?,
                src5_no=?, src5_expiry=?, license_class=?,
                license_expiry=?, is_active=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """
        return self.execute_update(query, (
            driver.full_name, driver.tc_no, driver.phone,
            driver.src5_no, driver.src5_expiry,
            driver.license_class, driver.license_expiry, int(driver.is_active), driver.id
        ))

    def get_driver(self, driver_id: int) -> Optional[Driver]:
        row = self.execute_one("SELECT * FROM drivers WHERE id=?", (driver_id,))
        return self._row_to_driver(row) if row else None

    def get_drivers(self, search: str = None, active_only: bool = True) -> List[Driver]:
        query = "SELECT * FROM drivers WHERE 1=1"
        params = []
        if active_only:
            query += " AND is_active=1"
        if search:
            query += " AND (full_name LIKE ? OR tc_no LIKE ? OR src5_no LIKE ?)"
            params.extend([f"%{search}%"] * 3)
        query += " ORDER BY full_name"
        rows = self.execute(query, tuple(params))
        return [self._row_to_driver(r) for r in rows]

    def _row_to_driver(self, row: sqlite3.Row) -> Driver:
        return Driver(
            id=row["id"], full_name=row["full_name"], tc_no=row["tc_no"] or "",
            phone=row["phone"] or "",
            src5_no=row["src5_no"] or "", src5_expiry=row["src5_expiry"] or "",
            license_class=row["license_class"] or "",
            license_expiry=row["license_expiry"] or "",
            is_active=bool(row["is_active"])
        )

    # --- VEHICLE CRUD ---

    def add_vehicle(self, vehicle: Vehicle) -> int:
        query = """
            INSERT INTO vehicles (plate, trailer_plate, adr_compliance_cert_no,
                adr_compliance_expiry, inspection_date, inspection_expiry, tank_info,
                vehicle_type, max_capacity)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        return self.execute_insert(query, (
            vehicle.plate, vehicle.trailer_plate, vehicle.adr_compliance_cert_no,
            vehicle.adr_compliance_expiry, vehicle.inspection_date,
            vehicle.inspection_expiry, vehicle.tank_info, vehicle.vehicle_type,
            vehicle.max_capacity
        ))

    def update_vehicle(self, vehicle: Vehicle) -> int:
        query = """
            UPDATE vehicles SET plate=?, trailer_plate=?, adr_compliance_cert_no=?,
                adr_compliance_expiry=?, inspection_date=?, inspection_expiry=?,
                tank_info=?, vehicle_type=?, max_capacity=?, is_active=?,
                updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """
        return self.execute_update(query, (
            vehicle.plate, vehicle.trailer_plate, vehicle.adr_compliance_cert_no,
            vehicle.adr_compliance_expiry, vehicle.inspection_date,
            vehicle.inspection_expiry, vehicle.tank_info, vehicle.vehicle_type,
            vehicle.max_capacity, int(vehicle.is_active), vehicle.id
        ))

    def get_vehicle(self, vehicle_id: int) -> Optional[Vehicle]:
        row = self.execute_one("SELECT * FROM vehicles WHERE id=?", (vehicle_id,))
        return self._row_to_vehicle(row) if row else None

    def get_vehicles(self, search: str = None, active_only: bool = True) -> List[Vehicle]:
        query = "SELECT * FROM vehicles WHERE 1=1"
        params = []
        if active_only:
            query += " AND is_active=1"
        if search:
            query += " AND (plate LIKE ? OR trailer_plate LIKE ?)"
            params.extend([f"%{search}%"] * 2)
        query += " ORDER BY plate"
        rows = self.execute(query, tuple(params))
        return [self._row_to_vehicle(r) for r in rows]

    def _row_to_vehicle(self, row: sqlite3.Row) -> Vehicle:
        return Vehicle(
            id=row["id"], plate=row["plate"], trailer_plate=row["trailer_plate"] or "",
            adr_compliance_cert_no=row["adr_compliance_cert_no"] or "",
            adr_compliance_expiry=row["adr_compliance_expiry"] or "",
            inspection_date=row["inspection_date"] or "",
            inspection_expiry=row["inspection_expiry"] or "",
            tank_info=row["tank_info"] or "", vehicle_type=row["vehicle_type"] or "",
            max_capacity=row["max_capacity"] or 0, is_active=bool(row["is_active"])
        )

    # --- CHEMICAL CRUD ---

    def add_chemical(self, chemical: Chemical) -> int:
        query = """
            INSERT INTO chemicals (un_number, classification_code,
                proper_shipping_name_tr, proper_shipping_name_en, class_code,
                packing_group, tunnel_code, transport_category, segregation_group,
                special_provisions, lq_allowed, eq_allowed,
                limited_quantity, excepted_quantity, hazard_labels)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        return self.execute_insert(query, (
            chemical.un_number, chemical.classification_code,
            chemical.proper_shipping_name_tr,
            chemical.proper_shipping_name_en, chemical.class_code,
            chemical.packing_group, chemical.tunnel_code, chemical.transport_category,
            chemical.segregation_group, chemical.special_provisions,
            int(chemical.lq_allowed), int(chemical.eq_allowed),
            chemical.limited_quantity, chemical.excepted_quantity,
            chemical.hazard_labels
        ))

    def count_chemicals(self) -> int:
        row = self.execute_one("SELECT COUNT(*) AS n FROM chemicals")
        return row["n"] if row else 0

    def _upsert_chemical(self, c: Chemical) -> bool:
        """True=yeni eklendi, False=mevcut guncellendi.
        Sadece firma envanteri ice aktarmada kullanilir (import_company_inventory_excel);
        bir firmanin kendi envanterinde ayni UN+siniflandirma+PG genelde tek
        gercek urunu temsil eder. UYARI: resmi ADR Tablo A icin bu anahtar
        GECERLI DEGIL (import_table_a_excel artik bunu kullanmiyor) -- Tablo
        A'da bu uclu ayni olup yalnizca ozel hukumle ayrisan gercek satirlar
        var, bkz. import_table_a_excel docstring'i."""
        existing = self.execute_one(
            """SELECT id FROM chemicals
               WHERE un_number=? AND classification_code=? AND packing_group=?""",
            (c.un_number, c.classification_code or "", c.packing_group or ""))
        if existing:
            self.execute_update(
                """UPDATE chemicals SET proper_shipping_name_tr=?, class_code=?,
                   tunnel_code=?, transport_category=?, special_provisions=?,
                   limited_quantity=?, excepted_quantity=?,
                   lq_allowed=?, eq_allowed=?, hazard_labels=?
                   WHERE un_number=? AND classification_code=? AND packing_group=?""",
                (c.proper_shipping_name_tr, c.class_code,
                 c.tunnel_code, c.transport_category, c.special_provisions,
                 c.limited_quantity, c.excepted_quantity, int(c.lq_allowed),
                 int(c.eq_allowed), c.hazard_labels,
                 c.un_number, c.classification_code or "", c.packing_group or ""))
            return False
        self.add_chemical(c)
        return True

    def update_chemical(self, chemical: Chemical) -> int:
        query = """
            UPDATE chemicals SET un_number=?, proper_shipping_name_tr=?,
                proper_shipping_name_en=?, class_code=?, packing_group=?,
                tunnel_code=?, transport_category=?, segregation_group=?,
                special_provisions=?, lq_allowed=?, eq_allowed=?,
                hazard_labels=?
            WHERE id=?
        """
        return self.execute_update(query, (
            chemical.un_number, chemical.proper_shipping_name_tr,
            chemical.proper_shipping_name_en, chemical.class_code,
            chemical.packing_group, chemical.tunnel_code, chemical.transport_category,
            chemical.segregation_group, chemical.special_provisions,
            int(chemical.lq_allowed), int(chemical.eq_allowed), chemical.hazard_labels, chemical.id
        ))

    def delete_chemical(self, chemical_id: int) -> int:
        return self.execute_delete("DELETE FROM chemicals WHERE id=?", (chemical_id,))

    def get_chemical(self, chemical_id: int) -> Optional[Chemical]:
        row = self.execute_one("SELECT * FROM chemicals WHERE id=?", (chemical_id,))
        return self._row_to_chemical(row) if row else None

    def get_chemical_by_un(self, un_number: str) -> Optional[Chemical]:
        row = self.execute_one("SELECT * FROM chemicals WHERE un_number=?", (un_number,))
        return self._row_to_chemical(row) if row else None

    def search_chemicals(self, query: str, limit: int = 50) -> List[Chemical]:
        query = query.strip().upper()
        if not query:
            return []

        # 4 haneli sayisal sorgu -> UN TAM eslesme: maddenin TUM
        # varyasyonlari (UN1950 = 12 satir) doner, yabanci UN gelmez.
        if query.isdigit():
            sql = """
                SELECT * FROM chemicals WHERE un_number = ?
                ORDER BY classification_code, packing_group
                LIMIT ?
            """
            params = (query.zfill(4), limit)
        else:
            sql = """
                SELECT * FROM chemicals
                WHERE proper_shipping_name_tr LIKE ?
                   OR proper_shipping_name_en LIKE ?
                   OR class_code LIKE ?
                ORDER BY
                    CASE WHEN proper_shipping_name_tr LIKE ? THEN 0 ELSE 1 END,
                    un_number, classification_code
                LIMIT ?
            """
            params = (f"%{query}%", f"%{query}%", f"{query}%", f"{query}%", limit)
        rows = self.execute(sql, params)
        chemicals = [self._row_to_chemical(r) for r in rows]

        if RAPIDFUZZ_AVAILABLE and not query.isdigit() and len(chemicals) < limit:
            all_chems = self.execute("SELECT * FROM chemicals", ())
            all_list = [self._row_to_chemical(r) for r in all_chems]

            search_texts = []
            for c in all_list:
                search_texts.append(
                    f"{c.un_number} {c.proper_shipping_name_tr} "
                    f"{c.proper_shipping_name_en} {c.class_code}"
                )

            results = process.extract(query, search_texts, scorer=fuzz.WRatio, limit=limit)
            fuzzy_ids = set()
            for match, score, idx in results:
                if score > 60:
                    fuzzy_ids.add(all_list[idx].id)

            existing_ids = {c.id for c in chemicals}
            for c in all_list:
                if c.id in fuzzy_ids and c.id not in existing_ids:
                    chemicals.append(c)

        return chemicals[:limit]

    def get_all_chemicals(self, search: str = None, class_filter: str = None,
                           limit: int = 500) -> List[Chemical]:
        """Tum kimyasallari filtreli getir (veritabani sayfasi icin)."""
        sql = "SELECT * FROM chemicals WHERE 1=1"
        params = []
        if search:
            s = f"%{search.upper()}%"
            sql += " AND (UPPER(un_number) LIKE ? OR UPPER(proper_shipping_name_tr) LIKE ? OR UPPER(proper_shipping_name_en) LIKE ?)"
            params.extend([s, s, s])
        if class_filter and class_filter != "Tumu":
            sql += " AND class_code=?"
            params.append(class_filter)
        sql += " ORDER BY un_number LIMIT ?"
        params.append(limit)
        rows = self.execute(sql, tuple(params))
        return [self._row_to_chemical(r) for r in rows]

    def _row_to_chemical(self, row: sqlite3.Row) -> Chemical:
        return Chemical(
            id=row["id"], un_number=row["un_number"] or "",
            proper_shipping_name_tr=row["proper_shipping_name_tr"] or "",
            proper_shipping_name_en=row["proper_shipping_name_en"] or "",
            class_code=row["class_code"] or "", packing_group=row["packing_group"] or "",
            tunnel_code=row["tunnel_code"] or "", transport_category=row["transport_category"] or "",
            segregation_group=row["segregation_group"] or "",
            special_provisions=row["special_provisions"] or "",
            lq_allowed=bool(row["lq_allowed"]), eq_allowed=bool(row["eq_allowed"]),
            classification_code=(row["classification_code"] if "classification_code" in row.keys() else "") or "",
            limited_quantity=(row["limited_quantity"] if "limited_quantity" in row.keys() else "") or "",
            excepted_quantity=(row["excepted_quantity"] if "excepted_quantity" in row.keys() else "") or "",
            hazard_labels=row["hazard_labels"] or ""
        )

    # --- SHIPMENT CRUD ---

    def set_shipment_validation(self, shipment_id: int, is_valid: bool,
                                errors_text: str = "") -> None:
        """Dogrulama sonucunu KALICI hale getirir: gecerli evrak 'Onaylandi'
        statusune yukselir, gecersiz evrak 'Taslak'a doner ve hata metni
        validation_errors alanina yazilir. (Eski surumde dogrulama yalnizca
        mesaj kutusunda gosterilip kayboluyordu.)"""
        self.execute_update(
            """UPDATE shipments SET status=?, is_validated=?, validation_errors=?,
               updated_at=CURRENT_TIMESTAMP WHERE id=?""",
            (DocumentStatus.VALIDATED.value if is_valid else DocumentStatus.DRAFT.value,
             int(is_valid), errors_text or "", shipment_id))

    def add_shipment(self, shipment: Shipment) -> int:
        query = """
            INSERT INTO shipments (document_no, document_date, status, sender_id,
                receiver_id, carrier_id, driver_id, vehicle_id, total_points,
                orange_plate_required, written_instructions_required,
                driver_adr_required, tunnel_restriction_code, exemption_type,
                is_validated, validation_errors, notes, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        return self.execute_insert(query, (
            shipment.document_no, shipment.document_date, shipment.status,
            shipment.sender_id, shipment.receiver_id, shipment.carrier_id,
            shipment.driver_id, shipment.vehicle_id, shipment.total_points,
            int(shipment.orange_plate_required), int(shipment.written_instructions_required),
            int(shipment.driver_adr_required), shipment.tunnel_restriction_code,
            shipment.exemption_type, int(shipment.is_validated),
            shipment.validation_errors, shipment.notes, shipment.created_by
        ))

    def update_shipment(self, shipment: Shipment) -> int:
        query = """
            UPDATE shipments SET document_no=?, document_date=?, status=?,
                sender_id=?, receiver_id=?, carrier_id=?, driver_id=?, vehicle_id=?,
                total_points=?, orange_plate_required=?, written_instructions_required=?,
                driver_adr_required=?, tunnel_restriction_code=?, exemption_type=?,
                is_validated=?, validation_errors=?, notes=?,
                updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """
        return self.execute_update(query, (
            shipment.document_no, shipment.document_date, shipment.status,
            shipment.sender_id, shipment.receiver_id, shipment.carrier_id,
            shipment.driver_id, shipment.vehicle_id, shipment.total_points,
            int(shipment.orange_plate_required), int(shipment.written_instructions_required),
            int(shipment.driver_adr_required), shipment.tunnel_restriction_code,
            shipment.exemption_type, int(shipment.is_validated),
            shipment.validation_errors, shipment.notes, shipment.id
        ))

    def get_shipment(self, shipment_id: int) -> Optional[Shipment]:
        row = self.execute_one("SELECT * FROM shipments WHERE id=?", (shipment_id,))
        return self._row_to_shipment(row) if row else None

    def get_shipments(self, status: str = None, date_from: str = None,
                      date_to: str = None, search: str = None,
                      limit: int = 100, offset: int = 0) -> List[Shipment]:
        query = "SELECT * FROM shipments WHERE 1=1"
        params = []
        if status:
            query += " AND status=?"
            params.append(status)
        if date_from:
            query += " AND document_date >= ?"
            params.append(date_from)
        if date_to:
            query += " AND document_date <= ?"
            params.append(date_to)
        if search:
            query += " AND document_no LIKE ?"
            params.append(f"%{search}%")
        query += " ORDER BY document_date DESC, id DESC LIMIT ? OFFSET ?"
        params.extend([limit, offset])
        rows = self.execute(query, tuple(params))
        return [self._row_to_shipment(r) for r in rows]

    def delete_shipment(self, shipment_id: int) -> int:
        return self.execute_delete("DELETE FROM shipments WHERE id=?", (shipment_id,))

    def get_shipments_with_details(self, status: str = None, date_from: str = None,
                                    date_to: str = None, search: str = None,
                                    limit: int = 500) -> List[dict]:
        """
        Tek JOIN sorgusuyla sevkiyat + firma + surucu + arac bilgilerini getirir.
        N+1 sorgu sorununu cozer: 500 sevkiyat icin 2000 sorgu yerine 1 sorgu.
        """
        sql = """
            SELECT
                s.id, s.document_no, s.document_date, s.status,
                s.total_points, s.orange_plate_required, s.exemption_type,
                s.tunnel_restriction_code,
                sender.name   AS sender_name,
                receiver.name AS receiver_name,
                d.full_name   AS driver_name,
                v.plate       AS vehicle_plate
            FROM shipments s
            LEFT JOIN companies sender   ON sender.id   = s.sender_id
            LEFT JOIN companies receiver ON receiver.id = s.receiver_id
            LEFT JOIN drivers d          ON d.id        = s.driver_id
            LEFT JOIN vehicles v         ON v.id        = s.vehicle_id
            WHERE 1=1
        """
        params = []
        if status:
            sql += " AND s.status=?"
            params.append(status)
        if date_from:
            sql += " AND s.document_date >= ?"
            params.append(date_from)
        if date_to:
            sql += " AND s.document_date <= ?"
            params.append(date_to)
        if search:
            sql += """ AND (
                s.document_no LIKE ? OR
                sender.name   LIKE ? OR
                receiver.name LIKE ? OR
                d.full_name   LIKE ? OR
                v.plate       LIKE ?
            )"""
            term = f"%{search}%"
            params.extend([term, term, term, term, term])
        sql += " ORDER BY s.document_date DESC, s.id DESC LIMIT ?"
        params.append(limit)
        rows = self.execute(sql, tuple(params))
        return [dict(r) for r in rows]

    def _row_to_shipment(self, row: sqlite3.Row) -> Shipment:
        return Shipment(
            id=row["id"], document_no=row["document_no"] or "",
            document_date=row["document_date"] or "", status=row["status"] or "Taslak",
            sender_id=row["sender_id"] or 0, receiver_id=row["receiver_id"] or 0,
            carrier_id=row["carrier_id"] or 0, driver_id=row["driver_id"] or 0,
            vehicle_id=row["vehicle_id"] or 0, total_points=row["total_points"] or 0,
            orange_plate_required=bool(row["orange_plate_required"]),
            written_instructions_required=bool(row["written_instructions_required"]),
            driver_adr_required=bool(row["driver_adr_required"]),
            tunnel_restriction_code=row["tunnel_restriction_code"] or "",
            exemption_type=row["exemption_type"] or "Yok",
            is_validated=bool(row["is_validated"]),
            validation_errors=row["validation_errors"] or "",
            notes=row["notes"] or ""
        )

    # --- SHIPMENT ITEM CRUD ---

    def add_shipment_item(self, item: ShipmentItem) -> int:
        query = """
            INSERT INTO shipment_items (shipment_id, chemical_id, un_number,
                proper_name, class_code, packing_group, packaging_type, packaging_count,
                net_quantity, gross_quantity, unit, is_lq, is_eq, lq_max_per_package,
                eq_max_per_package, notes, tunnel_code, transport_category, segregation_group,
                classification_code)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        return self.execute_insert(query, (
            item.shipment_id, item.chemical_id, item.un_number, item.proper_name,
            item.class_code, item.packing_group, item.packaging_type, item.packaging_count,
            item.net_quantity, item.gross_quantity, item.unit, int(item.is_lq),
            int(item.is_eq), item.lq_max_per_package, item.eq_max_per_package, item.notes,
            item.tunnel_code, item.transport_category, item.segregation_group,
            item.classification_code
        ))

    def get_shipment_items(self, shipment_id: int) -> List[ShipmentItem]:
        rows = self.execute("SELECT * FROM shipment_items WHERE shipment_id=?", (shipment_id,))
        return [self._row_to_shipment_item(r) for r in rows]

    def delete_shipment_items(self, shipment_id: int) -> int:
        return self.execute_delete("DELETE FROM shipment_items WHERE shipment_id=?", (shipment_id,))

    def _row_to_shipment_item(self, row: sqlite3.Row) -> ShipmentItem:
        # transport_category, tunnel_code ve segregation_group DB'den okunur.
        # Boş/None gelirse chemicals tablosundan tamamlanır (UN 1202 gibi TC=3 olan maddeler).
        keys = row.keys() if hasattr(row, "keys") else []
        tc_raw = row["transport_category"] if "transport_category" in keys else ""
        tunnel_raw = row["tunnel_code"] if "tunnel_code" in keys else ""
        seg_raw = row["segregation_group"] if "segregation_group" in keys else ""

        item = ShipmentItem(
            id=row["id"], shipment_id=row["shipment_id"], chemical_id=row["chemical_id"] or 0,
            un_number=row["un_number"] or "", proper_name=row["proper_name"] or "",
            class_code=row["class_code"] or "", packing_group=row["packing_group"] or "",
            packaging_type=row["packaging_type"] or "", packaging_count=row["packaging_count"] or 0,
            net_quantity=row["net_quantity"] or 0, gross_quantity=row["gross_quantity"] or 0,
            unit=row["unit"] or "kg", is_lq=bool(row["is_lq"]), is_eq=bool(row["is_eq"]),
            lq_max_per_package=row["lq_max_per_package"] or 0,
            eq_max_per_package=row["eq_max_per_package"] or 0,
            notes=row["notes"] or "",
            transport_category=tc_raw or "",
            tunnel_code=tunnel_raw or "",
            segregation_group=seg_raw or "",
            classification_code=row["classification_code"] if "classification_code" in keys else "",
        )
        # TC boşsa chemicals tablosundan tamamla
        if not item.transport_category and item.chemical_id:
            chem = self.get_chemical(item.chemical_id)
            if chem:
                item.transport_category = chem.transport_category or ""
                if not item.tunnel_code:
                    item.tunnel_code = chem.tunnel_code or ""
                if not item.segregation_group:
                    item.segregation_group = chem.segregation_group or ""
        return item

    # --- AYARLAR ---

    def get_setting(self, key: str, default: str = "") -> str:
        row = self.execute_one("SELECT value FROM settings WHERE key=?", (key,))
        return row["value"] if row else default

    def set_setting(self, key: str, value: str, description: str = ""):
        self.execute_update(
            "INSERT OR REPLACE INTO settings (key, value, description) VALUES (?, ?, ?)",
            (key, value, description)
        )

    # --- ISTATISTIKLER ---

    def get_statistics(self) -> Dict[str, Any]:
        stats = {}
        cursor = self._get_conn().cursor()

        cursor.execute("SELECT COUNT(*) as count FROM shipments")
        stats["total_shipments"] = cursor.fetchone()["count"]

        cursor.execute("SELECT COUNT(*) as count FROM shipments WHERE status='Taslak'")
        stats["draft_shipments"] = cursor.fetchone()["count"]

        cursor.execute("SELECT COUNT(*) as count FROM companies")
        stats["total_companies"] = cursor.fetchone()["count"]

        cursor.execute("SELECT COUNT(*) as count FROM drivers WHERE is_active=1")
        stats["active_drivers"] = cursor.fetchone()["count"]

        cursor.execute("SELECT COUNT(*) as count FROM vehicles WHERE is_active=1")
        stats["active_vehicles"] = cursor.fetchone()["count"]

        cursor.execute("SELECT COUNT(*) as count FROM chemicals")
        stats["total_chemicals"] = cursor.fetchone()["count"]

        cursor.execute("""
            SELECT strftime('%Y-%m', document_date) as month, COUNT(*) as count
            FROM shipments
            WHERE document_date >= date('now', '-6 months')
            GROUP BY month
            ORDER BY month
        """)
        stats["monthly_shipments"] = {r["month"]: r["count"] for r in cursor.fetchall()}

        return stats

    # --- IMPORT/EXPORT ---

    def get_expiring_documents(self, days: int = 30) -> dict:
        from datetime import date, timedelta
        today = date.today().isoformat()
        limit = (date.today() + timedelta(days=days)).isoformat()
        # DÜZELTME: doğrudan conn.execute() kullanmak Pg alt sınıfındaki
        # kiracı sarmalayıcısını (SET LOCAL app.tenant_id) atlıyordu —
        # drivers/vehicles kiracıya özel tablolar olduğu için bu, RLS'in
        # her zaman varsayılan kiracıya (1) düşmesi anlamına geliyordu.
        # self.execute() üzerinden gitmek her iki motor için de doğru.
        drivers = self.execute("""
            SELECT full_name, src5_no, src5_expiry,
                   CAST(julianday(src5_expiry) - julianday('now') AS INTEGER) AS kalan
            FROM drivers WHERE is_active=1 AND src5_expiry BETWEEN ? AND ?
            ORDER BY src5_expiry""", (today, limit))
        vehicles = self.execute("""
            SELECT plate, adr_compliance_expiry, inspection_expiry,
                   CAST(julianday(adr_compliance_expiry) - julianday('now') AS INTEGER) AS adr_kalan,
                   CAST(julianday(inspection_expiry)     - julianday('now') AS INTEGER) AS mua_kalan
            FROM vehicles WHERE is_active=1
              AND (adr_compliance_expiry BETWEEN ? AND ? OR inspection_expiry BETWEEN ? AND ?)
            ORDER BY adr_compliance_expiry""", (today, limit, today, limit))
        return {"drivers":[dict(r) for r in drivers],"vehicles":[dict(r) for r in vehicles]}

    def get_class_breakdown(self, year=None) -> list:
        sql = """SELECT si.class_code, COUNT(DISTINCT si.shipment_id) AS sevkiyat_sayisi,
                        SUM(si.net_quantity) AS toplam_net_kg
                 FROM shipment_items si JOIN shipments s ON s.id=si.shipment_id"""
        p=[]
        if year: sql+=" WHERE strftime('%Y',s.document_date)=?"; p.append(str(year))
        sql+=" GROUP BY si.class_code ORDER BY toplam_net_kg DESC"
        # (aynı düzeltme: self.execute() — bkz. get_expiring_documents notu)
        return [dict(r) for r in self.execute(sql, p)]

    def get_top_senders(self, limit=10, year=None) -> list:
        sql = "SELECT c.name, COUNT(s.id) AS sevkiyat_sayisi FROM shipments s JOIN companies c ON c.id=s.sender_id"
        p=[]
        if year: sql+=" WHERE strftime('%Y',s.document_date)=?"; p.append(str(year))
        sql+=" GROUP BY s.sender_id ORDER BY sevkiyat_sayisi DESC LIMIT ?"; p.append(limit)
        return [dict(r) for r in self._get_conn().execute(sql,p).fetchall()]

    def get_top_chemicals(self, limit=10, year=None) -> list:
        sql = """SELECT si.un_number, si.class_code, COUNT(*) AS adet,
                        SUM(si.net_quantity) AS toplam_net_kg
                 FROM shipment_items si JOIN shipments s ON s.id=si.shipment_id"""
        p=[]
        if year: sql+=" WHERE strftime('%Y',s.document_date)=?"; p.append(str(year))
        sql+=" GROUP BY si.un_number ORDER BY toplam_net_kg DESC LIMIT ?"; p.append(limit)
        return [dict(r) for r in self._get_conn().execute(sql,p).fetchall()]

    def backup_now(self) -> str:
        import shutil
        backup_dir = Path(self.db_path).parent / "backups"
        backup_dir.mkdir(exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = backup_dir / f"adr_backup_{ts}.db"
        shutil.copy2(self.db_path, dest)
        return str(dest)

    def list_backups(self) -> list:
        backup_dir = Path(self.db_path).parent / "backups"
        if not backup_dir.exists(): return []
        files = sorted(backup_dir.glob("adr_backup_*.db"), reverse=True)
        return [{"name":f.name,"path":str(f),
                 "size":f"{f.stat().st_size/1024:.1f} KB",
                 "date":datetime.fromtimestamp(f.stat().st_mtime).strftime("%d.%m.%Y %H:%M")}
                for f in files]

    def restore_backup(self, backup_path: str):
        import shutil
        if self.connection: self.connection.close(); self.connection=None
        shutil.copy2(backup_path, self.db_path); self._get_conn()

    def get_company_logo_b64(self) -> str:
        return self.get_setting("doc_company_logo_b64") or ""

    def set_company_logo_b64(self, b64: str):
        self.set_setting("doc_company_logo_b64", b64, "Şirket logosu base64")


    # ------------------------------------------------------------------
    # [v4.2] Gercek Excel kaynaklari: ADR Tablo A + firma envanteri
    # (main.py surumunden birebir tasindi, gercek dosyalarla dogrulandi)
    # ------------------------------------------------------------------
    @staticmethod
    def _xl_clean(val) -> str:
        if val is None:
            return ""
        s = str(val).replace("\n", " ").replace("\r", " ").strip()
        import re as _re
        s = _re.sub(r" {2,}", " ", s)
        return "" if s.lower() == "nan" else s

    @staticmethod
    def _xl_un(val) -> str:
        raw = DatabaseManager._xl_clean(val)
        if raw.endswith(".0"):
            raw = raw[:-2]
        import re as _re
        m = _re.match(r"^(?:UN\s*)?(\d{1,4})$", raw, _re.IGNORECASE)
        return m.group(1).zfill(4) if m else ""

    @staticmethod
    def _xl_category_tunnel(raw: str) -> Tuple[str, str]:
        """'2 (D/E)' benzeri hucreyi (kategori, tunel) olarak ayirir."""
        import re as _re
        text = DatabaseManager._xl_clean(raw)
        cat = ""
        m = _re.search(r"\b([0-4])\b", text)
        if m:
            cat = m.group(1)
        tunnel = ""
        m = _re.search(r"\(([A-E][^\)]*)\)", text)
        if m:
            tunnel = m.group(1).strip()
        return cat, tunnel

    def import_table_a_excel(self, xlsx_path: str) -> int:
        """Resmi ADR Tablo A Excel'ini (cok satirli baslik, sutun 7a/7b
        dahil) uygulamanin kimyasal veritabanina aktarir.

        ONEMLI: Her satir kendi kaydi olarak eklenir, birlestirme YAPILMAZ.
        Resmi Tablo A'da (UN, siniflandirma kodu, paketleme grubu) uclusu
        AYNI olup yalnizca ozel hukum (6. sutun) ile ayrilan gercekten farkli
        satirlar var (orn. UN1133 F1 PG II: 640C ve 640D varyantlari). Bu
        uclu daha once yanlislikla "birincil anahtar" sayilip byle satirlar
        birbirinin uzerine yaziliyordu (2939 gecerli satirdan 66'si kayboluyor,
        2873 kaliyordu). Tekrar calistirma idempotent DEGILDIR; temiz yukleme
        icin once kimyasal tablosunu bosaltin (Ayarlar sayfasinda uyari var).
        """
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]

        # İDEMPOTENSLİK (66-satır düzeltmesinin tamamlayıcısı): eski UNIQUE
        # kısıtı yanlış anahtar yüzünden 66 geçerli satırı yutuyordu ama yan
        # etki olarak yeniden-içe-aktarmayı da tekilleştiriyordu. Kısıt
        # kalktığı için tekilleştirme artık TAM SATIR İMZASI ile yapılır:
        # 640C/640D gibi yalnız özel hükümle ayrışan varyantlar ayrı imza
        # üretir (korunur), birebir aynı satırın ikinci kez eklenmesi ise
        # atlanır. Böylece Tablo A'yı tekrar yüklemek güvenlidir.
        mevcut_imzalar = {
            (r["un_number"], r["classification_code"], r["class_code"],
             r["packing_group"], r["special_provisions"],
             r["limited_quantity"], r["excepted_quantity"],
             r["tunnel_code"], r["transport_category"],
             r["proper_shipping_name_tr"])
            for r in self.execute(
                """SELECT un_number, classification_code, class_code,
                          packing_group, special_provisions, limited_quantity,
                          excepted_quantity, tunnel_code, transport_category,
                          proper_shipping_name_tr FROM chemicals""")
        }

        imported = 0
        for row in ws.iter_rows(min_row=5, values_only=True):
            un = self._xl_un(row[0] if len(row) > 0 else None)
            if not un:
                continue
            name = self._xl_clean(row[24] if len(row) > 24 else None) \
                or self._xl_clean(row[1] if len(row) > 1 else None)
            if not name:
                continue
            cat, tunnel = self._xl_category_tunnel(
                row[17] if len(row) > 17 else "")
            lq_text = self._xl_clean(row[7] if len(row) > 7 else None)
            eq_code = self._xl_clean(row[8] if len(row) > 8 else None).upper()
            special = " | ".join(filter(None, [
                self._xl_clean(row[6] if len(row) > 6 else None),
                self._xl_clean(row[21] if len(row) > 21 else None),
            ]))
            c = Chemical(
                un_number=un,
                proper_shipping_name_tr=name,
                class_code=self._xl_clean(row[2] if len(row) > 2 else None),
                classification_code=self._xl_clean(row[3] if len(row) > 3 else None),
                packing_group=self._xl_clean(row[4] if len(row) > 4 else None),
                tunnel_code=tunnel,
                transport_category=cat,
                special_provisions=special,
                limited_quantity=lq_text,
                excepted_quantity=eq_code,
                lq_allowed=_adr_engine().parse_lq_limit(lq_text)[0] > 0,
                eq_allowed=_adr_engine().eq_limits(eq_code)[0] > 0,
                hazard_labels=self._xl_clean(row[5] if len(row) > 5 else None),
            )
            imza = (c.un_number, c.classification_code, c.class_code,
                    c.packing_group, c.special_provisions,
                    c.limited_quantity, c.excepted_quantity,
                    c.tunnel_code, c.transport_category,
                    c.proper_shipping_name_tr)
            if imza in mevcut_imzalar:
                continue
            mevcut_imzalar.add(imza)
            self.add_chemical(c)
            imported += 1
        return imported

    def import_company_inventory_excel(self, xlsx_path: str) -> int:
        """Firmaya ozel kimyasal envanter Excel'ini aktarir.

        Baslik satiri 'UN NUMARASI' hucresi aranarak otomatik bulunur, boylece
        ust bilgi satirlari/birlestirilmis hucreler sorun cikarmaz. Envanterde
        EQ kodu yoksa, ayni UN daha once Tablo A'dan yuklendiyse EQ oradan
        tamamlanir (uc veri kaynaginin birlesme kurali)."""
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)

        imported = 0
        seen: set = set()
        for ws in wb.worksheets:
            header_row, cols = None, {}
            for r_idx, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
                cells = [self._xl_clean(c).upper() for c in row]
                if any("UN NUMARASI" in c for c in cells):
                    header_row = r_idx
                    for i, c in enumerate(cells):
                        if "UN NUMARASI" in c: cols["un"] = i
                        elif c.startswith("SINIFI"): cols["sinif"] = i
                        elif "PAKETLEME GRUBU" in c: cols["pg"] = i
                        elif "UYGUN SEVKİYAT" in c or "UYGUN SEVKIYAT" in c: cols["ad"] = i
                        elif "SINIRLI MİKTAR" in c or "SINIRLI MIKTAR" in c: cols["lq"] = i
                        elif "ÖZEL HÜKÜM" in c or "OZEL HUKUM" in c: cols["ozel"] = i
                        elif "TAŞIMA KATEGORİSİ" in c or "TASIMA KATEGORISI" in c: cols["kat"] = i
                        elif "KİMYASAL ADI" in c or "KIMYASAL ADI" in c: cols["ticari"] = i
                    break
            if header_row is None or "un" not in cols:
                continue

            def get(row, key):
                i = cols.get(key, -1)
                return self._xl_clean(row[i]) if 0 <= i < len(row) else ""

            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                un = self._xl_un(get(row, "un"))
                if not un:
                    continue  # tehlikesiz / UN'siz satirlar atlanir
                name = get(row, "ad") or get(row, "ticari")
                if not name:
                    continue
                if un in seen:
                    continue
                seen.add(un)

                cat, tunnel = self._xl_category_tunnel(get(row, "kat"))
                lq_text = get(row, "lq")
                ticari = get(row, "ticari")
                notes = f"Ticari ad: {ticari}" if (ticari and ticari != name) else ""

                # EQ envanterde yok: Tablo A'dan (varsa) tamamla
                eq_code = ""
                existing = self.execute_one(
                    "SELECT excepted_quantity FROM chemicals WHERE un_number=?", (un,))
                if existing and existing["excepted_quantity"]:
                    eq_code = existing["excepted_quantity"]

                c = Chemical(
                    un_number=un,
                    proper_shipping_name_tr=name,
                    class_code=get(row, "sinif"),
                    packing_group=get(row, "pg"),
                    tunnel_code=tunnel,
                    transport_category=cat,
                    special_provisions=" | ".join(filter(None, [get(row, "ozel"), notes])),
                    limited_quantity=lq_text,
                    excepted_quantity=eq_code,
                    lq_allowed=_adr_engine().parse_lq_limit(lq_text)[0] > 0,
                    eq_allowed=_adr_engine().eq_limits(eq_code)[0] > 0,
                )
                self._upsert_chemical(c)
                ticari = get(row, "ticari") or ""
                affected = self.execute_update(
                    """INSERT OR IGNORE INTO company_products
                       (trade_name, un_number, classification_code, packing_group,
                        limited_quantity, excepted_quantity, transport_category,
                        tunnel_code, source_file)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (ticari or name, un, c.classification_code, c.packing_group,
                     c.limited_quantity, c.excepted_quantity, c.transport_category,
                     c.tunnel_code, os.path.basename(xlsx_path)))
                if affected:
                    imported += 1
        return imported

    def search_company_products(self, query: str) -> list:
        q = query.strip().upper()
        if not q:
            return []
        rows = self.execute(
            """SELECT cp.*, c.class_code, c.hazard_labels
               FROM company_products cp
               LEFT JOIN chemicals c ON c.un_number=cp.un_number
                   AND c.classification_code=cp.classification_code
               WHERE UPPER(cp.trade_name) LIKE ?
                  OR cp.un_number LIKE ?
               ORDER BY cp.trade_name LIMIT 100""",
            (f"%{q}%", f"%{q}%"))
        return [dict(r) for r in rows] if rows else []


    def import_json_data(self, json_path: str):
        if not os.path.exists(json_path):
            return 0

        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        imported = 0

        for key in ["sender_companies", "receiver_companies"]:
            for comp in data.get(key, []):
                company = Company(
                    type="sender" if "sender" in key else "receiver",
                    name=comp.get("name", ""),
                    address=comp.get("address", ""),
                    city=comp.get("city", ""),
                    phone=comp.get("phone", ""),
                    email=comp.get("email", ""),
                    contact_person=comp.get("contact_person", "")
                )
                self.add_company(company)
                imported += 1

        for drv in data.get("drivers", []):
            driver = Driver(
                full_name=drv.get("name", ""),
                src5_no=drv.get("src5_no", ""),
                phone=drv.get("phone", "")
            )
            self.add_driver(driver)
            imported += 1

        for pkg in data.get("packaging_types", []):
            self.execute_insert(
                "INSERT OR IGNORE INTO packaging_types (code, name_tr, name_en) VALUES (?, ?, ?)",
                (pkg.get("code", ""), pkg.get("name_tr", ""), pkg.get("name_en", ""))
            )
            imported += 1

        return imported

    def export_to_json(self, output_path: str):
        data = {
            "export_date": datetime.now().isoformat(),
            "version": APP_VERSION,
            "companies": [],
            "drivers": [],
            "vehicles": [],
            "shipments": []
        }

        for row in self.execute("SELECT * FROM companies", ()):
            data["companies"].append(dict(row))
        for row in self.execute("SELECT * FROM drivers", ()):
            data["drivers"].append(dict(row))
        for row in self.execute("SELECT * FROM vehicles", ()):
            data["vehicles"].append(dict(row))
        for row in self.execute("SELECT * FROM shipments", ()):
            data["shipments"].append(dict(row))

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        return output_path


# =============================================================================
# ADR ENGINE - MEVZUAT MOTORU
# =============================================================================


# =============================================================================
# ADR 1.10.3 — EMNİYET PLANI HESAPLAMA MOTORU
# =============================================================================

