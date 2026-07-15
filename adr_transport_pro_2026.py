#!/usr/bin/env python3
"""
ADR TRANSPORT PRO 2026
Profesyonel Tehlikeli Madde Tasima Evrak Yonetim Sistemi
v4.1 - SQLite + PyQt6 + Tam Profesyonel Ozellikler

Degisiklikler v4.1:
  - ADR Veritabani sayfasi tam CRUD arayuzu eklendi
  - Yazdir onizleme ekrani profesyonel HTML tabanli yapiya gecirildi

Gereksinimler:
    pip install PyQt6 reportlab openpyxl rapidfuzz

Mimari:
    - MVC Pattern
    - Service Layer
    - SQLite Veritabani
    - PyQt6 Arayuz
    - ADR Mevzuat Motoru
    - Otomatik Yedekleme
"""
# ── Güvenlik & Lisans ──────────────────────────────────────────────────────────
import uuid
import socket
import hmac
import struct
import base64
import platform
import subprocess
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
import sys
import os
import json
import sqlite3
import logging
import logging.handlers
import qrcode
import io

import base64
from io import BytesIO


def _turkce_hata_metni(exc: Exception) -> str:
    """Kullanıcıya gösterilecek hata metnini üretir.

    Python/kütüphane istisnalarının ham metni (str(exc)) genellikle
    İngilizcedir (örn. "list index out of range", "[Errno 2] No such
    file or directory"). Kullanıcıya HER ZAMAN Türkçe bir açıklama
    gösterilmesi için yaygın istisna türleri burada Türkçeye çevrilir;
    tanınmayan durumlarda teknik ayrıntı yerine genel bir Türkçe mesaj
    ve istisna sınıfı adı gösterilir (ham İngilizce metin asla
    kullanıcı arayüzüne sızmaz — ayrıntı yalnızca log dosyasına yazılır).
    """
    if isinstance(exc, FileNotFoundError):
        return f"Dosya bulunamadı: {getattr(exc, 'filename', '') or ''}".strip()
    if isinstance(exc, PermissionError):
        return "Dosyaya erişim izni yok (başka bir programda açık olabilir)."
    if isinstance(exc, IsADirectoryError):
        return "Seçilen konum bir dosya değil, klasör."
    if isinstance(exc, (OSError, IOError)):
        return "Dosya okuma/yazma sırasında bir sistem hatası oluştu."
    if isinstance(exc, ValueError):
        return "Girilen değerlerden biri geçersiz veya beklenen biçimde değil."
    if isinstance(exc, KeyError):
        return "Beklenen bir alan bulunamadı (veri eksik olabilir)."
    if isinstance(exc, IndexError):
        return "Veri listesinde beklenen bir öğe bulunamadı."
    if isinstance(exc, ImportError):
        return "Gerekli bir kütüphane kurulu değil (kurulum sayfasına bakın)."
    try:
        import sqlite3 as _sqlite3
        if isinstance(exc, _sqlite3.Error):
            return "Veritabanı işlemi sırasında bir hata oluştu."
    except ImportError:
        pass
    return f"Beklenmeyen bir hata oluştu ({type(exc).__name__}). Ayrıntı günlük kaydına yazıldı."


def generate_qr_html(data: str, size: int = 60) -> str:
    """
    Verilen veri için base64 gömülü QR kod <img> etiketi döndürür.
    vCard formatındaki verilerle profesyonel kartvizit QR'ı oluşturur.
    """
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=5,
        border=1,
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    img_str = base64.b64encode(buffered.getvalue()).decode()
    return (
        f'<img src="data:image/png;base64,{img_str}" '
        f'width="{size}" height="{size}" '
        f'style="width:{size}px;height:{size}px;display:block;">'
    )

def _setup_logging():
    from pathlib import Path as _Path
    log_dir = _Path.home() / ".adr_transport_pro" / "logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    fmt = logging.Formatter(
        "%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S")
    fh = logging.handlers.RotatingFileHandler(
        log_dir / "adr_error.log",
        maxBytes=2*1024*1024, backupCount=3, encoding="utf-8")
    fh.setLevel(logging.WARNING)
    fh.setFormatter(fmt)
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    root = logging.getLogger()
    root.setLevel(logging.DEBUG)
    if not root.handlers:
        root.addHandler(fh)
        root.addHandler(ch)
    return logging.getLogger("ADRTransport")

logger = _setup_logging()

import re
import copy
import shutil
import hashlib
import tempfile
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass, asdict
from enum import Enum
from pathlib import Path


# =============================================================================
# PYQT6 IMPORTS
# =============================================================================
try:
    from PyQt6.QtWidgets import (
        QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
        QGridLayout, QSplitter, QStackedWidget, QLabel, QLineEdit,
        QComboBox, QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
        QTreeWidget, QTreeWidgetItem, QTabWidget, QTableView, QDialog, QDialogButtonBox,
        QMessageBox, QFileDialog, QProgressBar, QStatusBar, QToolBar,
        QMenuBar, QMenu, QFrame, QScrollArea, QGroupBox, QFormLayout, QSpinBox,
        QDoubleSpinBox, QCheckBox, QRadioButton, QButtonGroup, QTextEdit,
        QPlainTextEdit, QDateEdit, QTimeEdit, QDateTimeEdit, QSlider,
        QDockWidget, QListWidget, QListWidgetItem, QInputDialog,
        QColorDialog, QFontDialog, QStyledItemDelegate,
        QAbstractItemView, QSizePolicy, QSpacerItem, QWizard, QWizardPage,
        QKeySequenceEdit, QSystemTrayIcon, QSplashScreen,
        QGraphicsView, QGraphicsScene, QGraphicsRectItem, QGraphicsTextItem,
        QGraphicsPixmapItem, QCompleter
    )
    from PyQt6.QtCore import (
        Qt, QTimer, QThread, pyqtSignal, QSettings, QDate, QTime,
        QDateTime, QSize, QPoint, QRect, QUrl, QProcess, QMutex, QWaitCondition
    )
    from PyQt6.QtGui import (
        QFont, QColor, QIcon, QPixmap, QPainter, QPen, QBrush, QLinearGradient,
        QPalette, QKeySequence, QCursor, QFontDatabase,
        QImage, QTransform, QTextDocument, QTextCursor,
        QAction, QShortcut
    )
    from PyQt6.QtPrintSupport import QPrintPreviewDialog, QPrintDialog, QPrinter
    from PyQt6.QtGui import QPageLayout, QPageSize
    from PyQt6.QtCore import QMarginsF

    print("PyQt6 basariyla yuklendi!")

except ImportError as e:
    print(f"HATA: PyQt6 veya alt moduller bulunamadi: {e}")
    print("Kurulum: pip install PyQt6")
    sys.exit(1)

# =============================================================================
# DIGER MODULLER
# =============================================================================
try:
    from rapidfuzz import fuzz, process
    RAPIDFUZZ_AVAILABLE = True
except ImportError:
    RAPIDFUZZ_AVAILABLE = False
    print("UYARI: rapidfuzz bulunamadi. Fuzzy arama devre disi. Kurulum: pip install rapidfuzz")

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.lib.colors import HexColor, black, white, red, green, yellow, orange
    from reportlab.lib.units import mm, cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("UYARI: reportlab bulunamadi. PDF ciktisi devre disi.")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("UYARI: openpyxl bulunamadi. Excel ciktisi devre disi.")

try:
    from adr_csv_importer import import_csv_to_db
    CSV_IMPORTER_AVAILABLE = True
except ImportError:
    CSV_IMPORTER_AVAILABLE = False
    print("BILGI: adr_csv_importer.py bulunamadi. CSV aktarimi devre disi.")



# =============================================================================
# KONFIGURASYON VE SABITLER
# =============================================================================

APP_NAME = "ADR Transport Pro 2026"
APP_VERSION = "4.1.0"
APP_ORGANIZATION = "ADRSoft"

ADR_VERSION = "2025"
MAX_1136_POINTS = 1000

TUNNEL_HIERARCHY = {
    "A": 1,
    "B": 2,
    "C": 3,
    "D": 4,
    "E": 5,
}

TC_POINTS = {"0": 0, "1": 50, "2": 3, "3": 1, "4": 0}

CLASS_COLORS = {
    "1":   ("#FF0000", "#FFFFFF"),
    "2.1": ("#FF0000", "#FFFFFF"),
    "2.2": ("#00AA00", "#FFFFFF"),
    "2.3": ("#FFFFFF", "#000000"),
    "3":   ("#FF0000", "#FFFFFF"),
    "4.1": ("#FF0000", "#FFFFFF"),
    "4.2": ("#FF0000", "#FFFFFF"),
    "4.3": ("#0000FF", "#FFFFFF"),
    "5.1": ("#FFCC00", "#000000"),
    "5.2": ("#FFCC00", "#000000"),
    "6.1": ("#FFFFFF", "#000000"),
    "6.2": ("#FFFFFF", "#000000"),
    "7":   ("#FF00FF", "#FFFFFF"),
    "8":   ("#000000", "#FFFFFF"),
    "9":   ("#CCCCCC", "#000000"),
    "LQ":  ("#00FF00", "#000000"),
    "EQ":  ("#00FFFF", "#000000"),
}

INCOMPATIBILITY_MATRIX = {
    "Asitler":              ["Bazlar", "Siyanurler", "Yanici Maddeler", "Yukseltgenler"],
    "Bazlar":               ["Asitler", "Siyanurler", "Yanici Maddeler"],
    "Yukseltgenler":        ["Yanici Maddeler", "Asitler", "Organik Peroksitler"],
    "Yanici Maddeler":      ["Yukseltgenler", "Asitler", "Bazlar", "Organik Peroksitler"],
    "Organik Peroksitler":  ["Yanici Maddeler", "Yukseltgenler", "Asitler"],
    "Siyanurler":           ["Asitler", "Bazlar"],
    "2.3 Sinifi Zehirli Gazlar": ["Asitler", "Bazlar", "Yanici Maddeler"],
    "4.3 Sinifi Su ile Tepki Veren": ["Asitler", "Su ile Tepki Veren"],
    "Su ile Tepki Veren":   ["4.3 Sinifi Su ile Tepki Veren", "Asitler"],
}
# ── Güvenlik Sabitleri ────────────────────────────────────────────────────────
# BU DEĞERLERİ SADECE SEN BİLİRSİN — ASLA PAYLAŞMA
_LICENSE_SALT   = b"ADR_PRO_2026_GIZLI_SALT_DEGISTIR"   # 32+ karakter olsun
_HMAC_SECRET    = b"ADR_HMAC_IMZA_GIZLI_ANAHTARI_2026"
MAX_FAILED_LOGINS = 5          # Bu kadar hatalı girişten sonra hesap kilitlenir
SESSION_TIMEOUT_MIN = 480      # 8 saat sonra oturum kapanır
# =============================================================================
# ENUMLAR
# =============================================================================

class DocumentStatus(Enum):
    DRAFT     = "Taslak"
    VALIDATED = "Onaylandi"
    PRINTED   = "Yazdirildi"
    ARCHIVED  = "Arsivlendi"
    CANCELLED = "Iptal Edildi"

class WarningLevel(Enum):
    INFO     = "Bilgi"
    WARNING  = "Uyari"
    ERROR    = "Hata"
    CRITICAL = "Kritik"

class ExemptionType(Enum):
    """ADR Muafiyet Türleri - ADR Book 1.1.3"""
    NONE                    = "Yok"
    # 1.1.3.1 Genel (Tam) Muafiyetler
    PERSONAL_USE            = "ADR 1.1.3.1 - Kisisel Kullanim"
    MACHINERY_EQUIPMENT     = "ADR 1.1.3.1 - Makine/Ekipman Icindeki Maddeler"
    EMERGENCY_TRANSPORT     = "ADR 1.1.3.1 - Acil Durum Tasimasi"
    AGRICULTURE_FORESTRY    = "ADR 1.1.3.1 - Tarim ve Ormancilik"
    # 1.1.3.2 Gazlarin Tasinmasi
    GAS_TRANSPORT           = "ADR 1.1.3.2 - Gazlarin Tasinmasi"
    # 1.1.3.3 Sivi Yakitlar
    LIQUID_FUEL             = "ADR 1.1.3.3 - Sivi Yakitlar"
    # 1.1.3.6 Miktar Muafiyeti (1000 Puan Kurali)
    ADR_1_1_3_6             = "ADR 1.1.3.6 - Miktar Muafiyeti (1000 Puan)"
    # 3.4 Sinirli Miktar (LQ)
    LQ                      = "ADR 3.4 - Sinirli Miktar (LQ)"
    # 3.5 Istisnai Miktar (EQ)
    EQ                      = "ADR 3.5 - Istisnai Miktar (EQ)"
    # Ozel Hukumler (3.3)
    SPECIAL_PROVISION       = "ADR 3.3 - Ozel Hukumler"
    # Bos Ambalaj
    EMPTY_PACKAGING         = "Bos Ambalaj"
    # Numune
    SAMPLE                  = "Numune"
    # Kismi Muafiyet
    PARTIAL_EXEMPTION       = "Kismi Muafiyet"

# =============================================================================
# VERI MODELLERI (DATACLASSES)
# =============================================================================

@dataclass
class Company:
    id: Optional[int] = None
    type: str = "sender"
    name: str = ""
    tax_number: str = ""
    tax_office: str = ""
    mersis_no: str = ""
    address: str = ""
    city: str = ""
    district: str = ""
    phone: str = ""
    email: str = ""
    contact_person: str = ""
    is_favorite: bool = False
    logo_path: str = ""  # PDF antet arka plani icin firma logosu
    created_at: str = ""
    updated_at: str = ""

@dataclass
class Driver:
    id: Optional[int] = None
    full_name: str = ""
    tc_no: str = ""
    phone: str = ""
    adr_certificate_no: str = ""
    adr_certificate_expiry: str = ""
    src5_no: str = ""
    src5_expiry: str = ""
    license_class: str = ""
    license_expiry: str = ""
    is_active: bool = True
    created_at: str = ""
    updated_at: str = ""

@dataclass
class Vehicle:
    id: Optional[int] = None
    plate: str = ""
    trailer_plate: str = ""
    adr_compliance_cert_no: str = ""
    adr_compliance_expiry: str = ""
    inspection_date: str = ""
    inspection_expiry: str = ""
    tank_info: str = ""
    vehicle_type: str = ""
    max_capacity: float = 0.0
    is_active: bool = True
    created_at: str = ""
    updated_at: str = ""

@dataclass
class Chemical:
    id: Optional[int] = None
    un_number: str = ""
    classification_code: str = ""  # Tablo A 3b: 1.4G, F1, 5T vb. (varyant ayirici)
    proper_shipping_name_tr: str = ""
    proper_shipping_name_en: str = ""
    class_code: str = ""
    packing_group: str = ""
    tunnel_code: str = ""
    transport_category: str = ""
    segregation_group: str = ""
    special_provisions: str = ""
    lq_allowed: bool = False
    eq_allowed: bool = False
    limited_quantity: str = ""   # Tablo A 7a: "1 L", "5 kg", "0"
    excepted_quantity: str = ""  # Tablo A 7b: E0..E5
    hazard_labels: str = ""

@dataclass
class ShipmentItem:
    id: Optional[int] = None
    shipment_id: Optional[int] = None
    chemical_id: int = 0
    un_number: str = ""
    proper_name: str = ""
    class_code: str = ""
    packing_group: str = ""
    packaging_type: str = ""
    packaging_count: int = 0
    net_quantity: float = 0.0
    gross_quantity: float = 0.0
    unit: str = "kg"
    is_lq: bool = False
    is_eq: bool = False
    lq_max_per_package: float = 0.0
    eq_max_per_package: float = 0.0
    notes: str = ""
    tunnel_code: str = ""
    segregation_group: str = ""
    classification_code: str = ""
    transport_category: str = ""   # [DÜZELTİLDİ] "2" varsayılanı kaldırıldı
    special_provisions: str = ""

@dataclass
class Shipment:
    id: Optional[int] = None
    document_no: str = ""
    document_date: str = ""
    status: str = "Taslak"
    sender_id: int = 0
    receiver_id: int = 0
    carrier_id: int = 0
    driver_id: int = 0
    vehicle_id: int = 0
    total_points: float = 0.0
    orange_plate_required: bool = False
    written_instructions_required: bool = False
    driver_adr_required: bool = False
    tunnel_restriction_code: str = ""
    exemption_type: str = "Yok"
    is_validated: bool = False
    validation_errors: str = ""
    notes: str = ""
    created_at: str = ""
    updated_at: str = ""
    created_by: str = ""

@dataclass
class ValidationResult:
    is_valid: bool = False
    errors: List[Tuple[WarningLevel, str]] = None
    warnings: List[Tuple[WarningLevel, str]] = None
    info: List[Tuple[WarningLevel, str]] = None

    def __post_init__(self):
        if self.errors   is None: self.errors   = []
        if self.warnings is None: self.warnings = []
        if self.info     is None: self.info     = []

@dataclass
class ADRReport:
    total_points: float = 0.0
    orange_plate_required: bool = False
    written_instructions_required: bool = False
    driver_adr_required: bool = False
    tunnel_code: str = "E"
    exemption_type: str = "Yok"
    segregation_warnings: List[str] = None
    compatibility_errors: List[str] = None
    lq_status: str = ""
    eq_status: str = ""
    # ADR 1.10.3 — Emniyet Planı
    security_plan_required: bool = False
    security_plan_reasons: List[str] = None   # gerektiren kalemler
    security_plan_exempt: bool = False        # 1.10.4 muafiyeti var
    errors: List[Tuple[WarningLevel, str]] = None
    warnings: List[Tuple[WarningLevel, str]] = None
    info: List[Tuple[WarningLevel, str]] = None

    def __post_init__(self):
        if self.segregation_warnings  is None: self.segregation_warnings  = []
        if self.compatibility_errors  is None: self.compatibility_errors  = []
        if self.security_plan_reasons is None: self.security_plan_reasons = []
        if self.errors   is None: self.errors   = []
        if self.warnings is None: self.warnings = []
        if self.info     is None: self.info     = []

# =============================================================================
# GÜVENLİK YARDIMCI FONKSİYONLARI
# =============================================================================


def _get_machine_fingerprint() -> str:
    """Çok katmanlı donanım parmak izi: CPU+Anakart+MAC+Hostname."""
    import subprocess as _sp
    parts = []
    if platform.system() == "Windows":
        for cmd in ["wmic cpu get processorid", "wmic baseboard get serialnumber"]:
            try:
                out = _sp.check_output(cmd, shell=True,
                    stderr=_sp.DEVNULL, timeout=3).decode(errors="ignore")
                vals = [l.strip() for l in out.splitlines()
                        if l.strip() and "Id" not in l and "Serial" not in l and "Number" not in l]
                if vals:
                    parts.append(vals[0])
            except Exception:
                pass
    mac = hex(uuid.getnode()).replace("0x", "").upper()
    parts.append(mac)
    parts.append(socket.gethostname().upper())
    parts.append(platform.system() + platform.release())
    raw = "|".join(filter(None, parts)).encode("utf-8")
    digest = hmac.new(_HMAC_SECRET, raw, "sha256").hexdigest()
    return digest[:32].upper()
 
 
def _hash_password(password: str, salt: bytes = None) -> tuple:
    """PBKDF2-HMAC-SHA256 ile şifre hashler. (salt, hash) döndürür."""
    if salt is None:
        salt = os.urandom(32)
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt + _LICENSE_SALT,
        iterations=390_000,   # OWASP 2023 minimum önerisi
    )
    key = base64.urlsafe_b64encode(kdf.derive(password.encode("utf-8")))
    return salt, key.decode("utf-8")
 
 
def _verify_password(password: str, stored_salt: bytes, stored_hash: str) -> bool:
    """Şifre doğrular. Sabit zamanlı karşılaştırma (timing attack önleme)."""
    _, computed = _hash_password(password, stored_salt)
    return hmac.compare_digest(computed, stored_hash)
 
 
def _get_local_ip() -> str:
    """Çalışan makinenin yerel IP adresini döndürür."""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"
 
 
# =============================================================================
# LİSANS YÖNETİCİSİ
# =============================================================================
 
class LicenseManager:
    """
    Lisans doğrulama sistemi.
 
    Lisans dosyası şifreli JSON'dur:
      - machine_fp   : Makine parmak izi (başka PC'de çalışmaz)
      - max_users    : İzin verilen maksimum kullanıcı sayısı
      - expiry       : Lisans son tarihi (YYYY-MM-DD)
      - customer     : Müşteri adı
      - serial       : Seri numarası
 
    Lisans oluşturmak için: LicenseManager.generate_license(...)
    """
 
    LICENSE_FILE = Path.home() / ".adr_transport_pro" / "license.adrlic"
    _fernet: "Fernet | None" = None
 
    @classmethod
    def _get_fernet(cls) -> "Fernet":
        if cls._fernet is None:
            fp  = _get_machine_fingerprint().encode()
            kdf = PBKDF2HMAC(
                algorithm=hashes.SHA256(), length=32,
                salt=_LICENSE_SALT, iterations=100_000,
            )
            key = base64.urlsafe_b64encode(kdf.derive(fp))
            cls._fernet = Fernet(key)
        return cls._fernet
 
    @classmethod
    def generate_license(cls, customer: str, max_users: int,
                         expiry: str, machine_fp: str = None) -> str:
        """
        ─── SADECE GELİŞTİRİCİ KULLANIR ───
        Yeni bir lisans dosyası üretir.
 
        Kullanım (terminalden):
            python -c "
            from adr_security_module import *
            # Müşterinin makine parmak izini öğren:
            print(_get_machine_fingerprint())
            # Sonra lisans üret:
            lic = LicenseManager.generate_license(
                customer='Ahmet Nakliyat Ltd.',
                max_users=3,
                expiry='2027-12-31',
                machine_fp='BURAYA_MUSTERININ_PARMAK_IZI'
            )
            print(lic)
            "
 
        machine_fp: None verilirse MEVCUT makinenin parmak izi kullanılır.
        """
        if machine_fp is None:
            machine_fp = _get_machine_fingerprint()
 
        serial = base64.urlsafe_b64encode(os.urandom(12)).decode("utf-8")
        payload = {
            "machine_fp": machine_fp,
            "max_users":  max_users,
            "expiry":     expiry,
            "customer":   customer,
            "serial":     serial,
            "issued_at":  datetime.now().strftime("%Y-%m-%d"),
            "product":    "ADR Transport Pro 2026",
        }
        raw      = json.dumps(payload).encode("utf-8")
        # Lisans şifreleme anahtarı müşteri parmak izine bağlı
        fp_bytes = machine_fp.encode()
        kdf = PBKDF2HMAC(
            algorithm=hashes.SHA256(), length=32,
            salt=_LICENSE_SALT, iterations=100_000,
        )
        key     = base64.urlsafe_b64encode(kdf.derive(fp_bytes))
        fernet  = Fernet(key)
        token   = fernet.encrypt(raw).decode("utf-8")
        return token
 
    @classmethod
    def validate(cls) -> dict:
        """
        Lisansı doğrular.
        Döndürür: {
          "valid": bool,
          "customer": str,
          "max_users": int,
          "expiry": str,
          "days_left": int,
          "serial": str,
          "error": str  (sadece geçersizse)
        }
        """
        result = {
            "valid": False, "customer": "", "max_users": 1,
            "expiry": "", "days_left": 0, "serial": "", "error": ""
        }
 
        if not cls.LICENSE_FILE.exists():
            result["error"] = "Lisans dosyası bulunamadı."
            return result
 
        try:
            raw_token = cls.LICENSE_FILE.read_text(encoding="utf-8").strip()
            fernet    = cls._get_fernet()
            decrypted = fernet.decrypt(raw_token.encode("utf-8"))
            payload   = json.loads(decrypted)
        except Exception:
            result["error"] = "Lisans dosyası geçersiz veya bozuk."
            return result
 
        # Makine parmak izi kontrolü
        current_fp = _get_machine_fingerprint()
        if payload.get("machine_fp") != current_fp:
            result["error"] = "Bu lisans bu bilgisayara ait değil."
            return result
 
        # Son tarih kontrolü
        try:
            expiry_dt = datetime.strptime(payload["expiry"], "%Y-%m-%d")
            days_left = (expiry_dt - datetime.now()).days
            if days_left < 0:
                result["error"] = f"Lisans süresi dolmuş ({payload['expiry']})."
                return result
        except Exception:
            result["error"] = "Lisans tarihi okunamadı."
            return result
 
        result.update({
            "valid":     True,
            "customer":  payload.get("customer", ""),
            "max_users": payload.get("max_users", 1),
            "expiry":    payload.get("expiry", ""),
            "days_left": days_left,
            "serial":    payload.get("serial", ""),
        })
        return result
 
    @classmethod
    def install_license(cls, token: str) -> bool:
        """Lisans dosyasını yükler. Geçerli değilse False döner."""
        cls.LICENSE_FILE.parent.mkdir(parents=True, exist_ok=True)
        cls.LICENSE_FILE.write_text(token.strip(), encoding="utf-8")
        result = cls.validate()
        if not result["valid"]:
            cls.LICENSE_FILE.unlink(missing_ok=True)
            return False
        return True
 
 
# =============================================================================
# KULLANICI YÖNETİMİ (SecurityManager)
# =============================================================================
 
class UserRole:
    ADMIN    = "admin"      # Her şey: kullanıcı ekle/sil, tüm veriler
    OPERATOR = "operator"   # Evrak oluştur, kaydet, PDF al
    VIEWER   = "viewer"     # Sadece görüntüle, yazdır
 
 
class SecurityManager:
    """
    Kullanıcı kimlik doğrulama, oturum yönetimi, IP/erişim kaydı.
 
    Tablolar: users, user_sessions, security_log
    Ana programa DatabaseManager.__init__'e eklenir.
    """
 
    def __init__(self, db_conn: sqlite3.Connection):
        self.conn = db_conn
        self._current_user: dict | None = None
        self._session_id:   str | None  = None
        self._session_start: datetime | None = None
        self._setup_tables()
 
    def _setup_tables(self):
        cur = self.conn.cursor()
        cur.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                username      TEXT NOT NULL UNIQUE,
                password_salt BLOB NOT NULL,
                password_hash TEXT NOT NULL,
                role          TEXT NOT NULL DEFAULT 'operator',
                full_name     TEXT DEFAULT '',
                email         TEXT DEFAULT '',
                is_active     INTEGER DEFAULT 1,
                failed_logins         INTEGER DEFAULT 0,
                locked_until          TEXT DEFAULT NULL,
                must_change_password  INTEGER DEFAULT 0,
                last_login    TEXT DEFAULT NULL,
                last_ip       TEXT DEFAULT NULL,
                created_at    TEXT DEFAULT (datetime('now')),
                created_by    TEXT DEFAULT 'system'
            );
 
            CREATE TABLE IF NOT EXISTS user_sessions (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id  TEXT NOT NULL UNIQUE,
                user_id     INTEGER NOT NULL,
                username    TEXT NOT NULL,
                ip_address  TEXT NOT NULL,
                machine_fp  TEXT NOT NULL,
                login_at    TEXT NOT NULL,
                last_active TEXT NOT NULL,
                logout_at   TEXT DEFAULT NULL,
                is_active   INTEGER DEFAULT 1,
                FOREIGN KEY (user_id) REFERENCES users(id)
            );
 
            CREATE TABLE IF NOT EXISTS security_log (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                event_type TEXT NOT NULL,
                username   TEXT DEFAULT '',
                ip_address TEXT DEFAULT '',
                machine_fp TEXT DEFAULT '',
                detail     TEXT DEFAULT '',
                ts         TEXT DEFAULT (datetime('now','localtime'))
            );
        """)
        self.conn.commit()
        self._ensure_default_admin()
 
    def _ensure_default_admin(self):
        """İlk çalıştırmada varsayılan admin oluşturur."""
        cur = self.conn.cursor()
        try:
            cur.execute("ALTER TABLE users ADD COLUMN must_change_password INTEGER DEFAULT 0")
            self.conn.commit()
        except Exception:
            pass
        count = cur.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        if count == 0:
            salt, hsh = _hash_password("Admin2026!")
            self.conn.execute(
                """INSERT INTO users
                   (username,password_salt,password_hash,role,full_name,
                    created_by,must_change_password)
                   VALUES (?,?,?,?,?,?,?)""",
                ("admin", salt, hsh, UserRole.ADMIN,
                 "Sistem Yöneticisi", "system", 1))
            self.conn.commit()
            self._log("SYSTEM","admin","127.0.0.1","","Varsayılan admin oluşturuldu (şifre değişimi zorunlu)")
 
    @staticmethod
    def validate_password_complexity(password: str) -> tuple:
        import re as _re
        if len(password) < 8:
            return False, "Şifre en az 8 karakter olmalıdır."
        if not _re.search(r"[A-Z]", password):
            return False, "En az bir büyük harf gereklidir."
        if not _re.search(r"[a-z]", password):
            return False, "En az bir küçük harf gereklidir."
        if not _re.search(r"\d", password):
            return False, "En az bir rakam gereklidir."
        return True, ""

    def create_user(self, username: str, password: str, role: str,
                    full_name: str = "", email: str = "",
                    created_by: str = "admin", must_change: bool = False) -> bool:
        """Yeni kullanıcı oluşturur. Başarısızsa False döner."""
        ok, msg = self.validate_password_complexity(password)
        if not ok:
            return False
        salt, hsh = _hash_password(password)
        try:
            self.conn.execute(
                """INSERT INTO users
                   (username,password_salt,password_hash,role,full_name,email,created_by)
                   VALUES (?,?,?,?,?,?,?)""",
                (username, salt, hsh, role, full_name, email, created_by)
            )
            self.conn.commit()
            self._log("USER_CREATED", username, _get_local_ip(), "", f"Rol: {role}")
            return True
        except sqlite3.IntegrityError:
            return False
 
    def login(self, username: str, password: str) -> dict:
        """
        Giriş denemesi. Döndürür:
          {"success": bool, "error": str, "user": dict | None,
           "locked_until": str | None}
        """
        ip = _get_local_ip()
        fp = _get_machine_fingerprint()
 
        cur  = self.conn.cursor()
        row  = cur.execute(
            "SELECT * FROM users WHERE username=?", (username,)
        ).fetchone()
 
        if not row:
            self._log("LOGIN_FAIL", username, ip, fp, "Kullanıcı bulunamadı")
            return {"success": False, "error": "Kullanıcı adı veya şifre hatalı."}
 
        user = dict(row)
 
        # Hesap aktif mi?
        if not user["is_active"]:
            self._log("LOGIN_FAIL", username, ip, fp, "Hesap devre dışı")
            return {"success": False, "error": "Bu hesap devre dışı bırakılmış."}
 
        # Kilitli mi?
        if user["locked_until"]:
            try:
                lock_dt = datetime.strptime(user["locked_until"], "%Y-%m-%d %H:%M:%S")
                if datetime.now() < lock_dt:
                    mins = int((lock_dt - datetime.now()).seconds / 60) + 1
                    self._log("LOGIN_FAIL", username, ip, fp, "Hesap kilitli")
                    return {
                        "success": False,
                        "error": f"Hesap kilitli. {mins} dakika sonra tekrar deneyin.",
                        "locked_until": user["locked_until"]
                    }
            except Exception:
                pass
 
        # Şifre doğrula
        salt = user["password_salt"]
        if isinstance(salt, str):
            salt = salt.encode("utf-8")
 
        if not _verify_password(password, salt, user["password_hash"]):
            failed = user["failed_logins"] + 1
            locked_until = None
            if failed >= MAX_FAILED_LOGINS:
                # 30 dakika kilitle
                lock_dt = datetime.now() + timedelta(minutes=30)
                locked_until = lock_dt.strftime("%Y-%m-%d %H:%M:%S")
                self._log("ACCOUNT_LOCKED", username, ip, fp,
                          f"{failed} hatalı deneme — 30 dk kilitlendi")
 
            self.conn.execute(
                "UPDATE users SET failed_logins=?, locked_until=? WHERE username=?",
                (failed, locked_until, username)
            )
            self.conn.commit()
            self._log("LOGIN_FAIL", username, ip, fp,
                      f"Hatalı şifre ({failed}/{MAX_FAILED_LOGINS})")
            kalan = MAX_FAILED_LOGINS - failed
            if kalan > 0:
                return {"success": False,
                        "error": f"Şifre hatalı. {kalan} deneme hakkınız kaldı."}
            else:
                return {"success": False,
                        "error": "Hesabınız kilitlendi. 30 dakika sonra tekrar deneyin."}
 
        # Başarılı giriş
        session_id = str(uuid.uuid4())
        now_str    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
 
        self.conn.execute(
            """INSERT INTO user_sessions
               (session_id,user_id,username,ip_address,machine_fp,login_at,last_active)
               VALUES (?,?,?,?,?,?,?)""",
            (session_id, user["id"], username, ip, fp, now_str, now_str)
        )
        self.conn.execute(
            """UPDATE users SET failed_logins=0, locked_until=NULL,
               last_login=?, last_ip=? WHERE username=?""",
            (now_str, ip, username)
        )
        self.conn.commit()
 
        self._current_user  = user
        self._session_id    = session_id
        self._session_start = datetime.now()
 
        self._log("LOGIN_OK", username, ip, fp,
                  f"Rol: {user['role']} | Oturum: {session_id[:8]}")
 
        return {"success": True, "error": "", "user": user,
                "session_id": session_id}
 
    def logout(self):
        """Aktif oturumu kapatır."""
        if self._session_id:
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.conn.execute(
                "UPDATE user_sessions SET logout_at=?, is_active=0 WHERE session_id=?",
                (now_str, self._session_id)
            )
            self.conn.commit()
            self._log("LOGOUT", self._current_user["username"] if self._current_user else "",
                      _get_local_ip(), _get_machine_fingerprint(), "")
        self._current_user  = None
        self._session_id    = None
        self._session_start = None
 
    def check_session_timeout(self) -> bool:
        """Oturum süresi dolmuş mu? Dolmuşsa True döner."""
        if not self._session_start:
            return True
        elapsed = (datetime.now() - self._session_start).seconds / 60
        return elapsed > SESSION_TIMEOUT_MIN
 
    def ping_session(self):
        """Son aktif zamanı günceller (her UI etkileşiminde çağrılır)."""
        if self._session_id:
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.conn.execute(
                "UPDATE user_sessions SET last_active=? WHERE session_id=?",
                (now_str, self._session_id)
            )
            self.conn.commit()
            self._session_start = datetime.now()
 
    def active_session_count(self) -> int:
        """Şu an açık oturum sayısı."""
        row = self.conn.execute(
            "SELECT COUNT(*) FROM user_sessions WHERE is_active=1"
        ).fetchone()
        return row[0] if row else 0
 
    def get_all_sessions(self) -> list:
        """Tüm oturumları döndürür (yönetici ekranı için)."""
        rows = self.conn.execute(
            """SELECT s.session_id, s.username, s.ip_address, s.machine_fp,
                      s.login_at, s.last_active, s.is_active,
                      u.role, u.full_name
               FROM user_sessions s
               JOIN users u ON u.username = s.username
               ORDER BY s.login_at DESC LIMIT 200"""
        ).fetchall()
        return [dict(r) for r in rows]
 
    def get_security_log(self, limit: int = 500) -> list:
        """Güvenlik log kayıtları."""
        rows = self.conn.execute(
            "SELECT * FROM security_log ORDER BY ts DESC LIMIT ?", (limit,)
        ).fetchall()
        return [dict(r) for r in rows]
 
    def change_password(self, username: str, old_pass: str, new_pass: str) -> bool:
        """Şifre değiştirme."""
        result = self.login(username, old_pass)
        if not result["success"]:
            return False
        if len(new_pass) < 8:
            return False
        salt, hsh = _hash_password(new_pass)
        self.conn.execute(
            "UPDATE users SET password_salt=?, password_hash=? WHERE username=?",
            (salt, hsh, username)
        )
        self.conn.commit()
        self._log("PASSWORD_CHANGED", username, _get_local_ip(), "", "")
        return True
 
    def get_users(self) -> list:
        rows = self.conn.execute(
            "SELECT id,username,role,full_name,email,is_active,failed_logins,"
            "last_login,last_ip,created_at FROM users ORDER BY username"
        ).fetchall()
        return [dict(r) for r in rows]
 
    @property
    def current_user(self) -> dict | None:
        return self._current_user
 
    @property
    def current_role(self) -> str:
        return (self._current_user or {}).get("role", "")
 
    def can(self, action: str) -> bool:
        """Yetki kontrolü. action: 'delete','edit','manage_users','view' """
        role = self.current_role
        if role == UserRole.ADMIN:
            return True
        if role == UserRole.OPERATOR:
            return action in ("view", "edit", "create", "print")
        if role == UserRole.VIEWER:
            return action in ("view", "print")
        return False
 
    def _log(self, event_type: str, username: str, ip: str, fp: str, detail: str):
        try:
            self.conn.execute(
                """INSERT INTO security_log
                   (event_type,username,ip_address,machine_fp,detail)
                   VALUES (?,?,?,?,?)""",
                (event_type, username, ip, fp, detail)
            )
            self.conn.commit()
        except Exception:
            pass
 
 
# =============================================================================
# GİRİŞ EKRANI (LoginDialog)
# =============================================================================
 
class LoginDialog(QDialog):
    """
    Program açılışında gösterilen güvenlik ekranı.
    Lisans + kullanıcı giriş kontrolü.
    """
 
    def __init__(self, security: SecurityManager, license_info: dict, parent=None):
        super().__init__(parent)
        self.security     = security
        self.license_info = license_info
        self.logged_in    = False
        self.setWindowTitle("ADR Transport Pro 2026 — Giriş")
        self.setFixedSize(440, 560)
        self.setWindowFlags(
            Qt.WindowType.Dialog |
            Qt.WindowType.MSWindowsFixedSizeDialogHint
        )
        self._setup_ui()
 
    def _setup_ui(self):
        self.setStyleSheet("""
            QDialog {
                background-color: #1E1E2E;
                color: #CDD6F4;
                font-family: 'Segoe UI', Arial;
            }
            QLabel { background: transparent; }
            QLabel#title {
                font-size: 14pt; font-weight: bold; color: #89B4FA;
            }
            QLabel#subtitle { font-size: 9pt; color: #A6ADC8; }
            QLabel#lbl_license_ok  { color: #A6E3A1; font-size: 8.5pt; }
            QLabel#lbl_license_err { color: #F38BA8; font-size: 8.5pt; }
            QLineEdit {
                background: #313244; border: 1px solid #45475A;
                border-radius: 6px; padding: 8px 12px;
                color: #CDD6F4; font-size: 10pt;
            }
            QLineEdit:focus { border-color: #89B4FA; }
            QPushButton#btn_login {
                background: #89B4FA; color: #1E1E2E;
                border: none; border-radius: 6px;
                font-size: 11pt; font-weight: bold; padding: 10px;
            }
            QPushButton#btn_login:hover { background: #B4BEFE; }
            QPushButton#btn_login:disabled { background: #45475A; color: #585B70; }
            QPushButton#btn_license {
                background: transparent; color: #A6ADC8;
                border: 1px solid #45475A; border-radius: 4px;
                font-size: 8.5pt; padding: 4px 10px;
            }
            QPushButton#btn_license:hover { border-color: #89B4FA; color: #89B4FA; }
            QLabel#err { color: #F38BA8; font-size: 8.5pt; }
        """)
 
        layout = QVBoxLayout(self)
        layout.setContentsMargins(36, 28, 36, 28)
        layout.setSpacing(14)
 
        # Logo / başlık
        lbl_icon = QLabel("🛡")
        lbl_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_icon.setStyleSheet("font-size: 36pt; background: transparent;")
        layout.addWidget(lbl_icon)
 
        lbl_title = QLabel("ADR Transport Pro 2026")
        lbl_title.setObjectName("title")
        lbl_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(lbl_title)
 
        lbl_sub = QLabel("Tehlikeli Madde Taşıma Yönetim Sistemi")
        lbl_sub.setObjectName("subtitle")
        lbl_sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(lbl_sub)
 
        # Lisans durumu
        lic = self.license_info
        if lic["valid"]:
            lic_text = (f"✓  {lic['customer']}  |  "
                        f"Maks. {lic['max_users']} kullanıcı  |  "
                        f"{lic['days_left']} gün kaldı  ({lic['expiry']})")
            lbl_lic = QLabel(lic_text)
            lbl_lic.setObjectName("lbl_license_ok")
        else:
            lbl_lic = QLabel(f"✗  Lisans Hatası: {lic['error']}")
            lbl_lic.setObjectName("lbl_license_err")
        lbl_lic.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_lic.setWordWrap(True)
        layout.addWidget(lbl_lic)
 
        # IP bilgisi
        ip_lbl = QLabel(f"Bu bilgisayar: {_get_local_ip()}")
        ip_lbl.setObjectName("subtitle")
        ip_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(ip_lbl)
 
        layout.addSpacing(6)
 
        # Form
        self.txt_user = QLineEdit()
        self.txt_user.setPlaceholderText("Kullanıcı adı")
        layout.addWidget(self.txt_user)
 
        pass_w = QWidget()
        pass_hl = QHBoxLayout(pass_w); pass_hl.setContentsMargins(0,0,0,0); pass_hl.setSpacing(0)
        self.txt_pass = QLineEdit()
        self.txt_pass.setPlaceholderText("Şifre")
        self.txt_pass.setEchoMode(QLineEdit.EchoMode.Password)
        self.txt_pass.returnPressed.connect(self._do_login)
        self.txt_pass.setStyleSheet(
            "background:#313244;border:1px solid #45475A;border-radius:6px 0 0 6px;"
            "padding:8px 12px;color:#CDD6F4;font-size:10pt;")
        btn_eye = QPushButton("👁"); btn_eye.setFixedSize(40,40); btn_eye.setCheckable(True)
        btn_eye.setStyleSheet(
            "QPushButton{background:#313244;border:1px solid #45475A;"
            "border-left:none;border-radius:0 6px 6px 0;font-size:14pt;color:#585B70;}"
            "QPushButton:hover{color:#CDD6F4;}QPushButton:checked{color:#89B4FA;}")
        btn_eye.toggled.connect(lambda c: self.txt_pass.setEchoMode(
            QLineEdit.EchoMode.Normal if c else QLineEdit.EchoMode.Password))
        pass_hl.addWidget(self.txt_pass); pass_hl.addWidget(btn_eye)
        layout.addWidget(pass_w)
 
        self.lbl_err = QLabel("")
        self.lbl_err.setObjectName("err")
        self.lbl_err.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_err.setWordWrap(True)
        layout.addWidget(self.lbl_err)
 
        btn_login = QPushButton("Giriş Yap")
        btn_login.setObjectName("btn_login")
        btn_login.setMinimumHeight(44)
        btn_login.clicked.connect(self._do_login)
        btn_login.setEnabled(True)  # Lisans olmadan da girise izin ver
        layout.addWidget(btn_login)
 
        btn_license = QPushButton("Lisans Yükle / Yenile")
        btn_license.setObjectName("btn_license")
        btn_license.clicked.connect(self._load_license)
        layout.addWidget(btn_license)
 
        # Makine parmak izi (lisans almak için gerekli)
        fp = _get_machine_fingerprint()
        fp_lbl = QLabel(f"Makine Kodu: {fp[:16]}…")
        fp_lbl.setObjectName("subtitle")
        fp_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        fp_lbl.setToolTip(f"Tam parmak izi (lisans almak için sağlayıcıya gönderin):\\n{fp}")
        layout.addWidget(fp_lbl)
 
        btn_copy_fp = QPushButton("Makine Kodunu Kopyala")
        btn_copy_fp.setObjectName("btn_license")
        btn_copy_fp.clicked.connect(lambda: (
            QApplication.clipboard().setText(fp),
            QMessageBox.information(self, "Kopyalandı",
                "Makine kodu panoya kopyalandı.\\nLisans almak için sağlayıcınıza gönderin.")
        ))
        layout.addWidget(btn_copy_fp)
 
        self.txt_user.setFocus()
 
    def _do_login(self):
        username = self.txt_user.text().strip()
        password = self.txt_pass.text()
        if not username or not password:
            self.lbl_err.setText("Kullanıcı adı ve şifre gerekli.")
            return
 
        # Eski/takılı kalmış oturumları temizle (program çökmüşse oturum açık kalabilir)
        self.security.conn.execute(
            "UPDATE user_sessions SET is_active=0, logout_at=datetime('now','localtime') "
            "WHERE is_active=1 AND datetime(last_active) < datetime('now','-2 hours')"
        )
        self.security.conn.commit()

        # Aktif oturum sayısı kontrolü — lisans geçerliyse sınırla, değilse serbest
        active = self.security.active_session_count()
        max_u  = self.license_info.get("max_users", 999) if self.license_info.get("valid") else 999
        if active >= max_u:
            self.lbl_err.setText(
                f"Maksimum kullanıcı sınırına ulaşıldı ({max_u} kullanıcı).\n"
                "Başka bir kullanıcının oturumu kapanmasını bekleyin.")
            return
 
        result = self.security.login(username, password)
        if result["success"]:
            try:
                user_row = self.security.conn.execute(
                    "SELECT must_change_password FROM users WHERE username=?",
                    (username,)).fetchone()
                if user_row and user_row[0]:
                    dlg = _ForcePasswordChangeDialog(self.security, username, self)
                    if dlg.exec() != QDialog.DialogCode.Accepted:
                        self.lbl_err.setText("Şifrenizi değiştirmeniz zorunludur.")
                        self.security.logout()
                        return
            except Exception:
                pass
            self.logged_in = True
            self.accept()
        else:
            self.lbl_err.setText(result["error"])
            self.txt_pass.clear()
            self.txt_pass.setFocus()
 
    def _load_license(self):
        token, ok = QInputDialog.getMultiLineText(
            self, "Lisans Yükle",
            "Sağlayıcınızdan aldığınız lisans kodunu yapıştırın:"
        )
        if ok and token.strip():
            if LicenseManager.install_license(token.strip()):
                QMessageBox.information(self, "Başarılı", "Lisans başarıyla yüklendi. Program yeniden başlatılacak.")
                QApplication.quit()
            else:
                QMessageBox.critical(self, "Hata",
                    "Lisans geçersiz veya bu bilgisayara ait değil.\\n"
                    "Makine kodunuzu sağlayıcıya gönderin.")
 
 
# =============================================================================
# YÖNETİCİ PANELİ (SecurityAdminPage)
# =============================================================================
 
class SecurityAdminPage(QWidget):
    """
    Kullanıcı yönetimi + güvenlik log + aktif oturum izleme.
    Sadece admin rolündeki kullanıcılar görebilir.
    """
 
    def __init__(self, security: SecurityManager, license_info: dict, parent=None):
        super().__init__(parent)
        self.security     = security
        self.license_info = license_info
        self._setup_ui()
 
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
 
        title = QLabel("GÜVENLİK YÖNETİM PANELİ")
        title.setObjectName("title")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)
 
        tabs = QTabWidget()
 
        # ── Kullanıcılar ──────────────────────────────────────────────────
        user_tab = QWidget()
        ul = QVBoxLayout(user_tab)
 
        # Lisans özeti
        lic = self.license_info
        lic_str = (f"Lisans: {lic.get('customer','')}  |  "
                   f"Maks. {lic.get('max_users',1)} kullanıcı  |  "
                   f"Son: {lic.get('expiry','')}  |  "
                   f"Seri: {lic.get('serial','')[:12]}…")
        ul.addWidget(QLabel(lic_str))
 
        self.tbl_users = QTableWidget()
        self.tbl_users.setColumnCount(8)
        self.tbl_users.setHorizontalHeaderLabels([
            "Kullanıcı", "Rol", "Ad Soyad", "Son Giriş", "Son IP",
            "Durum", "Hatalı Giriş", "Oluşturma"
        ])
        self.tbl_users.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tbl_users.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tbl_users.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        ul.addWidget(self.tbl_users)
 
        btn_row = QHBoxLayout()
        btn_add = QPushButton("Kullanıcı Ekle")
        btn_add.setObjectName("primary")
        btn_add.clicked.connect(self._add_user)
        btn_row.addWidget(btn_add)
 
        btn_pass = QPushButton("Şifre Sıfırla")
        btn_pass.clicked.connect(self._reset_password)
        btn_row.addWidget(btn_pass)
 
        btn_toggle = QPushButton("Aktif/Pasif")
        btn_toggle.clicked.connect(self._toggle_user)
        btn_row.addWidget(btn_toggle)
 
        btn_row.addStretch()
        ul.addLayout(btn_row)
        tabs.addTab(user_tab, "Kullanıcılar")
 
        # ── Aktif Oturumlar ───────────────────────────────────────────────
        sess_tab = QWidget()
        sl = QVBoxLayout(sess_tab)
 
        self.tbl_sessions = QTableWidget()
        self.tbl_sessions.setColumnCount(7)
        self.tbl_sessions.setHorizontalHeaderLabels([
            "Kullanıcı", "Rol", "IP Adresi", "Makine Kodu", "Giriş", "Son Aktif", "Durum"
        ])
        self.tbl_sessions.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tbl_sessions.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        sl.addWidget(self.tbl_sessions)
 
        btn_kick = QPushButton("Oturumu Sonlandır")
        btn_kick.setObjectName("danger")
        btn_kick.clicked.connect(self._kick_session)
        sl.addWidget(btn_kick)
 
        tabs.addTab(sess_tab, "Aktif Oturumlar")
 
        # ── Güvenlik Logu ────────────────────────────────────────────────
        log_tab = QWidget()
        ll = QVBoxLayout(log_tab)
 
        self.tbl_log = QTableWidget()
        self.tbl_log.setColumnCount(6)
        self.tbl_log.setHorizontalHeaderLabels([
            "Zaman", "Olay", "Kullanıcı", "IP", "Makine", "Detay"
        ])
        self.tbl_log.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tbl_log.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        ll.addWidget(self.tbl_log)
 
        tabs.addTab(log_tab, "Güvenlik Logu")
 
        layout.addWidget(tabs)
 
        # İlk yükleme
        self._refresh()
        tabs.currentChanged.connect(lambda: self._refresh())
 
    def _refresh(self):
        self._load_users()
        self._load_sessions()
        self._load_log()
 
    def _load_users(self):
        users = self.security.get_users()
        self.tbl_users.setRowCount(len(users))
        ROLE_LABEL = {
            UserRole.ADMIN:    "Yönetici",
            UserRole.OPERATOR: "Operatör",
            UserRole.VIEWER:   "Görüntüleyici",
        }
        for i, u in enumerate(users):
            self.tbl_users.setItem(i, 0, QTableWidgetItem(u["username"]))
            self.tbl_users.setItem(i, 1, QTableWidgetItem(ROLE_LABEL.get(u["role"], u["role"])))
            self.tbl_users.setItem(i, 2, QTableWidgetItem(u["full_name"] or ""))
            self.tbl_users.setItem(i, 3, QTableWidgetItem(u["last_login"] or "—"))
            self.tbl_users.setItem(i, 4, QTableWidgetItem(u["last_ip"] or "—"))
            status_item = QTableWidgetItem("Aktif" if u["is_active"] else "Pasif")
            status_item.setForeground(
                QColor("#A6E3A1") if u["is_active"] else QColor("#F38BA8"))
            self.tbl_users.setItem(i, 5, status_item)
            self.tbl_users.setItem(i, 6, QTableWidgetItem(str(u["failed_logins"])))
            self.tbl_users.setItem(i, 7, QTableWidgetItem(u["created_at"] or ""))
 
    def _load_sessions(self):
        sessions = self.security.get_all_sessions()
        self.tbl_sessions.setRowCount(len(sessions))
        for i, s in enumerate(sessions):
            self.tbl_sessions.setItem(i, 0, QTableWidgetItem(s["username"]))
            self.tbl_sessions.setItem(i, 1, QTableWidgetItem(s["role"]))
            self.tbl_sessions.setItem(i, 2, QTableWidgetItem(s["ip_address"]))
            self.tbl_sessions.setItem(i, 3, QTableWidgetItem(s["machine_fp"][:12] + "…"))
            self.tbl_sessions.setItem(i, 4, QTableWidgetItem(s["login_at"]))
            self.tbl_sessions.setItem(i, 5, QTableWidgetItem(s["last_active"]))
            durum = QTableWidgetItem("Açık" if s["is_active"] else "Kapalı")
            durum.setForeground(
                QColor("#A6E3A1") if s["is_active"] else QColor("#A6ADC8"))
            self.tbl_sessions.setItem(i, 6, durum)
 
    def _load_log(self):
        logs = self.security.get_security_log(limit=300)
        self.tbl_log.setRowCount(len(logs))
        EVENT_COLOR = {
            "LOGIN_OK":        "#A6E3A1",
            "LOGIN_FAIL":      "#F9E2AF",
            "ACCOUNT_LOCKED":  "#F38BA8",
            "LOGOUT":          "#A6ADC8",
            "USER_CREATED":    "#89B4FA",
            "PASSWORD_CHANGED":"#89DCEB",
        }
        for i, row in enumerate(logs):
            self.tbl_log.setItem(i, 0, QTableWidgetItem(row["ts"]))
            ev_item = QTableWidgetItem(row["event_type"])
            ev_item.setForeground(QColor(EVENT_COLOR.get(row["event_type"], "#CDD6F4")))
            self.tbl_log.setItem(i, 1, ev_item)
            self.tbl_log.setItem(i, 2, QTableWidgetItem(row["username"]))
            self.tbl_log.setItem(i, 3, QTableWidgetItem(row["ip_address"]))
            self.tbl_log.setItem(i, 4, QTableWidgetItem(row["machine_fp"][:10] + "…"
                                                          if row["machine_fp"] else ""))
            self.tbl_log.setItem(i, 5, QTableWidgetItem(row["detail"]))
 
    def _add_user(self):
        dlg = _AddUserDialog(self.security, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self._load_users()
 
    def _reset_password(self):
        row = self.tbl_users.currentRow()
        if row < 0:
            return
        username = self.tbl_users.item(row, 0).text()
        new_pass, ok = QInputDialog.getText(
            self, "Şifre Sıfırla", f"'{username}' için yeni şifre:",
            QLineEdit.EchoMode.Password
        )
        if ok and new_pass:
            if len(new_pass) < 8:
                QMessageBox.warning(self, "Hata", "Şifre en az 8 karakter olmalı.")
                return
            salt, hsh = _hash_password(new_pass)
            self.security.conn.execute(
                "UPDATE users SET password_salt=?,password_hash=?,failed_logins=0,locked_until=NULL WHERE username=?",
                (salt, hsh, username)
            )
            self.security.conn.commit()
            self.security._log("PASSWORD_RESET", username, _get_local_ip(), "", "Admin sıfırladı")
            QMessageBox.information(self, "Tamam", f"'{username}' şifresi güncellendi.")
 
    def _toggle_user(self):
        row = self.tbl_users.currentRow()
        if row < 0:
            return
        username = self.tbl_users.item(row, 0).text()
        if username == "admin":
            QMessageBox.warning(self, "İzin Yok", "Admin hesabı devre dışı bırakılamaz.")
            return
        cur_status = self.tbl_users.item(row, 5).text()
        new_status = 0 if cur_status == "Aktif" else 1
        self.security.conn.execute(
            "UPDATE users SET is_active=? WHERE username=?", (new_status, username))
        self.security.conn.commit()
        self._load_users()
 
    def _kick_session(self):
        row = self.tbl_sessions.currentRow()
        if row < 0:
            return
        username = self.tbl_sessions.item(row, 0).text()
        self.security.conn.execute(
            """UPDATE user_sessions SET is_active=0,
               logout_at=datetime('now','localtime')
               WHERE username=? AND is_active=1""",
            (username,)
        )
        self.security.conn.commit()
        self.security._log("SESSION_KICKED", username, _get_local_ip(), "", "Admin sonlandırdı")
        self._load_sessions()
 
 
class _AddUserDialog(QDialog):
    def __init__(self, security: SecurityManager, parent=None):
        super().__init__(parent)
        self.security = security
        self.setWindowTitle("Kullanıcı Ekle")
        self.setFixedSize(360, 340)
        layout = QFormLayout(self)
 
        self.txt_user  = QLineEdit(); layout.addRow("Kullanıcı Adı:", self.txt_user)
        self.txt_name  = QLineEdit(); layout.addRow("Ad Soyad:",      self.txt_name)
        self.txt_email = QLineEdit(); layout.addRow("E-posta:",        self.txt_email)
        self.txt_pass  = QLineEdit()
        self.txt_pass.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addRow("Şifre (min 8):", self.txt_pass)
 
        self.cmb_role = QComboBox()
        self.cmb_role.addItems(["Operatör", "Görüntüleyici", "Yönetici"])
        layout.addRow("Rol:", self.cmb_role)
 
        self.lbl_err = QLabel("")
        self.lbl_err.setStyleSheet("color: #F38BA8;")
        layout.addRow(self.lbl_err)
 
        btn = QPushButton("Ekle")
        btn.setObjectName("primary")
        btn.clicked.connect(self._create)
        layout.addRow(btn)
 
    def _create(self):
        ROLE_MAP = {"Operatör": UserRole.OPERATOR,
                    "Görüntüleyici": UserRole.VIEWER,
                    "Yönetici": UserRole.ADMIN}
        role = ROLE_MAP[self.cmb_role.currentText()]
        ok = self.security.create_user(
            username  = self.txt_user.text().strip(),
            password  = self.txt_pass.text(),
            role      = role,
            full_name = self.txt_name.text().strip(),
            email     = self.txt_email.text().strip(),
            created_by= self.security.current_user["username"]
                        if self.security.current_user else "admin"
        )
        if ok:
            self.accept()
        else:
            self.lbl_err.setText("Kullanıcı adı zaten var veya şifre çok kısa.")
# =============================================================================
# DATABASE MANAGER - SQLITE
# =============================================================================

class DatabaseManager:
    """SQLite veritabani yonetimi - her islemde yeni baglanti."""

    def __init__(self, db_path: str = None):
        if db_path is None:
            app_dir = Path.home() / ".adr_transport_pro"
            app_dir.mkdir(exist_ok=True)
            db_path = str(app_dir / "adr_database.db")
        self.db_path = db_path
        self.connection = None
        self.init_database()
        self._setup_backup_system()

    @property
    def conn(self) -> sqlite3.Connection:
        """SecurityManager uyumluluğu icin kisayol."""
        return self._get_conn()

    def _get_conn(self) -> sqlite3.Connection:
        if self.connection is None:
            self.connection = sqlite3.connect(
                self.db_path, timeout=30, check_same_thread=False
            )
            self.connection.row_factory = sqlite3.Row
            self.connection.execute("PRAGMA journal_mode=WAL")
            self.connection.execute("PRAGMA synchronous=NORMAL")
            self.connection.execute("PRAGMA busy_timeout=30000")
            self.connection.execute("PRAGMA cache_size=-8000")
        return self.connection

    def init_database(self):
        conn = self._get_conn()
        cursor = conn.cursor()

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                type TEXT NOT NULL CHECK(type IN ('sender', 'receiver', 'carrier')),
                name TEXT NOT NULL,
                tax_number TEXT,
                tax_office TEXT,
                mersis_no TEXT,
                address TEXT,
                city TEXT,
                district TEXT,
                phone TEXT,
                email TEXT,
                contact_person TEXT,
                is_favorite INTEGER DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS drivers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                full_name TEXT NOT NULL,
                tc_no TEXT UNIQUE,
                phone TEXT,
                adr_certificate_no TEXT,
                adr_certificate_expiry TEXT,
                src5_no TEXT,
                src5_expiry TEXT,
                license_class TEXT,
                license_expiry TEXT,
                is_active INTEGER DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS vehicles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plate TEXT NOT NULL UNIQUE,
                trailer_plate TEXT,
                adr_compliance_cert_no TEXT,
                adr_compliance_expiry TEXT,
                inspection_date TEXT,
                inspection_expiry TEXT,
                tank_info TEXT,
                vehicle_type TEXT,
                max_capacity REAL DEFAULT 0,
                is_active INTEGER DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS chemicals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                un_number TEXT NOT NULL,
                classification_code TEXT DEFAULT '',
                proper_shipping_name_tr TEXT,
                proper_shipping_name_en TEXT,
                class_code TEXT,
                packing_group TEXT,
                tunnel_code TEXT,
                transport_category TEXT,
                segregation_group TEXT,
                special_provisions TEXT,
                lq_allowed INTEGER DEFAULT 0,
                eq_allowed INTEGER DEFAULT 0,
                limited_quantity TEXT DEFAULT '',
                excepted_quantity TEXT DEFAULT '',
                hazard_labels TEXT,
                UNIQUE(un_number, classification_code, packing_group)
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS shipments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                document_no TEXT NOT NULL UNIQUE,
                document_date TEXT,
                status TEXT DEFAULT 'Taslak',
                sender_id INTEGER,
                receiver_id INTEGER,
                carrier_id INTEGER,
                driver_id INTEGER,
                vehicle_id INTEGER,
                total_points REAL DEFAULT 0,
                orange_plate_required INTEGER DEFAULT 0,
                written_instructions_required INTEGER DEFAULT 0,
                driver_adr_required INTEGER DEFAULT 0,
                tunnel_restriction_code TEXT,
                exemption_type TEXT DEFAULT 'Yok',
                is_validated INTEGER DEFAULT 0,
                validation_errors TEXT,
                notes TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                created_by TEXT,
                FOREIGN KEY (sender_id)   REFERENCES companies(id),
                FOREIGN KEY (receiver_id) REFERENCES companies(id),
                FOREIGN KEY (carrier_id)  REFERENCES companies(id),
                FOREIGN KEY (driver_id)   REFERENCES drivers(id),
                FOREIGN KEY (vehicle_id)  REFERENCES vehicles(id)
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS shipment_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                shipment_id INTEGER NOT NULL,
                chemical_id INTEGER,
                un_number TEXT,
                proper_name TEXT,
                class_code TEXT,
                packing_group TEXT,
                packaging_type TEXT,
                packaging_count INTEGER DEFAULT 1,
                net_quantity REAL DEFAULT 0,
                gross_quantity REAL DEFAULT 0,
                unit TEXT DEFAULT 'kg',
                is_lq INTEGER DEFAULT 0,
                is_eq INTEGER DEFAULT 0,
                lq_max_per_package REAL DEFAULT 0,
                eq_max_per_package REAL DEFAULT 0,
                notes TEXT,
                FOREIGN KEY (shipment_id) REFERENCES shipments(id) ON DELETE CASCADE,
                FOREIGN KEY (chemical_id) REFERENCES chemicals(id)
            )
        """)

        # Migration: eski DB'lere tunnel_code / transport_category / segregation_group / classification_code ekle
        for col, default in [
            ("tunnel_code",         "''"),
            ("transport_category",  "''"),   # [DÜZELTİLDİ] '2' varsayılanı kaldırıldı
            ("segregation_group",   "''"),
            ("classification_code", "''"),
        ]:
            try:
                cursor.execute(
                    f"ALTER TABLE shipment_items ADD COLUMN {col} TEXT DEFAULT {default}"
                )
            except Exception:
                pass  # Kolon zaten varsa hata görmezden gel

        # Mevcut DB'de transport_category DEFAULT '2' ile eklenmiş BOŞLUK sorununu
        # çözmek için: artık ALTER TABLE DEFAULT '' kullanılıyor (yukarıda düzeltildi).
        # NOT: TC=2 geçerli bir ADR taşıma kategorisidir (UN 1203 Benzin gibi).
        #      Bu değerleri silmek YANLIŞ olur — UPDATE ifadeleri KALDIRILDI.

        # ------------------------------------------------------------------
        # [v4.2 MIGRASYON] chemicals: limit alanlari + bilesik anahtar
        # ------------------------------------------------------------------
        chem_cols = {row[1] for row in cursor.execute(
            "PRAGMA table_info(chemicals)").fetchall()}
        for col in ("classification_code", "limited_quantity", "excepted_quantity"):
            if col not in chem_cols:
                cursor.execute(
                    f"ALTER TABLE chemicals ADD COLUMN {col} TEXT DEFAULT ''")

        # Eski semada UNIQUE(un_number) varsa tabloyu yeniden insa et:
        # UN1950 gibi 12 varyantli maddeler icin bilesik anahtar sarttir.
        table_sql = cursor.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='chemicals'"
        ).fetchone()
        if table_sql and "un_number TEXT NOT NULL UNIQUE" in (table_sql[0] or ""):
            cursor.execute("ALTER TABLE chemicals RENAME TO chemicals_eski")
            cursor.execute("""
                CREATE TABLE chemicals (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    un_number TEXT NOT NULL,
                    classification_code TEXT DEFAULT '',
                    proper_shipping_name_tr TEXT,
                    proper_shipping_name_en TEXT,
                    class_code TEXT,
                    packing_group TEXT,
                    tunnel_code TEXT,
                    transport_category TEXT,
                    segregation_group TEXT,
                    special_provisions TEXT,
                    lq_allowed INTEGER DEFAULT 0,
                    eq_allowed INTEGER DEFAULT 0,
                    limited_quantity TEXT DEFAULT '',
                    excepted_quantity TEXT DEFAULT '',
                    hazard_labels TEXT,
                    UNIQUE(un_number, classification_code, packing_group)
                )
            """)
            cursor.execute("""
                INSERT OR IGNORE INTO chemicals
                    (id, un_number, classification_code, proper_shipping_name_tr,
                     proper_shipping_name_en, class_code, packing_group, tunnel_code,
                     transport_category, segregation_group, special_provisions,
                     lq_allowed, eq_allowed, limited_quantity, excepted_quantity,
                     hazard_labels)
                SELECT id, un_number, classification_code, proper_shipping_name_tr,
                       proper_shipping_name_en, class_code, packing_group, tunnel_code,
                       transport_category, segregation_group, special_provisions,
                       lq_allowed, eq_allowed, limited_quantity, excepted_quantity,
                       hazard_labels
                FROM chemicals_eski
            """)
            cursor.execute("DROP TABLE chemicals_eski")

        # DÜZELTME (Umut'un tespiti: "UN1202 için 3 seçenek, UN1950 için
        # daha fazla seçenek çıkması gerekirken tek seçenek çıkıyor"):
        # UNIQUE(un_number, classification_code, packing_group) kısıtı
        # yukarıdaki göçle doğru bir bileşik anahtar olarak eklenmişti
        # AMA resmi Tablo A'da bu üçlü AYNI olup yalnızca özel hüküm
        # (6. sütun) ile ayrışan GERÇEKTEN FARKLI satırlar var (ör.
        # UN1133 F1 PG II: 640C ve 640D varyantları). Bu üçlü YANLIŞLIKLA
        # "birincil anahtar" sayılıp böyle satırlar birbirinin üzerine
        # yazılıyordu — 2939 geçerli satırdan 66'sı kayboluyor, 2873
        # kalıyordu (bu fonksiyonun karşılığı web tarafında zaten
        # düzeltilmişti; masaüstüne hiç geri taşınmamıştı). Kısıt tamamen
        # kaldırılıyor; tekilleştirme artık import_table_a_excel()
        # içinde TAM SATIR İMZASI ile yapılıyor (aşağıya bakınız).
        table_sql2 = cursor.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='chemicals'"
        ).fetchone()
        if table_sql2 and "UNIQUE(un_number, classification_code, packing_group)" \
                in (table_sql2[0] or ""):
            cursor.execute("ALTER TABLE chemicals RENAME TO chemicals_eski2")
            cursor.execute("""
                CREATE TABLE chemicals (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    un_number TEXT NOT NULL,
                    classification_code TEXT DEFAULT '',
                    proper_shipping_name_tr TEXT,
                    proper_shipping_name_en TEXT,
                    class_code TEXT,
                    packing_group TEXT,
                    tunnel_code TEXT,
                    transport_category TEXT,
                    segregation_group TEXT,
                    special_provisions TEXT,
                    lq_allowed INTEGER DEFAULT 0,
                    eq_allowed INTEGER DEFAULT 0,
                    limited_quantity TEXT DEFAULT '',
                    excepted_quantity TEXT DEFAULT '',
                    hazard_labels TEXT
                )
            """)
            cursor.execute("""
                INSERT INTO chemicals
                    (id, un_number, classification_code, proper_shipping_name_tr,
                     proper_shipping_name_en, class_code, packing_group, tunnel_code,
                     transport_category, segregation_group, special_provisions,
                     lq_allowed, eq_allowed, limited_quantity, excepted_quantity,
                     hazard_labels)
                SELECT id, un_number, classification_code, proper_shipping_name_tr,
                       proper_shipping_name_en, class_code, packing_group, tunnel_code,
                       transport_category, segregation_group, special_provisions,
                       lq_allowed, eq_allowed, limited_quantity, excepted_quantity,
                       hazard_labels
                FROM chemicals_eski2
            """)
            cursor.execute("DROP TABLE chemicals_eski2")

        # companies.logo_path (PDF antet)
        comp_cols = {row[1] for row in cursor.execute(
            "PRAGMA table_info(companies)").fetchall()}
        if "logo_path" not in comp_cols:
            cursor.execute(
                "ALTER TABLE companies ADD COLUMN logo_path TEXT DEFAULT ''")

        # Firma urun katalogu: ayni ADR karsiligini paylasan ticari urunler
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS company_products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_name TEXT NOT NULL DEFAULT '',
                trade_name TEXT NOT NULL,
                un_number TEXT NOT NULL,
                classification_code TEXT DEFAULT '',
                packing_group TEXT DEFAULT '',
                limited_quantity TEXT DEFAULT '',
                excepted_quantity TEXT DEFAULT '',
                transport_category TEXT DEFAULT '',
                tunnel_code TEXT DEFAULT '',
                special_provisions TEXT DEFAULT '',
                source_file TEXT DEFAULT '',
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(trade_name, un_number, classification_code)
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS packaging_types (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name_tr TEXT,
                name_en TEXT,
                description TEXT,
                max_gross_mass REAL,
                un_tested INTEGER DEFAULT 0
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT DEFAULT CURRENT_TIMESTAMP,
                user_name TEXT,
                action TEXT,
                table_name TEXT,
                record_id INTEGER,
                old_values TEXT,
                new_values TEXT,
                ip_address TEXT
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT,
                description TEXT
            )
        """)

        cursor.execute('CREATE INDEX IF NOT EXISTS idx_companies_name   ON companies(name)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_companies_type   ON companies(type)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_chemicals_un     ON chemicals(un_number)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_chemicals_name   ON chemicals(proper_shipping_name_tr)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_shipments_docno  ON shipments(document_no)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_shipments_date   ON shipments(document_date)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_shipment_items_shipment ON shipment_items(shipment_id)')

        conn.commit()

    def _setup_backup_system(self):
        backup_dir = Path(self.db_path).parent / "backups"
        backup_dir.mkdir(exist_ok=True)

        today = datetime.now().strftime("%Y%m%d")
        backup_file = backup_dir / f"adr_backup_{today}.db"

        if not backup_file.exists():
            try:
                if self.connection:
                    self.connection.close()
                    self.connection = None
                shutil.copy2(self.db_path, backup_file)
            except Exception:
                pass
            finally:
                self._get_conn()

        cutoff = datetime.now() - timedelta(days=30)
        for f in backup_dir.glob("adr_backup_*.db"):
            try:
                file_date = datetime.strptime(f.stem.split("_")[-1], "%Y%m%d")
                if file_date < cutoff:
                    f.unlink()
            except:
                pass

    def execute(self, query: str, params: tuple = ()) -> List[sqlite3.Row]:
        cursor = self._get_conn().cursor()
        cursor.execute(query, params)
        return cursor.fetchall()

    def execute_one(self, query: str, params: tuple = ()) -> Optional[sqlite3.Row]:
        cursor = self._get_conn().cursor()
        cursor.execute(query, params)
        return cursor.fetchone()

    def execute_insert(self, query: str, params: tuple = ()) -> int:
        conn = self._get_conn()
        cursor = conn.cursor()
        cursor.execute(query, params)
        conn.commit()
        return cursor.lastrowid

    def execute_update(self, query: str, params: tuple = ()) -> int:
        conn = self._get_conn()
        cursor = conn.cursor()
        cursor.execute(query, params)
        conn.commit()
        return cursor.rowcount

    def execute_delete(self, query: str, params: tuple = ()) -> int:
        return self.execute_update(query, params)

    def close(self):
        if self.connection:
            try:
                self.connection.commit()
                self.connection.close()
            except Exception:
                pass
            finally:
                self.connection = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    # --- COMPANY CRUD ---

    def add_company(self, company: Company) -> int:
        existing = self.execute_one(
            "SELECT id FROM companies WHERE LOWER(name)=LOWER(?) AND type=?",
            (company.name.strip(), company.type)
        )
        if existing:
            return existing["id"]

        query = """
            INSERT INTO companies (type, name, tax_number, tax_office, mersis_no,
                address, city, district, phone, email, contact_person, is_favorite)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        return self.execute_insert(query, (
            company.type, company.name, company.tax_number, company.tax_office,
            company.mersis_no, company.address, company.city, company.district,
            company.phone, company.email, company.contact_person, int(company.is_favorite)
        ))

    def update_company(self, company: Company) -> int:
        query = """
            UPDATE companies SET type=?, name=?, tax_number=?, tax_office=?,
                mersis_no=?, address=?, city=?, district=?, phone=?, email=?,
                contact_person=?, is_favorite=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """
        return self.execute_update(query, (
            company.type, company.name, company.tax_number, company.tax_office,
            company.mersis_no, company.address, company.city, company.district,
            company.phone, company.email, company.contact_person, int(company.is_favorite),
            company.id
        ))

    def delete_company(self, company_id: int) -> int:
        return self.execute_delete("DELETE FROM companies WHERE id=?", (company_id,))

    def get_company(self, company_id: int) -> Optional[Company]:
        row = self.execute_one("SELECT * FROM companies WHERE id=?", (company_id,))
        if row:
            return self._row_to_company(row)
        return None

    def get_companies(self, company_type: str = None, search: str = None,
                      favorite_only: bool = False) -> List[Company]:
        query = "SELECT * FROM companies WHERE 1=1"
        params = []
        if company_type:
            query += " AND type=?"
            params.append(company_type)
        if favorite_only:
            query += " AND is_favorite=1"
        if search:
            query += " AND (name LIKE ? OR tax_number LIKE ? OR city LIKE ?)"
            params.extend([f"%{search}%"] * 3)
        query += " ORDER BY name"
        rows = self.execute(query, tuple(params))
        return [self._row_to_company(r) for r in rows]

    def _row_to_company(self, row: sqlite3.Row) -> Company:
        return Company(
            id=row["id"], type=row["type"], name=row["name"],
            tax_number=row["tax_number"] or "", tax_office=row["tax_office"] or "",
            mersis_no=row["mersis_no"] or "", address=row["address"] or "",
            city=row["city"] or "", district=row["district"] or "",
            phone=row["phone"] or "", email=row["email"] or "",
            contact_person=row["contact_person"] or "",
            is_favorite=bool(row["is_favorite"]),
            created_at=row["created_at"] or "", updated_at=row["updated_at"] or ""
        )

    # --- DRIVER CRUD ---

    def add_driver(self, driver: Driver) -> int:
        query = """
            INSERT INTO drivers (full_name, tc_no, phone, adr_certificate_no,
                adr_certificate_expiry, src5_no, src5_expiry, license_class, license_expiry)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        return self.execute_insert(query, (
            driver.full_name, driver.tc_no, driver.phone, driver.adr_certificate_no,
            driver.adr_certificate_expiry, driver.src5_no, driver.src5_expiry,
            driver.license_class, driver.license_expiry
        ))

    def update_driver(self, driver: Driver) -> int:
        query = """
            UPDATE drivers SET full_name=?, tc_no=?, phone=?, adr_certificate_no=?,
                adr_certificate_expiry=?, src5_no=?, src5_expiry=?, license_class=?,
                license_expiry=?, is_active=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """
        return self.execute_update(query, (
            driver.full_name, driver.tc_no, driver.phone, driver.adr_certificate_no,
            driver.adr_certificate_expiry, driver.src5_no, driver.src5_expiry,
            driver.license_class, driver.license_expiry, int(driver.is_active), driver.id
        ))

    def get_driver(self, driver_id: int) -> Optional[Driver]:
        row = self.execute_one("SELECT * FROM drivers WHERE id=?", (driver_id,))
        return self._row_to_driver(row) if row else None

    def get_drivers(self, search: str = None, active_only: bool = True) -> List[Driver]:
        query = "SELECT * FROM drivers WHERE 1=1"
        params = []
        if active_only:
            query += " AND is_active=1"
        if search:
            query += " AND (full_name LIKE ? OR tc_no LIKE ? OR src5_no LIKE ?)"
            params.extend([f"%{search}%"] * 3)
        query += " ORDER BY full_name"
        rows = self.execute(query, tuple(params))
        return [self._row_to_driver(r) for r in rows]

    def _row_to_driver(self, row: sqlite3.Row) -> Driver:
        return Driver(
            id=row["id"], full_name=row["full_name"], tc_no=row["tc_no"] or "",
            phone=row["phone"] or "", adr_certificate_no=row["adr_certificate_no"] or "",
            adr_certificate_expiry=row["adr_certificate_expiry"] or "",
            src5_no=row["src5_no"] or "", src5_expiry=row["src5_expiry"] or "",
            license_class=row["license_class"] or "",
            license_expiry=row["license_expiry"] or "",
            is_active=bool(row["is_active"])
        )

    # --- VEHICLE CRUD ---

    def add_vehicle(self, vehicle: Vehicle) -> int:
        query = """
            INSERT INTO vehicles (plate, trailer_plate, adr_compliance_cert_no,
                adr_compliance_expiry, inspection_date, inspection_expiry, tank_info,
                vehicle_type, max_capacity)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        return self.execute_insert(query, (
            vehicle.plate, vehicle.trailer_plate, vehicle.adr_compliance_cert_no,
            vehicle.adr_compliance_expiry, vehicle.inspection_date,
            vehicle.inspection_expiry, vehicle.tank_info, vehicle.vehicle_type,
            vehicle.max_capacity
        ))

    def update_vehicle(self, vehicle: Vehicle) -> int:
        query = """
            UPDATE vehicles SET plate=?, trailer_plate=?, adr_compliance_cert_no=?,
                adr_compliance_expiry=?, inspection_date=?, inspection_expiry=?,
                tank_info=?, vehicle_type=?, max_capacity=?, is_active=?,
                updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """
        return self.execute_update(query, (
            vehicle.plate, vehicle.trailer_plate, vehicle.adr_compliance_cert_no,
            vehicle.adr_compliance_expiry, vehicle.inspection_date,
            vehicle.inspection_expiry, vehicle.tank_info, vehicle.vehicle_type,
            vehicle.max_capacity, int(vehicle.is_active), vehicle.id
        ))

    def get_vehicle(self, vehicle_id: int) -> Optional[Vehicle]:
        row = self.execute_one("SELECT * FROM vehicles WHERE id=?", (vehicle_id,))
        return self._row_to_vehicle(row) if row else None

    def get_vehicles(self, search: str = None, active_only: bool = True) -> List[Vehicle]:
        query = "SELECT * FROM vehicles WHERE 1=1"
        params = []
        if active_only:
            query += " AND is_active=1"
        if search:
            query += " AND (plate LIKE ? OR trailer_plate LIKE ?)"
            params.extend([f"%{search}%"] * 2)
        query += " ORDER BY plate"
        rows = self.execute(query, tuple(params))
        return [self._row_to_vehicle(r) for r in rows]

    def _row_to_vehicle(self, row: sqlite3.Row) -> Vehicle:
        return Vehicle(
            id=row["id"], plate=row["plate"], trailer_plate=row["trailer_plate"] or "",
            adr_compliance_cert_no=row["adr_compliance_cert_no"] or "",
            adr_compliance_expiry=row["adr_compliance_expiry"] or "",
            inspection_date=row["inspection_date"] or "",
            inspection_expiry=row["inspection_expiry"] or "",
            tank_info=row["tank_info"] or "", vehicle_type=row["vehicle_type"] or "",
            max_capacity=row["max_capacity"] or 0, is_active=bool(row["is_active"])
        )

    # --- CHEMICAL CRUD ---

    def add_chemical(self, chemical: Chemical) -> int:
        query = """
            INSERT INTO chemicals (un_number, classification_code,
                proper_shipping_name_tr, proper_shipping_name_en, class_code,
                packing_group, tunnel_code, transport_category, segregation_group,
                special_provisions, lq_allowed, eq_allowed,
                limited_quantity, excepted_quantity, hazard_labels)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        return self.execute_insert(query, (
            chemical.un_number, chemical.classification_code,
            chemical.proper_shipping_name_tr,
            chemical.proper_shipping_name_en, chemical.class_code,
            chemical.packing_group, chemical.tunnel_code, chemical.transport_category,
            chemical.segregation_group, chemical.special_provisions,
            int(chemical.lq_allowed), int(chemical.eq_allowed),
            chemical.limited_quantity, chemical.excepted_quantity,
            chemical.hazard_labels
        ))

    def count_chemicals(self) -> int:
        row = self.execute_one("SELECT COUNT(*) AS n FROM chemicals")
        return row["n"] if row else 0

    def _upsert_chemical(self, c: Chemical) -> bool:
        """True=yeni eklendi, False=mevcut guncellendi.
        Birincil anahtar: UN + siniflandirma_kodu + paketleme_grubu."""
        existing = self.execute_one(
            """SELECT id FROM chemicals
               WHERE un_number=? AND classification_code=? AND packing_group=?""",
            (c.un_number, c.classification_code or "", c.packing_group or ""))
        if existing:
            self.execute_update(
                """UPDATE chemicals SET proper_shipping_name_tr=?, class_code=?,
                   tunnel_code=?, transport_category=?, special_provisions=?,
                   limited_quantity=?, excepted_quantity=?,
                   lq_allowed=?, eq_allowed=?, hazard_labels=?
                   WHERE un_number=? AND classification_code=? AND packing_group=?""",
                (c.proper_shipping_name_tr, c.class_code,
                 c.tunnel_code, c.transport_category, c.special_provisions,
                 c.limited_quantity, c.excepted_quantity, int(c.lq_allowed),
                 int(c.eq_allowed), c.hazard_labels,
                 c.un_number, c.classification_code or "", c.packing_group or ""))
            return False
        self.add_chemical(c)
        return True

    def update_chemical(self, chemical: Chemical) -> int:
        query = """
            UPDATE chemicals SET un_number=?, proper_shipping_name_tr=?,
                proper_shipping_name_en=?, class_code=?, packing_group=?,
                tunnel_code=?, transport_category=?, segregation_group=?,
                special_provisions=?, lq_allowed=?, eq_allowed=?,
                hazard_labels=?
            WHERE id=?
        """
        return self.execute_update(query, (
            chemical.un_number, chemical.proper_shipping_name_tr,
            chemical.proper_shipping_name_en, chemical.class_code,
            chemical.packing_group, chemical.tunnel_code, chemical.transport_category,
            chemical.segregation_group, chemical.special_provisions,
            int(chemical.lq_allowed), int(chemical.eq_allowed), chemical.hazard_labels, chemical.id
        ))

    def delete_chemical(self, chemical_id: int) -> int:
        return self.execute_delete("DELETE FROM chemicals WHERE id=?", (chemical_id,))

    def get_chemical(self, chemical_id: int) -> Optional[Chemical]:
        row = self.execute_one("SELECT * FROM chemicals WHERE id=?", (chemical_id,))
        return self._row_to_chemical(row) if row else None

    def get_chemical_by_un(self, un_number: str) -> Optional[Chemical]:
        row = self.execute_one("SELECT * FROM chemicals WHERE un_number=?", (un_number,))
        return self._row_to_chemical(row) if row else None

    def search_chemicals(self, query: str, limit: int = 50) -> List[Chemical]:
        query = query.strip().upper()
        if not query:
            return []

        # 4 haneli sayisal sorgu -> UN TAM eslesme: maddenin TUM
        # varyasyonlari (UN1950 = 12 satir) doner, yabanci UN gelmez.
        if query.isdigit():
            sql = """
                SELECT * FROM chemicals WHERE un_number = ?
                ORDER BY classification_code, packing_group
                LIMIT ?
            """
            params = (query.zfill(4), limit)
        else:
            sql = """
                SELECT * FROM chemicals
                WHERE proper_shipping_name_tr LIKE ?
                   OR proper_shipping_name_en LIKE ?
                   OR class_code LIKE ?
                ORDER BY
                    CASE WHEN proper_shipping_name_tr LIKE ? THEN 0 ELSE 1 END,
                    un_number, classification_code
                LIMIT ?
            """
            params = (f"%{query}%", f"%{query}%", f"{query}%", f"{query}%", limit)
        rows = self.execute(sql, params)
        chemicals = [self._row_to_chemical(r) for r in rows]

        if RAPIDFUZZ_AVAILABLE and not query.isdigit() and len(chemicals) < limit:
            all_chems = self.execute("SELECT * FROM chemicals", ())
            all_list = [self._row_to_chemical(r) for r in all_chems]

            search_texts = []
            for c in all_list:
                search_texts.append(
                    f"{c.un_number} {c.proper_shipping_name_tr} "
                    f"{c.proper_shipping_name_en} {c.class_code}"
                )

            results = process.extract(query, search_texts, scorer=fuzz.WRatio, limit=limit)
            fuzzy_ids = set()
            for match, score, idx in results:
                if score > 60:
                    fuzzy_ids.add(all_list[idx].id)

            existing_ids = {c.id for c in chemicals}
            for c in all_list:
                if c.id in fuzzy_ids and c.id not in existing_ids:
                    chemicals.append(c)

        return chemicals[:limit]

    def get_all_chemicals(self, search: str = None, class_filter: str = None,
                           limit: int = 500) -> List[Chemical]:
        """Tum kimyasallari filtreli getir (veritabani sayfasi icin)."""
        sql = "SELECT * FROM chemicals WHERE 1=1"
        params = []
        if search:
            s = f"%{search.upper()}%"
            sql += " AND (UPPER(un_number) LIKE ? OR UPPER(proper_shipping_name_tr) LIKE ? OR UPPER(proper_shipping_name_en) LIKE ?)"
            params.extend([s, s, s])
        if class_filter and class_filter != "Tumu":
            sql += " AND class_code=?"
            params.append(class_filter)
        sql += " ORDER BY un_number LIMIT ?"
        params.append(limit)
        rows = self.execute(sql, tuple(params))
        return [self._row_to_chemical(r) for r in rows]

    def _row_to_chemical(self, row: sqlite3.Row) -> Chemical:
        return Chemical(
            id=row["id"], un_number=row["un_number"] or "",
            proper_shipping_name_tr=row["proper_shipping_name_tr"] or "",
            proper_shipping_name_en=row["proper_shipping_name_en"] or "",
            class_code=row["class_code"] or "", packing_group=row["packing_group"] or "",
            tunnel_code=row["tunnel_code"] or "", transport_category=row["transport_category"] or "",
            segregation_group=row["segregation_group"] or "",
            special_provisions=row["special_provisions"] or "",
            lq_allowed=bool(row["lq_allowed"]), eq_allowed=bool(row["eq_allowed"]),
            classification_code=(row["classification_code"] if "classification_code" in row.keys() else "") or "",
            limited_quantity=(row["limited_quantity"] if "limited_quantity" in row.keys() else "") or "",
            excepted_quantity=(row["excepted_quantity"] if "excepted_quantity" in row.keys() else "") or "",
            hazard_labels=row["hazard_labels"] or ""
        )

    # --- SHIPMENT CRUD ---

    def set_shipment_validation(self, shipment_id: int, is_valid: bool,
                                errors_text: str = "") -> None:
        """Dogrulama sonucunu KALICI hale getirir: gecerli evrak 'Onaylandi'
        statusune yukselir, gecersiz evrak 'Taslak'a doner ve hata metni
        validation_errors alanina yazilir. (Eski surumde dogrulama yalnizca
        mesaj kutusunda gosterilip kayboluyordu.)"""
        self.execute_update(
            """UPDATE shipments SET status=?, is_validated=?, validation_errors=?,
               updated_at=CURRENT_TIMESTAMP WHERE id=?""",
            (DocumentStatus.VALIDATED.value if is_valid else DocumentStatus.DRAFT.value,
             int(is_valid), errors_text or "", shipment_id))

    def add_shipment(self, shipment: Shipment) -> int:
        query = """
            INSERT INTO shipments (document_no, document_date, status, sender_id,
                receiver_id, carrier_id, driver_id, vehicle_id, total_points,
                orange_plate_required, written_instructions_required,
                driver_adr_required, tunnel_restriction_code, exemption_type,
                is_validated, validation_errors, notes, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        return self.execute_insert(query, (
            shipment.document_no, shipment.document_date, shipment.status,
            shipment.sender_id, shipment.receiver_id, shipment.carrier_id,
            shipment.driver_id, shipment.vehicle_id, shipment.total_points,
            int(shipment.orange_plate_required), int(shipment.written_instructions_required),
            int(shipment.driver_adr_required), shipment.tunnel_restriction_code,
            shipment.exemption_type, int(shipment.is_validated),
            shipment.validation_errors, shipment.notes, shipment.created_by
        ))

    def update_shipment(self, shipment: Shipment) -> int:
        query = """
            UPDATE shipments SET document_no=?, document_date=?, status=?,
                sender_id=?, receiver_id=?, carrier_id=?, driver_id=?, vehicle_id=?,
                total_points=?, orange_plate_required=?, written_instructions_required=?,
                driver_adr_required=?, tunnel_restriction_code=?, exemption_type=?,
                is_validated=?, validation_errors=?, notes=?,
                updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """
        return self.execute_update(query, (
            shipment.document_no, shipment.document_date, shipment.status,
            shipment.sender_id, shipment.receiver_id, shipment.carrier_id,
            shipment.driver_id, shipment.vehicle_id, shipment.total_points,
            int(shipment.orange_plate_required), int(shipment.written_instructions_required),
            int(shipment.driver_adr_required), shipment.tunnel_restriction_code,
            shipment.exemption_type, int(shipment.is_validated),
            shipment.validation_errors, shipment.notes, shipment.id
        ))

    def get_shipment(self, shipment_id: int) -> Optional[Shipment]:
        row = self.execute_one("SELECT * FROM shipments WHERE id=?", (shipment_id,))
        return self._row_to_shipment(row) if row else None

    def get_shipments(self, status: str = None, date_from: str = None,
                      date_to: str = None, search: str = None,
                      limit: int = 100, offset: int = 0) -> List[Shipment]:
        query = "SELECT * FROM shipments WHERE 1=1"
        params = []
        if status:
            query += " AND status=?"
            params.append(status)
        if date_from:
            query += " AND document_date >= ?"
            params.append(date_from)
        if date_to:
            query += " AND document_date <= ?"
            params.append(date_to)
        if search:
            query += " AND document_no LIKE ?"
            params.append(f"%{search}%")
        query += " ORDER BY document_date DESC, id DESC LIMIT ? OFFSET ?"
        params.extend([limit, offset])
        rows = self.execute(query, tuple(params))
        return [self._row_to_shipment(r) for r in rows]

    def delete_shipment(self, shipment_id: int) -> int:
        return self.execute_delete("DELETE FROM shipments WHERE id=?", (shipment_id,))

    def get_shipments_with_details(self, status: str = None, date_from: str = None,
                                    date_to: str = None, search: str = None,
                                    limit: int = 500) -> List[dict]:
        """
        Tek JOIN sorgusuyla sevkiyat + firma + surucu + arac bilgilerini getirir.
        N+1 sorgu sorununu cozer: 500 sevkiyat icin 2000 sorgu yerine 1 sorgu.
        """
        sql = """
            SELECT
                s.id, s.document_no, s.document_date, s.status,
                s.total_points, s.orange_plate_required, s.exemption_type,
                s.tunnel_restriction_code,
                sender.name   AS sender_name,
                receiver.name AS receiver_name,
                d.full_name   AS driver_name,
                v.plate       AS vehicle_plate
            FROM shipments s
            LEFT JOIN companies sender   ON sender.id   = s.sender_id
            LEFT JOIN companies receiver ON receiver.id = s.receiver_id
            LEFT JOIN drivers d          ON d.id        = s.driver_id
            LEFT JOIN vehicles v         ON v.id        = s.vehicle_id
            WHERE 1=1
        """
        params = []
        if status:
            sql += " AND s.status=?"
            params.append(status)
        if date_from:
            sql += " AND s.document_date >= ?"
            params.append(date_from)
        if date_to:
            sql += " AND s.document_date <= ?"
            params.append(date_to)
        if search:
            sql += """ AND (
                s.document_no LIKE ? OR
                sender.name   LIKE ? OR
                receiver.name LIKE ? OR
                d.full_name   LIKE ? OR
                v.plate       LIKE ?
            )"""
            term = f"%{search}%"
            params.extend([term, term, term, term, term])
        sql += " ORDER BY s.document_date DESC, s.id DESC LIMIT ?"
        params.append(limit)
        rows = self.execute(sql, tuple(params))
        return [dict(r) for r in rows]

    def _row_to_shipment(self, row: sqlite3.Row) -> Shipment:
        return Shipment(
            id=row["id"], document_no=row["document_no"] or "",
            document_date=row["document_date"] or "", status=row["status"] or "Taslak",
            sender_id=row["sender_id"] or 0, receiver_id=row["receiver_id"] or 0,
            carrier_id=row["carrier_id"] or 0, driver_id=row["driver_id"] or 0,
            vehicle_id=row["vehicle_id"] or 0, total_points=row["total_points"] or 0,
            orange_plate_required=bool(row["orange_plate_required"]),
            written_instructions_required=bool(row["written_instructions_required"]),
            driver_adr_required=bool(row["driver_adr_required"]),
            tunnel_restriction_code=row["tunnel_restriction_code"] or "",
            exemption_type=row["exemption_type"] or "Yok",
            is_validated=bool(row["is_validated"]),
            validation_errors=row["validation_errors"] or "",
            notes=row["notes"] or ""
        )

    # --- SHIPMENT ITEM CRUD ---

    def add_shipment_item(self, item: ShipmentItem) -> int:
        query = """
            INSERT INTO shipment_items (shipment_id, chemical_id, un_number,
                proper_name, class_code, packing_group, packaging_type, packaging_count,
                net_quantity, gross_quantity, unit, is_lq, is_eq, lq_max_per_package,
                eq_max_per_package, notes, tunnel_code, transport_category, segregation_group,
                classification_code)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        return self.execute_insert(query, (
            item.shipment_id, item.chemical_id, item.un_number, item.proper_name,
            item.class_code, item.packing_group, item.packaging_type, item.packaging_count,
            item.net_quantity, item.gross_quantity, item.unit, int(item.is_lq),
            int(item.is_eq), item.lq_max_per_package, item.eq_max_per_package, item.notes,
            item.tunnel_code, item.transport_category, item.segregation_group,
            item.classification_code
        ))

    def get_shipment_items(self, shipment_id: int) -> List[ShipmentItem]:
        rows = self.execute("SELECT * FROM shipment_items WHERE shipment_id=?", (shipment_id,))
        return [self._row_to_shipment_item(r) for r in rows]

    def delete_shipment_items(self, shipment_id: int) -> int:
        return self.execute_delete("DELETE FROM shipment_items WHERE shipment_id=?", (shipment_id,))

    def _row_to_shipment_item(self, row: sqlite3.Row) -> ShipmentItem:
        # transport_category, tunnel_code ve segregation_group DB'den okunur.
        # Boş/None gelirse chemicals tablosundan tamamlanır (UN 1202 gibi TC=3 olan maddeler).
        keys = row.keys() if hasattr(row, "keys") else []
        tc_raw = row["transport_category"] if "transport_category" in keys else ""
        tunnel_raw = row["tunnel_code"] if "tunnel_code" in keys else ""
        seg_raw = row["segregation_group"] if "segregation_group" in keys else ""

        item = ShipmentItem(
            id=row["id"], shipment_id=row["shipment_id"], chemical_id=row["chemical_id"] or 0,
            un_number=row["un_number"] or "", proper_name=row["proper_name"] or "",
            class_code=row["class_code"] or "", packing_group=row["packing_group"] or "",
            packaging_type=row["packaging_type"] or "", packaging_count=row["packaging_count"] or 0,
            net_quantity=row["net_quantity"] or 0, gross_quantity=row["gross_quantity"] or 0,
            unit=row["unit"] or "kg", is_lq=bool(row["is_lq"]), is_eq=bool(row["is_eq"]),
            lq_max_per_package=row["lq_max_per_package"] or 0,
            eq_max_per_package=row["eq_max_per_package"] or 0,
            notes=row["notes"] or "",
            transport_category=tc_raw or "",
            tunnel_code=tunnel_raw or "",
            segregation_group=seg_raw or "",
            classification_code=row["classification_code"] if "classification_code" in keys else "",
        )
        # TC boşsa chemicals tablosundan tamamla
        if not item.transport_category and item.chemical_id:
            chem = self.get_chemical(item.chemical_id)
            if chem:
                item.transport_category = chem.transport_category or ""
                if not item.tunnel_code:
                    item.tunnel_code = chem.tunnel_code or ""
                if not item.segregation_group:
                    item.segregation_group = chem.segregation_group or ""
        return item

    # --- AYARLAR ---

    def get_setting(self, key: str, default: str = "") -> str:
        row = self.execute_one("SELECT value FROM settings WHERE key=?", (key,))
        return row["value"] if row else default

    def set_setting(self, key: str, value: str, description: str = ""):
        self.execute_update(
            "INSERT OR REPLACE INTO settings (key, value, description) VALUES (?, ?, ?)",
            (key, value, description)
        )

    # --- ISTATISTIKLER ---

    def get_statistics(self) -> Dict[str, Any]:
        stats = {}
        cursor = self._get_conn().cursor()

        cursor.execute("SELECT COUNT(*) as count FROM shipments")
        stats["total_shipments"] = cursor.fetchone()["count"]

        cursor.execute("SELECT COUNT(*) as count FROM shipments WHERE status='Taslak'")
        stats["draft_shipments"] = cursor.fetchone()["count"]

        cursor.execute("SELECT COUNT(*) as count FROM companies")
        stats["total_companies"] = cursor.fetchone()["count"]

        cursor.execute("SELECT COUNT(*) as count FROM drivers WHERE is_active=1")
        stats["active_drivers"] = cursor.fetchone()["count"]

        cursor.execute("SELECT COUNT(*) as count FROM vehicles WHERE is_active=1")
        stats["active_vehicles"] = cursor.fetchone()["count"]

        cursor.execute("SELECT COUNT(*) as count FROM chemicals")
        stats["total_chemicals"] = cursor.fetchone()["count"]

        cursor.execute("""
            SELECT strftime('%Y-%m', document_date) as month, COUNT(*) as count
            FROM shipments
            WHERE document_date >= date('now', '-6 months')
            GROUP BY month
            ORDER BY month
        """)
        stats["monthly_shipments"] = {r["month"]: r["count"] for r in cursor.fetchall()}

        return stats

    # --- IMPORT/EXPORT ---

    def get_expiring_documents(self, days: int = 30) -> dict:
        from datetime import date, timedelta
        today = date.today().isoformat()
        limit = (date.today() + timedelta(days=days)).isoformat()
        conn  = self._get_conn()
        drivers = conn.execute("""
            SELECT full_name, src5_no, src5_expiry,
                   CAST(julianday(src5_expiry) - julianday('now') AS INTEGER) AS kalan
            FROM drivers WHERE is_active=1 AND src5_expiry BETWEEN ? AND ?
            ORDER BY src5_expiry""", (today, limit)).fetchall()
        vehicles = conn.execute("""
            SELECT plate, adr_compliance_expiry, inspection_expiry,
                   CAST(julianday(adr_compliance_expiry) - julianday('now') AS INTEGER) AS adr_kalan,
                   CAST(julianday(inspection_expiry)     - julianday('now') AS INTEGER) AS mua_kalan
            FROM vehicles WHERE is_active=1
              AND (adr_compliance_expiry BETWEEN ? AND ? OR inspection_expiry BETWEEN ? AND ?)
            ORDER BY adr_compliance_expiry""", (today, limit, today, limit)).fetchall()
        return {"drivers":[dict(r) for r in drivers],"vehicles":[dict(r) for r in vehicles]}

    def get_class_breakdown(self, year=None) -> list:
        sql = """SELECT si.class_code, COUNT(DISTINCT si.shipment_id) AS sevkiyat_sayisi,
                        SUM(si.net_quantity) AS toplam_net_kg
                 FROM shipment_items si JOIN shipments s ON s.id=si.shipment_id"""
        p=[]
        if year: sql+=" WHERE strftime('%Y',s.document_date)=?"; p.append(str(year))
        sql+=" GROUP BY si.class_code ORDER BY toplam_net_kg DESC"
        return [dict(r) for r in self._get_conn().execute(sql,p).fetchall()]

    def get_top_senders(self, limit=10, year=None) -> list:
        sql = "SELECT c.name, COUNT(s.id) AS sevkiyat_sayisi FROM shipments s JOIN companies c ON c.id=s.sender_id"
        p=[]
        if year: sql+=" WHERE strftime('%Y',s.document_date)=?"; p.append(str(year))
        sql+=" GROUP BY s.sender_id ORDER BY sevkiyat_sayisi DESC LIMIT ?"; p.append(limit)
        return [dict(r) for r in self._get_conn().execute(sql,p).fetchall()]

    def get_top_chemicals(self, limit=10, year=None) -> list:
        sql = """SELECT si.un_number, si.class_code, COUNT(*) AS adet,
                        SUM(si.net_quantity) AS toplam_net_kg
                 FROM shipment_items si JOIN shipments s ON s.id=si.shipment_id"""
        p=[]
        if year: sql+=" WHERE strftime('%Y',s.document_date)=?"; p.append(str(year))
        sql+=" GROUP BY si.un_number ORDER BY toplam_net_kg DESC LIMIT ?"; p.append(limit)
        return [dict(r) for r in self._get_conn().execute(sql,p).fetchall()]

    def backup_now(self) -> str:
        import shutil
        backup_dir = Path(self.db_path).parent / "backups"
        backup_dir.mkdir(exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = backup_dir / f"adr_backup_{ts}.db"
        shutil.copy2(self.db_path, dest)
        return str(dest)

    def list_backups(self) -> list:
        backup_dir = Path(self.db_path).parent / "backups"
        if not backup_dir.exists(): return []
        files = sorted(backup_dir.glob("adr_backup_*.db"), reverse=True)
        return [{"name":f.name,"path":str(f),
                 "size":f"{f.stat().st_size/1024:.1f} KB",
                 "date":datetime.fromtimestamp(f.stat().st_mtime).strftime("%d.%m.%Y %H:%M")}
                for f in files]

    def restore_backup(self, backup_path: str):
        import shutil
        if self.connection: self.connection.close(); self.connection=None
        shutil.copy2(backup_path, self.db_path); self._get_conn()

    def get_company_logo_b64(self) -> str:
        return self.get_setting("doc_company_logo_b64") or ""

    def set_company_logo_b64(self, b64: str):
        self.set_setting("doc_company_logo_b64", b64, "Şirket logosu base64")


    # ------------------------------------------------------------------
    # [v4.2] Gercek Excel kaynaklari: ADR Tablo A + firma envanteri
    # (main.py surumunden birebir tasindi, gercek dosyalarla dogrulandi)
    # ------------------------------------------------------------------
    @staticmethod
    def _xl_clean(val) -> str:
        if val is None:
            return ""
        s = str(val).replace("\n", " ").replace("\r", " ").strip()
        import re as _re
        s = _re.sub(r" {2,}", " ", s)
        return "" if s.lower() == "nan" else s

    @staticmethod
    def _xl_un(val) -> str:
        raw = DatabaseManager._xl_clean(val)
        if raw.endswith(".0"):
            raw = raw[:-2]
        import re as _re
        m = _re.match(r"^(?:UN\s*)?(\d{1,4})$", raw, _re.IGNORECASE)
        return m.group(1).zfill(4) if m else ""

    @staticmethod
    def _xl_category_tunnel(raw: str) -> Tuple[str, str]:
        """'2 (D/E)' benzeri hucreyi (kategori, tunel) olarak ayirir."""
        import re as _re
        text = DatabaseManager._xl_clean(raw)
        cat = ""
        m = _re.search(r"\b([0-4])\b", text)
        if m:
            cat = m.group(1)
        tunnel = ""
        m = _re.search(r"\(([A-E][^\)]*)\)", text)
        if m:
            tunnel = m.group(1).strip()
        return cat, tunnel

    def import_table_a_excel(self, xlsx_path: str) -> int:
        """Resmi ADR Tablo A Excel'ini (cok satirli baslik, sutun 7a/7b
        dahil) uygulamanin kimyasal veritabanina aktarir.

        DUZELTME (Umut'un tespiti: arama sonucunda UN basina yalnizca
        1 secenek cikiyordu, olmasi gerekenden cok azdi): eskiden bu
        fonksiyon self._upsert_chemical(c) cagiriyordu -- o da (UN,
        siniflandirma kodu, paketleme grubu) uclusunu "birincil anahtar"
        sayiyordu. Ama resmi Tablo A'da bu uclu AYNI olup yalnizca ozel
        hukum (6. sutun) ile ayrisan GERCEKTEN FARKLI satirlar var (or.
        UN1133 F1 PG II: 640C ve 640D varyantlari) -- bu satirlar
        yanlislikla birbirinin uzerine yaziliyordu (2939 gecerli satirdan
        66'si kayboluyor, 2873 kaliyordu). Tekillestirme artik TAM SATIR
        IMZASI ile yapiliyor: 640C/640D gibi varyantlar korunur, yalnizca
        birebir ayni satirin ikinci kez eklenmesi atlanir. Boylece Tablo
        A'yi tekrar yuklemek de guvenlidir (idempotent).
        """
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]

        mevcut_imzalar = {
            (r["un_number"], r["classification_code"], r["class_code"],
             r["packing_group"], r["special_provisions"],
             r["limited_quantity"], r["excepted_quantity"],
             r["tunnel_code"], r["transport_category"],
             r["proper_shipping_name_tr"])
            for r in self.execute(
                """SELECT un_number, classification_code, class_code,
                          packing_group, special_provisions, limited_quantity,
                          excepted_quantity, tunnel_code, transport_category,
                          proper_shipping_name_tr FROM chemicals""")
        }

        imported = 0
        for row in ws.iter_rows(min_row=5, values_only=True):
            un = self._xl_un(row[0] if len(row) > 0 else None)
            if not un:
                continue
            name = self._xl_clean(row[24] if len(row) > 24 else None) \
                or self._xl_clean(row[1] if len(row) > 1 else None)
            if not name:
                continue
            cat, tunnel = self._xl_category_tunnel(
                row[17] if len(row) > 17 else "")
            lq_text = self._xl_clean(row[7] if len(row) > 7 else None)
            eq_code = self._xl_clean(row[8] if len(row) > 8 else None).upper()
            special = " | ".join(filter(None, [
                self._xl_clean(row[6] if len(row) > 6 else None),
                self._xl_clean(row[21] if len(row) > 21 else None),
            ]))
            c = Chemical(
                un_number=un,
                proper_shipping_name_tr=name,
                class_code=self._xl_clean(row[2] if len(row) > 2 else None),
                classification_code=self._xl_clean(row[3] if len(row) > 3 else None),
                packing_group=self._xl_clean(row[4] if len(row) > 4 else None),
                tunnel_code=tunnel,
                transport_category=cat,
                special_provisions=special,
                limited_quantity=lq_text,
                excepted_quantity=eq_code,
                lq_allowed=ADREngine.parse_lq_limit(lq_text)[0] > 0,
                eq_allowed=ADREngine.eq_limits(eq_code)[0] > 0,
                hazard_labels=self._xl_clean(row[5] if len(row) > 5 else None),
            )
            imza = (c.un_number, c.classification_code, c.class_code,
                    c.packing_group, c.special_provisions,
                    c.limited_quantity, c.excepted_quantity,
                    c.tunnel_code, c.transport_category,
                    c.proper_shipping_name_tr)
            if imza in mevcut_imzalar:
                continue
            mevcut_imzalar.add(imza)
            self.add_chemical(c)
            imported += 1
        return imported

    def import_company_inventory_excel(self, xlsx_path: str) -> int:
        """Firmaya ozel kimyasal envanter Excel'ini aktarir.

        Baslik satiri 'UN NUMARASI' hucresi aranarak otomatik bulunur, boylece
        ust bilgi satirlari/birlestirilmis hucreler sorun cikarmaz. Envanterde
        EQ kodu yoksa, ayni UN daha once Tablo A'dan yuklendiyse EQ oradan
        tamamlanir (uc veri kaynaginin birlesme kurali)."""
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)

        imported = 0
        seen: set = set()
        for ws in wb.worksheets:
            header_row, cols = None, {}
            for r_idx, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
                cells = [self._xl_clean(c).upper() for c in row]
                if any("UN NUMARASI" in c for c in cells):
                    header_row = r_idx
                    for i, c in enumerate(cells):
                        if "UN NUMARASI" in c: cols["un"] = i
                        elif c.startswith("SINIFI"): cols["sinif"] = i
                        elif "PAKETLEME GRUBU" in c: cols["pg"] = i
                        elif "UYGUN SEVKİYAT" in c or "UYGUN SEVKIYAT" in c: cols["ad"] = i
                        elif "SINIRLI MİKTAR" in c or "SINIRLI MIKTAR" in c: cols["lq"] = i
                        elif "ÖZEL HÜKÜM" in c or "OZEL HUKUM" in c: cols["ozel"] = i
                        elif "TAŞIMA KATEGORİSİ" in c or "TASIMA KATEGORISI" in c: cols["kat"] = i
                        elif "KİMYASAL ADI" in c or "KIMYASAL ADI" in c: cols["ticari"] = i
                    break
            if header_row is None or "un" not in cols:
                continue

            def get(row, key):
                i = cols.get(key, -1)
                return self._xl_clean(row[i]) if 0 <= i < len(row) else ""

            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                un = self._xl_un(get(row, "un"))
                if not un:
                    continue  # tehlikesiz / UN'siz satirlar atlanir
                name = get(row, "ad") or get(row, "ticari")
                if not name:
                    continue
                if un in seen:
                    continue
                seen.add(un)

                cat, tunnel = self._xl_category_tunnel(get(row, "kat"))
                lq_text = get(row, "lq")
                ticari = get(row, "ticari")
                notes = f"Ticari ad: {ticari}" if (ticari and ticari != name) else ""

                # EQ envanterde yok: Tablo A'dan (varsa) tamamla
                eq_code = ""
                existing = self.execute_one(
                    "SELECT excepted_quantity FROM chemicals WHERE un_number=?", (un,))
                if existing and existing["excepted_quantity"]:
                    eq_code = existing["excepted_quantity"]

                c = Chemical(
                    un_number=un,
                    proper_shipping_name_tr=name,
                    class_code=get(row, "sinif"),
                    packing_group=get(row, "pg"),
                    tunnel_code=tunnel,
                    transport_category=cat,
                    special_provisions=" | ".join(filter(None, [get(row, "ozel"), notes])),
                    limited_quantity=lq_text,
                    excepted_quantity=eq_code,
                    lq_allowed=ADREngine.parse_lq_limit(lq_text)[0] > 0,
                    eq_allowed=ADREngine.eq_limits(eq_code)[0] > 0,
                )
                self._upsert_chemical(c)
                ticari = get(row, "ticari") or ""
                affected = self.execute_update(
                    """INSERT OR IGNORE INTO company_products
                       (trade_name, un_number, classification_code, packing_group,
                        limited_quantity, excepted_quantity, transport_category,
                        tunnel_code, source_file)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (ticari or name, un, c.classification_code, c.packing_group,
                     c.limited_quantity, c.excepted_quantity, c.transport_category,
                     c.tunnel_code, os.path.basename(xlsx_path)))
                if affected:
                    imported += 1
        return imported

    def search_company_products(self, query: str) -> list:
        q = query.strip().upper()
        if not q:
            return []
        rows = self.execute(
            """SELECT cp.*, c.class_code, c.hazard_labels
               FROM company_products cp
               LEFT JOIN chemicals c ON c.un_number=cp.un_number
                   AND c.classification_code=cp.classification_code
               WHERE UPPER(cp.trade_name) LIKE ?
                  OR cp.un_number LIKE ?
               ORDER BY cp.trade_name LIMIT 100""",
            (f"%{q}%", f"%{q}%"))
        return [dict(r) for r in rows] if rows else []


    def import_json_data(self, json_path: str):
        if not os.path.exists(json_path):
            return 0

        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        imported = 0

        for key in ["sender_companies", "receiver_companies"]:
            for comp in data.get(key, []):
                company = Company(
                    type="sender" if "sender" in key else "receiver",
                    name=comp.get("name", ""),
                    address=comp.get("address", ""),
                    city=comp.get("city", ""),
                    phone=comp.get("phone", ""),
                    email=comp.get("email", ""),
                    contact_person=comp.get("contact_person", "")
                )
                self.add_company(company)
                imported += 1

        for drv in data.get("drivers", []):
            driver = Driver(
                full_name=drv.get("name", ""),
                src5_no=drv.get("src5_no", ""),
                phone=drv.get("phone", "")
            )
            self.add_driver(driver)
            imported += 1

        for pkg in data.get("packaging_types", []):
            self.execute_insert(
                "INSERT OR IGNORE INTO packaging_types (code, name_tr, name_en) VALUES (?, ?, ?)",
                (pkg.get("code", ""), pkg.get("name_tr", ""), pkg.get("name_en", ""))
            )
            imported += 1

        return imported

    def export_to_json(self, output_path: str):
        data = {
            "export_date": datetime.now().isoformat(),
            "version": APP_VERSION,
            "companies": [],
            "drivers": [],
            "vehicles": [],
            "shipments": []
        }

        for row in self.execute("SELECT * FROM companies", ()):
            data["companies"].append(dict(row))
        for row in self.execute("SELECT * FROM drivers", ()):
            data["drivers"].append(dict(row))
        for row in self.execute("SELECT * FROM vehicles", ()):
            data["vehicles"].append(dict(row))
        for row in self.execute("SELECT * FROM shipments", ()):
            data["shipments"].append(dict(row))

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        return output_path


# =============================================================================
# ADR ENGINE - MEVZUAT MOTORU
# =============================================================================


# =============================================================================
# ADR 1.10.3 — EMNİYET PLANI HESAPLAMA MOTORU
# =============================================================================

class SecurityPlanEngine:
    """
    ADR 1.10.3 kapsamında emniyet planı gereksinimi hesaplar.

    Karar hiyerarşisi (ADR 2023 Bölüm 1.10):
    ┌─────────────────────────────────────────────────────────────────┐
    │ 1. Madde 1.10.4 listesinde mi?                                 │
    │    EVET → 1.1.3.6 muafiyeti KULLANAMAZ, devam et              │
    │    HAYIR → Toplam puan < 1000 ise (1.1.3.6) emniyet planı YOK │
    ├─────────────────────────────────────────────────────────────────┤
    │ 2. Sınıf 7 radyoaktif? → Tablo 1.10.3.1.3 eşik hesabı         │
    ├─────────────────────────────────────────────────────────────────┤
    │ 3. Diğer sınıflar → Tablo 1.10.3.1.2 eşik karşılaştırması     │
    │    'a' = Bu taşıma modu için uygulanamaz (muaf sayılır)        │
    │    'b' = Miktar ne olursa olsun muaf                           │
    │    0   = Herhangi bir miktarda EMNİYET PLANI GEREKLİ           │
    │    N   = Miktar > N litre/kg → EMNİYET PLANI GEREKLİ          │
    └─────────────────────────────────────────────────────────────────┘
    """

    # ── ADR 1.10.4: Bu UN numaraları 1.1.3.6 muafiyetini KULLANAMAZ ──────
    # Kaynak: ADR 2023, Madde 1.10.4 (açıkça sayılan UN numaraları)
    EXCLUDED_FROM_1136: frozenset = frozenset({
        29, 30, 59, 65, 73, 104, 237, 255, 267, 288, 289, 290,
        360, 361, 364, 365, 366, 439, 440, 441, 455, 456, 500
    })

    # ── Tablo 1.10.3.1.2 ─────────────────────────────────────────────────
    # key: (sinif, kosul)
    # value: (tank_litre, dokme_kg, ambalaj_kg)
    # None = 'a' (uygulanamaz), -1 = 'b' (muaf)
    THRESHOLDS: dict = {
        # Sınıf 1
        ("1", "1.1"):              (None, None, 0),
        ("1", "1.2"):              (None, None, 0),
        ("1", "1.3C"):             (None, None, 0),
        ("1", "1.4_special"):      (None, None, 0),   # özel UN listesi (aşağıda)
        ("1", "1.5"):              (0,    None, 0),
        # Sınıf 2 — alevlenebilir gazlar (F, FC sınıf kodu)
        ("2", "F"):                (3000, None, -1),
        ("2", "FC"):               (3000, None, -1),
        # Sınıf 2 — zehirli gazlar (T, TF, TC, TO, TFC, TOC)
        ("2", "T"):                (0,    None, 0),
        ("2", "TF"):               (0,    None, 0),
        ("2", "TC"):               (0,    None, 0),
        ("2", "TO"):               (0,    None, 0),
        ("2", "TFC"):              (0,    None, 0),
        ("2", "TOC"):              (0,    None, 0),
        # Sınıf 3
        ("3", "PGI"):              (3000, None, -1),
        ("3", "PGII"):             (3000, None, -1),
        ("3", "desensitized"):     (0,    None, 0),
        # Sınıf 4.1
        ("4.1", "desensitized"):   (None, None, 0),
        # Sınıf 4.2
        ("4.2", "PGI"):            (3000, None, -1),
        # Sınıf 4.3
        ("4.3", "PGI"):            (3000, None, -1),
        # Sınıf 5.1
        ("5.1", "PGI_liquid"):     (3000, None, -1),
        ("5.1", "perchlorate_AN"): (3000, 3000, -1),
        # Sınıf 6.1
        ("6.1", "PGI"):            (0,    None, 0),
        # Sınıf 6.2
        ("6.2", "catA"):           (None, 0,    0),
        # Sınıf 8
        ("8", "PGI"):              (3000, None, -1),
    }

    # UN 1.4'teki özel tehlikeli numara listesi (ADR Tablo 1.10.3.1.2: 0104,0237,0255,0267,0289,0361,0365,0366,0440,0441,0455,0456,0500,0512,0513)
    CLASS14_HIGH_CONSEQUENCE: frozenset = frozenset({
        104, 237, 255, 267, 289, 361, 365, 366, 440, 441, 455, 456, 500, 512, 513
    })

    # Amonyum nitrat / perklorat UN numaraları (5.1 özel satır)
    PERCHLORATE_AN_UN: frozenset = frozenset({
        1942, 2067, 2068, 2069, 2070, 2426, 3375,   # AN / AN gübre
        1481, 1482, 1483, 3506,                       # Perklorat
    })

    # UN 6.2 Kategori A: UN 2814 (insan pat.), UN 2900 (hayvansal pat.), UN 3549 (tibbi atik Kat.A)
    UN_6_2_CAT_A: frozenset = frozenset({2814, 2900, 3549})

    # ── Tablo 1.10.3.1.3 — Radyonüklid eşik değerleri (TBq) ─────────────
    RADIONUCLIDE_THRESHOLDS: dict = {
        "Am-241": 0.6,  "Au-198": 2,    "Cd-109": 200,  "Cf-252": 0.2,
        "Cm-244": 0.5,  "Co-57":  7,    "Co-60":  0.3,  "Cs-137": 1,
        "Fe-55":  8000, "Ge-68":  7,    "Gd-153": 10,   "Ir-192": 0.8,
        "Ni-63":  600,  "Pd-103": 900,  "Pm-147": 400,  "Po-210": 0.6,
        "Pu-238": 0.6,  "Pu-239": 0.6,  "Ra-226": 0.4,  "Ru-106": 3,
        "Se-75":  2,    "Sr-90":  10,   "Tl-204": 200,  "Tm-170": 200,
        "Yb-169": 3,
    }
    CLASS7_GENERIC_THRESHOLD_TBq: float = 3000.0   # Tabloda olmayan nüklidler

    # ─────────────────────────────────────────────────────────────────────
    # YARDIMCI: Bir ShipmentItem için Tablo 1.10.3.1.2 anahtarını çıkar
    # ─────────────────────────────────────────────────────────────────────
    @classmethod
    def _get_table_key(cls, item: "ShipmentItem") -> str:
        """
        ShipmentItem bilgilerinden Tablo 1.10.3.1.2 satır anahtarını üretir.
        Döndürülen değer THRESHOLDS dict'indeki 'kosul' parçasıdır.
        """
        cls_raw_parts = (item.class_code or "").strip().split()
        cls_raw = cls_raw_parts[0] if cls_raw_parts else ""
        # [GUVENLIK] Bazı içe aktarılan kayıtlarda paketleme grubu Arap
        # rakamıyla ("1","2","3") gelebiliyor; PG karşılaştırmaları güvenlik
        # planı kararını etkilediğinden Roma rakamına normalize edilir.
        pg_raw  = (item.packing_group or "").strip().upper()
        pg      = {"1": "I", "2": "II", "3": "III"}.get(pg_raw, pg_raw)
        un_int  = 0
        try:
            un_int = int((item.un_number or "").strip().lstrip("0") or "0")
        except ValueError:
            pass

        # Sınıf 1 alt grupları
        # ADR Tablo 1.10.3.1.2: "Sınıf" sütununda 1, "Sınıflandırma kodu" sütununda 1.4S / 1.3C vb.
        # class_code: "1", "1.1"…"1.6" | classification_code: "1.4S", "1.3C", "1.1A" vb.
        if cls_raw.startswith("1.") or cls_raw == "1":
            import re as _re
            cc_full = (getattr(item, "classification_code", "") or "").strip().upper()
            # classification_code doluysa alt sınıf + uyumluluk grubunu buradan al
            if cc_full:
                m = _re.match(r"(1\.[0-9])([A-Z]*)$", cc_full)
                if m:
                    sub_cc = m.group(1)   # "1.4"
                    compat = m.group(2)   # "S", "B", "G", "C", ""
                    if sub_cc == "1.4":
                        if compat == "S":
                            return None       # 1.4S → emniyet planı gerektirmez
                        if un_int in cls.CLASS14_HIGH_CONSEQUENCE:
                            return "1.4_special"
                        return None           # diğer 1.4 → muaf
                    if sub_cc == "1.3":
                        return "1.3C" if compat == "C" else None   # sadece 1.3C kapsam içi
                    if sub_cc in ("1.1", "1.2"):
                        return sub_cc
                    if sub_cc == "1.5":
                        return "1.5"
                    return None
            # classification_code boşsa class_code'a göre değerlendir (eski kayıtlar)
            sub = cls_raw
            if sub == "1.4":
                if un_int in cls.CLASS14_HIGH_CONSEQUENCE:
                    return "1.4_special"
                return None
            if sub in ("1.1", "1.2"):
                return sub
            if sub == "1.3":
                return "1.3C"   # muhafazakâr: tüm 1.3 kapsam içi sayılır
            if sub == "1.5":
                return "1.5"
            return None   # 1.6 ve diğerleri → muaf

        # Sınıf 2
        if cls_raw == "2":
            # classification_code alanı yoksa class_code'da bazen yazıyor
            # önce special_provisions/notes içinde ara, yoksa default F
            cc = getattr(item, "classification_code", "").strip().upper()
            if not cc:
                # notes veya proper_name içinde geçen harf kombinasyonu
                search_in = (item.notes or "") + " " + (item.proper_name or "")
                for code in ("TOC","TFC","TF","TC","TO","FC","T","F"):
                    if code in search_in.upper():
                        cc = code
                        break
            # Aerosol muafiyeti: ADR 1.10.3.1.2 — sınıflandırma kodunda A harfi olan gazlar kapsam dışı
            if "A" in cc:
                return None   # aerosol → tablo dışı
            if cc in ("T","TF","TC","TO","TFC","TOC"):
                return cc
            # Alevlenebilir (F, FC) veya bilinmiyor → muhafazakâr: F
            return cc if cc in ("F","FC") else "F"

        # Sınıf 3
        if cls_raw == "3":
            # Duyarsızlaştırılmış patlayıcı UN numaraları (seçili liste)
            desens = {2852,3343,3357,3379,3380,3474,3475,3476,3477,3478,3479}
            if un_int in desens:
                return "desensitized"
            # ADR Tablo 1.10.3.1.2: Sınıf 3 için yalnızca PGI ve PGII yer alır.
            # PG III ürünler (örn. UN 1202 mazot) tablo kapsamı dışındadır → muaf.
            if pg == "I":
                return "PGI"
            if pg == "II":
                return "PGII"
            return None   # PG III veya belirsiz → tabloda yok → muaf

        # Sınıf 4.1
        if cls_raw == "4.1":
            # Duyarsızlaştırılmış patlayıcılar (bazı 4.1 maddeleri)
            desens_41 = {2555,2556,2557,3317,3319,3344,3380,3474,3475,3476,3477}
            if un_int in desens_41:
                return "desensitized"
            return None   # diğer 4.1 → tabloda yok

        # Sınıf 4.2
        if cls_raw == "4.2":
            return "PGI" if pg == "I" else None

        # Sınıf 4.3
        if cls_raw == "4.3":
            return "PGI" if pg == "I" else None

        # Sınıf 5.1
        if cls_raw == "5.1":
            if un_int in cls.PERCHLORATE_AN_UN:
                return "perchlorate_AN"
            return "PGI_liquid" if pg == "I" else None

        # Sınıf 6.1
        if cls_raw == "6.1":
            return "PGI" if pg == "I" else None

        # Sınıf 6.2
        if cls_raw == "6.2":
            if un_int == 2900:
                # UN 2900: hayvansal patojenlerde Kategori A → kapsam içi.
                # Kullanici "hayvansal malzeme hariç" istisnasini bilmeli;
                # kod muhafazakar davranarak catA döndürür.
                return "catA"
            if un_int in cls.UN_6_2_CAT_A:   # 2814, 3549
                return "catA"
            return None

        # Sınıf 8
        if cls_raw == "8":
            return "PGI" if pg == "I" else None

        return None

    # ─────────────────────────────────────────────────────────────────────
    # ANA HESAPLAMA FONKSİYONU
    # ─────────────────────────────────────────────────────────────────────
    @classmethod
    def check(
        cls,
        items: List["ShipmentItem"],
        transport_mode: str = "ambalaj",       # "tank" | "dokme" | "ambalaj"
        total_1136_points: float = None,       # zaten hesaplanmışsa ver
        radionuclides: dict = None,            # {nüklid: Ai_TBq} Sınıf 7 için
    ) -> dict:
        """
        ADR 1.10.3 emniyet planı gereksinim sonucu döndürür.

        Dönüş:
        {
          "required":   bool,
          "exempt":     bool,          # 1.10.4 / 1.1.3.6 muafiyeti
          "reasons":    [str],         # gereklilik sebepleri
          "details":    [str],         # tüm kalem detayları
          "class7_ratio": float|None,  # ∑(Ai/Ti) Sınıf 7 için
        }
        """
        result = {
            "required": False, "exempt": False,
            "reasons": [], "details": [], "class7_ratio": None
        }
        if not items:
            result["details"].append("Sevkiyat boş — değerlendirme yapılamadı.")
            return result

        mode_idx = {"tank": 0, "dokme": 1, "ambalaj": 2}.get(transport_mode, 2)

        # Ambalaj türü → taşıma modu eşleme yardımcısı
        # Her kalemin kendi packaging_type'ından mode_idx otomatik belirlenir.
        def _mode_idx_for_item(item) -> int:
            pt = (getattr(item, "packaging_type", "") or "").strip().lower()
            if pt == "tank":
                return 0   # tank sütunu
            if pt in ("dökme", "dokme"):
                return 1   # dökme sütunu
            return 2       # ambalaj sütunu (IBC, Varil, Bidon, Kutu, Çuval, Kompozit …)

        # ── 1. 1.1.3.6 Muafiyet Kontrolü (1.10.4) ────────────────────────
        # Eğer hiçbir kalem 1.10.4 listesindeyse VE toplam puan < 1000 ise → muaf
        has_excluded_un = any(
            cls._un_int(item) in cls.EXCLUDED_FROM_1136 for item in items
        )
        pts = total_1136_points if total_1136_points is not None else 0.0
        if not has_excluded_un and pts < 1000:
            result["exempt"] = True
            result["details"].append(
                f"✅ ADR 1.10.4 muafiyeti: Toplam 1.1.3.6 puanı {pts:.0f} < 1000 "
                f"ve listede kısıtlı UN yok → Emniyet planı GEREKMİYOR."
            )
            return result

        if has_excluded_un:
            result["details"].append(
                "⚠ 1.10.4 özel liste: Bazı UN numaraları 1.1.3.6 muafiyetinden yararlanamaz."
            )

        # ── 2. Sınıf 7 — Radyonüklid eşik hesabı ─────────────────────────
        class7_items = [i for i in items if (i.class_code or "").startswith("7")]
        if class7_items:
            ratio = cls._calc_class7_ratio(radionuclides or {})
            result["class7_ratio"] = ratio
            if ratio >= 1.0:
                result["required"] = True
                result["reasons"].append(
                    f"Sınıf 7: Radyoaktif eşik aşıldı (∑Ai/Ti = {ratio:.3f} ≥ 1)"
                )
            else:
                result["details"].append(
                    f"Sınıf 7: Eşik altında (∑Ai/Ti = {ratio:.3f} < 1) → bu sınıf için muaf"
                )

        # ── 3. Tablo 1.10.3.1.2 — Sınıf 1-8 ─────────────────────────────
        for item in items:
            _cls_parts = (item.class_code or "").strip().split()
            cls_raw = _cls_parts[0] if _cls_parts else ""
            if cls_raw.startswith("7"):
                continue   # Sınıf 7 zaten üstte işlendi

            un_str  = item.un_number or "—"
            qty_kg  = float(item.net_quantity or 0)
            # litre → kg dönüşüm gerektiren birimler için yaklaşık 1:1
            qty_l   = qty_kg   # tanklar litreyle kısıtlanır; basitlik için eşit

            # Her kalemin ambalaj türünden taşıma modu belirlenir
            item_mode_idx = _mode_idx_for_item(item)
            item_mode_str = {0: "Tank", 1: "Dökme", 2: "Ambalaj"}[item_mode_idx]

            key = cls._get_table_key(item)
            if key is None:
                result["details"].append(
                    f"  UN{un_str} ({cls_raw} PG{item.packing_group}) [{item_mode_str}]: "
                    f"Tablo 1.10.3.1.2 kapsamı dışında → muaf"
                )
                continue

            # Patlayicilar icin THRESHOLDS anahtari ("1", alt_grup) seklinde
            # cls_raw "1.1", "1.2" vb. olabilir → ana sinif "1" olarak normalize et
            threshold_cls = "1" if cls_raw.startswith("1.") else cls_raw
            threshold_row = cls.THRESHOLDS.get((threshold_cls, key))
            if threshold_row is None:
                result["details"].append(
                    f"  UN{un_str} [{item_mode_str}]: Tablo satırı bulunamadı → muaf kabul edildi"
                )
                continue

            limit = threshold_row[item_mode_idx]

            if limit is None:           # 'a': bu mod için uygulanamaz
                result["details"].append(
                    f"  UN{un_str} ({cls_raw}): {item_mode_str} modu için uygulanamaz (a) → muaf"
                )
            elif limit == -1:           # 'b': miktar ne olursa muaf
                result["details"].append(
                    f"  UN{un_str} ({cls_raw}): Miktar ne olursa olsun muaf (b) [{item_mode_str}]"
                )
            elif limit == 0:            # Her miktarda gerekli
                result["required"] = True
                result["reasons"].append(
                    f"UN{un_str} ({cls_raw} PG{item.packing_group}) [{item_mode_str}]: "
                    f"Herhangi bir miktarda emniyet planı gerekli (limit=0)"
                )
            else:                       # Eşik karşılaştırması
                val = qty_l if item_mode_idx == 0 else qty_kg
                unit_str = "litre" if item_mode_idx == 0 else "kg"
                if val > limit:
                    result["required"] = True
                    result["reasons"].append(
                        f"UN{un_str} ({cls_raw} PG{item.packing_group}) [{item_mode_str}]: "
                        f"{val:.0f} {unit_str} > {limit} {unit_str} eşiği aşıldı"
                    )
                else:
                    result["details"].append(
                        f"  UN{un_str} [{item_mode_str}]: {val:.0f} {unit_str} ≤ {limit} {unit_str} → muaf"
                    )

        if not result["required"] and not result["exempt"]:
            result["exempt"] = True
            result["details"].append("✅ Tüm kalemler eşik altında — Emniyet planı GEREKMİYOR.")

        return result

    # ─────────────────────────────────────────────────────────────────────
    # YARDIMCI METODlar
    # ─────────────────────────────────────────────────────────────────────
    @staticmethod
    def _un_int(item: "ShipmentItem") -> int:
        try:
            return int((item.un_number or "").strip().lstrip("0") or "0")
        except ValueError:
            return 0

    @classmethod
    def _calc_class7_ratio(cls, radionuclides: dict) -> float:
        """
        ADR 1.10.3.1.4 formülü: ∑(Ai / Ti)
        radionuclides = {"Cs-137": 1.5, "Co-60": 0.4, ...}  (TBq cinsinden)
        """
        total = 0.0
        for nuclide, activity_tbq in radionuclides.items():
            threshold = cls.RADIONUCLIDE_THRESHOLDS.get(nuclide,
                            cls.CLASS7_GENERIC_THRESHOLD_TBq)
            total += activity_tbq / threshold
        return total

    @classmethod
    def generate_plan_template(cls, shipment: "Shipment" = None,
                                items: List["ShipmentItem"] = None,
                                sample: bool = False) -> str:
        """
        ADR 1.10.3.2.2 gereksinimlerine uygun emniyet planı HTML şablonu üretir.
        sample=True → gerçekçi doldurulmuş örnek döner.
        """
        from datetime import datetime as _dt
        now    = _dt.now().strftime("%d.%m.%Y")
        doc_no = getattr(shipment, "document_no", "—") if shipment else "—"

        if sample:
            # ── Doldurulmuş örnek (TMGD perspektifli, UN 1203 Benzin / Tank taşıma) ──
            ep_no = "EP-" + _dt.now().strftime("%Y") + "-001"
            return f"""<!DOCTYPE html><html lang="tr"><head>
<meta charset="UTF-8">
<style>
  body{{font-family:Arial,Helvetica,sans-serif;font-size:10pt;margin:20mm;color:#1a1a1a;line-height:1.5;}}
  h1{{font-size:14pt;border-bottom:2pt solid #1E3A5F;color:#1E3A5F;padding-bottom:4px;margin-bottom:6px;}}
  h2{{font-size:11pt;color:#1E3A5F;margin-top:14px;margin-bottom:4px;}}
  table{{border-collapse:collapse;width:100%;margin:8px 0;}}
  th{{background:#1E3A5F;color:#fff;padding:5px 8px;text-align:left;font-size:9pt;}}
  td{{border:1px solid #ccc;padding:5px 8px;font-size:9pt;vertical-align:top;}}
  .section{{border:1px solid #1E3A5F;border-radius:3px;padding:8px 12px;margin-bottom:12px;}}
  .note{{font-size:8pt;color:#555;font-style:italic;}}
  .sig-box{{border:1px solid #aaa;min-height:80px;padding:8px;}}
  ol{{margin:6px 0 0 16px;padding:0;}} li{{margin-bottom:4px;}}
</style>
</head><body>

<h1>ADR EMNİYET PLANI — ADR 1.10.3.2</h1>
<p style="font-size:9pt;color:#555;">
  <strong>Belge No: {ep_no}</strong> &nbsp;|&nbsp; Düzenleme Tarihi: <strong>{now}</strong>
  &nbsp;|&nbsp; ADR Yönetmeliği Madde 1.10.3.2.2 kapsamında hazırlanmıştır.
</p>
<p class="note" style="background:#FEF9E7;padding:6px 10px;border-left:3px solid #F4D03F;border-radius:0 3px 3px 0;">
  ⚠ Bu belge gerçekçi bir örnek şablondur. Lütfen kendi firma, kişi ve güzergah
  bilgilerinizle güncelleyiniz.
</p>

<div class="section">
<h2>(a) Emniyet Sorumluluğu Dağılımı</h2>
<table>
  <tr><th>Görev</th><th>Sorumlu Kişi / Birim</th><th>İletişim</th></tr>
  <tr><td><strong>Emniyet Koordinatörü</strong></td><td>Selim Kaya</td><td>0532 XXX XX 01 · skaya@firma.com.tr</td></tr>
  <tr><td><strong>Gönderici Yetkilisi</strong></td><td>Mustafa Usta</td><td>0533 XXX XX 02 · musta@firma.com.tr</td></tr>
  <tr><td><strong>Taşıyıcı Yetkilisi</strong></td><td>Nizamettin Üstündağ</td><td>0530 XXX XX 03 · nizamettin@lojistik.com</td></tr>
  <tr><td><strong>Alıcı Yetkilisi</strong></td><td>Cem Sever</td><td>0555 XXX XX 04 · csever@alici.com.tr</td></tr>
</table>
</div>

<div class="section">
<h2>(b) Taşınan Tehlikeli Mal Kayıtları</h2>
<table>
  <tr><th>UN No</th><th>Uygun Sevkiyat Adı</th><th>Sınıf</th><th>PG</th><th>Miktar</th></tr>
  <tr><td><strong>UN 1203</strong></td><td>BENZİN veya GAZOLİN veya PETROL</td><td>3</td><td>II</td><td>3 050 litre</td></tr>
</table>
</div>

<div class="section">
<h2>(c) Emniyet Riski Değerlendirmesi</h2>
<p class="note">Duraklamalar, depolama, aktarma noktaları ve güzergah riskleri:</p>
<ul style="margin:6px 0 0 16px;padding:0;font-size:9pt;">
  <li><strong>Malzeme Karakteristiği:</strong> UN 1203, düşük parlama noktasına sahip, yüksek derecede yanıcı bir sıvıdır.
      Sabotaj veya terör eylemleri durumunda kitle yangın çıkarma aracı olarak kötüye kullanılma riski yüksektir (Yüksek Ciddi Risk).</li>
  <li><strong>Güzergah Riskleri:</strong> Şehir içi yoğun trafik, tünel ve köprü geçişlerinde saldırı ya da gasp girişimleri.</li>
  <li><strong>Duraklama / Mola Riskleri:</strong> Sürücünün zorunlu molaları esnasında (takograf süreleri), güvenliği sağlanmamış
      tesislerde aracın gözetimsiz kalması ve vana kapaklarının zorlanması riski.</li>
  <li><strong>Aktarma Noktaları:</strong> Dolum ve boşaltım tesislerinde yetkisiz personelin sahaya girmesi ve operasyona müdahalesi.</li>
</ul>
</div>

<div class="section">
<h2>(d) Riski Azaltma Önlemleri</h2>
<table>
  <tr><th style="width:22%;">Önlem</th><th>Uygulama Detayı</th></tr>
  <tr>
    <td><strong>Eğitim (ADR 1.3)</strong></td>
    <td>Tüm personel ADR Bölüm 1.3 ve 1.10 kapsamında Emniyet Farkındalık Eğitimi almıştır.
    Eğitim kayıtları İK departmanında saklanmaktadır. Sürücü geçerli SRC-5 (Tank) belgesine sahiptir.</td>
  </tr>
  <tr>
    <td><strong>Emniyet Politikası</strong></td>
    <td>Araca görevli personel dışında yolcu alınması yasaktır. Araç terk edildiğinde kapılar kilitlenmeli,
    kontak anahtarı araç üzerinde bırakılmamalıdır.</td>
  </tr>
  <tr>
    <td><strong>Güzergah Seçimi</strong></td>
    <td>Yalnızca Emniyet Koordinatörü tarafından onaylanmış ana arterler (otoyollar) kullanılır.
    Molalar aydınlatılmış, kameralı ve güvenli tır parklarında verilir.</td>
  </tr>
  <tr>
    <td><strong>Erişim Kontrolü</strong></td>
    <td>Tank menhol kapakları ve boşaltım vanaları mühürlü ve kilitlidir.
    Yükleme/boşaltma tesislerinde kimlik ibrazı ve kartlı geçiş zorunludur.</td>
  </tr>
  <tr>
    <td><strong>İzleme Teçhizatı</strong></td>
    <td>Araçta 7/24 GPS araç takip sistemi ve kabin içi Panik Butonu bulunmaktadır.
    Rota dışına çıkıldığında merkeze anında uyarı gönderilmektedir.</td>
  </tr>
</table>
</div>

<div class="section">
<h2>(e) Olay Raporlama Prosedürü</h2>
<p class="note">Emniyet ihlali veya olaylar için bildirim zinciri:</p>
<ol style="font-size:9pt;">
  <li><strong>Sürücü:</strong> Tehlike sezinlediğinde aracı güvenli şekilde terk eder/kilitler ve Panik Butonuna basar.</li>
  <li><strong>Kolluk Kuvvetleri:</strong> Sürücü derhal <strong>112</strong> Acil Çağrı Merkezi'ni arayarak konum ve
      yük bilgisini (UN 1203 – Benzin) bildirir.</li>
  <li><strong>Emniyet Koordinatörü:</strong> Sürücü eş zamanlı olarak Selim Kaya'yı bilgilendirir.</li>
  <li><strong>Kurum İçi Kriz Masası:</strong> Koordinatör; Gönderici, Taşıyıcı ve TMGD'ye acil kodla bilgi verir.
      Olay Ulaştırma ve Altyapı Bakanlığı'na yazılı olarak raporlanır.</li>
</ol>
</div>

<div class="section">
<h2>(f) Plan Gözden Geçirme</h2>
<table>
  <tr><th>Gözden Geçirme Tarihi</th><th>Yapılan Değişiklik</th><th>Onaylayan</th></tr>
  <tr><td>{now}</td><td>İlk taslağın oluşturulması ve güzergah analizlerinin eklenmesi.</td><td>Selim Kaya</td></tr>
  <tr><td></td><td></td><td></td></tr>
</table>
</div>

<div class="section">
<h2>İmza ve Onay</h2>
<table>
  <tr>
    <td style="width:50%;padding:12px;vertical-align:top;">
      <strong>Gönderici / Yükleyen Yetkilisi</strong><br><br>
      <span style="font-size:9pt;color:#555;">Ad Soyad:</span><br>
      <div class="sig-box">&nbsp;</div>
      <span style="font-size:9pt;color:#555;">İmza / Kaşe:</span><br>
      <div class="sig-box" style="min-height:100px;">&nbsp;</div>
      <span style="font-size:9pt;color:#555;">Tarih: {now}</span>
    </td>
    <td style="width:50%;padding:12px;vertical-align:top;border-left:1px solid #ccc;">
      <strong>Taşıyıcı Yetkilisi</strong><br><br>
      <span style="font-size:9pt;color:#555;">Ad Soyad:</span><br>
      <div class="sig-box">&nbsp;</div>
      <span style="font-size:9pt;color:#555;">İmza / Kaşe:</span><br>
      <div class="sig-box" style="min-height:100px;">&nbsp;</div>
      <span style="font-size:9pt;color:#555;">Tarih: {now}</span>
    </td>
  </tr>
</table>
</div>

<p class="note" style="margin-top:16px;">
  Bu plan ADR 1.10.3.2.2 maddelerini (a)–(f) karşılamaktadır.
  Periyodik gözden geçirme zorunludur (ADR 1.10.3.2.2/f).
</p>
</body></html>"""

        madde_listesi = ""
        if items:
            for it in items:
                madde_listesi += (
                    f"<tr><td>UN{it.un_number}</td><td>{it.proper_name}</td>"
                    f"<td>{it.class_code}</td><td>{it.packing_group}</td>"
                    f"<td>{it.net_quantity} {it.unit}</td></tr>"
                )
        else:
            madde_listesi = '<tr><td colspan="5">—</td></tr>'

        return f"""<!DOCTYPE html><html lang="tr"><head>
<meta charset="UTF-8">
<style>
  body{{font-family:Arial,Helvetica,sans-serif;font-size:10pt;margin:20mm;color:#1a1a1a;}}
  h1{{font-size:14pt;border-bottom:2pt solid #1E3A5F;color:#1E3A5F;padding-bottom:4px;}}
  h2{{font-size:11pt;color:#1E3A5F;margin-top:14px;}}
  table{{border-collapse:collapse;width:100%;margin:8px 0;}}
  th{{background:#1E3A5F;color:#fff;padding:5px 8px;text-align:left;font-size:9pt;}}
  td{{border:1px solid #ccc;padding:4px 8px;font-size:9pt;}}
  .field{{border-bottom:1px solid #999;min-height:18px;display:block;margin:2px 0 8px;}}
  .req{{color:#c0392b;font-weight:bold;}}
  .note{{font-size:8pt;color:#555;font-style:italic;}}
  .section{{border:1px solid #1E3A5F;border-radius:3px;padding:8px 12px;margin-bottom:10px;}}
</style>
</head><body>

<h1>ADR EMNİYET PLANI — ADR 1.10.3.2</h1>
<p style="font-size:9pt;color:#555;">
  Belge No: <strong>{doc_no}</strong> &nbsp;|&nbsp; Düzenleme Tarihi: <strong>{now}</strong>
  &nbsp;|&nbsp; ADR Yönetmeliği Madde 1.10.3.2.2 kapsamında hazırlanmıştır.
</p>

<div class="section">
<h2>(a) Emniyet Sorumluluğu Dağılımı</h2>
<table>
  <tr><th>Görev</th><th>Sorumlu Kişi / Birim</th><th>İletişim</th></tr>
  <tr><td>Emniyet koordinatörü</td><td class="field">&nbsp;</td><td class="field">&nbsp;</td></tr>
  <tr><td>Gönderici yetkilisi</td><td class="field">&nbsp;</td><td class="field">&nbsp;</td></tr>
  <tr><td>Taşıyıcı yetkilisi</td><td class="field">&nbsp;</td><td class="field">&nbsp;</td></tr>
  <tr><td>Alıcı yetkilisi</td><td class="field">&nbsp;</td><td class="field">&nbsp;</td></tr>
</table>
</div>

<div class="section">
<h2>(b) Taşınan Tehlikeli Mal Kayıtları</h2>
<table>
  <tr><th>UN No</th><th>Uygun Sevkiyat Adı</th><th>Sınıf</th><th>PG</th><th>Miktar</th></tr>
  {madde_listesi}
</table>
</div>

<div class="section">
<h2>(c) Emniyet Riski Değerlendirmesi</h2>
<p class="note">Duraklamalar, depolama, aktarma noktaları ve güzergah riskleri:</p>
<span class="field">&nbsp;<br>&nbsp;<br>&nbsp;</span>
</div>

<div class="section">
<h2>(d) Riski Azaltma Önlemleri</h2>
<table>
  <tr><th>Önlem</th><th>Uygulama Detayı</th></tr>
  <tr><td>Eğitim (ADR 1.3)</td><td class="field">&nbsp;</td></tr>
  <tr><td>Emniyet politikası</td><td class="field">&nbsp;</td></tr>
  <tr><td>Güzergah seçimi</td><td class="field">&nbsp;</td></tr>
  <tr><td>Erişim kontrolü</td><td class="field">&nbsp;</td></tr>
  <tr><td>İzleme teçhizatı</td><td class="field">&nbsp;</td></tr>
</table>
</div>

<div class="section">
<h2>(e) Olay Raporlama Prosedürü</h2>
<p class="note">Emniyet ihlali veya olaylar için bildirim zinciri:</p>
<span class="field">&nbsp;<br>&nbsp;</span>
</div>

<div class="section">
<h2>(f) Plan Gözden Geçirme</h2>
<table>
  <tr><th>Gözden Geçirme Tarihi</th><th>Yapılan Değişiklik</th><th>Onaylayan</th></tr>
  <tr><td class="field">&nbsp;</td><td class="field">&nbsp;</td><td class="field">&nbsp;</td></tr>
  <tr><td class="field">&nbsp;</td><td class="field">&nbsp;</td><td class="field">&nbsp;</td></tr>
</table>
</div>

<div class="section">
<h2>İmza ve Onay</h2>
<table>
  <tr>
    <td style="width:50%;padding:12px;">
      Gönderici / Yükleyen<br>
      Ad Soyad: <span class="field">&nbsp;</span>
      Tarih: <span class="field">&nbsp;</span>
      İmza / Kaşe: <span class="field">&nbsp;<br>&nbsp;</span>
    </td>
    <td style="width:50%;padding:12px;">
      Taşıyıcı Yetkilisi<br>
      Ad Soyad: <span class="field">&nbsp;</span>
      Tarih: <span class="field">&nbsp;</span>
      İmza / Kaşe: <span class="field">&nbsp;<br>&nbsp;</span>
    </td>
  </tr>
</table>
</div>

<p class="note" style="margin-top:16px;">
  Bu plan ADR 1.10.3.2.2 maddelerini (a)–(h) karşılamaktadır.
  Periyodik gözden geçirme zorunludur (ADR 1.10.3.2.2/f).
</p>
</body></html>"""


    # ─────────────────────────────────────────────────────────────────────
    # [v4.7] STATİK ENVANTER TARAMASI (Güvenlik Planı İnceleme Raporu)
    # ─────────────────────────────────────────────────────────────────────
    # Yukarıdaki check() metodu belirli bir SEVKİYATIN kalemlerini (miktar +
    # taşıma modu dahil) değerlendirir. Bu metod ise firmanın TÜM KİMYASAL
    # ENVANTERİNİ (belirli bir sevkiyattan bağımsız), yalnızca sınıf/PG/
    # sınıflandırma kodu bilgisine göre "Tablo 1.10.3.1.2 kapsamına
    # girip girmediğini" tarar — tıpkı örnek "Güvenlik Planı İnceleme
    # Raporu"nda olduğu gibi. Miktar burada değerlendirilmez çünkü envanter
    # taraması amacı, maddenin TÜRÜNE göre kapsam potansiyelini belirlemektir;
    # ambalaj sütunu ADR'de yalnızca 0 (her miktarda kapsam içi) veya b
    # (miktar ne olursa olsun muaf) değerlerini alır.
    @classmethod
    def screen_inventory_chemical(cls, chemical) -> dict:
        """Tek bir kimyasalı Tablo 1.10.3.1.2 kapsamında statik olarak
        değerlendirir. Dönüş: in_scope (True/False/"conditional"),
        table_key, sonuc_text (örnek rapordaki gibi Türkçe açıklama)."""
        from types import SimpleNamespace

        def _field(obj, key, default=""):
            """Chemical dataclass (getattr) VEYA sqlite3.Row / dict
            (anahtar erişimi) nesnelerinden güvenle alan okur. sqlite3.Row
            getattr ile ÇALIŞMAZ (sessizce default döner) — bu yüzden ham
            satır objelerinin yanlışlıkla boş veri üretmesini önler."""
            val = getattr(obj, key, None)
            if val not in (None, ""):
                return val
            try:
                if key in obj.keys():
                    return obj[key]
            except (AttributeError, TypeError):
                pass
            return default

        un_raw = str(_field(chemical, "un_number", "") or "")
        cls_raw_full = str(_field(chemical, "class_code", "") or "").strip()
        pg = str(_field(chemical, "packing_group", "") or "").strip().upper()
        cc = str(_field(chemical, "classification_code", "") or "").strip().upper()
        name = (_field(chemical, "proper_shipping_name_tr", "") or
                _field(chemical, "name", "") or
                _field(chemical, "proper_shipping_name_en", "") or "")

        fake_item = SimpleNamespace(
            class_code=cls_raw_full, packing_group=pg,
            classification_code=cc, un_number=un_raw,
            notes="", proper_name=name)

        base = {
            "un_number": un_raw, "class_code": cls_raw_full,
            "packing_group": pg, "classification_code": cc, "name": name,
        }

        key = cls._get_table_key(fake_item)
        if key is None:
            base.update(in_scope=False, table_key=None, sonuc_text=(
                f"Sınıf {cls_raw_full or chr(0x2014)}; bu madde ADR Tablo 1.10.3.1.2 "
                f"kapsamındaki sınıf/ambalajlama grubu kombinasyonlarından "
                f"hiçbirine girmemektedir. Bu nedenle güvenlik planı "
                f"hazırlanmasına gerek yoktur."))
            return base

        threshold_cls = "1" if cls_raw_full.startswith("1.") else cls_raw_full
        row = cls.THRESHOLDS.get((threshold_cls, key))
        if row is None:
            base.update(in_scope=False, table_key=key, sonuc_text=(
                f"Sınıf {cls_raw_full}; Tablo 1.10.3.1.2'de bu sınıf/kategori "
                f"için tanımlı bir satır bulunamadı. Bu nedenle güvenlik "
                f"planı hazırlanmasına gerek yoktur."))
            return base

        package_limit = row[2]

        descriptions = {
            ("1", "1.1"): "Sınıf 1.1 Patlayıcılar",
            ("1", "1.2"): "Sınıf 1.2 Patlayıcılar",
            ("1", "1.3C"): "Sınıf 1.3 Uyumluluk Grubu C Patlayıcılar",
            ("1", "1.4_special"): "Tablo 1.10.3.1.2'de özel olarak sayılan UN numaralı Sınıf 1.4 Patlayıcılar",
            ("1", "1.5"): "Sınıf 1.5 Patlayıcılar",
            ("2", "F"): "Alevlenebilir gazlar (F sınıflandırma kodu)",
            ("2", "FC"): "Alevlenebilir gazlar (FC sınıflandırma kodu)",
            ("2", "T"): "Zehirli gazlar (T sınıflandırma kodu)",
            ("2", "TF"): "Zehirli gazlar (TF sınıflandırma kodu)",
            ("2", "TC"): "Zehirli gazlar (TC sınıflandırma kodu)",
            ("2", "TO"): "Zehirli gazlar (TO sınıflandırma kodu)",
            ("2", "TFC"): "Zehirli gazlar (TFC sınıflandırma kodu)",
            ("2", "TOC"): "Zehirli gazlar (TOC sınıflandırma kodu)",
            ("3", "PGI"): "Ambalajlama grubu I'deki alevlenebilir sıvılar",
            ("3", "PGII"): "Ambalajlama grubu I ve II'deki alevlenebilir sıvılar",
            ("3", "desensitized"): "Duyarlılığı azaltılmış patlayıcılar (Sınıf 3)",
            ("4.1", "desensitized"): "Duyarlılığı azaltılmış patlayıcılar (Sınıf 4.1)",
            ("4.2", "PGI"): "Ambalajlama grubu I'deki kendiliğinden yanabilen maddeler",
            ("4.3", "PGI"): "Ambalajlama grubu I'deki maddeler (su ile temasta tehlikeli)",
            ("5.1", "PGI_liquid"): "Ambalajlama grubu I'deki yükseltgen sıvılar",
            ("5.1", "perchlorate_AN"): "Perkloratlar, amonyum nitrat, amonyum nitrat gübreler ve amonyum nitrat emülsiyonlar veya süspansiyonlar veya jeller",
            ("6.1", "PGI"): "Ambalajlama grubu I'deki zehirli maddeler",
            ("6.2", "catA"): "Kategori A'daki bulaşıcı maddeler (UN 2814 ve 2900 hayvansal malzemeler hariç)",
            ("8", "PGI"): "Ambalajlama grubu I'deki aşındırıcı maddeler",
        }
        desc = descriptions.get((threshold_cls, key), f"{cls_raw_full} sınıfı / {key}")

        if package_limit == 0:
            base.update(in_scope=True, table_key=key, sonuc_text=(
                f"Sınıf {cls_raw_full}; {desc} olarak sınıflandırılmıştır. "
                f"Bu madde Tablo 1.10.3.1.2 kapsamına GİRMEKTEDİR — ambalajlı "
                f"taşımada herhangi bir miktarda güvenlik planı değerlendirmesi "
                f"GEREKLİDİR."))
        elif package_limit == -1:
            base.update(in_scope=False, table_key=key, sonuc_text=(
                f"Sınıf {cls_raw_full}; {desc} olarak sınıflandırılmıştır. "
                f"Bu sınıfda ambalajlı taşımada miktar ne olursa olsun 1.10.3 "
                f"hükümleri uygulanmaz. Bu nedenle sınıfa girmemektedir; "
                f"güvenlik planı hazırlanmasına gerek yoktur."))
        elif package_limit is None:
            base.update(in_scope=False, table_key=key, sonuc_text=(
                f"Sınıf {cls_raw_full}; {desc} olarak sınıflandırılmıştır. "
                f"Ambalajlı taşıma modu için Tablo 1.10.3.1.2 uygulanamaz (a). "
                f"Bu nedenle güvenlik planı hazırlanmasına gerek yoktur."))
        else:
            base.update(in_scope="conditional", table_key=key, sonuc_text=(
                f"Sınıf {cls_raw_full}; {desc} olarak sınıflandırılmıştır. "
                f"Ambalajlı taşımada {package_limit} kg üzeri miktarlarda "
                f"güvenlik planı değerlendirmesi gereklidir; kesin sonuç için "
                f"sevkiyat miktarı esas alınmalıdır."))
        return base

    @classmethod
    def screen_inventory(cls, chemicals: list) -> dict:
        """Firmanın tüm kimyasal envanterini tarar ve özet + kalem bazlı
        sonuçları döndürür."""
        results = [cls.screen_inventory_chemical(c) for c in chemicals]
        in_scope = sum(1 for r in results if r["in_scope"] is True)
        conditional = sum(1 for r in results if r["in_scope"] == "conditional")
        exempt = sum(1 for r in results if r["in_scope"] is False)
        return {
            "results": results,
            "in_scope_count": in_scope,
            "conditional_count": conditional,
            "exempt_count": exempt,
            "total": len(results),
        }


    @classmethod
    def generate_inventory_review_html(cls, company_name: str, prepared_by: str,
                                       approved_by: str, screen_result: dict,
                                       date_str: str = "", validity_years: int = 2,
                                       logo_b64: str = "") -> str:
        """Örnek 'Güvenlik Planı İnceleme Raporu' formatında çok sayfalı HTML
        üretir: kapak + Tablo 1.10.3.1.2 referansı + madde bazlı değerlendirme
        + sonuç/imza sayfası. QTextDocument + QPrinter ile yazdırılabilir/
        PDF'e aktarılabilir (aynı doküman diğer evrak önizlemelerinde
        kullanılan mimariyle uyumludur)."""
        from datetime import datetime as _dt
        date_str = date_str or _dt.now().strftime("%d/%m/%Y")
        results = screen_result.get("results", [])
        in_scope_items = [r for r in results if r["in_scope"] is True]
        cond_items = [r for r in results if r["in_scope"] == "conditional"]
        exempt_items = [r for r in results if r["in_scope"] is False]

        NAVY = "#1E293B"
        MUTED = "#64748B"

        # ── Antet filigranı: varsa firma logosu (soluk), yoksa boş ──────
        watermark_b64 = ""
        try:
            # ShipmentEditorPage'deki paylaşılan filigran üreticisini
            # (logo + gerekiyorsa TASLAK yazısı) burada da kullan.
            watermark_b64 = ShipmentEditorPage._build_letterhead_watermark_b64(
                logo_b64, True)  # rapor onaylı belge niteliğinde, TASLAK yazısı yok
        except Exception:
            watermark_b64 = ""

        wrap_open = (
            f'<table width="100%" style="border-collapse:separate;'
            f'background-image:url(data:image/png;base64,{watermark_b64});'
            f'background-repeat:no-repeat;background-position:center top;">'
            f'<tr><td style="padding:0;">'
        ) if watermark_b64 else ""
        wrap_close = "</td></tr></table>" if watermark_b64 else ""

        def _esc(t):
            return (str(t or "")
                    .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))

        # ── Sayfa 1: Kapak ────────────────────────────────────────────
        cover = f"""
<div style="border:2px solid #333; padding:40px; margin:20px; page-break-after:always;">
  <table width="100%"><tr>
    <td style="font-size:16pt; font-weight:bold;">{_esc(company_name) or 'FİRMA ADI'}</td>
    <td align="right" style="color:{MUTED}; font-size:9pt;">METKA</td>
  </tr></table>
  <div style="height:220px;"></div>
  <p align="center" style="font-size:15pt; font-weight:bold; color:{NAVY};">
    {_esc(company_name) or 'FİRMA ADI'}
  </p>
  <p align="center" style="font-size:18pt; font-weight:bold; color:{NAVY}; margin-top:20px;">
    GÜVENLİK PLANI İNCELEME RAPORU
  </p>
  <div style="height:120px;"></div>
  <p align="center" style="font-size:11pt;">{date_str}</p>
  <p align="center" style="font-size:10pt; color:{MUTED};">Geçerlilik Süresi: {validity_years} Yıl</p>
  <div style="height:40px;"></div>
  <table width="100%" style="border-collapse:collapse;">
    <tr>
      <td width="50%" style="border:1px solid #999; padding:10px;">
        <b>Hazırlayan:</b> Tehlikeli Madde Güvenlik Danışmanı<br><br>
        {_esc(prepared_by) or '&nbsp;'}
      </td>
      <td width="50%" style="border:1px solid #999; padding:10px;">
        <b>Onaylayan:</b><br><br>
        {_esc(approved_by) or '&nbsp;'}
      </td>
    </tr>
  </table>
</div>"""

        # ── Sayfa 2: ADR Tablo 1.10.3.1.2 referans tablosu ──────────────
        table_rows = [
            ("1", "1.1", "Patlayıcılar", "a", "a", "0"),
            ("1", "1.2", "Patlayıcılar", "a", "a", "0"),
            ("1", "1.3", "Uyumluluk grubu C patlayıcılar", "a", "a", "0"),
            ("1", "1.4", "Patlayıcılar, UN No. 0104, 0237, 0255, 0267, 0289, 0361, "
                          "0365, 0366, 0440, 0441, 0455, 0456 ve 0500", "a", "a", "0"),
            ("1", "1.5", "Patlayıcılar", "0", "a", "0"),
            ("2", "", "Alevlenebilir gazlar (Yalnızca F harfi içeren sınıflandırma kodları)", "3000", "a", "b"),
            ("2", "", "Zehirli gazlar (T, TF, TC, TO, TFC veya TOC harflerini içeren "
                       "sınıflandırma kodları) aerosoller hariç", "0", "a", "0"),
            ("3", "", "Ambalajlama grubu I ve II'deki alevlenebilir sıvılar", "3000", "a", "b"),
            ("3", "", "Duyarlılığı azaltılmış patlayıcılar", "0", "a", "0"),
            ("4.1", "", "Duyarlılığı azaltılmış patlayıcılar", "a", "a", "0"),
            ("4.2", "", "Ambalajlama grubu I'deki maddeler", "3000", "a", "b"),
            ("4.3", "", "Ambalajlama grubu I'deki maddeler", "3000", "a", "b"),
            ("5.1", "", "Ambalajlama grubu I'deki yükseltgen sıvılar", "3000", "a", "b"),
            ("5.1", "", "Perkloratlar, amonyum nitrat, amonyum nitrat gübreler ve "
                        "amonyum nitrat emülsiyonlar veya süspansiyonlar veya jeller", "3000", "3000", "b"),
            ("6.1", "", "Ambalajlama grubu I'deki zehirli maddeler", "0", "a", "0"),
            ("6.2", "", "Kategori A'daki bulaşıcı maddeler (UN No. 2814 ve 2900 "
                        "hayvansal malzemeler hariç)", "a", "0", "0"),
            ("8", "", "Ambalajlama grubu I'deki aşındırıcı maddeler", "3000", "a", "b"),
        ]
        rows_html = "".join(
            f'<tr><td>{s}</td><td>{sg}</td><td>{_esc(m)}</td>'
            f'<td align="center">{t}</td><td align="center">{bl}</td>'
            f'<td align="center">{am}</td></tr>'
            for s, sg, m, t, bl, am in table_rows
        )
        table_page = f"""
<div style="padding:20px; page-break-after:always;">
  <p>Ciddi sonuçlara neden olabilecek maddelerin incelemesi ADRBOOK 1.10.3.1.2
  referans alınarak sonuçlar oluşturulmuştur.</p>
  <p style="font-weight:bold; font-size:11pt;">ADR TABLO 1.10.3.1.2</p>
  <table width="100%" border="1" style="border-collapse:collapse; font-size:8pt;">
    <tr style="background:#DDDDDD; font-weight:bold;">
      <td>Sınıf</td><td>Alt Grup</td><td>Madde veya nesne</td>
      <td align="center">Tank (l)</td><td align="center">Dökme yük (kg)</td>
      <td align="center">Ambalajlar (kg)</td>
    </tr>
    {rows_html}
  </table>
  <p style="font-size:8pt; margin-top:10px;">
    a: İlgili değil<br>
    b: Miktar ne olursa olsun 1.10.3 hükümleri uygulanmaz.<br>
    c: Bu sütunda belirtilen bir değer, tanklarda taşıma için izin verilmişse geçerlidir.<br>
    d: Bu sütunda belirtilen bir değer, dökme yük taşıma için izin verilmişse geçerlidir.
  </p>
</div>"""

        # ── Sayfa 3: Madde bazlı değerlendirme ──────────────────────────
        def _row(r, color):
            return (f'<tr><td>UN{_esc(r["un_number"])}<br>'
                    f'<span style="font-size:7.5pt;color:{MUTED}">{_esc(r["name"])}</span></td>'
                    f'<td align="center">{_esc(r["classification_code"]) or chr(0x2014)}</td>'
                    f'<td style="font-size:8pt;">{_esc(r["sonuc_text"])}</td>'
                    f'<td align="center">{_esc(r["class_code"])}</td>'
                    f'<td align="center">{_esc(r["packing_group"]) or chr(0x2014)}</td></tr>')

        eval_rows = "".join(_row(r, "#DC2626") for r in in_scope_items)
        eval_rows += "".join(_row(r, "#D97706") for r in cond_items)
        eval_rows += "".join(_row(r, "#166534") for r in exempt_items)

        eval_page = f"""
<div style="padding:20px; page-break-after:always;">
  <p style="font-weight:bold; font-size:11pt;">
    CİDDİ SONUÇLARA NEDEN OLABİLECEK MADDELER İNCELEMESİ
    (ADR BOOK 1.10.3.1.2 referans alınarak sonuçlar oluşturulmuştur.)
  </p>
  <table width="100%" border="1" style="border-collapse:collapse; font-size:8pt;">
    <tr style="background:#DDDDDD; font-weight:bold;">
      <td>UN No / Madde</td><td>Sınıflandırma Kodu</td><td>SONUÇ</td>
      <td>SINIF</td><td>PG</td>
    </tr>
    {eval_rows if eval_rows else '<tr><td colspan="5" align="center">Envanterde kimyasal bulunamadı</td></tr>'}
  </table>
</div>"""

        # ── Sayfa 4: Sonuç ve imzalar ────────────────────────────────────
        if in_scope_items or cond_items:
            un_list = ", ".join(f"UN{_esc(r['un_number'])}" for r in (in_scope_items + cond_items))
            conclusion_text = (
                f"Yukarıda incelemeleri yapılmış olan ADR'ye tabi kimyasallardan "
                f"<b>{len(in_scope_items) + len(cond_items)} tanesi</b> "
                f"({un_list}) Tablo 1.10.3.1.2 kapsamına girmektedir. "
                f"Bu maddeler için <b>GÜVENLİK PLANI HAZIRLANMASI GEREKLİDİR.</b> "
                f"Diğer {len(exempt_items)} kimyasal için güvenlik planı hazırlanmasına "
                f"gerek yoktur."
            )
        else:
            conclusion_text = (
                "Ciddi sonuçlara neden olabilecek tehlikeli malların veya ciddi "
                "sonuçlara neden olabilecek radyoaktif malzemelerin taşınmasına dahil "
                "olan, dolduran, paketleyen, yükleyen, gönderen, alıcı, boşaltan ve "
                "tank-konteyner/taşınabilir tank işletmecisi herhangi bir acil durum "
                "oluştuğunda hemen organize olarak, düzenli bir şekilde müdahale etmek "
                "ve ortaya çıkabilecek olan zararları en az seviyeye indirebilmesi için "
                "güvenlik planı incelenmesi gerçekleştirilmiştir.<br><br>"
                "Bu hususta yukarıda incelemeleri yapılmış olan ADR'ye tabi kimyasallar "
                "için <b>GÜVENLİK PLANI HAZIRLANMASINA GEREK YOKTUR.</b>"
            )

        conclusion_page = f"""
<div style="padding:20px;">
  <p>{conclusion_text}</p>
  <div style="height:60px;"></div>
  <table width="100%" style="border-collapse:collapse;">
    <tr>
      <td width="50%" style="padding:10px;">
        Tehlikeli Madde Güvenlik Danışmanlığı<br><br>
        {_esc(prepared_by) or '&nbsp;'}
      </td>
      <td width="50%" style="padding:10px;">
        {_esc(company_name) or '&nbsp;'}<br><br>
        {_esc(approved_by) or '&nbsp;'}
      </td>
    </tr>
  </table>
  <p style="font-size:8pt; color:{MUTED}; margin-top:30px;">
    Değerlendirilen kimyasal sayısı: {len(results)} &nbsp;|&nbsp;
    Kapsam içi: {len(in_scope_items)} &nbsp;|&nbsp;
    Koşullu: {len(cond_items)} &nbsp;|&nbsp;
    Muaf: {len(exempt_items)}
  </p>
</div>"""

        return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><style>
body {{ font-family: 'Segoe UI', Arial, sans-serif; font-size:9.5pt; color:#1F2937; }}
</style></head><body>
{wrap_open}
{cover}
{table_page}
{eval_page}
{conclusion_page}
{wrap_close}
</body></html>"""


class ADREngine:
    """ADR mevzuat hesaplamalari ve dogrulamalari."""

    TC_POINTS = {"0": 0, "1": 50, "2": 3, "3": 1, "4": 0}

    @classmethod
    def calculate_1136_points(cls, items: List[ShipmentItem]) -> Tuple[float, bool, str]:
        """
        ADR 1.1.3.6 Miktar Muafiyeti (1000 Puan Kurali)

        Tasima Kategorileri ve Puanlari:
        - TC 0: 0 puan (EQ maddeler)
        - TC 1: 50 puan (Cok tehlikeli)
        - TC 2: 3 puan (Tehlikeli)
        - TC 3: 1 puan (Az tehlikeli)
        - TC 4: 0 puan (LQ maddeler)

        Toplam 1000 puani asarsa turuncu plaka zorunlu.
        """
        total = 0.0
        detail_lines = []
        # TC=4 olan kalemler (sınırsız taşınabilir) ayrı takip edilir
        tc4_items = []

        for item in items:
            if item.is_lq:
                points_per_unit = 0
                detail_lines.append(f"  UN{item.un_number}: LQ → 0 puan (sınırsız)")
            elif item.is_eq:
                points_per_unit = 0
                detail_lines.append(f"  UN{item.un_number}: EQ → 0 puan (sınırsız)")
            else:
                raw_tc = str(getattr(item, "transport_category", "")).strip()
                tc = raw_tc.split()[0] if raw_tc else ""

                # [GUVENLIK] Bos/gecersiz TC: hesaptan ATLANMAZ — en
                # kisitlayici carpan (x50) uygulanir ve acikca uyarilir.
                # Atlamak puani dusuk gosterip plakasiz sevkiyata yol acar.
                if not tc or tc not in cls.TC_POINTS:
                    detail_lines.append(
                        f"  ⚠ UN{item.un_number}: Taşıma Kategorisi boş/geçersiz"
                        f" ('{raw_tc}') — GÜVENLİ TARAF: x50 uygulandı,"
                        f" kategoriyi veritabanında tamamlayın!"
                    )
                    tc = "1"

                # [GUVENLIK] Kategori 0: miktar ne olursa olsun 1.1.3.6
                # muafiyeti MUMKUN DEGIL — turuncu plaka zorunlu.
                if tc == "0":
                    detail_lines.append(
                        f"  ⛔ UN{item.un_number}: Kategori 0 — muafiyet"
                        f" mümkün değil, TURUNCU PLAKA ZORUNLU"
                    )
                    total = float(MAX_1136_POINTS) + 1  # limiti asir
                    continue

                points_per_unit = cls.TC_POINTS[tc]

                # [DÜZELTİLDİ] TC=4: ADR 3.4/3.5 kapsamında sınırsız taşıma
                # 0 puan verir, 1000 puan limitine dahil edilmez
                if tc == "4":
                    tc4_items.append(item)
                    detail_lines.append(
                        f"  UN{item.un_number}: TC=4 → 0 puan"
                        f" (ADR 3.4 — sınırsız, limite dahil değil)"
                    )
                    continue  # toplama ekleme

                line_total = float(item.net_quantity) * points_per_unit
                detail_lines.append(
                    f"  UN{item.un_number}: TC={tc} → {points_per_unit} puan/kg"
                    f" × {item.net_quantity} {item.unit} = {line_total:.1f} puan"
                )

            total += float(item.net_quantity) * points_per_unit

        required = total > MAX_1136_POINTS

        # TC=4 notu varsa özete ekle
        if tc4_items:
            detail_lines.append(
                f"  ℹ TC=4 kalem sayısı: {len(tc4_items)} "
                f"— ADR 3.4 kapsamında puan hesabına dahil edilmedi"
            )

        # [DÜZELTİLDİ] Puan 0 ise (tüm kalemler TC=4, LQ, EQ veya boş TC)
        # "gerekmez" mesajı yerine bilgilendirici mesaj göster
        if total == 0 and not required:
            if tc4_items or any(i.is_lq or i.is_eq for i in items):
                msg = f"Puan hesabı 0 — tüm kalemler sınırsız kategoride (TC=4/LQ/EQ)"
            else:
                msg = f"Puan hesabı 0 — Taşıma Kategorisi girilmemiş kalemler var"
        elif required:
            msg = f"TURUNCU PLAKA ZORUNLU! ({total:.0f} / {MAX_1136_POINTS} Puan)"
        else:
            msg = f"Turuncu plaka gerekmez ({total:.0f} / {MAX_1136_POINTS} Puan)"

        detail_lines.insert(0, f"=== 1.1.3.6 PUAN HESAPLAMA ===")
        detail_lines.append(f"TOPLAM: {total:.1f} puan")
        detail_lines.append(f"SONUC: {msg}")

        return total, required, "\n".join(detail_lines)

    @classmethod
    def calculate_tunnel_restriction(cls, items: List[ShipmentItem]) -> str:
        if not items:
            return "E"

        most_restrictive = "E"
        for item in items:
            # "D/E", "(C/D)" gibi bilesik kodlar: en kisitlayici bilesen esas
            raw = (item.tunnel_code or "E").replace("(", "").replace(")", "")
            for code in raw.split("/"):
                code = code.strip().upper()
                if code in TUNNEL_HIERARCHY:
                    if TUNNEL_HIERARCHY[code] < TUNNEL_HIERARCHY[most_restrictive]:
                        most_restrictive = code

        return most_restrictive

    @classmethod
    def check_compatibility(cls, items: List[ShipmentItem]) -> List[str]:
        errors = []
        groups = set()

        CLASS_TO_SEGREGATION = {
            "3":   "Yanici Maddeler",
            "5.1": "Yukseltgenler",
            "5.2": "Organik Peroksitler",
            "8":   "Asitler",
            "4.3": "Su ile Tepki Veren",
            "2.3": "2.3 Sinifi Zehirli Gazlar",
        }

        for item in items:
            class_code = item.class_code or ""
            seg_group = CLASS_TO_SEGREGATION.get(class_code)
            if seg_group:
                groups.add(seg_group)
            if item.segregation_group:
                groups.add(item.segregation_group)

        for group in groups:
            incompatible = INCOMPATIBILITY_MATRIX.get(group, [])
            for other in groups:
                if other != group and other in incompatible:
                    errors.append(f"UYUMSUZ: {group} + {other} birlikte tasinamaz!")

        return list(set(errors))

    # ------------------------------------------------------------------
    # LQ (ADR 3.4) / EQ (ADR 3.5) — maddeye ozgu limitler
    # Sinif-bazli sabit tablolar KALDIRILDI: LQ limiti Tablo A sutun 7a'da
    # UN bazinda, EQ limiti sutun 7b E-kodu ile belirlenir.
    # ------------------------------------------------------------------
    EQ_TABLE = {
        "E0": (0, 0),
        "E1": (30, 1000),
        "E2": (30, 500),
        "E3": (30, 300),
        "E4": (1, 500),
        "E5": (1, 300),
    }

    @staticmethod
    def parse_date_flexible(text) -> Optional[datetime]:
        """Tarihi yaygin Turk ve ISO bicimlerinde ayristirir.
        Kabul: 2027-12-31, 31.12.2027, 31/12/2027, 31-12-2027.
        Uymazsa None doner; cagiran taraf ACIK hata uretmelidir."""
        raw = str(text or "").strip()
        if not raw:
            return None
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(raw, fmt)
            except ValueError:
                continue
        return None

    @classmethod
    def parse_lq_limit(cls, text) -> Tuple[float, str]:
        """Tablo A 7a metnini ("1 L", "5 kg", "500 ml", "0") ayristirir.
        Donus: (deger, birim) — L veya kg'a normalize; yasaksa (0.0, "")."""
        if text is None:
            return 0.0, ""
        raw = str(text).strip()
        if raw == "" or raw.lower() in ("nan", "-", "yok", "none", "0"):
            return 0.0, ""
        m = re.match(r"^\s*(\d+(?:[.,]\d+)?)\s*(l|lt|litre|ml|kg|g|gr)?",
                     raw, re.IGNORECASE)
        if not m:
            return 0.0, ""
        value = float(m.group(1).replace(",", "."))
        unit = (m.group(2) or "kg").lower()
        if unit == "ml":
            return value / 1000.0, "L"
        if unit in ("l", "lt", "litre"):
            return value, "L"
        if unit in ("g", "gr"):
            return value / 1000.0, "kg"
        return value, "kg"

    @classmethod
    def eq_limits(cls, code) -> Tuple[int, int]:
        """E-kodu -> (ic, dis) g/ml limiti; taninmayan kod = (0, 0)."""
        key = str(code or "").strip().upper()
        return cls.EQ_TABLE.get(key, (0, 0))

    @classmethod
    def is_lq_allowed(cls, chemical: Chemical) -> bool:
        limit, _ = cls.parse_lq_limit(getattr(chemical, "limited_quantity", ""))
        if limit > 0:
            return True
        if not str(getattr(chemical, "limited_quantity", "") or "").strip():
            return bool(chemical.lq_allowed)  # eski kayit: boolean'a saygi
        return False

    @classmethod
    def is_eq_allowed(cls, chemical: Chemical) -> bool:
        code = str(getattr(chemical, "excepted_quantity", "") or "").strip().upper()
        if code:
            return cls.eq_limits(code)[0] > 0
        return bool(chemical.eq_allowed)

    @classmethod
    def check_lq_eligibility(cls, chemical: Chemical, quantity: float,
                              unit: str = "") -> Tuple[bool, str, float]:
        """Ic ambalaj basina miktari MADDENIN KENDI 7a limitiyle karsilastirir."""
        max_qty, limit_unit = cls.parse_lq_limit(
            getattr(chemical, "limited_quantity", ""))

        if max_qty <= 0:
            if (not str(getattr(chemical, "limited_quantity", "") or "").strip()
                    and chemical.lq_allowed):
                return True, ("LQ isaretli (eski kayit, limit bilgisi yok - "
                              "Tablo A'dan dogrulayin)"), 0.0
            return False, "Bu madde LQ olarak tasinamaz (7a = 0)", 0.0

        VOLUME = {"l", "lt", "litre", "ml"}
        MASS = {"kg", "g", "gr"}
        unit_note = ""
        u = str(unit or "").strip().lower()
        if u:
            entered_family = "L" if u in VOLUME else ("kg" if u in MASS else "?")
            if entered_family == "?" or entered_family != limit_unit:
                unit_note = (f" | UYARI: birim uyusmuyor (limit {limit_unit}, "
                             f"girilen {unit}) - manuel dogrulayin")

        if quantity > max_qty:
            return False, (f"LQ limiti asildi: {quantity:g} > "
                           f"{max_qty:g} {limit_unit}{unit_note}"), max_qty
        return True, f"LQ uygun (Max: {max_qty:g} {limit_unit}/ic ambalaj){unit_note}", max_qty

    @classmethod
    def check_eq_eligibility(cls, chemical: Chemical,
                             inner_quantity_g_ml: float) -> Tuple[bool, str, float]:
        """Ic ambalaj basina miktari (g/ml) maddenin E-koduna gore denetler."""
        code = str(getattr(chemical, "excepted_quantity", "") or "").strip().upper()

        if not code:
            if chemical.eq_allowed:
                return True, ("EQ isaretli (eski kayit, E-kodu yok - "
                              "Tablo A'dan dogrulayin)"), 0.0
            return False, "Bu madde EQ olarak tasinamaz (E-kodu yok)", 0.0

        inner, outer = cls.eq_limits(code)
        if inner <= 0:
            return False, f"Bu madde EQ olarak tasinamaz (kod: {code})", 0.0

        if inner_quantity_g_ml > inner:
            return False, (f"EQ limiti asildi ({code}): {inner_quantity_g_ml:g} > "
                           f"{inner} g/ml ic ambalaj (dis max {outer} g/ml)"), float(inner)
        return True, (f"EQ uygun ({code}: ic {inner} g/ml, "
                      f"dis {outer} g/ml)"), float(inner)

    @classmethod
    def verify_chemicals_data(cls, items: List[ShipmentItem]) -> List[dict]:
        """
        Eksik veya geçersiz veri içeren kimyasalları tespit eder.

        ADR 1.1.3.6 puan hesaplamasını ve tünel kodu atamasını
        olumsuz etkileyebilecek sorunları raporlar.

        Returns:
            List[dict]: Her sorun için {'un': ..., 'sorun': ...} içeren liste.
                        Liste boşsa tüm veriler tamdır.
        """
        sorunlar = []
        for item in items:
            un = getattr(item, "un_number", "?")
            un_str = f"UN{un}"

            # Taşıma Kategorisi (TC) kontrolü
            raw_tc = str(getattr(item, "transport_category", "") or "").strip()
            tc = raw_tc.split()[0] if raw_tc else ""
            if not item.is_lq and not item.is_eq:
                if not tc:
                    sorunlar.append({
                        "un": un_str,
                        "sorun": "Taşıma Kategorisi (TC) boş — 1.1.3.6 puan hesaplanamaz."
                    })
                elif tc not in cls.TC_POINTS:
                    sorunlar.append({
                        "un": un_str,
                        "sorun": f"Taşıma Kategorisi geçersiz: '{tc}' (geçerli: 0,1,2,3,4)."
                    })

            # Tünel Kodu kontrolü
            tun = getattr(item, "tunnel_code", "") or ""
            valid_tunnels = {"A", "B", "B1", "B/D", "B/E", "C", "C/D", "C/E", "D", "D/E", "E"}
            if tun and tun.upper() not in {t.upper() for t in valid_tunnels}:
                sorunlar.append({
                    "un": un_str,
                    "sorun": f"Tünel kodu geçersiz: '{tun}'."
                })

            # Net miktar sıfır veya negatif kontrolü
            try:
                qty = float(item.net_quantity or 0)
            except (TypeError, ValueError):
                qty = 0
            if qty <= 0:
                sorunlar.append({
                    "un": un_str,
                    "sorun": f"Net miktar sıfır veya negatif: {item.net_quantity!r}."
                })

        return sorunlar

    @classmethod
    def generate_adr_report(cls, items: List[ShipmentItem],
                             driver: Driver = None,
                             vehicle: Vehicle = None,
                             packaging_types: List[str] = None) -> ADRReport:
        """
        ADR raporu oluştur - muafiyetleri doğru değerlendir.

        Muafiyet Hiyerarşisi:
        1. EQ (İstisnai Miktar) - En üst düzey muafiyet
        2. LQ (Sınırlı Miktar)
        3. ADR 1.1.3.6 (Miktar Muafiyeti - 1000 Puan)
        4. Özel Hükümler (SP kodları)
        5. Genel Muafiyetler (1.1.3.1)
        """
        report = ADRReport()
        packaging_types = packaging_types or []

        if not items:
            report.info.append((WarningLevel.INFO, "Sevkiyat bos"))
            return report

        # === TC DURUM SINIFLANDIRMASI ===
        aktif = [i for i in items if float(i.net_quantity or 0) > 0]
        lq_count     = sum(1 for i in aktif if i.is_lq)
        eq_count     = sum(1 for i in aktif if i.is_eq)
        tc4_count    = sum(1 for i in aktif
                          if not i.is_lq and not i.is_eq and
                          str(getattr(i, "transport_category", "")).strip().split()[0:1] == ["4"])
        empty_tc_count = sum(1 for i in aktif
                             if not i.is_lq and not i.is_eq and
                             not str(getattr(i, "transport_category", "")).strip())
        normal_count = len(aktif) - lq_count - eq_count - tc4_count - empty_tc_count

        # TC boş kalemler varsa: puan hesabı yapılamaz, rapor kısmi gösterilir
        has_empty_tc = empty_tc_count > 0
        # Hesaplanabilir kalem var mı (TC=4/LQ/EQ veya normal TC)
        has_calculable = (lq_count + eq_count + tc4_count + normal_count) > 0

        if has_empty_tc:
            report.warnings.append((WarningLevel.WARNING,
                f"{empty_tc_count} kalemin Taşıma Kategorisi boş — "
                "puan hesabına dahil edilmedi. Excel'i kontrol edin."))

        # === 1.1.3.6 PUAN HESAPLAMA ===
        total_points, orange_required, points_detail = cls.calculate_1136_points(items)
        report.total_points = total_points
        report.orange_plate_required = orange_required

        # Puan sadece hesaplanabilir kalem varsa göster
        if total_points > 0:
            if orange_required:
                report.errors.append((WarningLevel.ERROR,
                    f"TURUNCU PLAKA ZORUNLU! ({total_points:.0f} / {MAX_1136_POINTS} Puan)"))
            else:
                report.info.append((WarningLevel.INFO,
                    f"Turuncu plaka gerekmez ({total_points:.0f} / {MAX_1136_POINTS} Puan)"))
            report.info.append((WarningLevel.INFO, points_detail))
        elif tc4_count > 0 and not has_empty_tc:
            report.info.append((WarningLevel.INFO,
                f"TC=4 kalemler sınırsız taşınabilir (ADR 3.4) — puan limiti uygulanmaz"))
        elif lq_count > 0 or eq_count > 0:
            report.info.append((WarningLevel.INFO,
                "LQ/EQ muafiyetli kalemler — puan limiti uygulanmaz"))

        # === TUNEL KISITLAMA ===
        tunnel = cls.calculate_tunnel_restriction(items)
        report.tunnel_code = tunnel
        report.info.append((WarningLevel.INFO, f"Tunel kisitlama kodu: {tunnel}"))
        if tunnel == "A":
            report.errors.append((WarningLevel.CRITICAL,
                "Tunel kodu A: Bu sevkiyat tunelden gecemez!"))
        elif tunnel in ["B", "C"]:
            report.warnings.append((WarningLevel.WARNING,
                f"Tunel kodu {tunnel}: Kisitli tunel gecisi"))

        # === UYUMSUZLUK KONTROLU ===
        # DÜZELTME (web portu denetiminde bulundu, aynı hata burada da
        # vardı — satırı satırına taşındığı için): burada ÖNCEDEN
        # cls.check_compatibility(items) çağrılıyordu — sabit, HAYALİ bir
        # sözlüğe dayanan, GERÇEK bir ADR referansı olmayan basitleştirilmiş
        # bir kontroldü. GERÇEK motor (AnaDbChemicalAdapter + MixChecker,
        # bu dosyada zaten mevcut) burada ÇAĞRILAMAZ çünkü generate_adr_report
        # veritabanı bağlantısına erişemez (saf hesaplama fonksiyonu).
        # report.compatibility_errors artık BOŞ bırakılıyor; gerçek sonuç
        # _gercek_karisik_yukleme_kontrolu() ile ÇAĞIRAN taraflarca
        # (_update_preview, _build_print_html) ayrıca hesaplanıp yerleştirilir.
        report.compatibility_errors = []

        # === LQ / EQ DURUMU ===
        if lq_count > 0:
            report.lq_status = f"{lq_count} kalemden LQ (Sinirli Miktar) uygulandi"
            report.info.append((WarningLevel.INFO, report.lq_status))
        if eq_count > 0:
            report.eq_status = f"{eq_count} kalemden EQ (Istisnai Miktar) uygulandi"
            report.info.append((WarningLevel.INFO, report.eq_status))

        # === [v4.4] LQ/EQ İŞARETLİ KALEMLERDE MADDEYE ÖZGÜ LİMİT DENETİMİ ===
        # LQ/EQ kutucuğu işaretlenmiş olması muafiyeti otomatik geçerli
        # kılmaz — iç ambalaj başına miktar, maddenin kendi 7a/7b limitini
        # aşıyorsa muafiyet GEÇERSİZDİR ve kritik hata üretilmelidir.
        for it in aktif:
            per_pkg = (float(it.net_quantity) / it.packaging_count
                       if it.packaging_count else float(it.net_quantity))
            if it.is_lq and it.lq_max_per_package > 0 and per_pkg > it.lq_max_per_package:
                report.errors.append((WarningLevel.CRITICAL,
                    f"UN {it.un_number}: LQ limiti asiliyor "
                    f"({per_pkg:g} > {it.lq_max_per_package:g} / ic ambalaj). "
                    f"LQ muafiyeti GECERSIZ!"))
            if it.is_eq and it.eq_max_per_package > 0:
                per_pkg_g = per_pkg * 1000  # kg/L -> g/ml
                if per_pkg_g > it.eq_max_per_package:
                    report.errors.append((WarningLevel.CRITICAL,
                        f"UN {it.un_number}: EQ ic ambalaj limiti asiliyor "
                        f"({per_pkg_g:g} g/ml > {it.eq_max_per_package:g} g/ml). "
                        f"EQ muafiyeti GECERSIZ!"))

        if tc4_count > 0:
            report.info.append((WarningLevel.INFO,
                f"{tc4_count} kalem TC=4 (sınırsız, ADR 3.4 kapsamı)"))
        if normal_count > 0:
            report.info.append((WarningLevel.INFO,
                f"{normal_count} kalem normal ADR kurallarina tabi"))

        # === MUAFIYET DEĞERLENDİRMESİ ===
        # TC boş kalemler varken muafiyet belirsiz — hesaplama yapma
        if has_empty_tc and not has_calculable:
            report.exemption_type = ExemptionType.NONE.value
            report.info.append((WarningLevel.INFO,
                "Muafiyet belirlenemiyor — Taşıma Kategorisi eksik"))
        elif eq_count > 0 and eq_count == len(aktif):
            report.exemption_type = ExemptionType.EQ.value
            report.info.append((WarningLevel.INFO,
                "TUMU EQ: ADR 3.5 Istisnai Miktar muafiyeti uygulanir"))
        elif lq_count > 0 and lq_count == len(aktif):
            report.exemption_type = ExemptionType.LQ.value
            report.info.append((WarningLevel.INFO,
                "TUMU LQ: ADR 3.4 Sinirli Miktar muafiyeti uygulanir"))
        elif tc4_count > 0 and (tc4_count + lq_count + eq_count) == len(aktif):
            # Sadece TC=4 / LQ / EQ → sınırsız
            report.exemption_type = ExemptionType.ADR_1_1_3_6.value
            report.info.append((WarningLevel.INFO,
                "Tüm kalemler TC=4/LQ/EQ — ADR 3.4 kapsamında sınırsız taşıma"))
        elif total_points > 0 and total_points <= MAX_1136_POINTS and normal_count > 0:
            report.exemption_type = ExemptionType.ADR_1_1_3_6.value
            report.info.append((WarningLevel.INFO,
                f"ADR 1.1.3.6: Miktar muafiyeti uygulanir ({total_points:.0f} < {MAX_1136_POINTS} puan)"))
        elif any(item.special_provisions for item in items if hasattr(item, 'special_provisions')):
            report.exemption_type = ExemptionType.SPECIAL_PROVISION.value
            report.info.append((WarningLevel.INFO,
                "ADR 3.3: Ozel hukumler kapsaminda muafiyet degerlendirildi"))
        else:
            report.exemption_type = ExemptionType.NONE.value
            report.info.append((WarningLevel.INFO,
                "Muafiyet yok - Tam ADR uygulamasi zorunlu"))

        # === YAZILI TALİMAT ===
        # TC boş ve hesaplanabilir kalem yoksa yazılı talimat da belirsiz
        if not has_calculable and has_empty_tc:
            pass  # Gösterme
        else:
            has_class_1 = any((item.class_code or "").startswith("1") for item in items)
            has_class_7 = any(item.class_code == "7" for item in items)
            if has_class_1 or has_class_7 or orange_required:
                report.written_instructions_required = True
                report.warnings.append((WarningLevel.WARNING,
                    "Yazili talimat zorunlu (ADR 8.1.2.1)"))
            else:
                report.info.append((WarningLevel.INFO, "Yazili talimat gerekmez"))

        # === SÜRÜCÜ ADR SERTİFİKA ===
        if not has_calculable and has_empty_tc:
            pass  # TC boş, sertifika kontrolü yapma
        else:
            has_tank = any(pt == "Tank" for pt in packaging_types) if packaging_types else False
            if orange_required or any(
                (item.class_code or "").startswith("1") or
                item.class_code in ["2.3", "6.2", "7"]
                for item in items
            ):
                report.driver_adr_required = True
                if driver:
                    if not driver.adr_certificate_no:
                        report.errors.append((WarningLevel.CRITICAL,
                            "ADR sertifikali surucu zorunlu!"))
                    else:
                        expiry = cls.parse_date_flexible(driver.adr_certificate_expiry)
                        if expiry is None:
                            # [GUVENLIK] Bozuk tarih = kontrol yapilamiyor;
                            # gecerli varsayilamaz, KRITIK hata uretilir.
                            report.errors.append((WarningLevel.CRITICAL,
                                f"Surucu ADR sertifika tarihi okunamadi: "
                                f"'{driver.adr_certificate_expiry}' "
                                f"(desteklenen: 2027-12-31, 31.12.2027, 31/12/2027)"))
                        elif expiry < datetime.now():
                            report.errors.append((WarningLevel.CRITICAL,
                                f"Surucu ADR sertifikasi gecersiz! (Bitis: {driver.adr_certificate_expiry})"))
                        elif expiry < datetime.now() + timedelta(days=30):
                            report.warnings.append((WarningLevel.WARNING,
                                f"Surucu ADR sertifikasi yakinda bitiyor: {driver.adr_certificate_expiry}"))
                        else:
                            report.info.append((WarningLevel.INFO,
                                f"Surucu ADR sertifikasi gecerli: {driver.adr_certificate_expiry}"))
                    if not driver.src5_no:
                        report.errors.append((WarningLevel.CRITICAL, "SRC5 belgesi zorunlu!"))
                    else:
                        report.info.append((WarningLevel.INFO, f"SRC5 belgesi: {driver.src5_no}"))
                else:
                    report.errors.append((WarningLevel.CRITICAL, "Surucu bilgisi eksik!"))

            if has_tank:
                report.warnings.append((WarningLevel.WARNING,
                    "Tank tasimasi: T9/Tasit Uygunluk Belgesi zorunlu!"))
                if vehicle:
                    if not vehicle.adr_compliance_cert_no:
                        report.errors.append((WarningLevel.CRITICAL,
                            "T9/Tasit Uygunluk Belgesi eksik!"))
                    else:
                        report.info.append((WarningLevel.INFO,
                            f"T9/Tasit Uygunluk Belgesi: {vehicle.adr_compliance_cert_no}"))

        # === ADR 1.10.3 EMNİYET PLANI ===
        try:
            sp = SecurityPlanEngine.check(
                items,
                transport_mode="ambalaj",
                total_1136_points=report.total_points,
            )
            report.security_plan_required = sp["required"]
            report.security_plan_exempt   = sp["exempt"]
            report.security_plan_reasons  = sp["reasons"]
            if sp["required"]:
                report.warnings.append((WarningLevel.WARNING,
                    "🛡 ADR 1.10.3: EMNİYET PLANI GEREKLİ — "
                    + "; ".join(sp["reasons"][:2])))
            elif sp["exempt"]:
                report.info.append((WarningLevel.INFO,
                    "✅ ADR 1.10.3: Emniyet planı gerekmiyor (1.10.4 muafiyeti)"))
        except Exception:
            pass

        return report
    @classmethod
    def validate_shipment(cls, items: List[ShipmentItem],
                           sender: Company = None,
                           receiver: Company = None,
                           driver: Driver = None,
                           vehicle: Vehicle = None,
                           packaging_types: List[str] = None) -> ValidationResult:
        """
        Sevkiyat validasyonu - ADR kurallarına uygunluk kontrolü.

        Zorunlu alanlar:
        - Ambalaj türü (IBC, Varil, Bidon, Kutu, Çuval, Kompozit Ambalaj, Tank, Dökme)
        - Miktar (> 0)
        - UN numarası
        - Sınıf bilgisi
        """
        result = ValidationResult()
        packaging_types = packaging_types or []

        if not items:
            result.errors.append((WarningLevel.ERROR, "En az bir urun eklenmeli"))
            return result

        # === ZORUNLU ALAN KONTROLLERI ===
        VALID_PACKAGING_TYPES = [
            "IBC", "Varil", "Bidon", "Kutu", 
            "Çuval", "Kompozit Ambalaj", "Tank", "Dökme"
        ]

        for i, item in enumerate(items, 1):
            # UN numarası kontrolü
            if not item.un_number:
                result.errors.append((WarningLevel.ERROR, f"Satir {i}: UN numarasi eksik"))

            # Sınıf kontrolü
            if not item.class_code:
                result.errors.append((WarningLevel.ERROR, f"Satir {i}: Sinif bilgisi eksik"))

            # Miktar kontrolü (> 0)
            if item.net_quantity <= 0:
                result.errors.append((WarningLevel.ERROR, 
                    f"Satir {i}: Miktar 0'dan buyuk olmali (Girilen: {item.net_quantity})"))

            # Ambalaj türü kontrolü (ZORUNLU)
            if not item.packaging_type:
                result.errors.append((WarningLevel.ERROR, 
                    f"Satir {i}: Ambalaj turu zorunlu! (Secenekler: {', '.join(VALID_PACKAGING_TYPES)})"))
            elif item.packaging_type not in VALID_PACKAGING_TYPES:
                result.warnings.append((WarningLevel.WARNING, 
                    f"Satir {i}: '{item.packaging_type}' standart ambalaj turu degil. "
                    f"Beklenen: {', '.join(VALID_PACKAGING_TYPES)}"))

            # Ambalaj adeti kontrolü
            if item.packaging_count <= 0:
                result.warnings.append((WarningLevel.WARNING, 
                    f"Satir {i}: Ambalaj adeti 0'dan buyuk olmali"))

            # Birim kontrolü
            if not item.unit:
                result.warnings.append((WarningLevel.WARNING, 
                    f"Satir {i}: Birim (kg/lt/adet) belirtilmemis"))

        # === FIRMA KONTROLLERI ===
        # [GUVENLIK] None kontrolu: firma HIC secilmemisse de hata uretilir
        # (eski kod "sender and ..." ile None durumunu sessizce geciyordu)
        if sender is None or not sender.name:
            result.errors.append((WarningLevel.ERROR, "Gonderici firma secilmemis"))
        if receiver is None or not receiver.name:
            result.errors.append((WarningLevel.ERROR, "Alici firma secilmemis"))

        # === SURUCU KONTROLLERI ===
        if driver:
            # SRC5 kontrolü (her zaman zorunlu)
            if not driver.src5_no:
                result.errors.append((WarningLevel.ERROR, "SRC5 belgesi zorunlu!"))
            else:
                src5_expiry = cls.parse_date_flexible(driver.src5_expiry)
                if src5_expiry is None:
                    result.errors.append((WarningLevel.ERROR,
                        f"SRC5 bitis tarihi okunamadi: '{driver.src5_expiry}' "
                        f"(desteklenen: 2027-12-31, 31.12.2027, 31/12/2027)"))
                elif src5_expiry < datetime.now():
                    result.errors.append((WarningLevel.ERROR, "SRC5 belgesi gecersiz!"))
                elif src5_expiry < datetime.now() + timedelta(days=30):
                    result.warnings.append((WarningLevel.WARNING,
                        "SRC5 belgesi yakinda bitiyor!"))

            # ADR sertifika kontrolü (turuncu plaka veya tehlikeli sınıflar için)
            has_dangerous = any(
                item.class_code in ["1", "2.3", "6.2", "7"] or 
                (not item.is_lq and not item.is_eq)
                for item in items
            )
            if has_dangerous and not driver.adr_certificate_no:
                result.warnings.append((WarningLevel.WARNING, 
                    "ADR sertifikasi onerilir (tehlikeli madde tasimasi)"))
        else:
            result.errors.append((WarningLevel.ERROR, "Surucu secilmemis!"))

        # === ARAC KONTROLLERI ===
        if vehicle:
            # Muayene kontrolü
            if not str(vehicle.inspection_expiry or "").strip():
                result.warnings.append((WarningLevel.WARNING,
                    "Arac muayene tarihi kayitli degil"))
            else:
                inspection_expiry = cls.parse_date_flexible(vehicle.inspection_expiry)
                if inspection_expiry is None:
                    result.errors.append((WarningLevel.ERROR,
                        f"Arac muayene tarihi okunamadi: '{vehicle.inspection_expiry}' "
                        f"(desteklenen: 2027-12-31, 31.12.2027, 31/12/2027)"))
                elif inspection_expiry < datetime.now():
                    result.errors.append((WarningLevel.ERROR, "Arac muayenesi gecersiz!"))
                elif inspection_expiry < datetime.now() + timedelta(days=30):
                    result.warnings.append((WarningLevel.WARNING, "Arac muayenesi yakinda bitiyor"))

            # T9/Taşıt Uygunluk Belgesi kontrolü (sadece tank taşıması için)
            has_tank = any(pt == "Tank" for pt in packaging_types)
            if has_tank:
                if not vehicle.adr_compliance_cert_no:
                    result.errors.append((WarningLevel.CRITICAL, 
                        "Tank tasimasi: T9/Tasit Uygunluk Belgesi zorunlu!"))
                else:
                    adr_expiry = cls.parse_date_flexible(vehicle.adr_compliance_expiry)
                    if adr_expiry is None:
                        result.errors.append((WarningLevel.ERROR,
                            f"T9 gecerlilik tarihi okunamadi: "
                            f"'{vehicle.adr_compliance_expiry}' "
                            f"(desteklenen: 2027-12-31, 31.12.2027, 31/12/2027)"))
                    elif adr_expiry < datetime.now():
                        result.errors.append((WarningLevel.ERROR,
                            "T9/Tasit Uygunluk Belgesi gecersiz!"))
            else:
                # Tank değilse ADR uygunluk belgesi opsiyonel
                if vehicle.adr_compliance_cert_no:
                    result.info.append((WarningLevel.INFO, 
                        "T9/Tasit Uygunluk Belgesi mevcut (opsiyonel)"))
        else:
            result.errors.append((WarningLevel.ERROR, "Arac secilmemis!"))

        # === ADR RAPORU ===
        adr_report = cls.generate_adr_report(items, driver, vehicle, packaging_types)
        result.errors.extend(adr_report.errors)
        result.warnings.extend(adr_report.warnings)
        result.info.extend(adr_report.info)

        result.is_valid = len(result.errors) == 0

        return result
    @classmethod
    def get_class_color(cls, class_code: str) -> Tuple[str, str]:
        return CLASS_COLORS.get(class_code, ("#808080", "#FFFFFF"))

    @classmethod
    def format_document_number(cls) -> str:
        now = datetime.now()
        return f"ADR-{now.strftime('%Y%m%d')}-{now.strftime('%H%M%S')}"


# =============================================================================
# KIMYASAL EKLE / DUZENLE DIALOG
# =============================================================================

class ChemicalDialog(QDialog):
    """Kimyasal madde ekle veya duzenle dialog penceresi."""

    def __init__(self, parent=None, chemical: Chemical = None):
        super().__init__(parent)
        self.chemical = chemical
        self.setWindowTitle("Yeni Kimyasal Ekle" if not chemical else f"Duzenle: UN{chemical.un_number}")
        self.setMinimumWidth(560)
        self.setMinimumHeight(500)
        self._setup_ui()
        self.setWindowTitle("ADR Transport Pro 2026")
        self.setWindowIcon(QIcon("icon.ico"))
        if chemical:
            self._populate(chemical)

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        tab = QTabWidget()

        # --- Genel Bilgiler ---
        general_widget = QWidget()
        form = QFormLayout(general_widget)
        form.setSpacing(10)

        self.txt_un       = QLineEdit()
        self.txt_un.setPlaceholderText("ornek: 1203")
        self.txt_name_tr  = QLineEdit()
        self.txt_name_tr.setPlaceholderText("ornek: BENZIN")
        self.txt_name_en  = QLineEdit()
        self.txt_name_en.setPlaceholderText("ornek: GASOLINE")

        self.cmb_class = QComboBox()
        self.cmb_class.addItems([
            "",
            "1.1", "1.2", "1.3", "1.4", "1.5", "1.6",   # Patlayicilar — alt grup zorunlu
            "2.1", "2.2", "2.3",
            "3", "4.1", "4.2", "4.3",
            "5.1", "5.2", "6.1", "6.2",
            "7", "8", "9"
        ])

        self.cmb_pg = QComboBox()
        self.cmb_pg.addItems(["", "I", "II", "III"])

        # Sınıflandırma kodu (ADR Tablo 1.10.3.1.2 — Sınıf 1 alt grubu, Sınıf 2 gaz kodu vb.)
        # Örnekler: 1.4S, 1.4B, TF, FC, T, F …
        self.txt_classification_code = QLineEdit()
        self.txt_classification_code.setPlaceholderText("Ör: 1.4S  /  TF  /  FC")
        self.txt_classification_code.setToolTip(
            "ADR Liste Sütun 3b — Sınıflandırma kodu.\n"
            "Sınıf 1 için: 1.1A, 1.1B, 1.2A, 1.3C, 1.4S, 1.4B, 1.5D vb.\n"
            "Sınıf 2 için: T, TF, TC, TO, TFC, TOC, F, FC, A, AF vb.\n"
            "Diğer sınıflar için boş bırakılabilir."
        )
        self._row_classification_code = None   # form satır referansı

        self.txt_tunnel = QLineEdit()
        self.txt_tunnel.setPlaceholderText("ornek: D/E")

        self.cmb_tc = QComboBox()
        self.cmb_tc.addItems(["", "0", "1", "2", "3", "4"])

        self.txt_segregation = QComboBox()
        # ADR segregasyon grupları — Türkçe tam adlarıyla
        _seg_options = [
            "",
            "Yanıcı Sıvılar (Sınıf 3)",
            "Yanıcı Katılar (Sınıf 4.1)",
            "Kendiliğinden Yanabilir Maddeler (Sınıf 4.2)",
            "Su ile Tepkiyen Maddeler (Sınıf 4.3)",
            "Yükseltgen Maddeler (Sınıf 5.1)",
            "Organik Peroksitler (Sınıf 5.2)",
            "Zehirli Maddeler (Sınıf 6.1)",
            "Bulaşıcı Maddeler (Sınıf 6.2)",
            "Radyoaktif Maddeler (Sınıf 7)",
            "Aşındırıcı Maddeler (Sınıf 8)",
            "Çeşitli Tehlikeli Maddeler (Sınıf 9)",
            "Patlayıcılar (Sınıf 1)",
            "Gazlar — Yanıcı (Sınıf 2.1)",
            "Gazlar — Yanıcı Olmayan / Zehirsiz (Sınıf 2.2)",
            "Gazlar — Zehirli (Sınıf 2.3)",
            "Asitler",
            "Alkaliler",
            "Oksitleyiciler",
            "Organik Peroksitler",
        ]
        self.txt_segregation.addItems(_seg_options)
        self.txt_segregation.setEditable(True)

        # Tehlike etiketi — hem metin girişi hem görsel önizleme
        _hazard_container = QWidget()
        _hazard_layout = QVBoxLayout(_hazard_container)
        _hazard_layout.setContentsMargins(0, 0, 0, 0)
        _hazard_layout.setSpacing(4)

        self.txt_hazard_labels = QLineEdit()
        self.txt_hazard_labels.setPlaceholderText("örnek: 3, 8  veya  3+8")
        self.txt_hazard_labels.textChanged.connect(self._update_hazard_preview)
        self.cmb_class.currentTextChanged.connect(self._on_class_changed)

        # Görsel önizleme alanı — renkli kutularla ADR etiket renk kodları
        self.hazard_preview_widget = QWidget()
        self.hazard_preview_widget.setFixedHeight(52)
        self.hazard_preview_layout = QHBoxLayout(self.hazard_preview_widget)
        self.hazard_preview_layout.setContentsMargins(0, 0, 0, 0)
        self.hazard_preview_layout.setSpacing(4)
        self.hazard_preview_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)

        _hazard_layout.addWidget(self.txt_hazard_labels)
        _hazard_layout.addWidget(self.hazard_preview_widget)

        self.spin_flash = QDoubleSpinBox()
        self.spin_flash.setRange(-100, 400)
        self.spin_flash.setSpecialValueText("Belirtilmemis")
        self.spin_flash.setValue(-100)
        self.spin_flash.setSuffix(" °C")

        form.addRow("UN Numarasi *:",        self.txt_un)
        form.addRow("Teknik Ad (TR) *:",     self.txt_name_tr)
        form.addRow("Teknik Ad (EN):",       self.txt_name_en)
        form.addRow("Sinif *:",              self.cmb_class)
        self._row_classification_code = form.addRow(
            "Sınıflandırma Kodu:", self.txt_classification_code)
        form.addRow("Ambalaj Grubu (PG):",   self.cmb_pg)
        form.addRow("Tunel Kodu:",           self.txt_tunnel)
        form.addRow("Tasima Kategorisi:",    self.cmb_tc)
        form.addRow("Segregasyon Grubu:",    self.txt_segregation)
        form.addRow("Tehlike Etiketleri:",   _hazard_container)

        tab.addTab(general_widget, "Genel Bilgiler")

        # --- Muafiyet & Ozel ---
        detail_widget = QWidget()
        detail_scroll = QScrollArea()
        detail_scroll.setWidgetResizable(True)
        detail_scroll_inner = QWidget()
        detail_vbox = QVBoxLayout(detail_scroll_inner)
        detail_vbox.setSpacing(10)
        detail_scroll.setWidget(detail_scroll_inner)

        # LQ Kartı
        lq_frame = QFrame()
        lq_frame.setFrameShape(QFrame.Shape.StyledPanel)
        lq_frame.setStyleSheet("QFrame { border: 2px solid #A6E3A1; border-radius: 6px; background: #f0fff0; }")
        lq_inner = QVBoxLayout(lq_frame)
        lq_inner.setContentsMargins(8, 6, 8, 6)
        lq_title = QLabel("📦  LQ — Sınırlı Miktar (ADR Bölüm 3.4)")
        lq_title.setStyleSheet("font-weight: bold; font-size: 12px; color: #2d6a2d; border: none;")
        lq_desc = QLabel(
            "Sınırlı miktar kapsamında taşınabilecek madde. "
            "Ambalaj başına maksimum miktar ADR Tablo 3.4'te belirlidir.\n"
            "LQ kapsamındaki maddelerde turuncu plaka, belge ve tehlike etiketi gerekmez."
        )
        lq_desc.setWordWrap(True)
        lq_desc.setStyleSheet("font-size: 11px; color: #444; border: none;")
        self.chk_lq = QCheckBox("Bu madde LQ (Sınırlı Miktar) kapsamında taşınabilir")
        self.chk_lq.setStyleSheet("font-size: 12px; border: none;")

        self.spin_lq_max = QDoubleSpinBox()
        self.spin_lq_max.setRange(0, 10000)
        self.spin_lq_max.setSuffix(" kg/L")
        self.spin_lq_max.setSpecialValueText("Belirtilmemiş")
        self.spin_lq_max.setValue(0)
        self.spin_lq_max.setToolTip("Ambalaj başına maksimum net miktar")
        lq_max_row = QHBoxLayout()
        lq_max_row.addWidget(QLabel("Ambalaj başına maks:"))
        lq_max_row.addWidget(self.spin_lq_max)
        lq_max_row.addStretch()

        lq_inner.addWidget(lq_title)
        lq_inner.addWidget(lq_desc)
        lq_inner.addWidget(self.chk_lq)
        lq_inner.addLayout(lq_max_row)

        # EQ Kartı
        eq_frame = QFrame()
        eq_frame.setFrameShape(QFrame.Shape.StyledPanel)
        eq_frame.setStyleSheet("QFrame { border: 2px solid #89B4FA; border-radius: 6px; background: #f0f4ff; }")
        eq_inner = QVBoxLayout(eq_frame)
        eq_inner.setContentsMargins(8, 6, 8, 6)
        eq_title = QLabel("🔬  EQ — İstisnai Miktar (ADR Bölüm 3.5)")
        eq_title.setStyleSheet("font-weight: bold; font-size: 12px; color: #1a3d6e; border: none;")
        eq_desc = QLabel(
            "İstisnai miktar kapsamında taşınan madde. "
            "E1–E5 kodlarına göre maksimum miktarlar belirlenir.\n"
            "EQ kapsamındaki maddelerde ADR belgesi ve turuncu plaka gerekmez; "
            "yalnızca ambalaj üzerine EQ işareti zorunludur."
        )
        eq_desc.setWordWrap(True)
        eq_desc.setStyleSheet("font-size: 11px; color: #444; border: none;")
        self.chk_eq = QCheckBox("Bu madde EQ (İstisnai Miktar) kapsamında taşınabilir")
        self.chk_eq.setStyleSheet("font-size: 12px; border: none;")

        self.cmb_eq_code = QComboBox()
        self.cmb_eq_code.addItems(["", "E1 (≤1 ml/g)", "E2 (≤30 ml/g)", "E3 (≤30 ml, ≤3 g)", "E4 (≤30 g)", "E5 (≤1 g/ml)"])
        self.cmb_eq_code.setToolTip("ADR EQ kodu - ambalaj başına maksimum iç ambalaj miktarı")
        eq_code_row = QHBoxLayout()
        eq_code_row.addWidget(QLabel("EQ Kodu:"))
        eq_code_row.addWidget(self.cmb_eq_code)
        eq_code_row.addStretch()

        eq_inner.addWidget(eq_title)
        eq_inner.addWidget(eq_desc)
        eq_inner.addWidget(self.chk_eq)
        eq_inner.addLayout(eq_code_row)

        # Özel Hükümler Kartı
        sp_frame = QFrame()
        sp_frame.setFrameShape(QFrame.Shape.StyledPanel)
        sp_frame.setStyleSheet("QFrame { border: 2px solid #F9E2AF; border-radius: 6px; background: #fffdf0; }")
        sp_inner = QVBoxLayout(sp_frame)
        sp_inner.setContentsMargins(8, 6, 8, 6)
        sp_title = QLabel("📋  Özel Hükümler (ADR 3.3 — SP Kodları)")
        sp_title.setStyleSheet("font-weight: bold; font-size: 12px; color: #7a5c00; border: none;")
        sp_desc = QLabel(
            "Maddeye özgü özel hükümler, CV kodları ve S kodlarını girin.\n"
            "Örnek: SP274, SP598, CV13, S2"
        )
        sp_desc.setWordWrap(True)
        sp_desc.setStyleSheet("font-size: 11px; color: #444; border: none;")
        self.txt_special = QTextEdit()
        self.txt_special.setMinimumHeight(70)
        self.txt_special.setMaximumHeight(110)
        self.txt_special.setPlaceholderText(
            "Özel hükümler — SP kodları, CV kodları, S kodları...\n"
            "Örnek: SP274 | CV13 | S2"
        )

        sp_inner.addWidget(sp_title)
        sp_inner.addWidget(sp_desc)
        sp_inner.addWidget(self.txt_special)

        detail_vbox.addWidget(lq_frame)
        detail_vbox.addWidget(eq_frame)
        detail_vbox.addWidget(sp_frame)
        detail_vbox.addStretch()

        detail_tab_widget = QWidget()
        detail_tab_layout = QVBoxLayout(detail_tab_widget)
        detail_tab_layout.setContentsMargins(0, 0, 0, 0)
        detail_tab_layout.addWidget(detail_scroll)

        # LQ/EQ ve özel hükümler kartlarla gösteriliyor (yukarıda detail_tab_widget)

        tab.addTab(detail_tab_widget, "Muafiyet & Özel Hükümler")

        layout.addWidget(tab)

        # Butonlar
        btn_layout = QHBoxLayout()
        btn_save   = QPushButton("Kaydet")
        btn_save.setObjectName("success")
        btn_save.setMinimumHeight(36)
        btn_save.clicked.connect(self._on_save)
        btn_cancel = QPushButton("Iptal")
        btn_cancel.setMinimumHeight(36)
        btn_cancel.clicked.connect(self.reject)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_save)
        btn_layout.addWidget(btn_cancel)
        layout.addLayout(btn_layout)

    def _populate(self, c: Chemical):
        self.txt_un.setText(c.un_number)
        self.txt_name_tr.setText(c.proper_shipping_name_tr)
        self.txt_name_en.setText(c.proper_shipping_name_en)
        idx = self.cmb_class.findText(c.class_code)
        self.cmb_class.setCurrentIndex(max(0, idx))
        idx = self.cmb_pg.findText(c.packing_group)
        self.cmb_pg.setCurrentIndex(max(0, idx))
        self.txt_tunnel.setText(c.tunnel_code)
        idx = self.cmb_tc.findText(c.transport_category)
        self.cmb_tc.setCurrentIndex(max(0, idx))
        # Segregasyon: önce tam eşleşme, yoksa ilgili metni içeren seçenek
        seg_text = c.segregation_group or ""
        seg_idx = self.txt_segregation.findText(seg_text)
        if seg_idx >= 0:
            self.txt_segregation.setCurrentIndex(seg_idx)
        else:
            self.txt_segregation.setEditText(seg_text)
        self.txt_hazard_labels.setText(c.hazard_labels)
        self._update_hazard_preview(c.hazard_labels)
        self.chk_lq.setChecked(c.lq_allowed)
        self.chk_eq.setChecked(c.eq_allowed)
        self.txt_special.setPlainText(c.special_provisions)

    def _on_save(self):
        un = self.txt_un.text().strip()
        name_tr = self.txt_name_tr.text().strip()
        class_code = self.cmb_class.currentText()

        if not un:
            QMessageBox.warning(self, "Hata", "UN numarasi zorunlu!")
            return
        if not name_tr:
            QMessageBox.warning(self, "Hata", "Teknik ad (TR) zorunlu!")
            return
        if not class_code:
            QMessageBox.warning(self, "Hata", "Sinif secimi zorunlu!")
            return

        flash = None
        if self.spin_flash.value() > -100:
            flash = self.spin_flash.value()

        self.result_chemical = Chemical(
            id=self.chemical.id if self.chemical else None,
            un_number=un,
            proper_shipping_name_tr=name_tr,
            proper_shipping_name_en=self.txt_name_en.text().strip(),
            class_code=class_code,
            packing_group=self.cmb_pg.currentText(),
            tunnel_code=self.txt_tunnel.text().strip(),
            transport_category=self.cmb_tc.currentText(),
            segregation_group=self.txt_segregation.currentText().strip(),
            special_provisions=self.txt_special.toPlainText().strip(),
            lq_allowed=self.chk_lq.isChecked(),
            eq_allowed=self.chk_eq.isChecked(),
            hazard_labels=self.txt_hazard_labels.text().strip()
        )
        # classification_code Chemical'da alan yok; ayrı attr olarak sakla.
        # ShipmentEditorPage ShipmentItem oluştururken buradan okur.
        self.result_chemical._classification_code = self.txt_classification_code.text().strip()
        self.accept()

    def _on_class_changed(self, cls: str):
        """Sınıf seçilince sınıflandırma kodu alanını göster/gizle ve ipucu ver."""
        show = cls.startswith("1") or cls in ("2.1", "2.2", "2.3")
        self.txt_classification_code.setVisible(show)
        # Sınıf 1 için örnek değer ipucu
        if cls.startswith("1"):
            sub = cls  # "1.1", "1.4" vb.
            self.txt_classification_code.setPlaceholderText(
                f"Ör: {sub}S  /  {sub}B  /  {sub}G  (ADR Liste Sütun 3b)"
            )
        elif cls in ("2.1", "2.2", "2.3"):
            self.txt_classification_code.setPlaceholderText("Ör: F  /  FC  /  T  /  TF  /  A")
        if not show:
            self.txt_classification_code.clear()

    def _update_hazard_preview(self, text: str = ""):
        """
        Tehlike etiketi metnine göre renkli ADR etiket kutuları çizer.
        ADR 5.2.2 uyarınca her sınıfın rengi ve sembolü gösterilir.
        """
        # Renk, kısa kod, sembol
        _LABEL_STYLES = {
            "1":   ("#FF6600", "#FFFFFF", "💥 1"),
            "1.4": ("#FF6600", "#FFFFFF", "💥 1.4"),
            "1.5": ("#FF6600", "#FFFFFF", "💥 1.5"),
            "1.6": ("#FF6600", "#FFFFFF", "💥 1.6"),
            "2.1": ("#FF3333", "#FFFFFF", "🔥 2.1"),
            "2.2": ("#33AA33", "#FFFFFF", "⬛ 2.2"),
            "2.3": ("#FFFFFF", "#333333", "☠ 2.3"),
            "3":   ("#FF3333", "#FFFFFF", "🔥 3"),
            "4.1": ("#FF6600", "#FFFFFF", "🔥 4.1"),
            "4.2": ("#FF6600", "#FFFFFF", "🔥 4.2"),
            "4.3": ("#0055CC", "#FFFFFF", "💧 4.3"),
            "5.1": ("#FFCC00", "#333333", "⬛ 5.1"),
            "5.2": ("#FF3333", "#FFFFFF", "💥 5.2"),
            "6.1": ("#FFFFFF", "#333333", "☠ 6.1"),
            "6.2": ("#FFFFFF", "#333333", "⚕ 6.2"),
            "7":   ("#FFFF00", "#333333", "☢ 7"),
            "8":   ("#333333", "#FFFFFF", "⚗ 8"),
            "9":   ("#CCCCCC", "#333333", "⬛ 9"),
        }

        # Mevcut etiket kutularını temizle
        while self.hazard_preview_layout.count():
            item = self.hazard_preview_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        if not text.strip():
            return

        import re
        codes = re.split(r'[,;+\s]+', text.strip())
        for code in codes:
            code = code.strip()
            if not code:
                continue
            style = _LABEL_STYLES.get(code, ("#888888", "#FFFFFF", f"⬛ {code}"))
            bg, fg, symbol = style

            lbl = QLabel(symbol)
            lbl.setFixedSize(48, 48)
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setStyleSheet(
                f"background-color: {bg}; color: {fg}; "
                f"border: 2px solid #333; border-radius: 4px; "
                f"font-size: 11px; font-weight: bold;"
            )
            lbl.setToolTip(f"ADR Tehlike Etiketi — Sınıf {code}")
            self.hazard_preview_layout.addWidget(lbl)

    def get_chemical(self) -> Optional[Chemical]:
        return getattr(self, 'result_chemical', None)


# =============================================================================
# KAYITLI EXCEL VERİSİ SAYFASI
# =============================================================================

class ExcelChemicalPage(QWidget):
    """
    Excel dosyasından (VERİ sayfası) okunan kimyasal kayıtları listeler.
    Ürün adıyla arama yapılır, seçilen ürün Taşıma Evrakı sayfasına eklenir.
    Her ürün yalnızca bir kez eklenebilir.
    """

    # Taşıma evrakı sayfasına ürün eklemek için sinyal
    product_add_requested = pyqtSignal(dict)
    # Excel yüklendiğinde tüm ürün listesini bildiren sinyal
    excel_loaded = pyqtSignal(list)

    def __init__(self, db=None, parent=None):
        super().__init__(parent)
        self.parent_window = parent
        self.all_products = []       # Excel'den yüklenen tüm ürünler
        self.added_products = set()  # Eklenen ürün adları (tekrar önleme)
        self._setup_ui()
        self._load_excel_data()

    # ------------------------------------------------------------------ UI --
    def _setup_ui(self):
        # ── Ana renk paleti (ADR endüstriyel teması) ──────────────────────
        C = {
            "bg_page":    "#0D1117",   # Sayfa arka planı — derin siyah
            "bg_panel":   "#161B22",   # Panel arka planı
            "bg_card":    "#1C2128",   # Kart/grup arka planı
            "bg_row_alt": "#1A1F26",   # Tablo alternatif satır
            "bg_input":   "#21262D",   # Input alanı
            "border":     "#30363D",   # Çerçeve rengi
            "border_hi":  "#E6A817",   # Vurgulu çerçeve (ADR sarısı)
            "accent":     "#E6A817",   # ADR uyarı sarısı
            "accent2":    "#F0611A",   # Tehlike turuncusu
            "ok":         "#3FB950",   # Onay yeşili
            "text_hi":    "#F0F6FC",   # Birincil metin
            "text_lo":    "#8B949E",   # İkincil metin
            "text_acc":   "#E6A817",   # Vurgulu metin
            "added_bg":   "#1A2A1A",   # Eklenen satır arka planı
            "added_fg":   "#3FB950",   # Eklenen satır metni
            "sel_bg":     "#2D3748",   # Seçili satır
        }
        self._C = C

        self.setStyleSheet(f"""
            QWidget {{
                background-color: {C['bg_page']};
                color: {C['text_hi']};
                font-family: 'Segoe UI', 'Inter', 'Roboto', sans-serif;
                font-size: 10pt;
            }}
            QGroupBox {{
                background-color: {C['bg_card']};
                border: 1px solid {C['border']};
                border-radius: 6px;
                margin-top: 10px;
                padding-top: 6px;
                font-size: 9pt;
                font-weight: bold;
                color: {C['text_lo']};
                letter-spacing: 1px;
                text-transform: uppercase;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 10px;
                padding: 0 6px;
                background-color: {C['bg_card']};
                color: {C['text_lo']};
            }}
            QLineEdit {{
                background-color: {C['bg_input']};
                border: 1px solid {C['border']};
                border-radius: 4px;
                padding: 5px 10px;
                color: {C['text_hi']};
                font-size: 10pt;
                selection-background-color: {C['accent']};
                selection-color: #000000;
            }}
            QLineEdit:focus {{
                border: 1px solid {C['border_hi']};
            }}
            QLineEdit:placeholder {{
                color: {C['text_lo']};
            }}
            QPushButton {{
                background-color: {C['bg_input']};
                border: 1px solid {C['border']};
                border-radius: 4px;
                padding: 6px 14px;
                color: {C['text_hi']};
                font-size: 9pt;
                font-weight: 600;
            }}
            QPushButton:hover {{
                background-color: #2D333B;
                border-color: {C['text_lo']};
            }}
            QPushButton:pressed {{
                background-color: {C['bg_card']};
            }}
            QTableWidget {{
                background-color: {C['bg_panel']};
                alternate-background-color: {C['bg_row_alt']};
                border: 1px solid {C['border']};
                border-radius: 4px;
                gridline-color: {C['border']};
                color: {C['text_hi']};
                selection-background-color: {C['sel_bg']};
                selection-color: {C['text_hi']};
                outline: none;
            }}
            QTableWidget::item {{
                padding: 4px 8px;
                border: none;
            }}
            QTableWidget::item:selected {{
                background-color: #2D3748;
                color: {C['text_hi']};
            }}
            QHeaderView::section {{
                background-color: {C['bg_card']};
                color: {C['text_lo']};
                border: none;
                border-bottom: 2px solid {C['accent']};
                padding: 6px 8px;
                font-size: 8pt;
                font-weight: 700;
                letter-spacing: 1px;
                text-transform: uppercase;
            }}
            QListWidget {{
                background-color: {C['bg_panel']};
                border: 1px solid {C['border']};
                border-radius: 4px;
                color: {C['text_hi']};
                outline: none;
            }}
            QListWidget::item {{
                padding: 4px 8px;
                border-bottom: 1px solid {C['border']};
            }}
            QListWidget::item:selected {{
                background-color: {C['sel_bg']};
            }}
            QScrollBar:vertical {{
                background: {C['bg_card']};
                width: 8px;
                border-radius: 4px;
            }}
            QScrollBar::handle:vertical {{
                background: {C['border']};
                border-radius: 4px;
                min-height: 20px;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {C['text_lo']};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0;
            }}
            QLabel#header_title {{
                font-size: 13pt;
                font-weight: 700;
                color: {C['text_hi']};
                letter-spacing: 2px;
            }}
            QLabel#header_sub {{
                font-size: 8pt;
                color: {C['text_lo']};
                letter-spacing: 1px;
            }}
            QLabel#stat_box {{
                background-color: {C['bg_card']};
                border: 1px solid {C['border']};
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 9pt;
                color: {C['text_lo']};
            }}
            QLabel#file_badge {{
                background-color: {C['bg_card']};
                border: 1px solid {C['border']};
                border-left: 3px solid {C['accent']};
                border-radius: 4px;
                padding: 5px 10px;
                font-size: 9pt;
                color: {C['text_lo']};
            }}
            QPushButton#btn_load {{
                background-color: transparent;
                border: 1px solid {C['accent']};
                color: {C['accent']};
                border-radius: 4px;
                padding: 6px 16px;
                font-weight: 700;
                font-size: 9pt;
            }}
            QPushButton#btn_load:hover {{
                background-color: rgba(230,168,23,0.12);
            }}
            QPushButton#btn_add_main {{
                background-color: {C['ok']};
                border: none;
                border-radius: 4px;
                color: #0D1117;
                font-size: 10pt;
                font-weight: 700;
                padding: 8px 20px;
                letter-spacing: 1px;
            }}
            QPushButton#btn_add_main:hover {{
                background-color: #4DC762;
            }}
            QPushButton#btn_add_main:disabled {{
                background-color: #2D333B;
                color: {C['text_lo']};
            }}
            QPushButton#btn_reset {{
                background-color: transparent;
                border: 1px solid {C['border']};
                color: {C['text_lo']};
                border-radius: 4px;
                padding: 6px 14px;
                font-size: 9pt;
            }}
            QPushButton#btn_reset:hover {{
                border-color: {C['accent2']};
                color: {C['accent2']};
            }}
            QPushButton#btn_clear_search {{
                background-color: transparent;
                border: 1px solid {C['border']};
                color: {C['text_lo']};
                border-radius: 4px;
                padding: 5px 10px;
                font-size: 9pt;
            }}
            QPushButton#btn_clear_search:hover {{
                border-color: {C['text_lo']};
                color: {C['text_hi']};
            }}
            QFrame#divider {{
                background-color: {C['border']};
                max-height: 1px;
            }}
            QFrame#detail_card {{
                background-color: {C['bg_card']};
                border: 1px solid {C['border']};
                border-radius: 6px;
            }}
        """)

        # ── ANA LAYOUT (yatay splitter) ───────────────────────────────────
        outer = QHBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setHandleWidth(1)
        splitter.setStyleSheet(f"QSplitter::handle {{ background-color: {C['border']}; }}")
        outer.addWidget(splitter)

        # ── SOL PANEL: Başlık + Dosya + Arama + Tablo ─────────────────────
        left_widget = QWidget()
        left_widget.setStyleSheet(f"background-color: {C['bg_page']};")
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(16, 16, 8, 16)
        left_layout.setSpacing(10)

        # ── Başlık bandı ──────────────────────────────────────────────────
        header_frame = QFrame()
        header_frame.setStyleSheet(f"""
            QFrame {{
                background-color: {C['bg_panel']};
                border: 1px solid {C['border']};
                border-left: 4px solid {C['accent']};
                border-radius: 6px;
            }}
        """)
        hf_layout = QHBoxLayout(header_frame)
        hf_layout.setContentsMargins(14, 10, 14, 10)
        hf_layout.setSpacing(14)

        # İkon alanı (⚗ sembolü)
        icon_lbl = QLabel("⚗")
        icon_lbl.setStyleSheet(f"font-size: 22pt; color: {C['accent']}; background: transparent; border: none;")
        hf_layout.addWidget(icon_lbl)

        title_col = QVBoxLayout()
        title_col.setSpacing(1)
        lbl_title = QLabel("KAYITLI EXCEL VERİSİ")
        lbl_title.setObjectName("header_title")
        lbl_title.setStyleSheet(f"font-size: 13pt; font-weight: 700; color: {C['text_hi']}; letter-spacing: 2px; background: transparent; border: none;")
        lbl_sub = QLabel("ADR Kimyasal Madde Veritabanı — Excel Kaynağı")
        lbl_sub.setStyleSheet(f"font-size: 8pt; color: {C['text_lo']}; letter-spacing: 1px; background: transparent; border: none;")
        title_col.addWidget(lbl_title)
        title_col.addWidget(lbl_sub)
        hf_layout.addLayout(title_col, 1)

        # Stat rozetleri
        self.lbl_total_badge = self._make_stat_badge("TOPLAM", "0")
        self.lbl_filtered_badge = self._make_stat_badge("FİLTRELİ", "0")
        self.lbl_added_badge = self._make_stat_badge("EKLENEN", "0", C['ok'])
        hf_layout.addWidget(self.lbl_total_badge)
        hf_layout.addWidget(self.lbl_filtered_badge)
        hf_layout.addWidget(self.lbl_added_badge)

        left_layout.addWidget(header_frame)

        # ── Dosya yükleme satırı ──────────────────────────────────────────
        file_row = QHBoxLayout()
        file_row.setSpacing(8)

        self.lbl_file = QLabel("Dosya yüklenmedi — Excel seçmek için düğmeye tıklayın")
        self.lbl_file.setObjectName("file_badge")
        self.lbl_file.setMinimumHeight(32)
        file_row.addWidget(self.lbl_file, 1)

        btn_load = QPushButton("  EXCEL SEÇ")
        btn_load.setObjectName("btn_load")
        btn_load.setMinimumHeight(32)
        btn_load.setFixedWidth(130)
        btn_load.clicked.connect(self._browse_excel)
        file_row.addWidget(btn_load)
        left_layout.addLayout(file_row)

        # ── Arama satırı ─────────────────────────────────────────────────
        search_frame = QFrame()
        search_frame.setStyleSheet(f"""
            QFrame {{
                background-color: {C['bg_panel']};
                border: 1px solid {C['border']};
                border-radius: 5px;
            }}
        """)
        sf_layout = QHBoxLayout(search_frame)
        sf_layout.setContentsMargins(10, 0, 6, 0)
        sf_layout.setSpacing(6)

        lbl_search_icon = QLabel("🔍")
        lbl_search_icon.setStyleSheet("background: transparent; border: none; font-size: 11pt;")
        sf_layout.addWidget(lbl_search_icon)

        self.txt_search = QLineEdit()
        self.txt_search.setPlaceholderText("Kimyasal adı, sevkiyat adı veya UN numarası ile arayın...")
        self.txt_search.setStyleSheet(f"""
            QLineEdit {{
                background: transparent;
                border: none;
                padding: 7px 0;
                color: {C['text_hi']};
                font-size: 10pt;
            }}
            QLineEdit:focus {{ border: none; outline: none; }}
        """)
        self.txt_search.textChanged.connect(self._filter_table)
        sf_layout.addWidget(self.txt_search, 1)

        btn_clear = QPushButton("✕")
        btn_clear.setObjectName("btn_clear_search")
        btn_clear.setFixedSize(28, 28)
        btn_clear.setToolTip("Aramayı temizle")
        btn_clear.clicked.connect(self._clear_search)
        sf_layout.addWidget(btn_clear)

        left_layout.addWidget(search_frame)

        # ── Ana tablo ────────────────────────────────────────────────────
        self.tbl = QTableWidget()
        headers = [
            "KİMYASAL ADI", "UN NO", "SEVKİYAT ADI (TAŞIMA EVRAKI)",
            "SINIF", "PG", "TÜNEL", "DURUM"
        ]
        self.tbl.setColumnCount(len(headers))
        self.tbl.setHorizontalHeaderLabels(headers)
        self.tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        self.tbl.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.tbl.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.tbl.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.tbl.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        self.tbl.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        self.tbl.setColumnWidth(0, 210)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tbl.setAlternatingRowColors(True)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setShowGrid(False)
        self.tbl.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.tbl.doubleClicked.connect(self._add_selected)
        self.tbl.setMinimumWidth(540)
        self.tbl.verticalHeader().setDefaultSectionSize(30)
        left_layout.addWidget(self.tbl, 1)

        # ── Tablo alt: ipucu metni ────────────────────────────────────────
        hint_lbl = QLabel("  Çift tıklayarak veya sağ paneldeki düğmeyle ekleyin  ·  Her sevkiyat adı yalnızca bir kez eklenebilir")
        hint_lbl.setStyleSheet(f"color: {C['text_lo']}; font-size: 8pt; padding: 2px 0; background: transparent;")
        left_layout.addWidget(hint_lbl)

        splitter.addWidget(left_widget)

        # ── SAĞ PANEL: Detay kartı + Ekleme + Eklenenler ─────────────────
        right_widget = QWidget()
        right_widget.setStyleSheet(f"background-color: {C['bg_page']};")
        right_widget.setMinimumWidth(300)
        right_widget.setMaximumWidth(380)
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(8, 16, 16, 16)
        right_layout.setSpacing(10)

        # ── Detay başlığı ────────────────────────────────────────────────
        det_header = QLabel("⬡  ÜRÜN DETAYI")
        det_header.setStyleSheet("font-size:9pt;font-weight:700;letter-spacing:2px;color:#585B70;padding-bottom:6px;border-bottom:2px solid #313244;background:transparent;")
        right_layout.addWidget(det_header)

        # ── Detay kart ───────────────────────────────────────────────────
        self.detail_frame = QFrame()
        self.detail_frame.setObjectName("detail_card")
        self.detail_frame.setStyleSheet(f"""
            QFrame#detail_card {{
                background-color: {C['bg_card']};
                border: 1px solid {C['border']};
                border-radius: 6px;
            }}
        """)
        det_layout = QVBoxLayout(self.detail_frame)
        det_layout.setContentsMargins(14, 12, 14, 12)
        det_layout.setSpacing(0)

        self.det_labels = {}
        det_fields = [
            ("Kimyasal Adı",      True),
            ("UN Numarası",       True),
            ("Sevkiyat Adı",      True),
            ("Ambalaj Türü",      False),
            ("Sınıf",             False),
            ("PG",                False),
            ("Tünel Kodu",        False),
            ("Taşıma Kategorisi", False),
            ("Sınırlı Miktar",    False),
            ("Özel Hükümler",     False),
            ("Tank Kodu",         False),
            ("Ambalaj Talimatları", False),
        ]

        for i, (field, is_primary) in enumerate(det_fields):
            row_w = QWidget()
            row_w.setStyleSheet(f"""
                QWidget {{
                    background-color: transparent;
                    border-bottom: 1px solid {C['border'] if i < len(det_fields)-1 else 'transparent'};
                }}
            """)
            row_h = QHBoxLayout(row_w)
            row_h.setContentsMargins(0, 6, 0, 6)
            row_h.setSpacing(8)

            lbl_key = QLabel(field.upper())
            lbl_key.setFixedWidth(128)
            lbl_key.setStyleSheet(f"""
                color: {C['text_lo']};
                font-size: 7pt;
                font-weight: 700;
                letter-spacing: 0.5px;
                background: transparent;
                border: none;
            """)
            lbl_key.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
            row_h.addWidget(lbl_key)

            lbl_val = QLabel("—")
            lbl_val.setWordWrap(True)
            lbl_val.setStyleSheet(f"""
                color: {C['text_hi'] if is_primary else C['text_lo']};
                font-size: {'10pt' if is_primary else '9pt'};
                font-weight: {'600' if is_primary else '400'};
                background: transparent;
                border: none;
            """)
            row_h.addWidget(lbl_val, 1)

            det_layout.addWidget(row_w)
            self.det_labels[field] = lbl_val

        right_layout.addWidget(self.detail_frame)

        # ── UN rozeti ────────────────────────────────────────────────────
        self.un_badge_frame = QFrame()
        self.un_badge_frame.setStyleSheet(f"""
            QFrame {{
                background-color: {C['bg_card']};
                border: 1px solid {C['border']};
                border-left: 4px solid {C['accent2']};
                border-radius: 4px;
            }}
        """)
        ub_layout = QHBoxLayout(self.un_badge_frame)
        ub_layout.setContentsMargins(12, 8, 12, 8)

        lbl_un_icon = QLabel("⚠")
        lbl_un_icon.setStyleSheet(f"font-size: 14pt; color: {C['accent2']}; background: transparent; border: none;")
        ub_layout.addWidget(lbl_un_icon)

        un_text_col = QVBoxLayout()
        un_text_col.setSpacing(0)
        lbl_un_key = QLabel("UN NUMARASI")
        lbl_un_key.setStyleSheet(f"font-size: 7pt; font-weight: 700; color: {C['text_lo']}; letter-spacing: 1px; background: transparent; border: none;")
        self.lbl_un_val = QLabel("—")
        self.lbl_un_val.setStyleSheet(f"font-size: 14pt; font-weight: 700; color: {C['accent2']}; background: transparent; border: none;")
        un_text_col.addWidget(lbl_un_key)
        un_text_col.addWidget(self.lbl_un_val)
        ub_layout.addLayout(un_text_col, 1)

        self.lbl_class_badge = QLabel("—")
        self.lbl_class_badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_class_badge.setFixedSize(52, 52)
        self.lbl_class_badge.setStyleSheet(f"""
            background-color: {C['border']};
            color: {C['text_hi']};
            border-radius: 4px;
            font-size: 11pt;
            font-weight: 700;
            border: none;
        """)
        ub_layout.addWidget(self.lbl_class_badge)

        right_layout.addWidget(self.un_badge_frame)

        # ── Ekle butonu ───────────────────────────────────────────────────
        self.btn_add = QPushButton("TAŞIMA EVRAKINA EKLE  ➜")
        self.btn_add.setObjectName("btn_add_main")
        self.btn_add.setMinimumHeight(42)
        self.btn_add.clicked.connect(self._add_selected)
        right_layout.addWidget(self.btn_add)

        # ── Sıfırla butonu ───────────────────────────────────────────────
        btn_reset = QPushButton("Eklenenler Listesini Sıfırla")
        btn_reset.setObjectName("btn_reset")
        btn_reset.clicked.connect(self._reset_added)
        right_layout.addWidget(btn_reset)

        # ── Divider ──────────────────────────────────────────────────────
        div = QFrame()
        div.setFrameShape(QFrame.Shape.HLine)
        div.setStyleSheet(f"background-color: {C['border']}; max-height: 1px; border: none;")
        right_layout.addWidget(div)

        # ── Eklenen ürünler listesi ───────────────────────────────────────
        added_header = QLabel("EKLENEN ÜRÜNLER")
        added_header.setStyleSheet(f"""
            color: {C['text_lo']};
            font-size: 8pt;
            font-weight: 700;
            letter-spacing: 2px;
            background: transparent;
        """)
        right_layout.addWidget(added_header)

        self.list_added = QListWidget()
        self.list_added.setStyleSheet(f"""
            QListWidget {{
                background-color: {C['bg_panel']};
                border: 1px solid {C['border']};
                border-radius: 4px;
                color: {C['text_hi']};
                outline: none;
            }}
            QListWidget::item {{
                padding: 5px 10px;
                border-bottom: 1px solid {C['border']};
                font-size: 9pt;
            }}
            QListWidget::item:last {{ border-bottom: none; }}
        """)
        right_layout.addWidget(self.list_added, 1)

        splitter.addWidget(right_widget)
        splitter.setSizes([700, 320])

        # Tablo seçim bağlantısı
        self.tbl.selectionModel().selectionChanged.connect(self._on_selection_changed)

    def _make_stat_badge(self, label: str, value: str, color: str = None) -> QLabel:
        """Küçük istatistik rozeti oluşturur."""
        C = self._C
        c = color or C['text_lo']
        w = QFrame()
        w.setStyleSheet(f"""
            QFrame {{
                background-color: {C['bg_card']};
                border: 1px solid {C['border']};
                border-radius: 4px;
                min-width: 62px;
            }}
        """)
        lay = QVBoxLayout(w)
        lay.setContentsMargins(8, 6, 8, 6)
        lay.setSpacing(0)
        lbl_v = QLabel(value)
        lbl_v.setStyleSheet(f"font-size: 14pt; font-weight: 700; color: {c}; background: transparent; border: none;")
        lbl_v.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_k = QLabel(label)
        lbl_k.setStyleSheet(f"font-size: 7pt; color: {C['text_lo']}; letter-spacing: 1px; background: transparent; border: none;")
        lbl_k.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(lbl_v)
        lay.addWidget(lbl_k)
        # badge'ı dışarıdan güncelleyebilmek için val label'ı saklarız
        w._val_lbl = lbl_v
        return w


    # -------------------------------------------------------- Excel Yükleme --
    def _load_excel_data(self, filepath: str = None):
        """
        Excel dosyasını yükler. VERİ sayfasındaki D,E,F,G,H sütunlarını okur.
        Sütun indeksleri (0 tabanlı):
            B=1 Kimyasal Adı, C=2 Ambalaj, D=3 UN, E=4 Sevkiyat Adı,
            F=5 Sınıf, G=6 PG, H=7 Tünel Kodu
        """
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "Hata", "openpyxl modülü bulunamadı!\npip install openpyxl")
            return

        if not filepath:
            return

        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            sheet_name = None
            for name in wb.sheetnames:
                if "VERİ" in name.upper() or "VERI" in name.upper():
                    sheet_name = name
                    break
            if not sheet_name:
                QMessageBox.warning(self, "Uyarı",
                    "Excel dosyasında 'VERİ' adlı sayfa bulunamadı!")
                return

            ws = wb[sheet_name]
            products = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Sütun düzeni: A=No, B=Kimyasal Adı, C=Ambalaj Türü,
                #   D=UN No, E=Sevkiyat Adı, F=Sınıf, G=PG, H=Tünel Kodu,
                #   I=Taşıma Kategorisi, J=Sınırlı Miktar, K=Özel Hükümler,
                #   L=Tank Kodu, M=Ambalaj Talimatları
                if not row or row[1] is None:
                    continue
                val = lambda idx: str(row[idx]).strip() if idx < len(row) and row[idx] not in (None, "") else ""
                product = {
                    "kimyasal_adi":      val(1),
                    "ambalaj_turu":      val(2),
                    "un_numarasi":       val(3),
                    "sevkiyat_adi":      val(4),
                    "sinif":             val(5),
                    "paketleme_grubu":   val(6),
                    "tunel_kodu":        val(7),
                    "tasima_kategorisi": val(8),
                    "sinirli_miktar":    val(9),
                    "ozel_hukumler":     val(10),
                    "tank_kodu":         val(11),
                    "ambalaj_talimatlari": val(12),
                }
                if product["kimyasal_adi"] and product["kimyasal_adi"] != "LÜTFEN SEÇİNİZ":
                    products.append(product)

            wb.close()
            self.all_products = products
            self._populate_table(products)
            self.lbl_file.setText(f"  {os.path.basename(filepath)}  —  {len(products)} kayıt yüklendi")
            self._update_badges(len(products))
            self.excel_loaded.emit(products)  # ShipmentEditorPage'e bildir

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel dosyası okunamadı:\n{e}")

    def _browse_excel(self):
        filepath, _ = QFileDialog.getOpenFileName(
            self, "Excel Dosyası Seç", "",
            "Excel Dosyaları (*.xlsx *.xlsm *.xls)"
        )
        if filepath:
            self._load_excel_data(filepath)

    # ---------------------------------------------------------- Tablo doldur --
    # Sınıf → (arka plan rengi, metin rengi) eşlemesi
    _CLASS_COLORS = {
        "1":    ("#FF6B6B", "#000"),
        "2":    ("#74B9FF", "#000"),
        "3":    ("#FD7272", "#000"),
        "4.1":  ("#FDCB6E", "#000"),
        "4.2":  ("#E17055", "#fff"),
        "4.3":  ("#0984E3", "#fff"),
        "5.1":  ("#FDCB6E", "#000"),
        "5.2":  ("#E17055", "#fff"),
        "6.1":  ("#A29BFE", "#000"),
        "6.2":  ("#6C5CE7", "#fff"),
        "7":    ("#FD79A8", "#000"),
        "8":    ("#636E72", "#fff"),
        "9":    ("#B2BEC3", "#000"),
    }

    def _class_color(self, sinif: str):
        """Sınıf koduna göre (bg, fg) renk çifti döndürür."""
        s = sinif.strip().split(" ")[0].split("(")[0].strip()
        return self._CLASS_COLORS.get(s, ("#30363D", "#8B949E"))

    def _populate_table(self, products: list):
        C = self._C
        # Performans: sinyal ve sıralama kapalıyken toplu doldur
        self.tbl.setSortingEnabled(False)
        self.tbl.blockSignals(True)
        self.tbl.setUpdatesEnabled(False)
        try:
            self.tbl.setRowCount(len(products))
            for row_idx, p in enumerate(products):
                is_added = p["kimyasal_adi"] in self.added_products
                sinif = p.get("sinif", "")
                bg_class, fg_class = self._class_color(sinif)

                values = [
                    p.get("kimyasal_adi", ""),
                    p.get("un_numarasi", ""),
                    p.get("sevkiyat_adi", ""),
                    sinif,
                    p.get("paketleme_grubu", ""),
                    p.get("tunel_kodu", ""),
                    "✔ EKLENDİ" if is_added else "",
                ]
                for col_idx, val in enumerate(values):
                    item = QTableWidgetItem(val)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)

                    if is_added:
                        item.setForeground(QColor(C["added_fg"] if col_idx == 6 else C["text_lo"]))
                        if col_idx == 6:
                            item.setBackground(QColor(C["added_bg"]))
                            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    elif col_idx == 3:
                        item.setBackground(QColor(bg_class))
                        item.setForeground(QColor(fg_class))
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        fnt = item.font(); fnt.setBold(True); item.setFont(fnt)
                    elif col_idx == 1:
                        item.setForeground(QColor(C["accent"]))
                        fnt = item.font(); fnt.setBold(True); item.setFont(fnt)
                    elif col_idx == 6:
                        item.setForeground(QColor(C["text_lo"]))
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                    self.tbl.setItem(row_idx, col_idx, item)

                # Ürün verisini ilk sütunda sakla
                self.tbl.item(row_idx, 0).setData(Qt.ItemDataRole.UserRole, p)
        finally:
            self.tbl.blockSignals(False)
            self.tbl.setUpdatesEnabled(True)
            self.tbl.setSortingEnabled(False)


    def _filter_table(self, text: str):
        text = text.strip().upper()
        if not text:
            filtered = self.all_products
        else:
            filtered = [p for p in self.all_products
                        if text in p["kimyasal_adi"].upper()
                        or text in p.get("sevkiyat_adi", "").upper()
                        or text in p.get("un_numarasi", "").upper()]
        self._populate_table(filtered)
        self._update_badges(len(filtered))

    def _update_badges(self, filtered_count: int = None):
        total = len(self.all_products)
        added = len(self.added_products)
        filt  = filtered_count if filtered_count is not None else total
        if hasattr(self, 'lbl_total_badge'):
            self.lbl_total_badge._val_lbl.setText(str(total))
            self.lbl_filtered_badge._val_lbl.setText(str(filt))
            self.lbl_added_badge._val_lbl.setText(str(added))


    def _clear_search(self):
        self.txt_search.clear()

    # -------------------------------------------------- Seçim & Detay paneli --
    def _on_selection_changed(self):
        rows = self.tbl.selectionModel().selectedRows()
        if not rows:
            return
        p = self.tbl.item(rows[0].row(), 0).data(Qt.ItemDataRole.UserRole)
        if not p:
            return
        mapping = {
            "Kimyasal Adı":         p.get("kimyasal_adi", "—"),
            "Ambalaj Türü":         p.get("ambalaj_turu", "—"),
            "UN Numarası":          p.get("un_numarasi", "—"),
            "Sevkiyat Adı":         p.get("sevkiyat_adi", "—"),
            "Sınıf":                p.get("sinif", "—"),
            "PG":                   p.get("paketleme_grubu", "—"),
            "Tünel Kodu":           p.get("tunel_kodu", "—"),
            "Taşıma Kategorisi":    p.get("tasima_kategorisi", "—"),
            "Sınırlı Miktar":       p.get("sinirli_miktar", "—"),
            "Özel Hükümler":        p.get("ozel_hukumler", "—"),
            "Tank Kodu":            p.get("tank_kodu", "—"),
            "Ambalaj Talimatları":  p.get("ambalaj_talimatlari", "—"),
        }
        for field, val in mapping.items():
            if field in self.det_labels:
                self.det_labels[field].setText(val or "—")

        # UN rozeti güncelle
        un = p.get("un_numarasi", "")
        self.lbl_un_val.setText(f"UN {un}" if un else "—")

        # Sınıf rozeti güncelle
        sinif = p.get("sinif", "")
        bg, fg = self._class_color(sinif)
        self.lbl_class_badge.setText(sinif or "—")
        self.lbl_class_badge.setStyleSheet(f"""
            background-color: {bg};
            color: {fg};
            border-radius: 4px;
            font-size: 10pt;
            font-weight: 700;
            border: none;
        """)


    # -------------------------------------------------- Ekle / Sıfırla -------
    def _add_selected(self):
        rows = self.tbl.selectionModel().selectedRows()
        if not rows:
            QMessageBox.information(self, "Bilgi", "Lütfen eklemek istediğiniz ürünü seçin.")
            return
        p = self.tbl.item(rows[0].row(), 0).data(Qt.ItemDataRole.UserRole)
        if not p:
            return
        name = p["kimyasal_adi"]
        if name in self.added_products:
            QMessageBox.warning(self, "Tekrar",
                f"'{name}' zaten eklenmiş!\nHer ürün yalnızca bir kez eklenebilir.")
            return

        # Taşıma Evrakı sayfasına sinyal gönder
        self.product_add_requested.emit(p)
        self.added_products.add(name)

        # Eklenenler listesine ekle
        sev_adi = p.get("sevkiyat_adi", name)
        un_no   = p.get("un_numarasi", "")
        item = QListWidgetItem(f"  ✔  {name}  →  {sev_adi}  (UN {un_no})")
        item.setForeground(QColor(self._C["ok"]))
        self.list_added.addItem(item)
        self.list_added.scrollToBottom()

        # Tablo ve badgeleri güncelle
        self._filter_table(self.txt_search.text())
        self._update_badges()

    def _reset_added(self):
        reply = QMessageBox.question(
            self, "Sıfırla",
            "Eklenen ürünler listesi sıfırlansın mı?\n(Taşıma evrakındaki ürünler silinmez.)",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.added_products.clear()
            self.list_added.clear()
            # Tablo yeniden çizimi sırasında UI donmasını önle
            self.tbl.setUpdatesEnabled(False)
            try:
                self._populate_table(self.all_products)
            finally:
                self.tbl.setUpdatesEnabled(True)
            self._update_badges()



# =============================================================================
# ADR VERITABANI SAYFASI  (YENİ - v4.1)
# =============================================================================

class ADRDatabasePage(QWidget):
    """
    ADR kimyasal veritabani yonetim sayfasi.
    - Arama ve filtreleme
    - Yeni kayit ekleme
    - Kayit duzenleme
    - Kayit silme
    - Detay paneli
    """

    def __init__(self, db: DatabaseManager, parent=None):
        super().__init__(parent)
        self.db = db
        self.parent_window = parent
        self._setup_ui()
        self.__init_search_timer()
        self._load_chemicals()
    def _setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(12, 12, 12, 12)
        main_layout.setSpacing(8)

        # Baslik
        title = QLabel("ADR KİMYASAL VERİTABANI")
        title.setObjectName("title")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(title)

        # Filtre satiri
        filter_group = QGroupBox("Arama ve Filtreleme")
        filter_layout = QHBoxLayout(filter_group)

        filter_layout.addWidget(QLabel("Ara:"))
        self.txt_search = QLineEdit()
        self.txt_search.setPlaceholderText("UN No, Teknik Ad veya Sinif ile ara...")
        self.txt_search.setMinimumWidth(280)
        self.txt_search.textChanged.connect(self._on_search_changed)
        filter_layout.addWidget(self.txt_search)

        filter_layout.addWidget(QLabel("Sinif:"))
        self.cmb_class_filter = QComboBox()
        self.cmb_class_filter.setMinimumWidth(80)
        self.cmb_class_filter.addItems([
            "Tumu", "1", "2.1", "2.2", "2.3",
            "3", "4.1", "4.2", "4.3",
            "5.1", "5.2", "6.1", "6.2",
            "7", "8", "9"
        ])
        self.cmb_class_filter.currentIndexChanged.connect(self._load_chemicals)
        filter_layout.addWidget(self.cmb_class_filter)

        self.chk_lq_filter = QCheckBox("Sadece LQ")
        self.chk_lq_filter.toggled.connect(self._load_chemicals)
        filter_layout.addWidget(self.chk_lq_filter)

        self.chk_eq_filter = QCheckBox("Sadece EQ")
        self.chk_eq_filter.toggled.connect(self._load_chemicals)
        filter_layout.addWidget(self.chk_eq_filter)

        btn_refresh = QPushButton("Yenile")
        btn_refresh.clicked.connect(self._load_chemicals)
        filter_layout.addWidget(btn_refresh)

        filter_layout.addStretch()
        self.lbl_count = QLabel("0 kayit")
        self.lbl_count.setObjectName("subtitle")
        filter_layout.addWidget(self.lbl_count)

        main_layout.addWidget(filter_group)

        # Ana splitter: tablo | detay paneli
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # Sol: tablo
        table_widget = QWidget()
        table_layout = QVBoxLayout(table_widget)
        table_layout.setContentsMargins(0, 0, 0, 0)

        self.tbl_chemicals = QTableWidget()
        self.tbl_chemicals.setColumnCount(9)
        self.tbl_chemicals.setHorizontalHeaderLabels([
            "ID", "UN No", "Teknik Ad (TR)", "Sinif", "PG",
            "Tunel", "TC", "LQ", "EQ"
        ])
        self.tbl_chemicals.setColumnHidden(0, True)
        self.tbl_chemicals.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.tbl_chemicals.horizontalHeader().setStretchLastSection(False)
        self.tbl_chemicals.setColumnWidth(1,  70)
        self.tbl_chemicals.setColumnWidth(2, 260)
        self.tbl_chemicals.setColumnWidth(3,  60)
        self.tbl_chemicals.setColumnWidth(4,  45)
        self.tbl_chemicals.setColumnWidth(5,  65)
        self.tbl_chemicals.setColumnWidth(6,  40)
        self.tbl_chemicals.setColumnWidth(7,  40)
        self.tbl_chemicals.setColumnWidth(8,  40)
        self.tbl_chemicals.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tbl_chemicals.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tbl_chemicals.setAlternatingRowColors(True)
        self.tbl_chemicals.setStyleSheet("""
            QTableWidget { alternate-background-color: #252535; }
        """)
        self.tbl_chemicals.selectionModel().selectionChanged.connect(self._on_selection_changed)
        self.tbl_chemicals.doubleClicked.connect(self._edit_chemical)
        self.tbl_chemicals.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tbl_chemicals.customContextMenuRequested.connect(self._show_context_menu)
        table_layout.addWidget(self.tbl_chemicals)

        # Tablo alt butonlar
        tbl_btn_layout = QHBoxLayout()

        btn_add = QPushButton("+ Yeni Kimyasal")
        btn_add.setObjectName("success")
        btn_add.setMinimumHeight(34)
        btn_add.clicked.connect(self._add_chemical)
        tbl_btn_layout.addWidget(btn_add)

        btn_edit = QPushButton("Duzenle")
        btn_edit.setObjectName("primary")
        btn_edit.setMinimumHeight(34)
        btn_edit.clicked.connect(self._edit_chemical)
        tbl_btn_layout.addWidget(btn_edit)

        btn_copy = QPushButton("Kopyala")
        btn_copy.setMinimumHeight(34)
        btn_copy.clicked.connect(self._copy_chemical)
        tbl_btn_layout.addWidget(btn_copy)

        btn_delete = QPushButton("Sil")
        btn_delete.setObjectName("danger")
        btn_delete.setMinimumHeight(34)
        btn_delete.clicked.connect(self._delete_chemical)
        tbl_btn_layout.addWidget(btn_delete)

        tbl_btn_layout.addStretch()

        btn_export = QPushButton("Excel Aktar")
        btn_export.setMinimumHeight(34)
        btn_export.clicked.connect(self._export_excel)
        tbl_btn_layout.addWidget(btn_export)

        table_layout.addLayout(tbl_btn_layout)
        splitter.addWidget(table_widget)

        # Sag: detay paneli
        detail_widget = QWidget()
        detail_widget.setMinimumWidth(280)
        detail_widget.setMaximumWidth(340)
        detail_layout = QVBoxLayout(detail_widget)
        detail_layout.setContentsMargins(8, 0, 0, 0)

        detail_title = QLabel("MADDE DETAYI")
        detail_title.setObjectName("title")
        detail_title.setText("⬡  KİMYASAL DETAYI")
        detail_title.setStyleSheet("font-size:9pt;font-weight:700;letter-spacing:2px;color:#585B70;padding-bottom:6px;border-bottom:2px solid #313244;background:transparent;")
        detail_title.setAlignment(Qt.AlignmentFlag.AlignLeft)
        detail_layout.addWidget(detail_title)

        # Sinif rozeti
        self.lbl_class_badge = QLabel("")
        self.lbl_class_badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_class_badge.setMinimumHeight(48)
        self.lbl_class_badge.setStyleSheet(
            "font-size: 22pt; font-weight: bold; border-radius: 8px;"
            "background-color: #45475A; color: #CDD6F4;"
        )
        detail_layout.addWidget(self.lbl_class_badge)

        # Detay form
        detail_form_group = QGroupBox("Kimlik Bilgileri")
        detail_form = QFormLayout(detail_form_group)
        detail_form.setSpacing(6)

        self.dlbl_un       = QLabel("-")
        self.dlbl_un.setStyleSheet("font-weight: bold; font-size: 13pt; color: #89B4FA;")
        self.dlbl_name_tr  = QLabel("-")
        self.dlbl_name_tr.setWordWrap(True)
        self.dlbl_name_en  = QLabel("-")
        self.dlbl_name_en.setWordWrap(True)
        self.dlbl_class    = QLabel("-")
        self.dlbl_pg       = QLabel("-")
        self.dlbl_tunnel   = QLabel("-")
        self.dlbl_tc       = QLabel("-")
        self.dlbl_seg      = QLabel("-")
        self.dlbl_seg.setWordWrap(True)
        self.dlbl_flash    = QLabel("-")

        detail_form.addRow("UN No:", self.dlbl_un)
        detail_form.addRow("Teknik Ad (TR):", self.dlbl_name_tr)
        detail_form.addRow("Teknik Ad (EN):", self.dlbl_name_en)
        detail_form.addRow("Sinif:", self.dlbl_class)
        detail_form.addRow("Ambalaj Grubu:", self.dlbl_pg)
        detail_form.addRow("Tunel Kodu:", self.dlbl_tunnel)
        detail_form.addRow("Tasima Kategorisi:", self.dlbl_tc)
        detail_form.addRow("Tehlike Gurubu:", self.dlbl_seg)

        detail_layout.addWidget(detail_form_group)

        exemption_group = QGroupBox("Muafiyet Durumu")
        exemption_layout = QVBoxLayout(exemption_group)

        self.dlbl_lq = QLabel("LQ: -")
        self.dlbl_eq = QLabel("EQ: -")
        exemption_layout.addWidget(self.dlbl_lq)
        exemption_layout.addWidget(self.dlbl_eq)

        detail_layout.addWidget(exemption_group)

        special_group = QGroupBox("Ozel Hukumler")
        special_layout = QVBoxLayout(special_group)
        self.dlbl_special = QPlainTextEdit()
        self.dlbl_special.setReadOnly(True)
        self.dlbl_special.setMaximumHeight(80)
        special_layout.addWidget(self.dlbl_special)
        detail_layout.addWidget(special_group)

        detail_layout.addStretch()

        splitter.addWidget(detail_widget)
        splitter.setSizes([700, 300])

        main_layout.addWidget(splitter)

    # --- Timer ile arama (anlık tetikleme) ---
    def __init_search_timer(self):
        if not hasattr(self, '_search_timer'):
            self._search_timer = QTimer(self)
            self._search_timer.setSingleShot(True)
            self._search_timer.timeout.connect(self._load_chemicals)

    def _on_search_changed(self):
        self.__init_search_timer()
        self._search_timer.start(300)

    # --- Veri yukle ---
    def _load_chemicals(self):
        search = self.txt_search.text().strip()
        class_filter = self.cmb_class_filter.currentText()
        lq_only = self.chk_lq_filter.isChecked()
        eq_only = self.chk_eq_filter.isChecked()

        chemicals = self.db.get_all_chemicals(
            search=search if search else None,
            class_filter=class_filter,
            limit=1000
        )

        if lq_only:
            chemicals = [c for c in chemicals if c.lq_allowed]
        if eq_only:
            chemicals = [c for c in chemicals if c.eq_allowed]

        self.tbl_chemicals.blockSignals(True)
        self.tbl_chemicals.setRowCount(len(chemicals))

        for i, c in enumerate(chemicals):
            self.tbl_chemicals.setItem(i, 0, QTableWidgetItem(str(c.id)))
            self.tbl_chemicals.setItem(i, 1, QTableWidgetItem(f"UN{c.un_number}"))
            self.tbl_chemicals.setItem(i, 2, QTableWidgetItem(c.proper_shipping_name_tr))

            class_item = QTableWidgetItem(c.class_code)
            bg, fg = ADREngine.get_class_color(c.class_code)
            class_item.setBackground(QColor(bg))
            class_item.setForeground(QColor(fg))
            class_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.tbl_chemicals.setItem(i, 3, class_item)

            pg_item = QTableWidgetItem(c.packing_group)
            pg_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.tbl_chemicals.setItem(i, 4, pg_item)

            tunnel_item = QTableWidgetItem(c.tunnel_code or "-")
            tunnel_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if "A" in (c.tunnel_code or ""):
                tunnel_item.setBackground(QColor("#F38BA8"))
                tunnel_item.setForeground(QColor("#1E1E2E"))
            elif any(x in (c.tunnel_code or "") for x in ["B", "C"]):
                tunnel_item.setBackground(QColor("#F9E2AF"))
                tunnel_item.setForeground(QColor("#1E1E2E"))
            self.tbl_chemicals.setItem(i, 5, tunnel_item)

            tc_item = QTableWidgetItem(c.transport_category or "-")
            tc_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.tbl_chemicals.setItem(i, 6, tc_item)

            lq_item = QTableWidgetItem("✓" if c.lq_allowed else "")
            lq_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if c.lq_allowed:
                lq_item.setForeground(QColor("#A6E3A1"))
            self.tbl_chemicals.setItem(i, 7, lq_item)

            eq_item = QTableWidgetItem("✓" if c.eq_allowed else "")
            eq_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if c.eq_allowed:
                eq_item.setForeground(QColor("#89B4FA"))
            self.tbl_chemicals.setItem(i, 8, eq_item)

        self.tbl_chemicals.blockSignals(False)
        self.lbl_count.setText(f"{len(chemicals)} kayit")
        self._clear_detail()






    def _on_selection_changed(self, *_):
        row = self.tbl_chemicals.currentRow()
        if row < 0:
            self._clear_detail(); return
        id_item = self.tbl_chemicals.item(row, 0)
        if not id_item:
            self._clear_detail(); return
        c = self.db.get_chemical(int(id_item.text()))
        if c: self._show_detail(c)
        else: self._clear_detail()

    def _show_detail(self, c):
        bg, fg = ADREngine.get_class_color(c.class_code)
        self.lbl_class_badge.setText(f"Sinif {c.class_code}")
        self.lbl_class_badge.setStyleSheet(
            f"font-size:18pt;font-weight:bold;border-radius:8px;"
            f"background-color:{bg};color:{fg};")
        self.dlbl_un.setText(f"UN{c.un_number}")
        self.dlbl_name_tr.setText(c.proper_shipping_name_tr or "-")
        self.dlbl_name_en.setText(c.proper_shipping_name_en or "-")
        self.dlbl_class.setText(c.class_code or "-")
        self.dlbl_pg.setText(c.packing_group or "-")
        self.dlbl_tunnel.setText(c.tunnel_code or "-")
        self.dlbl_tc.setText(c.transport_category or "-")
        self.dlbl_seg.setText(c.segregation_group or "-")
        self.dlbl_lq.setText(f"LQ: {'Izinli' if c.lq_allowed else 'Izinsiz'}")
        self.dlbl_eq.setText(f"EQ: {'Izinli' if c.eq_allowed else 'Izinsiz'}")
        self.dlbl_special.setPlainText(c.special_provisions or "")

    def _clear_detail(self):
        self.lbl_class_badge.setText("")
        self.lbl_class_badge.setStyleSheet(
            "font-size:22pt;font-weight:bold;border-radius:8px;"
            "background-color:#45475A;color:#CDD6F4;")
        for lbl in (self.dlbl_un, self.dlbl_name_tr, self.dlbl_name_en,
                    self.dlbl_class, self.dlbl_pg, self.dlbl_tunnel,
                    self.dlbl_tc, self.dlbl_seg):
            lbl.setText("-")
            
            
            
            
            
            
        self.dlbl_lq.setText("LQ: -")
        self.dlbl_eq.setText("EQ: -")
        self.dlbl_special.setPlainText("")

    # --- CRUD islemleri ---

    def _get_selected_chemical(self) -> Optional[Chemical]:
        row = self.tbl_chemicals.currentRow()
        if row < 0:
            return None
        id_item = self.tbl_chemicals.item(row, 0)
        if not id_item:
            return None
        return self.db.get_chemical(int(id_item.text()))

    def _add_chemical(self):
        dialog = ChemicalDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            c = dialog.get_chemical()
            if c:
                try:
                    self.db.add_chemical(c)
                    self._load_chemicals()
                    if self.parent_window:
                        self.parent_window.statusbar.showMessage(
                            f"UN{c.un_number} eklendi.", 4000)
                except Exception as e:
                    QMessageBox.critical(self, "Hata",
                        f"Kaydetme hatasi:\n{str(e)}\n\nBu UN numarasi zaten mevcut olabilir.")

    def _edit_chemical(self):
        c = self._get_selected_chemical()
        if not c:
            QMessageBox.warning(self, "Uyari", "Lutfen duzenlenecek bir kayit secin!")
            return
        dialog = ChemicalDialog(self, chemical=c)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            updated = dialog.get_chemical()
            if updated:
                try:
                    self.db.update_chemical(updated)
                    self._load_chemicals()
                    if self.parent_window:
                        self.parent_window.statusbar.showMessage(
                            f"UN{updated.un_number} guncellendi.", 4000)
                except Exception as e:
                    QMessageBox.critical(self, "Hata", f"Guncelleme hatasi:\n{str(e)}")

    def _copy_chemical(self):
        c = self._get_selected_chemical()
        if not c:
            QMessageBox.warning(self, "Uyari", "Lutfen kopyalanacak bir kayit secin!")
            return
        new_un, ok = QInputDialog.getText(self, "UN No Girin",
            f"UN{c.un_number} kopyalanacak.\nYeni UN numarasini girin:",
            text=c.un_number + "_KOPYA")
        if ok and new_un.strip():
            c.id = None
            c.un_number = new_un.strip()
            try:
                self.db.add_chemical(c)
                self._load_chemicals()
                if self.parent_window:
                    self.parent_window.statusbar.showMessage(
                        f"UN{c.un_number} olarak kopyalandi.", 4000)
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Kopyalama hatasi:\n{str(e)}")

    def _delete_chemical(self):
        c = self._get_selected_chemical()
        if not c:
            QMessageBox.warning(self, "Uyari", "Lutfen silinecek bir kayit secin!")
            return
        reply = QMessageBox.question(self, "Sil",
            f"UN{c.un_number} - {c.proper_shipping_name_tr}\n\n"
            f"Bu kayit veritabanindan kalici olarak silinecek.\nEmin misiniz?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.db.delete_chemical(c.id)
            self._load_chemicals()
            if self.parent_window:
                self.parent_window.statusbar.showMessage(
                    f"UN{c.un_number} silindi.", 4000)

    def _show_context_menu(self, pos):
        c = self._get_selected_chemical()
        menu = QMenu(self)
        act_add    = menu.addAction("Yeni Kimyasal Ekle")
        act_add.triggered.connect(self._add_chemical)
        if c:
            act_edit   = menu.addAction(f"Duzenle: UN{c.un_number}")
            act_edit.triggered.connect(self._edit_chemical)
            act_copy   = menu.addAction("Kopyala")
            act_copy.triggered.connect(self._copy_chemical)
            menu.addSeparator()
            act_delete = menu.addAction(f"Sil: UN{c.un_number}")
            act_delete.triggered.connect(self._delete_chemical)
        menu.exec(self.tbl_chemicals.viewport().mapToGlobal(pos))

    def _export_excel(self):
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "Hata", "openpyxl modulu eksik! Kurulum: pip install openpyxl")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Excel Aktar",
            f"ADR_Kimyasallar_{datetime.now().strftime('%Y%m%d')}.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "ADR Kimyasallar"
            headers = ["Kimyasal Adi", "UN NUMARASI", "UYGUN SEVKIYAT ADI",
                       "SINIFI / ETIKETI", "PAKETLEME GRUBU", "TUNEL KODU",
                       "OZEL HUKUMLAR", "TASIMA KATEGORISI", "SINIRLI MIKTAR",
                       "ISTISNAI MIKTAR", "AMBALAJLAMA TALIMATLARI", "TANK KODU"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1E3A5F", fill_type="solid")

            chemicals = self.db.get_all_chemicals(limit=10000)
            for row, c in enumerate(chemicals, 2):
                ws.cell(row=row, column=1,  value=c.proper_shipping_name_tr)
                ws.cell(row=row, column=2,  value=f"UN{c.un_number}")
                ws.cell(row=row, column=3,  value=c.proper_shipping_name_en)
                ws.cell(row=row, column=4,  value=c.class_code)
                ws.cell(row=row, column=5,  value=c.packing_group)
                ws.cell(row=row, column=6,  value=c.tunnel_code)
                ws.cell(row=row, column=7,  value=c.special_provisions)
                ws.cell(row=row, column=8,  value=c.transport_category)
                ws.cell(row=row, column=9,  value="Evet" if c.lq_allowed else "Hayir")
                ws.cell(row=row, column=10, value="Evet" if c.eq_allowed else "Hayir")
                ws.cell(row=row, column=11, value=c.hazard_labels)
                ws.cell(row=row, column=12, value="")

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 22

            wb.save(path)
            QMessageBox.information(self, "Basarili", f"Excel olusturuldu: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel aktar hatasi: {str(e)}")

    def _import_excel(self):
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "Hata", "openpyxl modulu eksik! Kurulum: pip install openpyxl")
            return

        path, _ = QFileDialog.getOpenFileName(self, "Excel'den Veri Al",
            "", "Excel (*.xlsx *.xls)")
        if not path:
            return

        try:
            wb = load_workbook(path, data_only=True)
            ws = wb.active

            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).strip().upper() if cell.value else "")

            col_map = {}
            for idx, h in enumerate(headers, 1):
                if any(k in h for k in ["KIMYASAL", "ADI", "TEKNIK"]):
                    col_map['name_tr'] = idx
                elif "UN" in h and "NUMARA" in h:
                    col_map['un_number'] = idx
                elif "UYGUN" in h and "SEVKIYAT" in h:
                    col_map['name_en'] = idx
                elif "SINIF" in h or "ETIKET" in h:
                    col_map['class_code'] = idx
                elif "PAKETLEME" in h or "GRUBU" in h:
                    col_map['packing_group'] = idx
                elif "TUNEL" in h:
                    col_map['tunnel_code'] = idx
                elif "OZEL" in h and "HUKUM" in h:
                    col_map['special_provisions'] = idx
                elif "TASIMA" in h and "KATEGORI" in h:
                    col_map['transport_category'] = idx
                elif "SINIRLI" in h:
                    col_map['lq'] = idx
                elif "ISTISNAI" in h:
                    col_map['eq'] = idx
                elif "AMBALAJLAMA" in h:
                    col_map['hazard_labels'] = idx
                elif "TANK" in h:
                    col_map['tank_code'] = idx

            required = ['name_tr', 'un_number', 'class_code']
            missing = [r for r in required if r not in col_map]
            if missing:
                QMessageBox.warning(self, "Hata",
                    f"Excel dosyasinda zorunlu sutunlar bulunamadi: {', '.join(missing)}. "
                    f"Beklenen sutunlar: Kimyasal Adi, UN NUMARASI, UYGUN SEVKIYAT ADI, "
                    f"SINIFI / ETIKETI, PAKETLEME GRUBU, TUNEL KODU, OZEL HUKUMLAR, "
                    f"TASIMA KATEGORISI, SINIRLI MIKTAR, ISTISNAI MIKTAR, "
                    f"AMBALAJLAMA TALIMATLARI, TANK KODU")
                return

            imported = 0
            updated = 0
            errors = []

            for row_idx in range(2, ws.max_row + 1):
                try:
                    def get_val(col_key, default=""):
                        col_idx = col_map.get(col_key)
                        if not col_idx:
                            return default
                        val = ws.cell(row=row_idx, column=col_idx).value
                        return str(val).strip() if val is not None else default

                    un_number = get_val('un_number')
                    un_number = un_number.replace("UN", "").replace("un", "").strip()

                    if not un_number:
                        continue

                    name_tr = get_val('name_tr')
                    if not name_tr:
                        continue

                    class_code = get_val('class_code')
                    if not class_code:
                        continue

                    def parse_bool(val):
                        if not val:
                            return False
                        return val.upper() in ['EVET', 'E', 'YES', 'Y', 'TRUE', '1', 'V']

                    chemical = Chemical(
                        un_number=un_number,
                        proper_shipping_name_tr=name_tr,
                        proper_shipping_name_en=get_val('name_en'),
                        class_code=class_code,
                        packing_group=get_val('packing_group'),
                        tunnel_code=get_val('tunnel_code'),
                        transport_category=get_val('transport_category'),
                        special_provisions=get_val('special_provisions'),
                        lq_allowed=parse_bool(get_val('lq')),
                        eq_allowed=parse_bool(get_val('eq')),
                        hazard_labels=get_val('hazard_labels'),
                        segregation_group=""
                    )

                    existing = self.db.get_chemical_by_un(un_number)
                    if existing:
                        chemical.id = existing.id
                        self.db.update_chemical(chemical)
                        updated += 1
                    else:
                        self.db.add_chemical(chemical)
                        imported += 1

                except Exception as e:
                    errors.append(f"Satir {row_idx}: {str(e)}")

            self._load_chemicals()

            msg = f"Ici aktarma tamamlandi! Yeni eklenen: {imported}, Guncellenen: {updated}."
            if errors:
                msg += f" Hatalar ({len(errors)}): " + "; ".join(errors[:5])
                if len(errors) > 5:
                    msg += f" ... ve {len(errors) - 5} hata daha"

            QMessageBox.information(self, "Basarili", msg)

            if self.parent_window:
                self.parent_window.statusbar.showMessage(
                    f"Excel'den {imported} yeni, {updated} guncelleme yapildi.", 5000)

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel okuma hatasi: {str(e)}")



# =============================================================================
# YAZDIR ONIZLEME DIALOG  (YENİ - v4.1)
# =============================================================================

class PrintPreviewDialog(QDialog):
    """
    Profesyonel A4 önizleme + yazdırma + PDF dışa aktarma.
    - QWebEngineView OLMADAN, QTextEdit + QPrinter ile çalışır
    - Türkçe karakter desteği (UTF-8)
    - Sayfa sınırlarına uygun düzen
    - Kaşe/imza kutuları
    """

    def __init__(self, parent, html_content: str, title: str = "Evrak Önizleme"):
        super().__init__(parent)
        self.html_content = html_content
        self.setWindowTitle(title)
        self.setMinimumSize(940, 760)
        self.resize(1040, 820)
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # ── Araç çubuğu ───────────────────────────────────────────────────
        toolbar = QWidget()
        toolbar.setFixedHeight(48)
        toolbar.setStyleSheet("""
            QWidget { background-color: #1C2128; border-bottom: 1px solid #30363D; }
            QPushButton {
                background-color: #21262D; border: 1px solid #30363D;
                border-radius: 4px; color: #F0F6FC; padding: 5px 16px;
                font-size: 9pt; font-weight: 600; min-height: 28px;
            }
            QPushButton:hover { background-color: #2D333B; }
            QPushButton#btn_print {
                background-color: #1F3358; border-color: #89B4FA; color: #89B4FA;
            }
            QPushButton#btn_print:hover { background-color: #2A4570; }
            QPushButton#btn_pdf {
                background-color: #3A1A1A; border-color: #F38BA8; color: #F38BA8;
            }
            QPushButton#btn_pdf:hover { background-color: #4A2020; }
        """)
        tb_lay = QHBoxLayout(toolbar)
        tb_lay.setContentsMargins(12, 0, 12, 0)
        tb_lay.setSpacing(8)

        lbl_title = QLabel(self.windowTitle())
        lbl_title.setStyleSheet("color: #F0F6FC; font-size: 10pt; font-weight: 700; "
                                "background: transparent; border: none;")
        tb_lay.addWidget(lbl_title)
        tb_lay.addStretch()

        # Zoom
        tb_lay.addWidget(QLabel("Zoom:"))
        self.zoom_spin = QSpinBox()
        self.zoom_spin.setRange(50, 200)
        self.zoom_spin.setValue(100)
        self.zoom_spin.setSuffix("%")
        self.zoom_spin.setFixedWidth(70)
        self.zoom_spin.setStyleSheet(
            "background: #21262D; border: 1px solid #30363D; color: #F0F6FC; "
            "border-radius: 4px; padding: 3px 6px;")
        self.zoom_spin.valueChanged.connect(self._apply_zoom)
        tb_lay.addWidget(self.zoom_spin)

        btn_print = QPushButton("🖨  Yazdır")
        btn_print.setObjectName("btn_print")
        btn_print.clicked.connect(self._do_print)
        tb_lay.addWidget(btn_print)

        btn_pdf = QPushButton("  PDF Kaydet")
        btn_pdf.setObjectName("btn_pdf")
        btn_pdf.clicked.connect(self._do_save_pdf)
        tb_lay.addWidget(btn_pdf)

        btn_close = QPushButton("Kapat")
        btn_close.clicked.connect(self.accept)
        tb_lay.addWidget(btn_close)

        layout.addWidget(toolbar)

        # ── A4 Önizleme alanı (gri arka plan + beyaz A4 kağıt) ───────────
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { background-color: #3D444D; border: none; }")

        page_container = QWidget()
        page_container.setStyleSheet("background-color: #3D444D;")
        pc_lay = QVBoxLayout(page_container)
        pc_lay.setContentsMargins(24, 24, 24, 24)
        pc_lay.setAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignTop)

        # Beyaz A4 kağıt simülasyonu
        self.paper = QFrame()
        self.paper.setObjectName("paper")
        self.paper.setStyleSheet("""
            QFrame#paper {
                background-color: #FFFFFF;
                border: 1px solid #B0B8C4;
                border-radius: 2px;
            }
        """)
        # A4 oran: 210×297mm → 794×1123px @ 96dpi
        self.paper.setFixedWidth(794)
        self.paper_layout = QVBoxLayout(self.paper)
        self.paper_layout.setContentsMargins(0, 0, 0, 0)

        self.view = QTextEdit()
        self.view.setReadOnly(True)
        self.view.setFrameShape(QFrame.Shape.NoFrame)
        self.view.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.view.setStyleSheet(
            "QTextEdit { background-color: #FFFFFF; color: #000000; border: none; }")
        self.view.setHtml(self.html_content)
        self.paper_layout.addWidget(self.view)

        pc_lay.addWidget(self.paper)
        scroll.setWidget(page_container)
        layout.addWidget(scroll, 1)

    def _apply_zoom(self, pct: int):
        factor = pct / 100.0
        base_w = 794
        self.paper.setFixedWidth(int(base_w * factor))
        font = self.view.font()
        font.setPointSizeF(10 * factor)
        self.view.setFont(font)

    def _do_print(self):
        try:
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            page_layout = QPageLayout(
                QPageSize(QPageSize.PageSizeId.A4),
                QPageLayout.Orientation.Portrait,
                QMarginsF(15, 15, 15, 15),
                QPageLayout.Unit.Millimeter
            )
            printer.setPageLayout(page_layout)
            dialog = QPrintDialog(printer, self)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                doc = QTextDocument()
                doc.setDefaultStyleSheet(
                    "body { font-family: Arial, sans-serif; font-size: 10pt; }"
                    "table { border-collapse: collapse; width: 100%; }"
                )
                doc.setHtml(self.html_content)
                doc.setPageSize(printer.pageRect(QPrinter.Unit.Point).size())
                doc.print(printer)
        except Exception as e:
            QMessageBox.warning(self, "Yazdırma Hatası", str(e))

    def _do_save_pdf(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "PDF Olarak Kaydet", "ADR_Evrak.pdf", "PDF (*.pdf)")
        if not path:
            return
        try:
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(path)
            page_layout = QPageLayout(
                QPageSize(QPageSize.PageSizeId.A4),
                QPageLayout.Orientation.Portrait,
                QMarginsF(15, 15, 15, 15),
                QPageLayout.Unit.Millimeter
            )
            printer.setPageLayout(page_layout)
            doc = QTextDocument()
            doc.setHtml(self.html_content)
            doc.setPageSize(printer.pageRect(QPrinter.Unit.Point).size())
            doc.print(printer)
            QMessageBox.information(self, "Başarılı", f"PDF oluşturuldu:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "PDF Hatası", str(e))







class ReportsPage(QWidget):
    """Raporlar & Analiz — 5 sekme."""
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self._build_ui()
        self._refresh()

    def _build_ui(self):
        root = QVBoxLayout(self); root.setContentsMargins(0,0,0,0); root.setSpacing(0)
        hdr = QWidget(); hdr.setStyleSheet("background:#181825;border-bottom:1px solid #313244;")
        hl = QHBoxLayout(hdr); hl.setContentsMargins(20,12,20,12)
        ttl = QLabel("📊 Raporlar & Analiz"); ttl.setStyleSheet("font-size:14pt;font-weight:700;color:#CDD6F4;")
        hl.addWidget(ttl); hl.addStretch()
        hl.addWidget(QLabel("Yıl:"))
        self.cmb_year = QComboBox(); self.cmb_year.setFixedWidth(90)
        from datetime import date as _date
        self.cmb_year.addItem("Tümü", None)
        for y in range(_date.today().year, _date.today().year-6, -1):
            self.cmb_year.addItem(str(y), y)
        self.cmb_year.currentIndexChanged.connect(self._refresh); hl.addWidget(self.cmb_year)
        btn_r = QPushButton("↻ Yenile"); btn_r.setFixedWidth(90); btn_r.clicked.connect(self._refresh); hl.addWidget(btn_r)
        btn_x = QPushButton("📥 Excel'e Aktar"); btn_x.setFixedWidth(140); btn_x.clicked.connect(self._export_excel); hl.addWidget(btn_x)
        root.addWidget(hdr)
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("QTabWidget::pane{border:none;background:#1E1E2E;}QTabBar::tab{background:#181825;color:#585B70;padding:8px 16px;border:none;font-size:10pt;}QTabBar::tab:selected{background:#1E1E2E;color:#CDD6F4;border-bottom:2px solid #89B4FA;}")
        root.addWidget(self.tabs, 1)
        self.tabs.addTab(self._tab_dashboard(), "🏠 Özet")
        self.tabs.addTab(self._tab_class(),     "🧪 ADR Sınıfları")
        self.tabs.addTab(self._tab_senders(),   "🏭 Göndericiler")
        self.tabs.addTab(self._tab_chemicals(), "☣ UN Numaraları")
        self.tabs.addTab(self._tab_expiry(),    "⚠ Bitiş Tarihleri")

    def _tbl(self, headers):
        t = QTableWidget(0, len(headers))
        t.setHorizontalHeaderLabels(headers)
        t.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        t.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        for i in range(len(headers)):
            t.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
        return t

    def _sec(self, lay, txt):
        l=QLabel(f"⬡  {txt}"); l.setStyleSheet("font-size:9pt;font-weight:700;letter-spacing:2px;color:#585B70;border-bottom:2px solid #313244;margin-top:8px;"); lay.addWidget(l)

    def _tab_dashboard(self):
        w=QWidget(); lay=QVBoxLayout(w); lay.setContentsMargins(20,16,20,16); lay.setSpacing(12)
        kpi_row=QHBoxLayout(); kpi_row.setSpacing(10); self._kpi={}
        for key,lbl,ico,col in [
            ("total_shipments","Toplam Sevkiyat","📦","#89B4FA"),
            ("draft_shipments","Taslak","✏","#F9E2AF"),
            ("total_companies","Firma","🏢","#A6E3A1"),
            ("active_drivers","Aktif Sürücü","🧑","#CBA6F7"),
            ("active_vehicles","Aktif Araç","🚛","#FAB387"),
            ("total_chemicals","Kimyasal","🧪","#89DCEB")]:
            card=QFrame(); card.setStyleSheet(f"QFrame{{background:#24273A;border-radius:10px;border-left:4px solid {col};}}")
            cl=QVBoxLayout(card); cl.setContentsMargins(12,10,12,10)
            v=QLabel("—"); v.setStyleSheet(f"font-size:20pt;font-weight:800;color:{col};"); v.setAlignment(Qt.AlignmentFlag.AlignCenter)
            cl.addWidget(v); cl.addWidget(QLabel(f"{ico} {lbl}"))
            kpi_row.addWidget(card); self._kpi[key]=v
        lay.addLayout(kpi_row)
        self._sec(lay,"AYLIK SEVKİYAT (Son 6 Ay)")
        self.tbl_monthly=self._tbl(["Ay","Sevkiyat"]); self.tbl_monthly.setMaximumHeight(180); lay.addWidget(self.tbl_monthly)
        lay.addStretch(); return w

    def _tab_class(self):
        w=QWidget(); lay=QVBoxLayout(w); lay.setContentsMargins(20,16,20,16)
        self._sec(lay,"ADR SINIFINA GÖRE TOPLAM NET MİKTAR")
        self.tbl_class=self._tbl(["ADR Sınıfı","Sevkiyat Sayısı","Toplam Net (kg)"]); lay.addWidget(self.tbl_class,1); return w

    def _tab_senders(self):
        w=QWidget(); lay=QVBoxLayout(w); lay.setContentsMargins(20,16,20,16)
        self._sec(lay,"EN ÇOK TAŞIMA YAPILAN GÖNDERİCİLER (İlk 10)")
        self.tbl_senders=self._tbl(["Firma Adı","Sevkiyat Sayısı"]); lay.addWidget(self.tbl_senders,1); return w

    def _tab_chemicals(self):
        w=QWidget(); lay=QVBoxLayout(w); lay.setContentsMargins(20,16,20,16)
        self._sec(lay,"EN ÇOK TAŞINAN UN NUMARALARI (İlk 10)")
        self.tbl_chem=self._tbl(["UN No","ADR Sınıfı","Taşıma Sayısı","Toplam Net (kg)"]); lay.addWidget(self.tbl_chem,1); return w

    def _tab_expiry(self):
        w=QWidget(); lay=QVBoxLayout(w); lay.setContentsMargins(20,16,20,16); lay.setSpacing(10)
        hl=QHBoxLayout(); hl.addWidget(QLabel("Uyarı eşiği:"))
        self.spn_exp_days=QSpinBox(); self.spn_exp_days.setRange(7,180); self.spn_exp_days.setValue(30); self.spn_exp_days.setSuffix(" gün"); self.spn_exp_days.setFixedWidth(100)
        self.spn_exp_days.valueChanged.connect(self._refresh_expiry); hl.addWidget(self.spn_exp_days); hl.addStretch(); lay.addLayout(hl)
        self._sec(lay,"SÜRÜCÜ SRC5 BİTİŞ TARİHLERİ")
        self.tbl_drv_exp=self._tbl(["Sürücü","SRC5 No","Bitiş / Kalan"]); self.tbl_drv_exp.setMaximumHeight(180); lay.addWidget(self.tbl_drv_exp)
        self._sec(lay,"ARAÇ ADR / MUAYENE BİTİŞ TARİHLERİ")
        self.tbl_veh_exp=self._tbl(["Plaka","ADR Bitiş / Kalan","Muayene Bitiş / Kalan"]); lay.addWidget(self.tbl_veh_exp,1); return w

    def _refresh(self):
        year = self.cmb_year.currentData()
        try:
            stats = self.db.get_statistics()
            for k,lbl in self._kpi.items(): lbl.setText(str(stats.get(k,0)))
            self.tbl_monthly.setRowCount(0)
            for ay,cnt in stats.get("monthly_shipments",{}).items():
                r=self.tbl_monthly.rowCount(); self.tbl_monthly.insertRow(r)
                self.tbl_monthly.setItem(r,0,QTableWidgetItem(ay))
                self.tbl_monthly.setItem(r,1,QTableWidgetItem(str(cnt)))
        except Exception as e:
            pass
        try:
            CCOL={"1":"#F38BA8","2":"#FAB387","3":"#F9E2AF","4.1":"#A6E3A1","4.2":"#A6E3A1","4.3":"#89B4FA","5.1":"#CBA6F7","5.2":"#CBA6F7","6.1":"#89DCEB","8":"#B4BEFE","9":"#CDD6F4"}
            rows=self.db.get_class_breakdown(year=year); self.tbl_class.setRowCount(0)
            for row in rows:
                r=self.tbl_class.rowCount(); self.tbl_class.insertRow(r)
                cls=row.get("class_code") or "—"
                it=QTableWidgetItem(f"  Sınıf {cls}"); it.setForeground(QColor(CCOL.get(cls,"#CDD6F4")))
                self.tbl_class.setItem(r,0,it)
                self.tbl_class.setItem(r,1,QTableWidgetItem(str(row.get("sevkiyat_sayisi",0))))
                self.tbl_class.setItem(r,2,QTableWidgetItem(f"{row.get('toplam_net_kg') or 0:,.1f}"))
        except Exception as _e: logger.warning("Dashboard sınıf tablosu yüklenemedi: %s", _e)
        try:
            rows=self.db.get_top_senders(year=year); self.tbl_senders.setRowCount(0)
            for i,row in enumerate(rows):
                r=self.tbl_senders.rowCount(); self.tbl_senders.insertRow(r)
                self.tbl_senders.setItem(r,0,QTableWidgetItem(f"  {i+1}. {row.get('name','')}"))
                self.tbl_senders.setItem(r,1,QTableWidgetItem(str(row.get("sevkiyat_sayisi",0))))
        except Exception as _e: logger.warning("Dashboard gönderici tablosu yüklenemedi: %s", _e)
        try:
            rows=self.db.get_top_chemicals(year=year); self.tbl_chem.setRowCount(0)
            for row in rows:
                r=self.tbl_chem.rowCount(); self.tbl_chem.insertRow(r)
                self.tbl_chem.setItem(r,0,QTableWidgetItem(f"UN{row.get('un_number','')}"))
                self.tbl_chem.setItem(r,1,QTableWidgetItem(row.get("class_code") or "—"))
                self.tbl_chem.setItem(r,2,QTableWidgetItem(str(row.get("adet",0))))
                self.tbl_chem.setItem(r,3,QTableWidgetItem(f"{row.get('toplam_net_kg') or 0:,.1f}"))
        except Exception as _e: logger.warning("Dashboard kimyasal tablosu yüklenemedi: %s", _e)
        self._refresh_expiry()

    def _refresh_expiry(self):
        days=self.spn_exp_days.value()
        def _ks(k):
            if k is None: return "—"
            if k<=0: return f"⛔ {-k} gün önce doldu"
            if k<=15: return f"🔴 {k} gün"
            if k<=30: return f"🟠 {k} gün"
            if k<=90: return f"🟡 {k} gün"
            return f"🟢 {k} gün"
        def _kc(k):
            if k is None or k<=0: return "#F38BA8"
            if k<=30: return "#FAB387"
            if k<=90: return "#F9E2AF"
            return "#A6E3A1"
        try:
            exp=self.db.get_expiring_documents(days=days)
            self.tbl_drv_exp.setRowCount(0)
            for d in exp["drivers"]:
                r=self.tbl_drv_exp.rowCount(); self.tbl_drv_exp.insertRow(r)
                self.tbl_drv_exp.setItem(r,0,QTableWidgetItem(d.get("full_name","")))
                self.tbl_drv_exp.setItem(r,1,QTableWidgetItem(d.get("src5_no","")))
                k=d.get("kalan"); it=QTableWidgetItem(f"{d.get('src5_expiry','')} ({_ks(k)})")
                it.setForeground(QColor(_kc(k))); self.tbl_drv_exp.setItem(r,2,it)
            self.tbl_veh_exp.setRowCount(0)
            for v in exp["vehicles"]:
                r=self.tbl_veh_exp.rowCount(); self.tbl_veh_exp.insertRow(r)
                self.tbl_veh_exp.setItem(r,0,QTableWidgetItem(v.get("plate","")))
                ak=v.get("adr_kalan"); mk=v.get("mua_kalan")
                ait=QTableWidgetItem(f"{v.get('adr_compliance_expiry','')} ({_ks(ak)})"); ait.setForeground(QColor(_kc(ak)))
                mit=QTableWidgetItem(f"{v.get('inspection_expiry','')} ({_ks(mk)})"); mit.setForeground(QColor(_kc(mk)))
                self.tbl_veh_exp.setItem(r,1,ait); self.tbl_veh_exp.setItem(r,2,mit)
        except Exception as _e: logger.warning("Son kullanma tablosu yüklenemedi: %s", _e)

    def _export_excel(self):
        try: import openpyxl
        except ImportError:
            QMessageBox.warning(self,"Hata","pip install openpyxl"); return
        year=self.cmb_year.currentData(); year_str=str(year) if year else "Tümü"
        path,_=QFileDialog.getSaveFileName(self,"Excel Kaydet",f"ADR_Rapor_{year_str}.xlsx","Excel (*.xlsx)")
        if not path: return
        wb=openpyxl.Workbook()
        from openpyxl.styles import Font,PatternFill,Alignment
        hf=Font(bold=True,color="FFFFFF"); hfill=PatternFill("solid",fgColor="1E3A5F"); ctr=Alignment(horizontal="center")
        def _ws(title,headers,rows):
            ws=wb.create_sheet(title); ws.append(headers)
            for c in ws[1]: c.font=hf; c.fill=hfill; c.alignment=ctr
            for row in rows: ws.append(row)
            for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=max(len(str(c.value or "")) for c in col)+4
        _ws("ADR Sınıfları",["Sınıf","Sevkiyat","Toplam Net (kg)"],
            [(r["class_code"],r["sevkiyat_sayisi"],round(r["toplam_net_kg"] or 0,1)) for r in self.db.get_class_breakdown(year=year)])
        _ws("Göndericiler",["Firma","Sevkiyat"],
            [(r["name"],r["sevkiyat_sayisi"]) for r in self.db.get_top_senders(50,year=year)])
        _ws("UN Numaraları",["UN No","Sınıf","Adet","Toplam (kg)"],
            [(f"UN{r['un_number']}",r["class_code"],r["adet"],round(r["toplam_net_kg"] or 0,1)) for r in self.db.get_top_chemicals(50,year=year)])
        if "Sheet" in wb.sheetnames: del wb["Sheet"]
        wb.save(path); QMessageBox.information(self,"Başarılı",f"Kaydedildi:\n{path}")

class SafetyPlansPage(QWidget):
    """
    ADR 1.10.3 Emniyet Planları Sayfası.
    Sekmeler:
      0 — Gereksinim Sorgulama  (SecurityPlanEngine)
      1 — Plan Belgesi          (HTML şablon + yazdır/PDF)
      2 — Tablo 1.10.3.1.2      (referans)
      3 — Tablo 1.10.3.1.3      (radyonüklid eşikleri)
      4 — Acil Müdahale         (1.10 / 8.1.2.1)
    """

    TRANSPORT_MODES = [("ambalaj", "Ambalaj"),
                       ("tank",    "Tank (l)"),
                       ("dokme",   "Dökme Yük (kg)")]

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.parent_window = parent  # ADRTransportPro referansı
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 12, 16, 12)
        root.setSpacing(10)

        hdr = QHBoxLayout()
        title = QLabel("🛡 Emniyet Planları — ADR 1.10.3")
        title.setStyleSheet("font-size:14pt;font-weight:700;color:#89B4FA;")
        hdr.addWidget(title); hdr.addStretch()
        root.addLayout(hdr)

        self._tabs = QTabWidget()
        self._tabs.setStyleSheet("QTabBar::tab{padding:6px 14px;}")

        # ── Tab 0: Gereksinim Sorgulama ──────────────────────────────────
        q_widget = QWidget()
        ql = QVBoxLayout(q_widget); ql.setSpacing(10)

        info_bar = QLabel(
            "Bu araç, mevcut sevkiyatınızdaki kalemleri ADR Madde 1.10.3 "
            "(Tablo 1.10.3.1.2) hükümlerine göre değerlendirerek emniyet planı "
            "gerekliliğini hesaplar.")
        info_bar.setWordWrap(True)
        info_bar.setStyleSheet(
            "background:#1E3A5F;color:#BAD4F5;padding:8px 12px;"
            "border-radius:4px;font-size:9pt;")
        ql.addWidget(info_bar)

        # Taşıma modu seçimi (Sınıf 7 / radyoaktif kalemler için — diğerleri ambalaj türünden otomatik)
        mode_row = QHBoxLayout()
        lbl_mode = QLabel("Sınıf 7 modu:")
        lbl_mode.setToolTip("Radyoaktif (Sınıf 7) kalemler için taşıma modu.\n"
                            "Diğer tüm kalemler ambalaj türünden otomatik belirlenir.")
        mode_row.addWidget(lbl_mode)
        self._cmb_mode = QComboBox()
        for val, lbl in self.TRANSPORT_MODES:
            self._cmb_mode.addItem(lbl, val)
        self._cmb_mode.setFixedWidth(160)
        self._cmb_mode.setToolTip("Yalnızca Sınıf 7 (radyoaktif) kalemler için geçerlidir.\n"
                                  "Tank / Dökme / Ambalajlı kalemler kendi ambalaj türünden otomatik belirlenir.")
        mode_row.addWidget(self._cmb_mode)

        btn_from_shipment = QPushButton("🔄 Mevcut Sevkiyattan Hesapla")
        btn_from_shipment.setObjectName("success")
        btn_from_shipment.clicked.connect(self._calc_from_shipment)
        mode_row.addWidget(btn_from_shipment)
        mode_row.addStretch()
        ql.addLayout(mode_row)

        # Sonuç kutusu
        self._result_box = QFrame()
        self._result_box.setStyleSheet(
            "QFrame{background:#1C2128;border:1px solid #30363D;border-radius:6px;padding:8px;}")
        rb_l = QVBoxLayout(self._result_box)
        self._lbl_verdict = QLabel("Henüz hesaplanmadı")
        self._lbl_verdict.setStyleSheet("font-size:13pt;font-weight:700;")
        self._lbl_verdict.setAlignment(Qt.AlignmentFlag.AlignCenter)
        rb_l.addWidget(self._lbl_verdict)
        self._txt_reasons = QPlainTextEdit()
        self._txt_reasons.setReadOnly(True)
        self._txt_reasons.setMaximumHeight(100)
        self._txt_reasons.setStyleSheet("font-size:8.5pt;")
        rb_l.addWidget(self._txt_reasons)
        ql.addWidget(self._result_box)

        # Detay log
        lbl_det = QLabel("Hesaplama Detayı:")
        lbl_det.setStyleSheet("font-size:9pt;font-weight:600;color:#A6ADC8;")
        ql.addWidget(lbl_det)
        self._txt_detail = QPlainTextEdit()
        self._txt_detail.setReadOnly(True)
        self._txt_detail.setStyleSheet("font-size:8pt;font-family:'Courier New';")
        ql.addWidget(self._txt_detail, 1)

        self._tabs.addTab(q_widget, "🔍 Gereksinim Sorgula")

        # ── Tab 1: Plan Belgesi ──────────────────────────────────────────
        p_widget = QWidget()
        pl = QVBoxLayout(p_widget); pl.setSpacing(8)
        p_bar = QHBoxLayout()
        btn_gen = QPushButton("📄 Boş Plan Şablonu Oluştur")
        btn_gen.clicked.connect(self._gen_template)
        btn_sample = QPushButton("📋 Doldurulmuş Örnek")
        btn_sample.setToolTip("UN 1203 Benzin / Tank taşıma örneği — TMGD perspektifli gerçekçi şablon")
        btn_sample.clicked.connect(self._gen_sample)
        btn_preview = QPushButton("🖨 Önizle / Yazdır / PDF")
        btn_preview.setObjectName("primary")
        btn_preview.clicked.connect(self._preview_and_print)
        for b in (btn_gen, btn_sample, btn_preview):
            p_bar.addWidget(b)
        p_bar.addStretch()
        pl.addLayout(p_bar)

        hint = QLabel(
            "✏ Belgedeki boş/çizgili alanlara tıklayıp doğrudan yazabilirsiniz. "
            "Doldurduktan sonra 'Önizle / Yazdır / PDF' ile A4 çıktısını alın.")
        hint.setWordWrap(True)
        hint.setStyleSheet("font-size:8pt;color:#A6ADC8;padding:2px 2px 0;")
        pl.addWidget(hint)

        self._txt_plan = QTextEdit()
        self._txt_plan.setStyleSheet(
            "QTextEdit{background:#FFFFFF;color:#000000;"
            "border:1px solid #30363D;border-radius:4px;padding:4px;}")
        pl.addWidget(self._txt_plan, 1)
        self._tabs.addTab(p_widget, "📋 Plan Belgesi")

        # ── Tab 2: Envanter İnceleme Raporu (v4.7) ───────────────────────
        inv_widget = self._build_inventory_review_widget()
        self._tabs.addTab(inv_widget, "📋 Envanter İnceleme Raporu")

        # ── Tab 3: Tablo 1.10.3.1.2 ─────────────────────────────────────
        t1_widget = self._build_table_1_widget()
        self._tabs.addTab(t1_widget, "📊 Tablo 1.10.3.1.2")

        # ── Tab 4: Tablo 1.10.3.1.3 (Radyonüklid) ───────────────────────
        t3_widget = self._build_radionuclide_widget()
        self._tabs.addTab(t3_widget, "☢ Radyonüklid Eşikleri")

        # ── Tab 5: Acil Müdahale ─────────────────────────────────────────
        ac_widget = self._build_emergency_widget()
        self._tabs.addTab(ac_widget, "🚨 Acil Müdahale")

        root.addWidget(self._tabs, 1)

    # ── [v4.7] Envanter İnceleme Raporu ─────────────────────────────────
    def _build_inventory_review_widget(self) -> QWidget:
        """Firmanın kimyasal envanterini (chemicals + company_products
        tablolarından) Tablo 1.10.3.1.2 kapsamında statik olarak tarayıp,
        örnek 'Güvenlik Planı İnceleme Raporu' formatında çok sayfalı PDF
        üreten sekme."""
        w = QWidget()
        lv = QVBoxLayout(w); lv.setSpacing(10)

        info = QLabel(
            "Bu araç, firmanızın kayıtlı kimyasal envanterindeki her maddeyi "
            "ADR Tablo 1.10.3.1.2 kapsamında (sınıf/ambalajlama grubu bazında, "
            "miktardan bağımsız) statik olarak tarar ve kaç kimyasalın güvenlik "
            "planı değerlendirmesi kapsamına girdiğini/girmediğini gösteren "
            "resmi formatta bir PDF rapor üretir.")
        info.setWordWrap(True)
        info.setStyleSheet(
            "background:#1E3A5F;color:#BAD4F5;padding:8px 12px;"
            "border-radius:4px;font-size:9pt;")
        lv.addWidget(info)

        btn_row = QHBoxLayout()
        btn_scan = QPushButton("🔄 Envanteri Tara")
        btn_scan.setObjectName("success")
        btn_scan.clicked.connect(self._scan_inventory)
        btn_row.addWidget(btn_scan)
        btn_row.addStretch()
        lv.addLayout(btn_row)

        self._inv_summary = QLabel("Henüz taranmadı.")
        self._inv_summary.setStyleSheet("font-size:10pt;font-weight:600;")
        lv.addWidget(self._inv_summary)

        self._inv_table = QTableWidget(0, 5)
        self._inv_table.setHorizontalHeaderLabels(
            ["UN No", "Madde", "Sınıf / PG", "Durum", "Sonuç"])
        self._inv_table.horizontalHeader().setSectionResizeMode(
            4, QHeaderView.ResizeMode.Stretch)
        self._inv_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        lv.addWidget(self._inv_table, 1)

        form_box = QFrame()
        form_box.setStyleSheet(
            "QFrame{background:#1C2128;border:1px solid #30363D;"
            "border-radius:6px;padding:8px;}")
        form = QFormLayout(form_box)
        self._inv_company = QLineEdit()
        self._inv_prepared = QLineEdit()
        self._inv_approved = QLineEdit()
        self._inv_validity = QSpinBox()
        self._inv_validity.setRange(1, 10)
        self._inv_validity.setValue(2)
        self._inv_validity.setSuffix(" Yıl")
        form.addRow("Firma Adı:", self._inv_company)
        form.addRow("Hazırlayan (Danışman):", self._inv_prepared)
        form.addRow("Onaylayan:", self._inv_approved)
        form.addRow("Geçerlilik Süresi:", self._inv_validity)
        lv.addWidget(form_box)

        btn_pdf = QPushButton("🖨 PDF Rapor Oluştur")
        btn_pdf.setObjectName("primary")
        btn_pdf.clicked.connect(self._export_inventory_review_pdf)
        lv.addWidget(btn_pdf)

        self._inv_screen_result = None
        return w

    def _load_inventory_chemicals(self) -> list:
        """company_products'ta kayıtlı benzersiz (UN, kod, PG) kombinasyonlarını
        chemicals tablosundan tam kayıt olarak çeker. company_products boşsa
        (henüz firma envanteri içe aktarılmamışsa) tüm chemicals döner."""
        rows = self.db.execute(
            "SELECT DISTINCT un_number, classification_code, packing_group "
            "FROM company_products")
        if not rows:
            all_rows = self.db.execute("SELECT * FROM chemicals")
            return [self.db._row_to_chemical(r) for r in (all_rows or [])]
        result = []
        for r in rows:
            row = self.db.execute_one(
                "SELECT * FROM chemicals WHERE un_number=? AND "
                "classification_code=? AND packing_group=?",
                (r["un_number"], r["classification_code"] or "",
                 r["packing_group"] or ""))
            if row:
                result.append(self.db._row_to_chemical(row))
        return result

    def _scan_inventory(self):
        chemicals = self._load_inventory_chemicals()
        if not chemicals:
            QMessageBox.information(self, "Bilgi",
                "Envanterde kimyasal bulunamadı. Önce Tablo A / firma "
                "envanteri Excel içe aktarımını yapın.")
            return
        result = SecurityPlanEngine.screen_inventory(chemicals)
        self._inv_screen_result = result

        self._inv_summary.setText(
            f"Toplam {result['total']} kimyasal tarandı  —  "
            f"Kapsam İçi: {result['in_scope_count']}  |  "
            f"Koşullu: {result['conditional_count']}  |  "
            f"Muaf: {result['exempt_count']}")

        self._inv_table.setRowCount(len(result["results"]))
        for i, r in enumerate(result["results"]):
            status = ("KAPSAM İÇİ" if r["in_scope"] is True else
                      "KOŞULLU" if r["in_scope"] == "conditional" else "MUAF")
            color = ("#F38BA8" if r["in_scope"] is True else
                    "#F9E2AF" if r["in_scope"] == "conditional" else "#A6E3A1")
            self._inv_table.setItem(i, 0, QTableWidgetItem(f"UN{r['un_number']}"))
            self._inv_table.setItem(i, 1, QTableWidgetItem(r["name"]))
            self._inv_table.setItem(i, 2, QTableWidgetItem(
                f"{r['class_code']} / {r['packing_group'] or '—'}"))
            status_item = QTableWidgetItem(status)
            status_item.setForeground(QColor(color))
            self._inv_table.setItem(i, 3, status_item)
            self._inv_table.setItem(i, 4, QTableWidgetItem(r["sonuc_text"]))

        if result["in_scope_count"] or result["conditional_count"]:
            QMessageBox.warning(self, "Dikkat",
                f"{result['in_scope_count'] + result['conditional_count']} "
                f"kimyasal Tablo 1.10.3.1.2 kapsamına girmektedir — bu maddeler "
                f"için güvenlik planı değerlendirmesi gereklidir.")
        else:
            QMessageBox.information(self, "Tamam",
                "Taranan hiçbir kimyasal Tablo 1.10.3.1.2 kapsamına girmiyor — "
                "güvenlik planı hazırlanmasına gerek yoktur.")

    def _export_inventory_review_pdf(self):
        if not self._inv_screen_result:
            QMessageBox.warning(self, "Eksik", "Önce 'Envanteri Tara' butonuna basın.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Güvenlik Planı İnceleme Raporu",
            "guvenlik_plani_inceleme_raporu.pdf", "PDF (*.pdf)")
        if not path:
            return
        try:
            logo_b64 = self.db.get_company_logo_b64()
            html = SecurityPlanEngine.generate_inventory_review_html(
                company_name=self._inv_company.text(),
                prepared_by=self._inv_prepared.text(),
                approved_by=self._inv_approved.text(),
                screen_result=self._inv_screen_result,
                validity_years=self._inv_validity.value(),
                logo_b64=logo_b64,
            )
            doc = QTextDocument()
            doc.setHtml(html)
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(path)
            doc.print(printer)
            QMessageBox.information(self, "Tamam", f"Rapor kaydedildi:\n{path}")
        except Exception as exc:
            logging.getLogger(__name__).exception(
                "Güvenlik planı inceleme raporu oluşturulamadı")
            QMessageBox.critical(self, "Hata",
                f"Rapor oluşturulurken bir hata oluştu:\n{_turkce_hata_metni(exc)}")

    # ── Hesaplama ─────────────────────────────────────────────────────────
    def _calc_from_shipment(self):
        """Ana penceredeki aktif sevkiyat kalemlerini al ve hesapla."""
        items = []
        pts   = 0.0
        hata  = ""
        try:
            # Önce doğrudan referansa bak (en güvenilir yol)
            sp = getattr(self, 'shipment_page_ref', None)

            # Yoksa parent_window üzerinden dene
            if sp is None:
                main = getattr(self, 'parent_window', None) or self.parent()
                sp = getattr(main, 'shipment_page', None)

            if sp is None:
                hata = "shipment_page bulunamadı."
            else:
                sp._sync_items_from_table()
                items = sp.get_items()
                if items:
                    pts, _, _ = ADREngine.calculate_1136_points(items)
                # items boşsa pts=0.0 kalır, SecurityPlanEngine bunu raporlar
        except Exception as e:
            hata = str(e)
            items = []
            pts   = 0.0

        mode   = self._cmb_mode.currentData()
        result = SecurityPlanEngine.check(items, transport_mode=mode,
                                          total_1136_points=pts)
        if hata:
            result["details"].insert(0, f"⚠ Hata: {hata}")
        self._display_result(result, len(items))

    def _display_result(self, result: dict, item_count: int):
        if result["required"]:
            self._lbl_verdict.setText("🔴 EMNİYET PLANI GEREKLİ")
            self._lbl_verdict.setStyleSheet(
                "font-size:13pt;font-weight:700;color:#F38BA8;"
                "background:#2D0F0F;border-radius:6px;padding:6px;")
            self._txt_reasons.setPlainText(
                "\n".join(f"• {r}" for r in result["reasons"]) or "—")
        elif result["exempt"]:
            self._lbl_verdict.setText("✅ EMNİYET PLANI GEREKMİYOR")
            self._lbl_verdict.setStyleSheet(
                "font-size:13pt;font-weight:700;color:#A6E3A1;"
                "background:#0F2D0F;border-radius:6px;padding:6px;")
            self._txt_reasons.setPlainText("ADR 1.10.4 muafiyeti uygulanıyor.")
        else:
            self._lbl_verdict.setText("⚪ Belirlenemedi")
            self._lbl_verdict.setStyleSheet("font-size:13pt;font-weight:700;")
            self._txt_reasons.setPlainText("Sevkiyat verisi yok veya hesaplanamadı.")

        self._txt_detail.setPlainText(
            f"Değerlendirilen kalem sayısı: {item_count}\n"
            + "\n".join(result.get("details", [])))

    # ── Plan Belgesi ─────────────────────────────────────────────────────
    def _gen_template(self):
        try:
            main = self.parent_window
            sp   = getattr(main, 'shipment_page', None) if main else None
            shipment = getattr(sp, 'current_shipment', None) if sp else None
            items    = sp.get_items() if sp else []
        except Exception:
            shipment = None; items = []
        html = SecurityPlanEngine.generate_plan_template(shipment, items)
        self._txt_plan.setHtml(html)
        self._tabs.setCurrentIndex(1)

    def _gen_sample(self):
        """Doldurulmuş gerçekçi örnek şablonu yükle."""
        html = SecurityPlanEngine.generate_plan_template(sample=True)
        self._txt_plan.setHtml(html)
        self._tabs.setCurrentIndex(1)

    def _preview_and_print(self):
        """
        Kullanıcının doldurduğu (veya henüz boş) plan belgesini, mevcut
        PrintPreviewDialog ile A4 önizleme + yazdır + PDF kaydet olarak açar.
        QTextEdit içeriği (kullanıcı düzenlemeleri dahil) kullanılır.
        """
        if not self._txt_plan.toPlainText().strip():
            self._gen_template()
        html = self._txt_plan.toHtml()
        dlg = PrintPreviewDialog(self, html, "Emniyet Planı — Önizleme")
        dlg.exec()

    # ── Tablo 1.10.3.1.2 widget ─────────────────────────────────────────
    def _build_table_1_widget(self) -> QWidget:
        w = QWidget(); lay = QVBoxLayout(w)
        lbl = QLabel("Tablo 1.10.3.1.2 — Ciddi Sonuçlara Yol Açabilecek Tehlikeli Mallar")
        lbl.setStyleSheet("font-size:10pt;font-weight:700;color:#89B4FA;")
        lay.addWidget(lbl)
        note = QLabel("a = İlgili Değil (uygulanamaz)   b = Muaf (miktar ne olursa)   0 = Her miktarda zorunlu   N = N litre/kg üstünde zorunlu")
        note.setStyleSheet("font-size:8pt;color:#A6ADC8;")
        lay.addWidget(note)

        tbl = QTableWidget(0, 5)
        tbl.setHorizontalHeaderLabels(["Sınıf / Durum", "Koşul", "Tank Limiti (l)", "Dökme Limiti (kg)", "Ambalaj Limiti (kg)"])
        tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)

        rows = [
            ("1.1 Patlayıcılar",       "1.1",        "a",    "a",    "0"),
            ("1.2 Patlayıcılar",       "1.2",        "a",    "a",    "0"),
            ("1.3 Patlayıcılar (C)",   "1.3C",       "a",    "a",    "0"),
            ("1.4 (Özel UN)",          "1.4_special","a",    "a",    "0"),
            ("1.5 Patlayıcılar",       "1.5",        "0",    "a",    "0"),
            ("2 Alevlenir Gaz (F/FC)", "F/FC",       "3000", "a",    "b"),
            ("2 Zehirli Gaz (T,TF…)",  "T,TF,TC…",   "0",    "a",    "0"),
            ("3 Alevlenir Sıvı PGI",   "PGI",        "3000", "a",    "b"),
            ("3 Alevlenir Sıvı PGII",  "PGII",       "3000", "a",    "b"),
            ("3 Duyarsızlaştırılmış",  "desens.",    "0",    "a",    "0"),
            ("4.1 Duyarsızlaştırılmış","desens.",    "a",    "a",    "0"),
            ("4.2 PGI",                "PGI",        "3000", "a",    "b"),
            ("4.3 PGI",                "PGI",        "3000", "a",    "b"),
            ("5.1 PGI Sıvı Yükseltgen","PGI sıvı",  "3000", "a",    "b"),
            ("5.1 Perklorat/AN/Gübre", "perc./AN",   "3000", "3000", "b"),
            ("6.1 PGI Zehirli",        "PGI",        "0",    "a",    "0"),
            ("6.2 Kat.A Bulaşıcı",     "Kategori A", "a",    "0",    "0"),
            ("8 PGI Aşındırıcı",       "PGI",        "3000", "a",    "b"),
        ]
        tbl.setRowCount(len(rows))
        RED = QColor("#F38BA8"); GRN = QColor("#A6E3A1")
        for r, (cls_name, cond, tank, bulk, pkg) in enumerate(rows):
            tbl.setItem(r, 0, QTableWidgetItem(cls_name))
            tbl.setItem(r, 1, QTableWidgetItem(cond))
            for c, val in enumerate([tank, bulk, pkg], 2):
                it = QTableWidgetItem(val)
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if val == "0":    it.setForeground(RED)
                elif val == "b":  it.setForeground(GRN)
                elif val == "a":  it.setForeground(QColor("#585B70"))
                tbl.setItem(r, c, it)
        lay.addWidget(tbl, 1)
        return w

    # ── Radyonüklid widget ───────────────────────────────────────────────
    def _build_radionuclide_widget(self) -> QWidget:
        w = QWidget(); lay = QVBoxLayout(w)
        lbl = QLabel("Tablo 1.10.3.1.3 — Belirli Radyonüklidler İçin Taşıma Emniyet Eşiği")
        lbl.setStyleSheet("font-size:10pt;font-weight:700;color:#89B4FA;")
        lay.addWidget(lbl)
        note = QLabel("Tabloda yer almayan nüklidler için genel eşik: 3000 TBq  |  Karışımlar için: ∑(Ai / Ti) ≥ 1 → Emniyet planı zorunlu")
        note.setStyleSheet("font-size:8pt;color:#A6ADC8;"); note.setWordWrap(True)
        lay.addWidget(note)

        tbl = QTableWidget(0, 3)
        tbl.setHorizontalHeaderLabels(["Element", "Radyonüklid", "Eşik (TBq)"])
        tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)

        data = [
            ("Amerikyum","Am-241",0.6),("Altın","Au-198",2),("Kadmiyum","Cd-109",200),
            ("Kaliforniyum","Cf-252",0.2),("Küriyum","Cm-244",0.5),("Kobalt","Co-57",7),
            ("Kobalt","Co-60",0.3),("Sezyum","Cs-137",1),("Demir","Fe-55",8000),
            ("Germanyum","Ge-68",7),("Gadolinyum","Gd-153",10),("İridyum","Ir-192",0.8),
            ("Nikel","Ni-63",600),("Paladyum","Pd-103",900),("Prometyum","Pm-147",400),
            ("Polonyum","Po-210",0.6),("Plütonyum","Pu-238",0.6),("Plütonyum","Pu-239",0.6),
            ("Radyum","Ra-226",0.4),("Rutenyum","Ru-106",3),("Selenyum","Se-75",2),
            ("Stronsiyum","Sr-90",10),("Talyum","Tl-204",200),("Tulyum","Tm-170",200),
            ("İterbiyum","Yb-169",3),
        ]
        tbl.setRowCount(len(data))
        for r, (elem, nuc, val) in enumerate(data):
            tbl.setItem(r, 0, QTableWidgetItem(elem))
            tbl.setItem(r, 1, QTableWidgetItem(nuc))
            v_it = QTableWidgetItem(str(val))
            v_it.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            if val <= 1:    v_it.setForeground(QColor("#F38BA8"))
            elif val <= 10: v_it.setForeground(QColor("#FAB387"))
            tbl.setItem(r, 2, v_it)
        lay.addWidget(tbl, 1)
        return w

    # ── Acil Müdahale widget ─────────────────────────────────────────────
    def _build_emergency_widget(self) -> QWidget:
        w = QWidget(); scroll = QScrollArea(); scroll.setWidgetResizable(True)
        inner = QWidget(); lay = QVBoxLayout(inner); lay.setSpacing(8)
        outer = QVBoxLayout(w); outer.addWidget(scroll); scroll.setWidget(inner)

        hdr = QLabel("⚠ KAZA ANINDA İLK YAPILACAKLAR — ADR 8.1.2.1 / 1.10")
        hdr.setStyleSheet("font-size:11pt;font-weight:700;color:#F38BA8;"
                          "padding:8px;background:#2D1B1B;border-radius:4px;")
        lay.addWidget(hdr)

        steps = [
            ("1. DURUN — Motoru Kapatın",
             "Güvenli bir yerde durun, motoru kapatın. Uyarı ışıklarını açın. "
             "Emniyet üçgeni kurun (araç arızalıysa minimum 50 m)."),
            ("2. Değerlendirin — Yaklaşmayın",
             "Sızıntı, yangın veya kaza durumunu uzaktan değerlendirin. "
             "Tehlikeli maddeye yaklaşmayın. Rüzgarı arkanıza alarak durumlanın."),
            ("3. Tahliye Edin",
             "Yakın çevreyi tahliye edin. İzinsiz kişileri uzaklaştırın. "
             "Yangın: Minimum 300 m. Gaz kaçağı: Minimum 500 m güvenli alan sağlayın."),
            ("4. Yetkilileri Bilgilendirin",
             "İtfaiye: 110  |  Ambulans: 112  |  Polis: 155  |  Kriz Masası: 157\n"
             "UN numarasını, sınıfı, miktarı ve konumu bildirin."),
            ("5. ADR Belgelerini Teslim Edin",
             "Taşıma evraklarını, emniyet bilgi formlarını (SDS) ve emniyet planını "
             "müdahale ekiplerine teslim edin. Araç içinde bırakmayın."),
            ("6. Kendinizi Tehlikeye Atmayın",
             "Eğitim almadıysanız müdahale etmeyin. Yetkililerin talimatını bekleyin. "
             "Kişisel koruyucu ekipman olmadan tehlikeli bölgeye girmeyin."),
        ]
        for title_s, desc in steps:
            frame = QFrame()
            frame.setStyleSheet(
                "QFrame{background:#1E2535;border-radius:6px;border:1px solid #313244;}"
                "padding:4px;")
            fl = QVBoxLayout(frame); fl.setSpacing(3)
            t = QLabel(title_s); t.setStyleSheet("font-weight:700;color:#CBA6F7;font-size:10pt;")
            d = QLabel(desc); d.setWordWrap(True); d.setStyleSheet("color:#CDD6F4;font-size:9pt;")
            fl.addWidget(t); fl.addWidget(d); lay.addWidget(frame)

        # Yazılı talimat / Emniyet planı yasal dayanaği
        legal = QLabel(
            "📋 Yasal Dayanak: ADR 2023 — Bölüm 1.10 (Emniyet), "
            "5.4.3 (Yazılı Talimatlar), 8.1.2.1 (Belgeler), "
            "8.5 (Ek Güvenlik Hükümleri)")
        legal.setWordWrap(True)
        legal.setStyleSheet("font-size:8pt;color:#A6ADC8;margin-top:8px;"
                            "padding:6px;background:#1C2128;border-radius:4px;")
        lay.addWidget(legal)
        lay.addStretch()
        return w

class SettingsPage(QWidget):
    """Uygulama ayarları."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db=db; self._build_ui(); self._load()

    def _sec(self, lay, txt):
        l=QLabel(f"⬡  {txt}"); l.setStyleSheet("font-size:9pt;font-weight:700;letter-spacing:2px;color:#585B70;border-bottom:2px solid #313244;margin-top:8px;"); lay.addWidget(l)

    def _row(self, lay, label, widget, hint=""):
        hl=QHBoxLayout(); lbl=QLabel(label); lbl.setFixedWidth(200); lbl.setStyleSheet("color:#CDD6F4;font-size:10pt;")
        hl.addWidget(lbl); hl.addWidget(widget); lay.addLayout(hl)
        if hint: h=QLabel(hint); h.setStyleSheet("color:#585B70;font-size:8pt;margin-left:205px;"); lay.addWidget(h)

    def _build_ui(self):
        root=QVBoxLayout(self); root.setContentsMargins(0,0,0,0); root.setSpacing(0)
        hdr=QWidget(); hdr.setStyleSheet("background:#181825;border-bottom:1px solid #313244;")
        hl=QHBoxLayout(hdr); hl.setContentsMargins(20,12,20,12)
        ttl=QLabel("⚙ Ayarlar"); ttl.setStyleSheet("font-size:14pt;font-weight:700;color:#CDD6F4;")
        hl.addWidget(ttl); hl.addStretch()
        btn_s=QPushButton("💾 Tüm Ayarları Kaydet"); btn_s.setFixedWidth(180)
        btn_s.setStyleSheet("background:#89B4FA;color:#1E1E2E;font-weight:700;border-radius:6px;padding:6px;")
        btn_s.clicked.connect(self._save); hl.addWidget(btn_s); root.addWidget(hdr)
        tabs=QTabWidget(); tabs.setStyleSheet("QTabWidget::pane{border:none;background:#1E1E2E;}QTabBar::tab{background:#181825;color:#585B70;padding:8px 16px;border:none;font-size:10pt;}QTabBar::tab:selected{background:#1E1E2E;color:#CDD6F4;border-bottom:2px solid #89B4FA;}")
        root.addWidget(tabs,1)
        tabs.addTab(self._tab_doc(),    "📄 Belge & Logo")
        tabs.addTab(self._tab_warn(),   "🔔 Uyarı Eşikleri")
        tabs.addTab(self._tab_backup(), "💾 Yedekleme")
        tabs.addTab(self._tab_integ(),  "🔗 Entegrasyonlar")

    def _tab_doc(self):
        w=QWidget(); lay=QVBoxLayout(w); lay.setContentsMargins(24,16,24,16); lay.setSpacing(8)
        self._sec(lay,"FİRMA BİLGİLERİ")
        self.txt_co_name=QLineEdit(); self.txt_co_addr=QLineEdit(); self.txt_co_phone=QLineEdit()
        self.txt_co_email=QLineEdit(); self.txt_co_web=QLineEdit()
        self._row(lay,"Şirket Adı",self.txt_co_name,)
        self._row(lay,"Adres",self.txt_co_addr)
        self._row(lay,"Telefon",self.txt_co_phone,)
        self._row(lay,"E-posta",self.txt_co_email,)
        self._row(lay,"Web Sitesi",self.txt_co_web,)
        self._sec(lay,"ŞİRKET LOGOSU")
        logo_row=QHBoxLayout()
        self.lbl_logo=QLabel("Logo Yok"); self.lbl_logo.setFixedSize(125,50)
        self.lbl_logo.setStyleSheet("border:1px dashed #313244;border-radius:5px;background:#181825;color:#585B70;font-size:8pt;")
        self.lbl_logo.setAlignment(Qt.AlignmentFlag.AlignCenter); logo_row.addWidget(self.lbl_logo)
        bc=QVBoxLayout()
        b_sel=QPushButton("📂 Logo Seç (PNG/JPG)"); b_sel.setFixedWidth(160); b_sel.clicked.connect(self._sel_logo)
        b_clr=QPushButton("🗑 Logoyu Kaldır"); b_clr.setFixedWidth(160); b_clr.setStyleSheet("color:#F38BA8;"); b_clr.clicked.connect(self._clr_logo)
        bc.addWidget(b_sel); bc.addWidget(b_clr)
        hint=QLabel()
        hint.setStyleSheet("color:#585B70;font-size:6pt;"); hint.setWordWrap(True); bc.addWidget(hint)
        logo_row.addLayout(bc); logo_row.addStretch(); lay.addLayout(logo_row)
        self._sec(lay,"QR KOD")
        self.chk_qr=QCheckBox("PDF evraklarda QR kod göster"); self.chk_qr.setChecked(True); lay.addWidget(self.chk_qr)
        info_qr=QLabel()
        info_qr.setStyleSheet("color:#585B70;font-size:6pt;"); lay.addWidget(info_qr)
        self._sec(lay,"VARSAYILAN EVRAK NOTU")
        self.txt_note=QPlainTextEdit(); self.txt_note.setPlaceholderText("Her evrakın altına eklenecek not..."); self.txt_note.setMaximumHeight(100); lay.addWidget(self.txt_note)
        lay.addStretch(); return w

    def _tab_warn(self):
        w=QWidget(); lay=QVBoxLayout(w); lay.setContentsMargins(24,16,24,16); lay.setSpacing(8)
        self._sec(lay,"BİTİŞ TARİHİ UYARI EŞİKLERİ")
        lay.addWidget(QLabel("Belirtilen günden az süre kaldığında uyarı verilir."))
        def spn():
            s=QSpinBox(); s.setRange(1,365); s.setValue(30); s.setSuffix(" gün"); s.setFixedWidth(120); return s
        self.spn_src5=spn(); self.spn_adr=spn(); self.spn_insp=spn()
        self._row(lay,"SRC5 Belgesi",self.spn_src5); self._row(lay,"ADR Sertifikası",self.spn_adr); self._row(lay,"Araç Muayenesi",self.spn_insp)
        lay.addStretch(); return w

    def _tab_backup(self):
        w=QWidget(); lay=QVBoxLayout(w); lay.setContentsMargins(24,16,24,16); lay.setSpacing(10)
        self._sec(lay,"MANUEL YEDEKLEME")
        hl=QHBoxLayout()
        btn_b=QPushButton("📦 Şimdi Yedek Al"); btn_b.setFixedWidth(160)
        btn_b.setStyleSheet("background:#A6E3A1;color:#1E1E2E;font-weight:700;border-radius:6px;padding:6px;")
        btn_b.clicked.connect(self._do_backup); hl.addWidget(btn_b)
        self.lbl_bk=QLabel(""); self.lbl_bk.setStyleSheet("color:#A6E3A1;font-size:9pt;"); hl.addWidget(self.lbl_bk); hl.addStretch(); lay.addLayout(hl)
        self._sec(lay,"YEDEK LİSTESİ")
        self.tbl_bk=QTableWidget(0,3); self.tbl_bk.setHorizontalHeaderLabels(["Dosya","Tarih","Boyut"])
        for i in range(3): self.tbl_bk.horizontalHeader().setSectionResizeMode(i,QHeaderView.ResizeMode.Stretch)
        self.tbl_bk.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers); lay.addWidget(self.tbl_bk,1)
        hl2=QHBoxLayout()
        btn_r=QPushButton("↩ Seçili Yedeğe Geri Dön"); btn_r.setFixedWidth(200)
        btn_r.setStyleSheet("background:#F38BA8;color:#1E1E2E;font-weight:700;border-radius:6px;padding:6px;")
        btn_r.clicked.connect(self._do_restore); hl2.addWidget(btn_r)
        btn_ref=QPushButton("↻ Yenile"); btn_ref.clicked.connect(self._ref_bk); hl2.addWidget(btn_ref); hl2.addStretch(); lay.addLayout(hl2)
        self._ref_bk(); return w

    def _tab_integ(self):
        w=QWidget(); lay=QVBoxLayout(w); lay.setContentsMargins(24,16,24,16); lay.setSpacing(8)
        self._sec(lay,"U-ETDS ENTEGRASYONU (Geleceğe Hazırlık)")
        info=QLabel("Ulaştırma Elektronik Takip ve Denetim Sistemi. Bilgiler şifreli saklanır.")
        info.setStyleSheet("color:#585B70;font-size:9pt;"); info.setWordWrap(True); lay.addWidget(info)
        self.txt_uetds_u=QLineEdit(); self.txt_uetds_u.setPlaceholderText("Kullanıcı Adı / TC No")
        self.txt_uetds_p=QLineEdit(); self.txt_uetds_p.setPlaceholderText("Şifre"); self.txt_uetds_p.setEchoMode(QLineEdit.EchoMode.Password)
        self._row(lay,"Kullanıcı Adı",self.txt_uetds_u); self._row(lay,"Şifre",self.txt_uetds_p)
        note=QLabel("⚠ U-ETDS API entegrasyonu aktif değil — ileriki sürümde kullanılacak.")
        note.setStyleSheet("color:#F9E2AF;font-size:9pt;"); note.setWordWrap(True); lay.addWidget(note)
        lay.addStretch(); return w

    def _load(self):
        g=self.db.get_setting
        self.txt_co_name.setText(g("doc_company_name") or "")
        self.txt_co_addr.setText(g("doc_company_address") or "")
        self.txt_co_phone.setText(g("doc_company_phone") or "")
        self.txt_co_email.setText(g("doc_company_email") or "")
        self.txt_co_web.setText(g("doc_company_website") or "")
        self.txt_note.setPlainText(g("doc_default_note") or "")
        self.spn_src5.setValue(int(g("warn_days_src5") or 30))
        self.spn_adr.setValue(int(g("warn_days_adr_cert") or 30))
        self.spn_insp.setValue(int(g("warn_days_inspection") or 30))
        self.txt_uetds_u.setText(g("uetds_username") or "")
        self.txt_uetds_p.setText(g("uetds_password") or "")
        self.chk_qr.setChecked((g("doc_show_qr") or "1") == "1")
        self._ref_logo()

    def _save(self):
        s=self.db.set_setting
        s("doc_company_name", self.txt_co_name.text())
        s("doc_company_address", self.txt_co_addr.text())
        s("doc_company_phone", self.txt_co_phone.text())
        s("doc_company_email", self.txt_co_email.text())
        s("doc_company_website", self.txt_co_web.text())
        s("doc_default_note", self.txt_note.toPlainText())
        s("warn_days_src5", str(self.spn_src5.value()))
        s("warn_days_adr_cert", str(self.spn_adr.value()))
        s("warn_days_inspection", str(self.spn_insp.value()))
        s("uetds_username", self.txt_uetds_u.text())
        s("uetds_password", self.txt_uetds_p.text())
        s("doc_show_qr", "1" if self.chk_qr.isChecked() else "0")
        QMessageBox.information(self,"Kaydedildi","Ayarlar başarıyla kaydedildi.")

    def _ref_logo(self):
        b64=self.db.get_company_logo_b64()
        if b64:
            try:
                import base64 as _b; raw=_b.b64decode(b64); pix=QPixmap(); pix.loadFromData(raw)
                self.lbl_logo.setPixmap(pix.scaled(118,58,Qt.AspectRatioMode.KeepAspectRatio,Qt.TransformationMode.SmoothTransformation))
                self.lbl_logo.setText(""); return
            except Exception as _e:
                logger.warning("Logo görüntülenemedi: %s", _e)
        self.lbl_logo.setPixmap(QPixmap()); self.lbl_logo.setText("Logo Yok")

    def _sel_logo(self):
        path,_=QFileDialog.getOpenFileName(self,"Logo Seç","","Resim (*.png *.jpg *.jpeg)")
        if not path: return
        try:
            import base64 as _b
            b64=_b.b64encode(open(path,"rb").read()).decode()
            self.db.set_company_logo_b64(b64); self._ref_logo()
        except Exception as e:
                logging.getLogger(__name__).exception("Beklenmeyen hata")
                QMessageBox.critical(self, "Hata", _turkce_hata_metni(e))

    def _clr_logo(self):
        self.db.set_company_logo_b64(""); self._ref_logo()

    def _do_backup(self):
        try:
            path=self.db.backup_now(); self.lbl_bk.setText(f"✅ {Path(path).name}"); self._ref_bk()
        except Exception as e: self.lbl_bk.setText(f"❌ {e}"); self.lbl_bk.setStyleSheet("color:#F38BA8;")

    def _ref_bk(self):
        bks=self.db.list_backups(); self.tbl_bk.setRowCount(0)
        for b in bks:
            r=self.tbl_bk.rowCount(); self.tbl_bk.insertRow(r)
            self.tbl_bk.setItem(r,0,QTableWidgetItem(b["name"]))
            self.tbl_bk.setItem(r,1,QTableWidgetItem(b["date"]))
            self.tbl_bk.setItem(r,2,QTableWidgetItem(b["size"]))

    def _do_restore(self):
        row=self.tbl_bk.currentRow()
        if row<0: QMessageBox.warning(self,"Uyarı","Lütfen bir yedek seçin."); return
        name=self.tbl_bk.item(row,0).text()
        if QMessageBox.question(self,"Geri Yükle",
            f"'{name}' yedeğine geri yüklenecek.\nMevcut veriler silinecek! Devam?",
            QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No)!=QMessageBox.StandardButton.Yes: return
        bks=self.db.list_backups(); path=next((b["path"] for b in bks if b["name"]==name),None)
        if path:
            try: self.db.restore_backup(path); QMessageBox.information(self,"Başarılı","Yedek geri yüklendi. Programı yeniden başlatın.")
            except Exception as e:
                logging.getLogger(__name__).exception("Beklenmeyen hata")
                QMessageBox.critical(self, "Hata", _turkce_hata_metni(e))


# =============================================================================
# [v4.4] KARIŞIK YÜKLEME KONTROLÜ — adr_mix_pro entegrasyonu (ANA PROGRAM)
# =============================================================================
# main.py artık kullanılmıyor; bu entegrasyon doğrudan adr_transport_pro_2026.py
# içine, uygulamanın kendi SQL veritabanını (composite anahtar: UN +
# classification_code + packing_group) kaynak alacak şekilde inşa edilmiştir.
#
# Önemli tasarım kararı: Aynı UN numarasının birden fazla ADR Tablo A
# varyasyonu olabilir (örn. UN1950 AEROSOL = 12 varyasyon). Kullanıcı UN
# numarası elle girdiğinde ve birden fazla varyasyon varsa, SEÇENEKLER
# kullanıcıya gösterilir ve uygun olanı kendisi seçer (bkz. VariantPickerDialog).
# "Aktif Sevkiyattan Aktar" akışında ise kalemler zaten kendi
# classification_code/packing_group bilgisini taşıdığından belirsizlik yoktur.
# =============================================================================

from adr_mix_pro.constants import MAX_UN_ITEMS
from adr_mix_pro.core.checker import MixChecker
from adr_mix_pro.core.rule_engine import SegregationRuleEngine
from adr_mix_pro.models import ProductRecord
from adr_mix_pro.reports.excel_export import export_results_to_excel
from adr_mix_pro.reports.pdf_export import export_results_to_pdf
from adr_mix_pro.ui.result_detail_panel import ResultDetailPanel
from adr_mix_pro.ui.results_model import ResultsTableModel
from adr_mix_pro.validators import is_valid_un, normalize_un

_MIX_BASE_DIR = Path(__file__).resolve().parent
_MIX_RULE_FILE = _MIX_BASE_DIR / "resources" / "data" / "segregation_rules.csv"


class AnaDbChemicalAdapter:
    """Ana programın kendi SQL veritabanını (chemicals tablosu, composite
    anahtar UN+classification_code+packing_group) adr_mix_pro'nun beklediği
    ProductDatabase arayüzüne (try_get_record / all_records / search) çevirir.

    Klasik tek-UN-tek-ürün varsayımından farklı olarak, bu adaptör kayıtları
    ÖNCEDEN toptan yüklemez: her UN, kullanıcı veya "aktif sevkiyattan aktar"
    akışı tarafından register() ile AÇIKÇA hangi varyasyonun kullanılacağı
    belirtildikten sonra belleğe alınır. Böylece UN1950 gibi 12 varyasyonlu
    maddelerde yanlış/rastgele bir varyasyonun sessizce kullanılması engellenir.
    """

    def __init__(self, db: "DatabaseManager"):
        self.db = db
        self._records_by_un: Dict[str, ProductRecord] = {}

    @staticmethod
    def _extract_cv_codes(special_provisions: str) -> str:
        if not special_provisions:
            return ""
        codes = re.findall(r"CV\d+", special_provisions, re.IGNORECASE)
        return " ".join(sorted(set(c.upper() for c in codes)))

    def _row_to_record(self, row: "sqlite3.Row") -> ProductRecord:
        labels_raw = row["hazard_labels"] or ""
        labels = [l.strip() for l in labels_raw.replace(",", " ").split() if l.strip()]
        if not labels and row["class_code"]:
            labels = [str(row["class_code"]).strip()]
        return ProductRecord(
            un_no=row["un_number"],
            name=row["proper_shipping_name_tr"] or row["proper_shipping_name_en"] or "",
            hazard_class=row["class_code"] or "",
            classification_code=row["classification_code"] or "",
            packing_group=row["packing_group"] or "",
            labels=labels,
            special_provisions=row["special_provisions"] or "",
            transport_category=row["transport_category"] or "",
            cv_codes=self._extract_cv_codes(row["special_provisions"] or ""),
            tunnel_code=row["tunnel_code"] or "",
            raw=dict(row) if hasattr(row, "keys") else {},
        )

    def get_variants(self, un: str) -> List["sqlite3.Row"]:
        """Verilen UN icin TUM Tablo A varyasyonlarini (composite anahtar)
        dondurur; birden fazla olabilir (orn. UN1950 -> 12 satir)."""
        un = normalize_un(un)
        rows = self.db.execute(
            "SELECT * FROM chemicals WHERE un_number=? "
            "ORDER BY classification_code, packing_group",
            (un,))
        return rows or []

    def register_variant(self, un: str, classification_code: str = "",
                          packing_group: str = "") -> Optional[ProductRecord]:
        """Kullanicinin (veya sevkiyat kaleminin) sectigi TAM varyasyonu
        belleğe alır; bundan sonra try_get_record(un) bu kaydı döndürür."""
        un = normalize_un(un)
        row = self.db.execute_one(
            "SELECT * FROM chemicals WHERE un_number=? AND classification_code=? "
            "AND packing_group=?",
            (un, classification_code or "", packing_group or ""))
        if not row:
            rows = self.get_variants(un)
            row = rows[0] if rows else None
        if row:
            rec = self._row_to_record(row)
            self._records_by_un[un] = rec
            return rec
        return None

    def register_unknown(self, un: str) -> None:
        """Veritabanında bulunamayan UN icin bos/etiketsiz kayit -- checker
        bunu UNKNOWN olarak isaretler, cokme olmaz."""
        un = normalize_un(un)
        self._records_by_un[un] = ProductRecord(un_no=un, name="", labels=[])

    # --- adr_mix_pro ProductDatabase arayuzu -----------------------------
    def try_get_record(self, un: str) -> Optional[ProductRecord]:
        return self._records_by_un.get(normalize_un(un))

    def all_records(self) -> List[ProductRecord]:
        return list(self._records_by_un.values())

    def search(self, query: str, limit: int = 200) -> List[ProductRecord]:
        q = query.strip()
        if not q:
            return []
        if q.isdigit():
            rows = self.db.execute(
                "SELECT * FROM chemicals WHERE un_number=? LIMIT ?",
                (q.zfill(4), limit))
        else:
            rows = self.db.execute(
                "SELECT * FROM chemicals WHERE proper_shipping_name_tr LIKE ? "
                "OR proper_shipping_name_en LIKE ? LIMIT ?",
                (f"%{q}%", f"%{q}%", limit))
        return [self._row_to_record(r) for r in (rows or [])]


def _gercek_karisik_yukleme_kontrolu(db, items) -> list:
    """DÜZELTME (web portundaki denetimde bulunan hata — masaüstünde de
    aynı kusur vardı, satırı satırına taşındığı için): generate_adr_report
    İÇİNDE eskiden basitleştirilmiş, GERÇEK bir ADR referansı olmayan
    ADRTransportPro.check_compatibility() çağrılıyor, sonucu
    report.compatibility_errors'a yazılıyordu. Bu, hem canlı panelin
    metin önizlemesinde (_update_preview) hem YAZDIRILAN belgede
    (_build_print_html) gösteriliyordu — GERÇEK motor (AnaDbChemicalAdapter
    + MixChecker, zaten bu dosyada mevcut, MixLoadCheckPage'in kullandığı)
    hiç devreye girmiyordu.

    Bu fonksiyon, her iki gösterim noktasında da report.compatibility_errors
    ATANMADAN HEMEN ÖNCE çağrılır; MixLoadCheckPage'in kendi self.adapter/
    self.checker'ı FARKLI bir sınıfa ait olduğu için yeniden kullanılamaz —
    bu yüzden burada TAZE bir örnek kurulur (ucuz bir işlem, sorun değil).
    Herhangi bir hata durumunda (kural dosyası bulunamazsa vb.) BOŞ liste
    döner — evrak/panel üretimini asla çökertmez."""
    try:
        rule_engine = SegregationRuleEngine(_MIX_RULE_FILE)
        adapter = AnaDbChemicalAdapter(db)
        checker = MixChecker(adapter, rule_engine)
        if len(items) < 2:
            return []
        for it in items:
            adapter.register_variant(it.un_number, it.classification_code,
                                     it.packing_group)
        sonuclar, _eksikler = checker.check_all([it.un_number for it in items])
        return [
            f"UN{r.un1} ({r.name1}) + UN{r.un2} ({r.name2}): {r.reason} "
            f"[ADR {r.adr_reference}]"
            for r in sonuclar if r.status not in ("OK",)
        ]
    except Exception:
        return []


class VariantPickerDialog(QDialog):
    """Bir UN numarasinin birden fazla ADR Tablo A varyasyonu oldugunda
    kullaniciya secim yaptiran diyalog (orn. UN1950 AEROSOL -> 12 secenek)."""

    def __init__(self, un: str, variants: list, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"UN {un} — Varyasyon Seçin")
        self.setMinimumWidth(560)
        self.selected_row = None

        layout = QVBoxLayout(self)
        info = QLabel(
            f"UN {un} için ADR Tablo A'da {len(variants)} farklı varyasyon "
            f"bulundu. Taşınacak maddeyle eşleşen satırı seçin:")
        info.setWordWrap(True)
        layout.addWidget(info)

        self.table = QTableWidget(len(variants), 6)
        self.table.setHorizontalHeaderLabels(
            ["Sınıflandırma Kodu", "PG", "Taşıma Adı", "Tünel", "LQ (7a)", "EQ (7b)"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        for i, row in enumerate(variants):
            name = row["proper_shipping_name_tr"] or row["proper_shipping_name_en"] or ""
            self.table.setItem(i, 0, QTableWidgetItem(row["classification_code"] or "—"))
            self.table.setItem(i, 1, QTableWidgetItem(row["packing_group"] or "—"))
            self.table.setItem(i, 2, QTableWidgetItem(name))
            self.table.setItem(i, 3, QTableWidgetItem(row["tunnel_code"] or "—"))
            self.table.setItem(i, 4, QTableWidgetItem(row["limited_quantity"] or "—"))
            self.table.setItem(i, 5, QTableWidgetItem(row["excepted_quantity"] or "—"))
        self.table.resizeColumnsToContents()
        self.table.selectRow(0)
        layout.addWidget(self.table)

        self._variants = variants
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self._on_accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _on_accept(self):
        idx = self.table.currentRow()
        if 0 <= idx < len(self._variants):
            self.selected_row = self._variants[idx]
        self.accept()


class MixLoadCheckPage(QWidget):
    """Karışık Yükleme Kontrolü sayfası — ADR 7.5.2 / 7.5.4 uyumluluk
    kontrolü, ana programın kendi veritabanı ve sevkiyat akışıyla bütünleşik."""

    def __init__(self, db: "DatabaseManager", parent=None):
        super().__init__(parent)
        self.db = db
        self.parent_window = parent
        self.adapter = AnaDbChemicalAdapter(db)
        try:
            self.rule_engine = SegregationRuleEngine(_MIX_RULE_FILE)
        except Exception:
            self.rule_engine = None
        self.checker = MixChecker(self.adapter, self.rule_engine) if self.rule_engine else None
        self._results = []
        self._un_order: List[str] = []  # gorunum sirasi (tekrarsiz)
        self._build_ui()

    def _build_ui(self):
        root = QHBoxLayout(self)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        root.addWidget(splitter)

        left = QWidget()
        lv = QVBoxLayout(left)
        title = QLabel("KARIŞIK YÜKLEME KONTROLÜ")
        title.setStyleSheet("font-size:13pt;font-weight:bold;color:#89B4FA;")
        lv.addWidget(title)
        info = QLabel(
            "Aynı araca yüklenecek UN numaralarını ekleyin.\n"
            "ADR 7.5.2 (ayrım) ve 7.5.4 (gıda ayrımı) uyumluluğu "
            "ikili olarak kontrol edilir.")
        info.setWordWrap(True)
        lv.addWidget(info)

        if self.checker is None:
            err = QLabel(
                "Kural dosyası yüklenemedi:\n"
                f"{_MIX_RULE_FILE}\n\n"
                "Karışık yükleme kontrolü kullanılamıyor.")
            err.setStyleSheet("color:#F38BA8;")
            err.setWordWrap(True)
            lv.addWidget(err)
            splitter.addWidget(left)
            root_right = QWidget()
            splitter.addWidget(root_right)
            return

        row = QHBoxLayout()
        self.un_input = QLineEdit()
        self.un_input.setPlaceholderText("UN No (örn. 1203)")
        self.un_input.returnPressed.connect(self._add_un)
        btn_add = QPushButton("Ekle")
        btn_add.clicked.connect(self._add_un)
        row.addWidget(self.un_input)
        row.addWidget(btn_add)
        lv.addLayout(row)

        self.un_list = QListWidget()
        lv.addWidget(self.un_list, 1)

        row2 = QHBoxLayout()
        btn_del = QPushButton("Seçileni Sil")
        btn_del.clicked.connect(self._remove_selected)
        btn_clear = QPushButton("Temizle")
        btn_clear.clicked.connect(self._clear_all)
        row2.addWidget(btn_del)
        row2.addWidget(btn_clear)
        lv.addLayout(row2)

        self.btn_from_shipment = QPushButton("Aktif Sevkiyattan Aktar")
        self.btn_from_shipment.clicked.connect(self._import_from_shipment)
        lv.addWidget(self.btn_from_shipment)

        self.btn_check = QPushButton("KONTROL ET")
        self.btn_check.setMinimumHeight(40)
        self.btn_check.setStyleSheet(
            "background-color:#89B4FA;color:#1E1E2E;font-weight:bold;border-radius:6px;")
        self.btn_check.clicked.connect(self._run_check)
        lv.addWidget(self.btn_check)

        exp_row = QHBoxLayout()
        self.btn_excel = QPushButton("Excel'e Aktar")
        self.btn_excel.clicked.connect(self._export_excel)
        self.btn_excel.setEnabled(False)
        self.btn_pdf = QPushButton("PDF Raporu")
        self.btn_pdf.clicked.connect(self._export_pdf)
        self.btn_pdf.setEnabled(False)
        exp_row.addWidget(self.btn_excel)
        exp_row.addWidget(self.btn_pdf)
        lv.addLayout(exp_row)

        splitter.addWidget(left)

        right = QWidget()
        rv = QVBoxLayout(right)
        self.summary_label = QLabel("Henüz kontrol çalıştırılmadı.")
        self.summary_label.setWordWrap(True)
        rv.addWidget(self.summary_label)

        self.results_model = ResultsTableModel()
        self.results_view = QTableView()
        self.results_view.setModel(self.results_model)
        self.results_view.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self.results_view.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        self.results_view.selectionModel().currentRowChanged.connect(
            self._on_row_changed)
        rv.addWidget(self.results_view, 2)

        self.detail_panel = ResultDetailPanel()
        rv.addWidget(self.detail_panel, 1)

        splitter.addWidget(right)
        splitter.setSizes([340, 760])

    # ------------------------------------------------------------------
    def _current_uns(self) -> List[str]:
        return list(self._un_order)

    def _add_un(self, raw_un: str = None):
        raw = raw_un if isinstance(raw_un, str) and raw_un else self.un_input.text()
        stripped = raw.strip()
        if stripped[:2].upper() == "UN":
            stripped = stripped[2:].strip()
        if not is_valid_un(stripped):
            QMessageBox.warning(self, "Geçersiz UN",
                f"Geçersiz UN numarası: {raw}\nUN numaraları 4 hanelidir (örn. 1203).")
            return
        un = normalize_un(stripped)
        if un in self._un_order:
            self.un_input.clear()
            return
        if len(self._un_order) >= MAX_UN_ITEMS:
            QMessageBox.warning(self, "Sınır", f"En fazla {MAX_UN_ITEMS} UN eklenebilir.")
            return

        variants = self.adapter.get_variants(un)
        if not variants:
            self.adapter.register_unknown(un)
            self.un_list.addItem(f"{un} (veritabanında yok)")
        elif len(variants) == 1:
            rec = self.adapter.register_variant(un)
            self.un_list.addItem(f"{un} - {rec.display_name}")
        else:
            dlg = VariantPickerDialog(un, variants, self)
            if dlg.exec() != QDialog.DialogCode.Accepted or dlg.selected_row is None:
                return
            row = dlg.selected_row
            rec = self.adapter.register_variant(
                un, row["classification_code"] or "", row["packing_group"] or "")
            code = row["classification_code"] or "-"
            self.un_list.addItem(f"{un} [{code}] - {rec.display_name}")

        self._un_order.append(un)
        self.un_input.clear()

    def _remove_selected(self):
        for item in self.un_list.selectedItems():
            row = self.un_list.row(item)
            self.un_list.takeItem(row)
            if row < len(self._un_order):
                del self._un_order[row]

    def _clear_all(self):
        self.un_list.clear()
        self._un_order.clear()

    def _import_from_shipment(self):
        main_win = self.parent_window
        shipment_page = getattr(main_win, "shipment_page", None)
        items = getattr(shipment_page, "items", None) if shipment_page else None
        if not items:
            QMessageBox.information(self, "Bilgi", "Aktif sevkiyatta kalem bulunmuyor.")
            return
        for it in items:
            un = normalize_un(it.un_number)
            if not un or un in self._un_order:
                continue
            if len(self._un_order) >= MAX_UN_ITEMS:
                break
            # Kalem zaten kendi varyasyonunu tasidigi icin belirsizlik yok
            rec = self.adapter.register_variant(
                un, getattr(it, "classification_code", "") or "",
                getattr(it, "packing_group", "") or "")
            if rec is None:
                self.adapter.register_unknown(un)
                self.un_list.addItem(f"{un} (veritabanında yok)")
            else:
                code = getattr(it, "classification_code", "") or "-"
                self.un_list.addItem(f"{un} [{code}] - {rec.display_name}")
            self._un_order.append(un)

    def _run_check(self):
        uns = self._current_uns()
        if len(uns) < 2:
            QMessageBox.warning(self, "Eksik Veri",
                "Kontrol için en az 2 UN numarası ekleyin.")
            return
        results, missing = self.checker.check_all(uns)
        self._results = results
        self.results_model.set_results(results)
        self.btn_excel.setEnabled(bool(results))
        self.btn_pdf.setEnabled(bool(results))

        forbidden = sum(1 for r in results if r.status == "NO")
        special = sum(1 for r in results if r.status not in ("OK", "NO"))
        parts = [f"{len(results)} ikili kontrol edildi."]
        if forbidden:
            parts.append(f"YASAK: {forbidden}")
        if special:
            parts.append(f"Özel durum / manuel kontrol: {special}")
        if missing:
            parts.append("Veritabanında bulunamayan: " + ", ".join(missing))
        self.summary_label.setText("  |  ".join(parts))

        if forbidden:
            QMessageBox.critical(self, "Karışık Yükleme YASAK",
                f"{forbidden} ikili için karışık yükleme yasağı tespit edildi!\n"
                "Detaylar için tablodaki satırları inceleyin.")

    def _on_row_changed(self, current, _previous):
        row = current.row()
        if 0 <= row < len(self._results):
            self.detail_panel.set_result(self._results[row], self.adapter)

    def _export_excel(self):
        if not self._results:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Excel'e Aktar", "karisik_yukleme_raporu.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            export_results_to_excel(self._results, path)
            QMessageBox.information(self, "Tamam", f"Rapor kaydedildi:\n{path}")
        except Exception as exc:
            logging.getLogger(__name__).exception("Beklenmeyen hata")
            QMessageBox.critical(self, "Hata", _turkce_hata_metni(exc))

    def _export_pdf(self):
        if not self._results:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "PDF Raporu", "karisik_yukleme_raporu.pdf", "PDF (*.pdf)")
        if not path:
            return
        try:
            export_results_to_pdf(self._results, path)
            QMessageBox.information(self, "Tamam", f"Rapor kaydedildi:\n{path}")
        except Exception as exc:
            logging.getLogger(__name__).exception("Beklenmeyen hata")
            QMessageBox.critical(self, "Hata", _turkce_hata_metni(exc))


# =============================================================================
# PYQT6 ANA UYGULAMA
# =============================================================================

class ADRTransportPro(QMainWindow):

    
    def __init__(self, security: SecurityManager = None, parent=None):
        super().__init__(parent)
        self.security = security
        self.db = DatabaseManager()
        self.current_shipment = None
        self.current_items = []
        self._confirm_close = False

        self.setWindowTitle(f"{APP_NAME} v{APP_VERSION}")
        # Ekran boyutunu algıla ve pencereyi ekrana sığdır
        screen = QApplication.primaryScreen().geometry()
        screen_width = screen.width()
        screen_height = screen.height()

        # Ekranın %85'ini kullan, ama minimum 1200x700 olsun
        target_width = max(1200, int(screen_width * 0.85))
        target_height = max(700, int(screen_height * 0.85))

        self.setMinimumSize(1200, 700)
        self.resize(target_width, target_height)

        # Ekran ortala
        self.move(
            int((screen_width - target_width) / 2),
            int((screen_height - target_height) / 2)
        )

        self._setup_theme()
        self._create_menu_bar()
        self._create_tool_bar()
        self._create_status_bar()
        self._create_central_widget()
        self._setup_shortcuts()
        self.show()

    def _setup_theme(self):
        self.setStyleSheet("""
            QMainWindow { background-color: #1E1E2E; }
            QWidget {
                background-color: #1E1E2E;
                color: #CDD6F4;
                font-family: 'Segoe UI', sans-serif;
                font-size: 10pt;
            }
            QPushButton {
                background-color: #45475A;
                color: #CDD6F4;
                border: 1px solid #6C7086;
                border-radius: 4px;
                padding: 6px 12px;
                min-width: 80px;
            }
            QPushButton:hover  { background-color: #585B70; }
            QPushButton:pressed { background-color: #313244; }
            QPushButton#primary { background-color: #89B4FA; color: #1E1E2E; font-weight: bold; }
            QPushButton#primary:hover { background-color: #B4BEFE; }
            QPushButton#danger  { background-color: #F38BA8; color: #1E1E2E; }
            QPushButton#success { background-color: #A6E3A1; color: #1E1E2E; }
            QPushButton#warning { background-color: #F9E2AF; color: #1E1E2E; }
            QLineEdit, QComboBox, QSpinBox, QDoubleSpinBox, QDateEdit {
                background-color: #313244;
                color: #CDD6F4;
                border: 1px solid #6C7086;
                border-radius: 4px;
                padding: 4px 8px;
            }
            QLineEdit:focus, QComboBox:focus { border: 1px solid #89B4FA; }
            QTableWidget {
                background-color: #313244;
                color: #CDD6F4;
                border: 1px solid #6C7086;
                gridline-color: #45475A;
                alternate-background-color: #272838;
            }
            QTableWidget::item:selected { background-color: #89B4FA; color: #1E1E2E; }
            QHeaderView::section {
                background-color: #181825;
                color: #CDD6F4;
                padding: 6px;
                border: 1px solid #45475A;
                font-weight: bold;
            }
            QTabWidget::pane { border: 1px solid #45475A; background-color: #1E1E2E; }
            QTabBar::tab {
                background-color: #313244;
                color: #CDD6F4;
                padding: 8px 16px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            QTabBar::tab:selected { background-color: #89B4FA; color: #1E1E2E; font-weight: bold; }
            QTabBar::tab:hover:!selected { background-color: #45475A; }
            QGroupBox {
                border: 1px solid #45475A;
                border-radius: 6px;
                margin-top: 12px;
                padding-top: 12px;
                font-weight: bold;
                color: #89B4FA;
            }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; }
            QMenuBar { background-color: #181825; color: #CDD6F4; }
            QMenuBar::item:selected { background-color: #89B4FA; color: #1E1E2E; }
            QMenu { background-color: #181825; color: #CDD6F4; border: 1px solid #45475A; }
            QMenu::item:selected { background-color: #89B4FA; color: #1E1E2E; }
            QStatusBar { background-color: #181825; color: #CDD6F4; }
            QToolBar { background-color: #181825; border: none; spacing: 4px; }
            QLabel#title    { font-size: 14pt; font-weight: bold; color: #89B4FA; }
            QLabel#subtitle { font-size: 11pt; color: #A6ADC8; }
            QLabel#error    { color: #F38BA8; font-weight: bold; }
            QLabel#warning  { color: #F9E2AF; }
            QLabel#success  { color: #A6E3A1; }
            QProgressBar {
                border: 1px solid #45475A;
                border-radius: 4px;
                text-align: center;
                background-color: #313244;
            }
            QProgressBar::chunk { background-color: #89B4FA; border-radius: 4px; }
            QSplitter::handle { background-color: #45475A; }
            QScrollArea { border: none; }
            QTextEdit, QPlainTextEdit {
                background-color: #313244;
                color: #CDD6F4;
                border: 1px solid #6C7086;
                border-radius: 4px;
            }
            QListWidget {
                background-color: #313244;
                color: #CDD6F4;
                border: 1px solid #6C7086;
                border-radius: 4px;
            }
            QListWidget::item:selected { background-color: #89B4FA; color: #1E1E2E; }
        """)

    def _create_menu_bar(self):
        menubar = self.menuBar()

        file_menu = menubar.addMenu("Dosya")

        new_action = QAction("Yeni Evrak", self)
        new_action.setShortcut("Ctrl+N")
        new_action.triggered.connect(self._new_shipment)
        file_menu.addAction(new_action)

        open_action = QAction("Evrak Ac...", self)
        open_action.setShortcut("Ctrl+O")
        open_action.triggered.connect(self._open_shipment)
        file_menu.addAction(open_action)

        save_action = QAction("Kaydet", self)
        save_action.setShortcut("Ctrl+S")
        save_action.triggered.connect(self._save_shipment)
        file_menu.addAction(save_action)

        file_menu.addSeparator()

        export_pdf_action = QAction("PDF Olustur", self)
        export_pdf_action.triggered.connect(self._export_pdf)
        file_menu.addAction(export_pdf_action)

        export_excel_action = QAction("Excel Olustur", self)
        export_excel_action.triggered.connect(self._export_excel)
        file_menu.addAction(export_excel_action)

        print_action = QAction("Yazdir...", self)
        print_action.setShortcut("Ctrl+P")
        print_action.triggered.connect(self._print_document)
        file_menu.addAction(print_action)

        file_menu.addSeparator()

        logout_action = QAction("Oturumu Kapat / Kullanici Degistir", self)
        logout_action.setShortcut("Ctrl+Shift+L")
        logout_action.triggered.connect(self._logout_and_relogin)
        file_menu.addAction(logout_action)

        file_menu.addSeparator()

        exit_action = QAction("Cikis", self)
        exit_action.setShortcut("Alt+F4")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        db_menu = menubar.addMenu("Veritabani")

        import_action = QAction("JSON Aktar", self)
        import_action.triggered.connect(self._import_json)
        db_menu.addAction(import_action)

        # CSV Import - ADR A Tablosu
        csv_import_action = QAction("CSV Aktar (ADR A Tablosu)", self)
        csv_import_action.triggered.connect(self._import_csv_adr_table)
        db_menu.addAction(csv_import_action)

        export_json_action = QAction("JSON Yedekle", self)
        export_json_action.triggered.connect(self._export_json_backup)
        db_menu.addAction(export_json_action)

        db_menu.addSeparator()

        backup_action = QAction("Yedekleri Gor", self)
        backup_action.triggered.connect(self._show_backups)
        db_menu.addAction(backup_action)

        help_menu = menubar.addMenu("Yardim")

        about_action = QAction("Hakkinda", self)
        about_action.triggered.connect(self._show_about)
        help_menu.addAction(about_action)

    def _create_tool_bar(self):
        toolbar = QToolBar("Ana Arac Cubugu")
        self.addToolBar(toolbar)

        toolbar.addAction("Yeni",    self._new_shipment)
        toolbar.addAction("Kaydet",  self._save_shipment)
        toolbar.addAction("PDF",     self._export_pdf)
        toolbar.addAction("Yazdir",  self._print_document)
        toolbar.addSeparator()
        toolbar.addAction("Dogrula", self._validate_shipment)
        toolbar.addAction("Onizle",  self._preview_shipment)
        toolbar.addSeparator()
        toolbar.addAction("Raporlar", self._show_reports)

    def _create_status_bar(self):
        self.statusbar = QStatusBar()
        self.setStatusBar(self.statusbar)
        self.statusbar.showMessage("ADR Transport Pro 2026 - Hazir")

        self.status_db_label = QLabel("SQLite")
        self.status_db_label.setStyleSheet("color: #A6E3A1;")
        self.statusbar.addPermanentWidget(self.status_db_label)

        self.status_version_label = QLabel(f"v{APP_VERSION}")
        self.statusbar.addPermanentWidget(self.status_version_label)

    def _create_central_widget(self):
        central = QWidget()
        self.setCentralWidget(central)

        layout = QHBoxLayout(central)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        main_splitter = QSplitter(Qt.Orientation.Horizontal)

        self.nav_panel = self._create_nav_panel()
        main_splitter.addWidget(self.nav_panel)

        self.content_stack = QStackedWidget()
        self._setup_content_pages()
        main_splitter.addWidget(self.content_stack)

        self.adr_panel = self._create_adr_control_panel()
        main_splitter.addWidget(self.adr_panel)

        main_splitter.setSizes([200, 800, 400])
        layout.addWidget(main_splitter)

    def _create_nav_panel(self) -> QWidget:
        panel = QWidget()
        panel.setMaximumWidth(220)
        panel.setMinimumWidth(200)
        panel.setStyleSheet(f"""
            QWidget {{
                background-color: #161B22;
                border-right: 1px solid #30363D;
            }}
        """)
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # ── Logo / başlık bandı ───────────────────────────────────────────
        header = QWidget()
        header.setFixedHeight(64)
        header.setStyleSheet("""
            QWidget {
                background-color: #0D1117;
                border-bottom: 2px solid #E6A817;
                border-right: none;
            }
        """)
        hlay = QVBoxLayout(header)
        hlay.setContentsMargins(14, 8, 14, 8)
        hlay.setSpacing(1)
        lbl_logo = QLabel("⚗  ADR TRANSPORT")
        lbl_logo.setStyleSheet(
            "color: #E6A817; font-size: 11pt; font-weight: 800; "
            "letter-spacing: 1px; border: none; background: transparent;")
        lbl_ver = QLabel("PRO  2026")
        lbl_ver.setStyleSheet(
            "color: #8B949E; font-size: 8pt; letter-spacing: 3px; "
            "border: none; background: transparent;")
        hlay.addWidget(lbl_logo)
        hlay.addWidget(lbl_ver)
        layout.addWidget(header)

        # ── Menü alanı ───────────────────────────────────────────────────
        menu_widget = QWidget()
        menu_widget.setStyleSheet("background: transparent; border: none;")
        menu_layout = QVBoxLayout(menu_widget)
        menu_layout.setContentsMargins(8, 10, 8, 10)
        menu_layout.setSpacing(2)

        # Aktif sayfalar ve ikonları
        menus = [
            ("Tasima Evraki",        0,  " ",  True),
            ("Sevkiyatlar",          1,  " ",  True),
            ("Firma Yonetimi",       2,  " ",  True),
            ("Surucu Yonetimi",      3,  " ",  True),
            ("Arac Yonetimi",        4,  " ",  True),
            ("ADR Veritabani",       5,  "🗄",  True),
            ("Kayitli Excel Verisi", 6,  "📊",  True),
            ("Raporlar",             7,  "📈",  True),  # yapım aşamasında
            ("Emniyet Planlari",     8,  " ",  True),  # yapım aşamasında
            ("Ayarlar",              9,  "⚙",   True),  # yapım aşamasında
            ("Güvenlik",            10,  "🔒",   True),
            ("Karışık Yükleme",     11,  "⚠",   True),
        ]

        # Buton grubu — yalnızca bir buton aynı anda checked olabilir
        self.nav_btn_group = QButtonGroup(panel)
        self.nav_btn_group.setExclusive(True)
        self.nav_buttons = {}

        for text, index, icon, active in menus:
            btn = QPushButton(f"  {icon}  {text}")
            btn.setMinimumHeight(38)
            btn.setCheckable(True)

            if active:
                btn.setStyleSheet("""
                    QPushButton {
                        text-align: left;
                        padding: 0 10px;
                        background-color: transparent;
                        border: none;
                        border-radius: 6px;
                        color: #8B949E;
                        font-size: 9pt;
                        font-weight: 500;
                    }
                    QPushButton:hover {
                        background-color: #21262D;
                        color: #CDD6F4;
                    }
                    QPushButton:checked {
                        background-color: #1F3358;
                        color: #89B4FA;
                        font-weight: 700;
                        border-left: 3px solid #89B4FA;
                        padding-left: 7px;
                    }
                """)
                btn.clicked.connect(lambda checked, idx=index: self._switch_page(idx))
            else:
                # Pasif (yapım aşamasında) — tıklanamaz, soluk görünür
                btn.setCheckable(False)
                btn.setEnabled(False)
                btn.setStyleSheet("""
                    QPushButton {
                        text-align: left;
                        padding: 0 10px;
                        background-color: transparent;
                        border: none;
                        border-radius: 6px;
                        color: #3D444D;
                        font-size: 9pt;
                        font-weight: 400;
                    }
                """)
                # "yapım aşamasında" mini etiketi
                btn.setText(f"  {icon}  {text}  ·")

            self.nav_btn_group.addButton(btn, index)
            self.nav_buttons[index] = btn
            menu_layout.addWidget(btn)

        menu_layout.addStretch()
        layout.addWidget(menu_widget, 1)

        # ── Versiyon / alt bilgi ─────────────────────────────────────────
        footer = QWidget()
        footer.setFixedHeight(48)
        footer.setStyleSheet("""
            QWidget {
                background-color: #0D1117;
                border-top: 1px solid #30363D;
                border-right: none;
            }
        """)
        flay = QVBoxLayout(footer)
        flay.setContentsMargins(14, 8, 14, 8)
        lbl_footer = QLabel("ADR 2025 · IMDG · RID")
        lbl_footer.setStyleSheet(
            "color: #3D444D; font-size: 7pt; letter-spacing: 1px; "
            "border: none; background: transparent;")
        lbl_footer.setAlignment(Qt.AlignmentFlag.AlignCenter)
        flay.addWidget(lbl_footer)
        layout.addWidget(footer)

        return panel

    def _setup_content_pages(self):
        # Sayfa 0: Tasima Evraki
        self.shipment_page = ShipmentEditorPage(self.db, self)
        self.content_stack.addWidget(self.shipment_page)

        # Sayfa 1: Sevkiyat arsivi
        self.archive_page = ShipmentArchivePage(self.db, self)
        self.content_stack.addWidget(self.archive_page)

        # Sayfa 2: Firma Yönetimi (GERCEK SAYFA)
        self.company_page = CompanyManagementPage(self.db, self)
        self.content_stack.addWidget(self.company_page)

        # Sayfa 3: Sürücü Yönetimi (GERCEK SAYFA)
        self.driver_page = DriverManagementPage(self.db, self)
        self.content_stack.addWidget(self.driver_page)

        # Sayfa 4: Araç Yönetimi (GERCEK SAYFA)
        self.vehicle_page = VehicleManagementPage(self.db, self)
        self.content_stack.addWidget(self.vehicle_page)

        # Sayfa 5: ADR Veritabani (GERCEK SAYFA)
        self.adr_db_page = ADRDatabasePage(self.db, self)
        self.content_stack.addWidget(self.adr_db_page)

        # Sayfa 6: Kayıtlı Excel Verisi (YENİ)
        self.excel_page = ExcelChemicalPage(self.db, self)
        self.excel_page.product_add_requested.connect(self._on_excel_product_add)
        self.excel_page.excel_loaded.connect(self._on_excel_loaded)
        self.content_stack.addWidget(self.excel_page)
        
        # Sayfa 7: Raporlar (GERCEK SAYFA)
        self.reports_page = ReportsPage(self.db, self)
        self.content_stack.addWidget(self.reports_page)

        # Sayfa 8: Emniyet Planları (GERCEK SAYFA)
        self.safety_page = SafetyPlansPage(self.db, self)
        self.safety_page.shipment_page_ref = self.shipment_page  # doğrudan referans
        self.content_stack.addWidget(self.safety_page)

        # Sayfa 9: Ayarlar (GERCEK SAYFA)
        self.settings_page = SettingsPage(self.db, self)
        self.content_stack.addWidget(self.settings_page)

        # ── 10. SAYFA: GÜVENLİK MODÜLÜ PANELİ (YENİ EKLENEN KISIM) ──────────────────
        lic_info = LicenseManager.validate() 
        self.security_page = SecurityAdminPage(self.security, lic_info, self)
        self.content_stack.addWidget(self.security_page)
        # ──────────────────────────────────────────────────────────────────────────

        # Sayfa 11: Karışık Yükleme Kontrolü (main.py'den taşındı, v4.4)
        self.mix_load_page = MixLoadCheckPage(self.db, self)
        self.content_stack.addWidget(self.mix_load_page)

    def _create_adr_control_panel(self) -> QWidget:
        C = {
            "bg":      "#161B22",
            "card":    "#1C2128",
            "border":  "#30363D",
            "accent":  "#E6A817",
            "ok":      "#3FB950",
            "warn":    "#F9E2AF",
            "err":     "#F38BA8",
            "text_hi": "#F0F6FC",
            "text_lo": "#8B949E",
            "blue":    "#89B4FA",
        }

        panel = QWidget()
        panel.setMaximumWidth(390)
        panel.setMinimumWidth(320)
        panel.setStyleSheet(f"""
            QWidget {{
                background-color: {C['bg']};
                color: {C['text_hi']};
                font-family: 'Segoe UI', 'Inter', sans-serif;
                font-size: 9pt;
            }}
            QGroupBox {{
                background-color: {C['card']};
                border: 1px solid {C['border']};
                border-radius: 6px;
                margin-top: 8px;
                padding-top: 6px;
                color: {C['text_lo']};
                font-size: 7pt;
                font-weight: 700;
                letter-spacing: 1px;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 10px;
                padding: 0 5px;
                background: {C['card']};
            }}
            QProgressBar {{
                background-color: {C['border']};
                border: none;
                border-radius: 4px;
                height: 8px;
                text-align: center;
                font-size: 0pt;
            }}
            QProgressBar::chunk {{
                border-radius: 4px;
                background-color: {C['ok']};
            }}
            QListWidget {{
                background-color: {C['bg']};
                border: 1px solid {C['border']};
                border-radius: 4px;
                outline: none;
                font-size: 8.5pt;
            }}
            QListWidget::item {{
                padding: 3px 6px;
                border-bottom: 1px solid #21262D;
            }}
            QPlainTextEdit {{
                background-color: #0D1117;
                border: 1px solid {C['border']};
                border-radius: 4px;
                color: {C['text_lo']};
                font-family: 'Consolas','Courier New',monospace;
                font-size: 8pt;
            }}
            QScrollBar:vertical {{
                background: {C['card']};
                width: 6px;
                border-radius: 3px;
            }}
            QScrollBar::handle:vertical {{
                background: {C['border']};
                border-radius: 3px;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0;
            }}
        """)

        main_scroll = QScrollArea()
        main_scroll.setWidgetResizable(True)
        main_scroll.setFrameShape(QFrame.Shape.NoFrame)
        main_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        main_scroll.setStyleSheet("background: transparent; border: none;")

        scroll_widget = QWidget()
        scroll_widget.setStyleSheet("background: transparent;")
        layout = QVBoxLayout(scroll_widget)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)

        # ── Başlık bandı ─────────────────────────────────────────────────
        header = QFrame()
        header.setStyleSheet(f"""
            QFrame {{
                background-color: #0D1117;
                border: 1px solid {C['border']};
                border-left: 4px solid {C['accent']};
                border-radius: 5px;
            }}
        """)
        hlay = QHBoxLayout(header)
        hlay.setContentsMargins(10, 8, 10, 8)
        lbl_icon = QLabel("🛡")
        lbl_icon.setStyleSheet("font-size: 16pt; border: none; background: transparent;")
        hlay.addWidget(lbl_icon)
        title_col = QVBoxLayout()
        title_col.setSpacing(0)
        lbl_t = QLabel("ADR KONTROL MERKEZİ")
        lbl_t.setStyleSheet(f"font-size: 10pt; font-weight: 800; color: {C['text_hi']}; "
                            "letter-spacing: 1px; border: none; background: transparent;")
        lbl_s = QLabel("Canlı Uyumluluk Analizi")
        lbl_s.setStyleSheet(f"font-size: 8pt; color: {C['text_lo']}; "
                            "border: none; background: transparent;")
        title_col.addWidget(lbl_t)
        title_col.addWidget(lbl_s)
        hlay.addLayout(title_col, 1)
        layout.addWidget(header)

        # ── 1.1.3.6 Puan Hesabı ─────────────────────────────────────────
        pts_group = QGroupBox("1.1.3.6 · MİKTAR MUAFİYETİ")
        pts_lay = QVBoxLayout(pts_group)
        pts_lay.setSpacing(6)

        # Puan + bar
        puan_row = QHBoxLayout()
        self.lbl_total_points = QLabel("0")
        self.lbl_total_points.setStyleSheet(
            f"font-size: 22pt; font-weight: 800; color: {C['ok']}; "
            "border: none; background: transparent;")
        puan_row.addWidget(self.lbl_total_points)
        puan_suffix = QVBoxLayout()
        puan_suffix.setSpacing(0)
        lbl_of = QLabel(f"/ {MAX_1136_POINTS}")
        lbl_of.setStyleSheet(f"font-size: 10pt; color: {C['text_lo']}; border: none; background: transparent;")
        lbl_puan = QLabel("PUAN")
        lbl_puan.setStyleSheet(f"font-size: 7pt; letter-spacing: 2px; color: {C['text_lo']}; "
                               "border: none; background: transparent;")
        puan_suffix.addWidget(lbl_of)
        puan_suffix.addWidget(lbl_puan)
        puan_row.addLayout(puan_suffix)
        puan_row.addStretch()
        pts_lay.addLayout(puan_row)

        self.pbar_points = QProgressBar()
        self.pbar_points.setMinimum(0)
        self.pbar_points.setMaximum(100)
        self.pbar_points.setValue(0)
        self.pbar_points.setFixedHeight(8)
        pts_lay.addWidget(self.pbar_points)

        # Puan bar yüzde etiketi
        self.lbl_pbar_pct = QLabel("0 / 1000 puan  (0%)")
        self.lbl_pbar_pct.setStyleSheet(f"color: {C['text_lo']}; font-size: 8pt; "
                                        "border: none; background: transparent;")
        pts_lay.addWidget(self.lbl_pbar_pct)

        layout.addWidget(pts_group)

        # ── Durum rozetleri (3'lü grid) ───────────────────────────────────
        status_group = QGroupBox("DURUM GÖSTERGELERİ")
        sg_lay = QVBoxLayout(status_group)
        sg_lay.setSpacing(5)

        self.lbl_orange_plate = QLabel("✓  Turuncu Plaka: Gerekmez")
        self.lbl_orange_plate.setStyleSheet(
            f"font-size: 9.5pt; color: {C['ok']}; border: none; background: transparent;")
        self.lbl_orange_plate.setVisible(False)

        self.lbl_tunnel_badge = QLabel("Tünel: E  (Serbest)")
        self.lbl_tunnel_badge.setStyleSheet(
            f"font-size: 9pt; font-weight: 700; color: {C['ok']}; "
            f"background: #0A2800; border-radius: 4px; padding: 3px 8px; border: none;")
        self.lbl_tunnel_badge.setVisible(False)

        self.lbl_written_instructions = QLabel("✓  Yazılı Talimat: Gerekmez")
        self.lbl_written_instructions.setStyleSheet(
            f"color: {C['text_lo']}; border: none; background: transparent;")
        self.lbl_written_instructions.setVisible(False)

        self.lbl_driver_adr = QLabel("✓  ADR Sertifika: Gerekmez")
        self.lbl_driver_adr.setStyleSheet(
            f"color: {C['text_lo']}; border: none; background: transparent;")
        self.lbl_driver_adr.setVisible(False)

        self.lbl_exemption = QLabel("Muafiyet: Yok")
        self.lbl_exemption.setStyleSheet(
            f"color: {C['text_lo']}; font-size: 8pt; border: none; background: transparent;")
        self.lbl_exemption.setVisible(False)

        for w in [self.lbl_orange_plate, self.lbl_tunnel_badge,
                  self.lbl_written_instructions, self.lbl_driver_adr,
                  self.lbl_exemption]:
            sg_lay.addWidget(w)

        layout.addWidget(status_group)

        # ── Sürücü Sertifika Geri Sayım ──────────────────────────────────
        cert_group = QGroupBox("SÜRÜCÜ SERTİFİKA DURUMU")
        cert_lay = QVBoxLayout(cert_group)
        self.lbl_cert_countdown = QLabel(
            f"<span style='color:{C['text_lo']}'>Sürücü seçilmedi</span>")
        self.lbl_cert_countdown.setTextFormat(Qt.TextFormat.RichText)
        self.lbl_cert_countdown.setWordWrap(True)
        self.lbl_cert_countdown.setStyleSheet("border: none; background: transparent;")
        cert_lay.addWidget(self.lbl_cert_countdown)
        layout.addWidget(cert_group)

        # ── Uyarı ve Hatalar ─────────────────────────────────────────────
        warn_group = QGroupBox("UYARI VE HATALAR")
        warn_lay = QVBoxLayout(warn_group)
        self.list_warnings = QListWidget()
        self.list_warnings.setMinimumHeight(120)
        self.list_warnings.setMaximumHeight(220)
        warn_lay.addWidget(self.list_warnings)
        layout.addWidget(warn_group)

        # ── Canlı Evrak Önizleme ─────────────────────────────────────────
        preview_group = QGroupBox("CANLI EVRAK ÖNİZLEME")
        prev_lay = QVBoxLayout(preview_group)
        self.preview_text = QPlainTextEdit()
        self.preview_text.setReadOnly(True)
        self.preview_text.setMinimumHeight(180)
        self.preview_text.setMaximumHeight(320)
        prev_lay.addWidget(self.preview_text)
        layout.addWidget(preview_group)

        # ── Doğrula butonu ───────────────────────────────────────────────
        validate_btn = QPushButton("⚡  EVRAKI DOĞRULA  (Ctrl+Shift+V)")
        validate_btn.setMinimumHeight(40)
        validate_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {C['blue']};
                color: #0D1117;
                border: none;
                border-radius: 5px;
                font-size: 10pt;
                font-weight: 800;
                letter-spacing: 1px;
            }}
            QPushButton:hover {{ background-color: #A5C8FF; }}
            QPushButton:pressed {{ background-color: #6E9FE0; }}
        """)
        validate_btn.clicked.connect(self._validate_shipment)
        layout.addWidget(validate_btn)

        layout.addStretch()
        main_scroll.setWidget(scroll_widget)

        outer = QVBoxLayout(panel)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(main_scroll)

        return panel

    def _setup_shortcuts(self):

        QShortcut(QKeySequence("Ctrl+N"),       self, self._new_shipment)
        QShortcut(QKeySequence("Ctrl+S"),       self, self._save_shipment)
        QShortcut(QKeySequence("Ctrl+O"),       self, self._open_shipment)
        QShortcut(QKeySequence("Ctrl+P"),       self, self._print_document)
        QShortcut(QKeySequence("Ctrl+Shift+V"), self, self._validate_shipment)
        QShortcut(QKeySequence("F5"),           self, self._update_adr_panel)

    def _switch_page(self, index: int):
        self.content_stack.setCurrentIndex(index)
        if hasattr(self, 'nav_buttons') and index in self.nav_buttons:
            self.nav_buttons[index].setChecked(True)
        # Yönetim sayfalarını her açılışta yenile
        if index == 1 and hasattr(self, 'archive_page'):
            self.archive_page._load_shipments()
        elif index == 2 and hasattr(self, 'company_page'):
            self.company_page._load()
        elif index == 3 and hasattr(self, 'driver_page'):
            self.driver_page._load()
        elif index == 4 and hasattr(self, 'vehicle_page'):
            self.vehicle_page._load()
        elif index == 7 and hasattr(self, 'reports_page'):
            # Raporlar sayfasını yenile
            try:
                self.reports_page._refresh()
            except Exception:
                pass
        elif index == 9 and hasattr(self, 'settings_page'):
            # Ayarlar sayfasını yenile
            try:
                self.settings_page._load()
            except Exception:
                pass

    def showEvent(self, event):
        super().showEvent(event)
        # İlk açılışta "Tasima Evraki" butonunu aktif yap
        if hasattr(self, 'nav_buttons') and 0 in self.nav_buttons:
            self.nav_buttons[0].setChecked(True)

    def _on_excel_loaded(self, products: list):
        """Excel yüklendiğinde ShipmentEditorPage'in arama listesini güncelle."""
        if hasattr(self, 'shipment_page'):
            self.shipment_page.excel_products = products

    def _on_excel_product_add(self, product: dict):
        """
        ExcelChemicalPage'den gelen ürünü Taşıma Evrakı sayfasına ekler.
        Veriler: kimyasal_adi, un_numarasi, sevkiyat_adi, sinif, paketleme_grubu, tunel_kodu
        """
        # ShipmentEditorPage'in _add_item_from_excel metodunu çağır
        if hasattr(self, 'shipment_page'):
            self.shipment_page.add_item_from_excel(product)
            # Taşıma Evrakı sayfasına geç
            self.content_stack.setCurrentIndex(0)
            self._update_adr_panel()
            QMessageBox.information(
                self, "Eklendi",
                f"'{product.get('kimyasal_adi', '')}' taşıma evrakına eklendi."
            )

    def _update_statistics(self):
        stats = self.db.get_statistics()
        # İstatistik etiketleri ADR paneline taşındı, güvenli güncelle
        for attr, key, prefix in [
            ("stat_shipments", "total_shipments", "Sevkiyat"),
            ("stat_companies", "total_companies", "Firma"),
            ("stat_drivers",   "active_drivers",  "Surucu"),
            ("stat_vehicles",  "active_vehicles",  "Arac"),
            ("stat_chemicals", "total_chemicals",  "Kimyasal"),
        ]:
            if hasattr(self, attr):
                getattr(self, attr).setText(f"{prefix}: {stats.get(key, 0)}")

    def _update_adr_panel(self):
        # shipment_page henüz oluşturulmadıysa çık (init sırası koruması)
        if not hasattr(self, 'shipment_page') or self.shipment_page is None:
            return
        # Tablo hücrelerindeki güncel değerleri self.items'a yansıt
        self.shipment_page._sync_items_from_table()
        items   = self.shipment_page.get_items()
        driver  = self.shipment_page.get_selected_driver()
        vehicle = self.shipment_page.get_selected_vehicle()

        # Ürün eklenmemişse veya tüm net miktarlar sıfırsa durum etiketlerini gizle
        has_items    = len(items) > 0
        has_quantity = has_items and any(float(i.net_quantity or 0) > 0 for i in items)

        status_widgets = [self.lbl_orange_plate, self.lbl_written_instructions,
                          self.lbl_driver_adr, self.lbl_exemption, self.lbl_tunnel_badge]
        for w in status_widgets:
            w.setVisible(has_quantity)

        # Bilgi etiketi — yoksa oluştur
        if not hasattr(self, '_lbl_no_items'):
            self._lbl_no_items = QLabel()
            self._lbl_no_items.setStyleSheet(
                "color:#6C7086; font-size:9pt; border:none; background:transparent;")
            self._lbl_no_items.setAlignment(Qt.AlignmentFlag.AlignCenter)
            # status_group'un layout'una ekle (lbl_orange_plate ile aynı parent)
            self.lbl_orange_plate.parent().layout().addWidget(self._lbl_no_items)

        if not has_items:
            self._lbl_no_items.setText("— Ürün eklenmedi —")
            self._lbl_no_items.setVisible(True)
        elif not has_quantity:
            self._lbl_no_items.setText("— Net miktar girilmedi —")
            self._lbl_no_items.setVisible(True)
        else:
            self._lbl_no_items.setVisible(False)

        # [DÜZELTİLDİ] Uyarı listesi her zaman temizlenir — erken return'den ÖNCE.
        # Önceki kodda has_quantity=False durumunda return yapılıyor,
        # list_warnings.clear() hiç çağrılmıyor ve eski uyarılar ekranda kalıyordu.
        self.list_warnings.clear()

        if not has_quantity:
            return

        report = ADREngine.generate_adr_report(items, driver, vehicle)

        # --- Puan etiketi + progress bar ---
        # [DÜZELTİLDİ] Puan 0 ise miktar muafiyeti bölümünü gösterme
        # (TC boş, TC=4, LQ veya EQ gibi durumlarda puan 0 olur)
        pts_widget_visible = report.total_points > 0
        self.lbl_total_points.setVisible(pts_widget_visible)
        self.pbar_points.setVisible(pts_widget_visible)
        self.lbl_pbar_pct.setVisible(pts_widget_visible)

        if report.total_points > 0:
            pct = min(report.total_points / MAX_1136_POINTS, 1.0)
            self.lbl_total_points.setText(f"{report.total_points:.0f}")

            if report.total_points > MAX_1136_POINTS:
                pts_color = "#F38BA8"
                bar_style = "QProgressBar::chunk { background-color: #F38BA8; border-radius: 4px; }"
            elif report.total_points > MAX_1136_POINTS * 0.8:
                pts_color = "#F9E2AF"
                bar_style = "QProgressBar::chunk { background-color: #F9E2AF; border-radius: 4px; }"
            else:
                pts_color = "#3FB950"
                bar_style = "QProgressBar::chunk { background-color: #3FB950; border-radius: 4px; }"

            self.lbl_total_points.setStyleSheet(
                f"font-size: 22pt; font-weight: 800; color: {pts_color}; "
                "border: none; background: transparent;")
            self.pbar_points.setValue(int(pct * 100))
            self.pbar_points.setStyleSheet(bar_style)
            self.lbl_pbar_pct.setText(
                f"{report.total_points:.0f} / {MAX_1136_POINTS} puan  ({pct*100:.0f}%)")

        # --- Durum göstergeleri görünürlük kuralı ---
        # TC=4, LQ, EQ → puan=0 ama sınırsız taşıma geçerli → göstergeler gösterilir
        # Sadece TC boş kalemler varsa → veri eksik → göstergeler gizlenir
        # 0 < puan <= 1000 → göstergeler gösterilir
        # puan > 1000 → göstergeler gösterilir (turuncu plaka zorunlu zaten)

        has_tc4_lq_eq = any(
            i.is_lq or i.is_eq or
            str(getattr(i, "transport_category", "")).strip().split()[0:1] == ["4"]
            for i in items if float(i.net_quantity or 0) > 0
        )
        has_empty_tc = any(
            not i.is_lq and not i.is_eq and
            not str(getattr(i, "transport_category", "")).strip()
            for i in items if float(i.net_quantity or 0) > 0
        )

        # Gösterge göster: puan > 0 VEYA TC=4/LQ/EQ varken TC boş kalem yoksa
        status_valid = (
            report.total_points > 0 or
            (has_tc4_lq_eq and not has_empty_tc)
        )

        self.lbl_orange_plate.setVisible(status_valid)
        self.lbl_written_instructions.setVisible(status_valid)
        self.lbl_driver_adr.setVisible(status_valid)
        self.lbl_exemption.setVisible(status_valid)
        if hasattr(self, "lbl_tunnel_badge"):
            self.lbl_tunnel_badge.setVisible(status_valid)

        if not status_valid:
            if not hasattr(self, '_lbl_status_info'):
                self._lbl_status_info = QLabel()
                self._lbl_status_info.setStyleSheet(
                    "color:#6C7086; font-size:9pt; border:none; background:transparent;")
                self._lbl_status_info.setAlignment(Qt.AlignmentFlag.AlignCenter)
                self.lbl_orange_plate.parent().layout().addWidget(self._lbl_status_info)
            self._lbl_status_info.setText(
                "— Taşıma Kategorisi girilmemiş —\n"
                "Durum göstergesi hesaplanamaz")
            self._lbl_status_info.setVisible(True)
        else:
            if hasattr(self, '_lbl_status_info'):
                self._lbl_status_info.setVisible(False)

            # --- Turuncu plaka ---
            if report.orange_plate_required:
                self.lbl_orange_plate.setText("⚠  Turuncu Plaka: ZORUNLU!")
                self.lbl_orange_plate.setStyleSheet(
                    "font-size:11pt; font-weight:bold; color:#F38BA8;"
                    "background:#3A0010; border-radius:4px; padding:3px 6px;")
            else:
                self.lbl_orange_plate.setText("✓  Turuncu Plaka: Gerekmez")
                self.lbl_orange_plate.setStyleSheet(
                    "font-size:11pt; color:#A6E3A1;"
                    "background:#0A2800; border-radius:4px; padding:3px 6px;")

            # --- Yazılı talimat ---
            if report.written_instructions_required:
                self.lbl_written_instructions.setText("⚠  Yazılı Talimat: ZORUNLU")
                self.lbl_written_instructions.setStyleSheet(
                    "color:#F9E2AF; font-weight:bold;")
            else:
                self.lbl_written_instructions.setText("✓  Yazılı Talimat: Gerekmez")
                self.lbl_written_instructions.setStyleSheet("color:#A6ADC8;")

            # --- ADR sertifika ---
            if report.driver_adr_required:
                self.lbl_driver_adr.setText("⚠  ADR Sertifika: ZORUNLU")
                self.lbl_driver_adr.setStyleSheet(
                    "color:#F9E2AF; font-weight:bold;")
            else:
                self.lbl_driver_adr.setText("✓  ADR Sertifika: Gerekmez")
                self.lbl_driver_adr.setStyleSheet("color:#A6ADC8;")

            # --- Tünel kodu rozeti ---
            TUNNEL_STYLE = {
                "A": ("Tünel: A  (GEÇİŞ YASAK)", "#F38BA8", "#3A0010"),
                "B": ("Tünel: B  (Kısıtlı)",      "#F9E2AF", "#3A2800"),
                "C": ("Tünel: C  (Kısıtlı)",      "#F9E2AF", "#3A2800"),
                "D": ("Tünel: D  (Dikkatli)",      "#EF9F27", "#3A2800"),
                "E": ("Tünel: E  (Serbest)",       "#A6E3A1", "#0A2800"),
            }
            tc_tunnel = report.tunnel_code or "E"
            if hasattr(self, "lbl_tunnel_badge"):
                txt, bg, fg = TUNNEL_STYLE.get(tc_tunnel, (f"Tünel: {tc_tunnel}", "#45475A", "#CDD6F4"))
                self.lbl_tunnel_badge.setText(txt)
                self.lbl_tunnel_badge.setStyleSheet(
                    f"font-weight:bold; font-size:10pt;"
                    f"background:{bg}; color:{fg};"
                    f"border-radius:4px; padding:3px 8px;")

            # --- Muafiyet ---
            self.lbl_exemption.setText(f"Muafiyet: {report.exemption_type}")

        # --- Sürücü sertifika son tarih sayacı ---
        if hasattr(self, "lbl_cert_countdown") and driver:
            warnings_cert = []
            for cert_label, expiry_str in [
                ("ADR Sertifikası", driver.adr_certificate_expiry),
                ("SRC5",            driver.src5_expiry),
            ]:
                if expiry_str:
                    try:
                        exp = datetime.strptime(expiry_str, "%Y-%m-%d")
                        days = (exp - datetime.now()).days
                        if days < 0:
                            warnings_cert.append(
                                f"<span style='color:#F38BA8'>✗ {cert_label}: "
                                f"SÜRESİ DOLDU ({expiry_str})</span>")
                        elif days <= 30:
                            warnings_cert.append(
                                f"<span style='color:#F9E2AF'>⚠ {cert_label}: "
                                f"{days} gün kaldı ({expiry_str})</span>")
                        else:
                            warnings_cert.append(
                                f"<span style='color:#A6E3A1'>✓ {cert_label}: "
                                f"{days} gün ({expiry_str})</span>")
                    except Exception:
                        pass
            self.lbl_cert_countdown.setText("<br>".join(warnings_cert) if warnings_cert
                                             else "<span style='color:#A6ADC8'>Sürücü sertifika bilgisi yok</span>")
        elif hasattr(self, "lbl_cert_countdown"):
            self.lbl_cert_countdown.setText(
                "<span style='color:#A6ADC8'>Sürücü seçilmedi</span>")

        # --- Uyarı listesi --- (üstte erken return'den önce zaten temizlendi)
        for _level, msg in report.errors:
            item = QListWidgetItem(f"✗  {msg}")
            item.setForeground(QColor("#F38BA8"))
            fnt = item.font(); fnt.setBold(True); item.setFont(fnt)
            self.list_warnings.addItem(item)
        for _level, msg in report.warnings:
            item = QListWidgetItem(f"⚠  {msg}")
            item.setForeground(QColor("#F9E2AF"))
            self.list_warnings.addItem(item)
        for _level, msg in report.info:
            item = QListWidgetItem(f"ℹ  {msg}")
            item.setForeground(QColor("#A6ADC8"))
            self.list_warnings.addItem(item)

        self._update_preview(items, report)

    def _update_preview(self, items: List[ShipmentItem], report: ADRReport):
        preview = []
        preview.append("=" * 50)
        preview.append("ADR TEHLIKELI MADDE TASIMA EVRAKI")
        preview.append("=" * 50)
        preview.append("")

        doc_info = self.shipment_page.get_document_info()
        preview.append(f"Evrak No: {doc_info.get('document_no', '---')}")
        preview.append(f"Tarih: {doc_info.get('date', '---')}")
        preview.append("")

        sender = self.shipment_page.get_selected_sender()
        if sender:
            preview.append(f"GONDERICI: {sender.name}")
            preview.append(f"Adres: {sender.address}, {sender.city}")
            preview.append("")

        receiver = self.shipment_page.get_selected_receiver()
        if receiver:
            preview.append(f"ALICI: {receiver.name}")
            preview.append(f"Adres: {receiver.address}, {receiver.city}")
            preview.append("")

        preview.append("-" * 50)
        preview.append("TASINAN URUNLER:")
        preview.append("-" * 50)

        for i, item in enumerate(items, 1):
            preview.append(f"{i}. UN{item.un_number} {item.proper_name}")
            preview.append(f"   Sinif: {item.class_code} | PG: {item.packing_group} | Tunel: {item.tunnel_code or '-'}")
            preview.append(f"   Miktar: {item.net_quantity} {item.unit}")
            preview.append(f"   Ambalaj: {item.packaging_type} x {item.packaging_count}")
            if item.is_lq:
                preview.append("   [LQ - Limited Quantity]")
            if item.is_eq:
                preview.append("   [EQ - Excepted Quantity]")
            preview.append("")

        preview.append("-" * 50)
        preview.append("ADR KONTROL OZETI:")
        preview.append("-" * 50)
        preview.append(f"Toplam Puan: {report.total_points:.0f} / {MAX_1136_POINTS}")
        preview.append(f"Turuncu Plaka: {'EVET' if report.orange_plate_required else 'HAYIR'}")
        preview.append(f"Muafiyet: {report.exemption_type}")
        preview.append(f"Yazili Talimat: {'ZORUNLU' if report.written_instructions_required else 'Gerekmez'}")
        preview.append("")

        # DÜZELTME: GERÇEK motor sonucu burada (self.db erişimi olan bu
        # noktada) hesaplanıp report.compatibility_errors'a yerleştiriliyor.
        report.compatibility_errors = _gercek_karisik_yukleme_kontrolu(self.db, items)

        if report.compatibility_errors:
            preview.append("UYUMSUZLUKLAR:")
            for err in report.compatibility_errors:
                preview.append(f"  ! {err}")
            preview.append("")

        preview.append("=" * 50)

        self.preview_text.setPlainText("\n".join(preview))

    def _new_shipment(self):
        self.shipment_page.clear_all()
        self.current_items = []
        self._update_adr_panel()
        self.statusbar.showMessage("Yeni evrak olusturuldu", 3000)

    def _save_shipment(self):
        self.shipment_page.save_shipment()
        self._update_statistics()
        self.statusbar.showMessage("Evrak kaydedildi", 3000)

    def _open_shipment(self):
        self.archive_page.load_selected_shipment()

    def _validate_shipment(self):
        items   = self.shipment_page.get_items()
        sender  = self.shipment_page.get_selected_sender()
        receiver= self.shipment_page.get_selected_receiver()
        driver  = self.shipment_page.get_selected_driver()
        vehicle = self.shipment_page.get_selected_vehicle()

        # Ambalaj turleri listesini olustur
        packaging_types = [item.packaging_type for item in items]

        # ── ADREngine: Eksik kimyasal verisi kontrolü ─────────────────────
        data_sorunlar = ADREngine.verify_chemicals_data(items)
        if data_sorunlar:
            sorun_mesaj = "\n".join(
                f"  • {s['un']}: {s['sorun']}" for s in data_sorunlar
            )
            QMessageBox.warning(
                self, "Eksik / Hatalı Madde Verisi",
                f"Aşağıdaki maddelerde eksik veya geçersiz veri tespit edildi.\n"
                f"Bu durum ADR 1.1.3.6 puan hesaplamasını ve tünel kodu atamalarını "
                f"olumsuz etkileyebilir:\n\n{sorun_mesaj}\n\n"
                f"Kimyasal veritabanındaki kayıtları kontrol edin."
            )

        result = ADREngine.validate_shipment(items, sender, receiver, driver, vehicle, packaging_types)

        # [v4.2] Onay bir YASAM DONGUSU asamasidir: sonuc mesaj kutusunda
        # kaybolmaz — evrak kaydedilir ve statusu KALICI olarak guncellenir.
        error_text_all = "\n".join([msg for _, msg in result.errors])
        saved = False
        if items and hasattr(self.shipment_page, "_save_shipment_internal"):
            try:
                saved = bool(self.shipment_page._save_shipment_internal())
            except Exception:
                logging.getLogger(__name__).exception("Onay oncesi kayit basarisiz")
        sid = getattr(self.shipment_page, "current_shipment_id", None)
        if saved and sid:
            self.db.set_shipment_validation(sid, result.is_valid, error_text_all)
            if hasattr(self.shipment_page, "lbl_status"):
                if result.is_valid:
                    self.shipment_page.lbl_status.setText("ONAYLANDI")
                    self.shipment_page.lbl_status.setStyleSheet(
                        "color: #A6E3A1; font-weight: bold;")
                else:
                    self.shipment_page.lbl_status.setText("DOGRULAMA HATALI")
                    self.shipment_page.lbl_status.setStyleSheet(
                        "color: #F38BA8; font-weight: bold;")

        if result.is_valid:
            QMessageBox.information(self, "Dogrulama Basarili",
                "Evrak tum ADR kontrollerinden gecti"
                + (", kaydedildi ve ONAYLANDI statusune alindi." if saved and sid
                   else ". (Kalici onay icin evraki kaydedin.)")
                + " PDF olusturabilirsiniz.")
        else:
            error_text = "\n".join([msg for _, msg in result.errors])
            warning_text = "\n".join([msg for _, msg in result.warnings])
            info_text = "\n".join([msg for _, msg in result.info])

            msg = f"Evrakta hatalar var:\n\n{error_text}"
            if warning_text:
                msg += f"\n\nUYARILAR:\n{warning_text}"
            if info_text:
                msg += f"\n\nBILGILER:\n{info_text}"
            msg += "\n\nLutfen duzeltip tekrar deneyin."

            QMessageBox.warning(self, "Dogrulama Hatalari", msg)

        self._update_adr_panel()

    def _preview_shipment(self):
        self._update_adr_panel()

    def _export_pdf(self):
        self.shipment_page.export_pdf()

    def _export_excel(self):
        self.shipment_page.export_excel()

    def _print_document(self):
        self.shipment_page.print_document()

    def _import_json(self):
        path, _ = QFileDialog.getOpenFileName(self, "JSON Aktar", "", "JSON (*.json)")
        if path:
            count = self.db.import_json_data(path)
            QMessageBox.information(self, "Aktarma Tamamlandi",
                f"{count} kayit aktarildi.")
            self._update_statistics()

    def _import_csv_adr_table(self):
        """ADR A Tablosu CSV dosyasını içe aktar."""
        path, _ = QFileDialog.getOpenFileName(
            self, "ADR A Tablosu CSV Sec", "", "CSV (*.csv);;All Files (*)")
        if not path:
            return

        try:
            if not CSV_IMPORTER_AVAILABLE:
                QMessageBox.warning(self, "Modül Eksik",
                    "adr_csv_importer.py bulunamadı.\n"
                    "CSV aktarımı için bu dosyanın programa eklenmesi gerekir.")
                return
            # CSV importer fonksiyonlarını çağır
            result = import_csv_to_db(path, self.db.db_path)
            if result > 0:
                QMessageBox.information(self, "CSV Aktarma Basarili",
                    f"ADR A Tablosu basariyla aktarildi.\n{result} kayit isleme alindi.")
                self._update_statistics()
            else:
                QMessageBox.warning(self, "CSV Aktarma",
                    "CSV dosyasinda islenecek kayit bulunamadi.")
        except Exception as e:
            QMessageBox.critical(self, "CSV Aktarma Hatasi",
                f"CSV aktarimi sirasinda hata olustu:\n{str(e)}")

    def _export_json_backup(self):
        path, _ = QFileDialog.getSaveFileName(self, "JSON Yedekle",
            f"adr_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            "JSON (*.json)")
        if path:
            self.db.export_to_json(path)
            QMessageBox.information(self, "Yedekleme Tamamlandi",
                f"Yedek olusturuldu:\n{path}")

    def _show_backups(self):
        backup_dir = Path(self.db.db_path).parent / "backups"
        if backup_dir.exists():
            backups = sorted(backup_dir.glob("adr_backup_*.db"), reverse=True)
            msg = "Son 10 yedek:\n\n"
            for b in backups[:10]:
                size = b.stat().st_size / 1024
                msg += f"{b.name} ({size:.1f} KB)\n"
            QMessageBox.information(self, "Yedekler", msg)
        else:
            QMessageBox.information(self, "Yedekler", "Yedek bulunamadi.")

    def _show_reports(self):
        """Toolbar 'Raporlar' butonu — sol menüdeki Raporlar sayfasına geçer."""
        self._switch_page(7)

    def _show_about(self):
        QMessageBox.about(self, "Hakkinda",
            f"<h2>{APP_NAME}</h2>"
            f"<p>Versiyon: {APP_VERSION}</p>"
            f"<p>ADR Tehlikeli Madde Tasima Evrak Yonetim Sistemi</p>"
            f"<p>PyQt6 + SQLite + ReportLab</p>"
            f"<p>2026 ADRSoft</p>")

    def _logout_and_relogin(self):
        reply=QMessageBox.question(self,"Oturumu Kapat","Oturumunuz kapatilacak.\nDevam?",
            QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No,QMessageBox.StandardButton.No)
        if reply==QMessageBox.StandardButton.No: return
        try: self.security.logout()
        except Exception as _e: logger.warning("Güvenlik oturum kapatma hatası: %s", _e)
        if hasattr(self,'session_timer') and self.session_timer: self.session_timer.stop()
        self.hide()
        login_dlg=LoginDialog(self.security,LicenseManager.validate()); login_dlg.exec()
        if login_dlg.logged_in:
            if hasattr(self,'session_timer') and self.session_timer: self.session_timer.start(300000)
            self.show()
        else:
            try: self.db.close()
            except Exception as _e: logger.warning("DB kapatılamadı: %s", _e)
            QApplication.quit()

    def closeEvent(self, event):
        if event.spontaneous():
            reply = QMessageBox.question(self, "Cikis",
                "Uygulamadan cikmak istiyor musunuz?\nKaydedilmemis degisiklikler kaybolabilir.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.No:
                event.ignore()
                return
        self.db.close()
        event.accept()



# =============================================================================
# AMBALAJ TÜRÜ COMBOBOX DELEGATE
# =============================================================================

PACKAGING_TYPE_OPTIONS = [
    "IBC", "Varil", "Bidon", "Kutu",
    "Çuval", "Kompozit Ambalaj", "Tank", "Dökme"
]

class PackagingTypeDelegate(QStyledItemDelegate):
    """Tablo sütununda Ambalaj Türü için QComboBox sunar."""

    def createEditor(self, parent, option, index):
        combo = QComboBox(parent)
        combo.addItems(PACKAGING_TYPE_OPTIONS)
        # Tek tıkla açılması için popup'ı hemen tetikle
        QTimer.singleShot(0, combo.showPopup)
        return combo

    def setEditorData(self, editor, index):
        value = index.data(Qt.ItemDataRole.DisplayRole) or ""
        i = editor.findText(value)
        if i >= 0:
            editor.setCurrentIndex(i)

    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentText(), Qt.ItemDataRole.EditRole)

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)
        
UNIT_OPTIONS = ["kg", "lt", "m3", "ton", "adet"]

class UnitDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        combo = QComboBox(parent)
        combo.addItems(UNIT_OPTIONS)
        QTimer.singleShot(0, combo.showPopup)
        return combo
    def setEditorData(self, editor, index):
        value = index.data(Qt.ItemDataRole.DisplayRole) or "kg"
        i = editor.findText(value)
        editor.setCurrentIndex(i if i >= 0 else 0)
    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentText(), Qt.ItemDataRole.EditRole)
    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)

# =============================================================================
# SEVKIYAT EDITOR SAYFASI
# =============================================================================

class ShipmentEditorPage(QWidget):

    def __init__(self, db: DatabaseManager, parent=None):
        super().__init__(parent)
        self.db = db
        self.parent_window = parent
        self.current_shipment_id = None
        self.items = []
        self.undo_stack = []
        self.undo_index = -1
        # Excel'den yüklenen ürünler — MainWindow tarafından doldurulur
        self.excel_products: list = []

        self._setup_ui()
        self._init_new_document()

    def _setup_ui(self):
        outer_layout = QVBoxLayout(self)
        outer_layout.setContentsMargins(0, 0, 0, 0)
        outer_layout.setSpacing(0)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.Shape.NoFrame)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        inner_widget = QWidget()
        scroll_area.setWidget(inner_widget)
        outer_layout.addWidget(scroll_area)

        layout = QVBoxLayout(inner_widget)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        # Evrak Bilgileri
        doc_group = QGroupBox("Evrak Bilgileri")
        doc_layout = QGridLayout(doc_group)

        doc_layout.addWidget(QLabel("Evrak No:"), 0, 0)
        self.txt_doc_no = QLineEdit()
        self.txt_doc_no.setReadOnly(True)
        self.txt_doc_no.setPlaceholderText("Otomatik olusturulur")
        doc_layout.addWidget(self.txt_doc_no, 0, 1)

        doc_layout.addWidget(QLabel("Tarih:"), 0, 2)
        self.date_doc = QDateEdit()
        self.date_doc.setCalendarPopup(True)
        self.date_doc.setDate(QDate.currentDate())
        doc_layout.addWidget(self.date_doc, 0, 3)

        doc_layout.addWidget(QLabel("Durum:"), 0, 4)
        self.lbl_status = QLabel("TASLAK")
        self.lbl_status.setStyleSheet("color: #F9E2AF; font-weight: bold;")
        doc_layout.addWidget(self.lbl_status, 0, 5)

        layout.addWidget(doc_group)

        # Firma Bilgileri
        company_splitter = QSplitter(Qt.Orientation.Horizontal)

        sender_group = QGroupBox("Gonderici")
        sender_layout = QFormLayout(sender_group)

        self.cmb_sender = QComboBox()
        self.cmb_sender.setEditable(True)
        self.cmb_sender.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.cmb_sender.currentIndexChanged.connect(self._on_sender_changed)
        sender_layout.addRow("Firma:", self.cmb_sender)

        self.lbl_sender_tax = QLabel("-")
        sender_layout.addRow("Vergi No:", self.lbl_sender_tax)

        self.lbl_sender_address = QLabel("-")
        self.lbl_sender_address.setWordWrap(True)
        sender_layout.addRow("Adres:", self.lbl_sender_address)

        self.lbl_sender_city = QLabel("-")
        sender_layout.addRow("Sehir:", self.lbl_sender_city)

        btn_new_sender = QPushButton("Yeni Firma")
        btn_new_sender.clicked.connect(self._add_new_sender)
        sender_layout.addRow(btn_new_sender)

        company_splitter.addWidget(sender_group)

        receiver_group = QGroupBox("Alici")
        receiver_layout = QFormLayout(receiver_group)

        self.cmb_receiver = QComboBox()
        self.cmb_receiver.setEditable(True)
        self.cmb_receiver.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.cmb_receiver.currentIndexChanged.connect(self._on_receiver_changed)
        receiver_layout.addRow("Firma:", self.cmb_receiver)

        self.lbl_receiver_tax = QLabel("-")
        receiver_layout.addRow("Vergi No:", self.lbl_receiver_tax)

        self.lbl_receiver_address = QLabel("-")
        self.lbl_receiver_address.setWordWrap(True)
        receiver_layout.addRow("Adres:", self.lbl_receiver_address)

        self.lbl_receiver_city = QLabel("-")
        receiver_layout.addRow("Sehir:", self.lbl_receiver_city)

        btn_new_receiver = QPushButton("Yeni Firma")
        btn_new_receiver.clicked.connect(self._add_new_receiver)
        receiver_layout.addRow(btn_new_receiver)

        company_splitter.addWidget(receiver_group)

        carrier_group = QGroupBox("Tasiyici")
        carrier_layout = QFormLayout(carrier_group)

        self.cmb_carrier = QComboBox()
        self.cmb_carrier.setEditable(True)
        self.cmb_carrier.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        carrier_layout.addRow("Firma:", self.cmb_carrier)

        btn_new_carrier = QPushButton("Yeni Firma")
        btn_new_carrier.clicked.connect(self._add_new_carrier)
        carrier_layout.addRow(btn_new_carrier)

        company_splitter.addWidget(carrier_group)
        company_splitter.setSizes([350, 350, 300])
        layout.addWidget(company_splitter)

        # Surucu ve Arac
        transport_splitter = QSplitter(Qt.Orientation.Horizontal)

        driver_group = QGroupBox("Surucu")
        driver_layout = QFormLayout(driver_group)

        self.cmb_driver = QComboBox()
        self.cmb_driver.setEditable(True)
        self.cmb_driver.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.cmb_driver.currentIndexChanged.connect(self._on_driver_changed)
        driver_layout.addRow("Surucu:", self.cmb_driver)

        self.lbl_driver_src5 = QLabel("-")
        driver_layout.addRow("SRC5 No:", self.lbl_driver_src5)

        self.lbl_driver_adr = QLabel("-")
        driver_layout.addRow("ADR Belgesi:", self.lbl_driver_adr)

        self.lbl_driver_adr_expiry = QLabel("-")
        driver_layout.addRow("ADR Gecerlilik:", self.lbl_driver_adr_expiry)

        btn_new_driver = QPushButton("Yeni Surucu")
        btn_new_driver.clicked.connect(self._add_new_driver)
        driver_layout.addRow(btn_new_driver)

        transport_splitter.addWidget(driver_group)

        vehicle_group = QGroupBox("Arac")
        vehicle_layout = QFormLayout(vehicle_group)

        self.cmb_vehicle = QComboBox()
        self.cmb_vehicle.setEditable(True)
        self.cmb_vehicle.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.cmb_vehicle.currentIndexChanged.connect(self._on_vehicle_changed)
        vehicle_layout.addRow("Plaka:", self.cmb_vehicle)

        self.lbl_vehicle_trailer = QLabel("-")
        vehicle_layout.addRow("Dorse:", self.lbl_vehicle_trailer)

        self.lbl_vehicle_adr = QLabel("-")
        vehicle_layout.addRow("ADR Uygunluk:", self.lbl_vehicle_adr)

        self.lbl_vehicle_inspection = QLabel("-")
        vehicle_layout.addRow("Muayene:", self.lbl_vehicle_inspection)

        btn_new_vehicle = QPushButton("Yeni Arac")
        btn_new_vehicle.clicked.connect(self._add_new_vehicle)
        vehicle_layout.addRow(btn_new_vehicle)

        transport_splitter.addWidget(vehicle_group)
        transport_splitter.setSizes([400, 400])
        layout.addWidget(transport_splitter)

        # Urun Tablosu
        items_group = QGroupBox("Tasinan Urunler")
        items_layout = QVBoxLayout(items_group)

        add_layout = QHBoxLayout()

        self.txt_un_search = QLineEdit()
        self.txt_un_search.setPlaceholderText("UN No veya isim ara (ornek: 1203, Benzin)...")
        self.txt_un_search.setMinimumWidth(250)
        self.txt_un_search.textChanged.connect(self._on_un_search)
        add_layout.addWidget(self.txt_un_search)

        self.cmb_chemical_results = QComboBox()
        self.cmb_chemical_results.setMinimumWidth(300)
        add_layout.addWidget(self.cmb_chemical_results)

        btn_add_item = QPushButton("Ekle")
        btn_add_item.setObjectName("success")
        btn_add_item.clicked.connect(self._add_item)
        add_layout.addWidget(btn_add_item)

        add_layout.addStretch()
        items_layout.addLayout(add_layout)

        self.tbl_items = QTableWidget()
        self.tbl_items.setColumnCount(12)
        self.tbl_items.setHorizontalHeaderLabels([
            "#", "UN No", "Teknik Isim", "Sinif", "PG", "Tunel kodu",
            "Ambalaj Turu", "Adet", "Net Miktar", "Birim", "LQ", "EQ"
        ])
        self.tbl_items.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.tbl_items.horizontalHeader().setStretchLastSection(True)
        self.tbl_items.setColumnWidth(0,  35)
        self.tbl_items.setColumnWidth(1,  70)
        self.tbl_items.setColumnWidth(2, 200)
        self.tbl_items.setColumnWidth(3,  55)
        self.tbl_items.setColumnWidth(4,  40)
        self.tbl_items.setColumnWidth(5,  80)
        self.tbl_items.setColumnWidth(6, 110)
        self.tbl_items.setColumnWidth(7,  50)
        self.tbl_items.setColumnWidth(8,  80)
        self.tbl_items.setColumnWidth(9,  55)   # Birim
        self.tbl_items.setColumnWidth(10, 35)   # LQ
        self.tbl_items.setColumnWidth(11, 35)   # EQ
        self.tbl_items.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.tbl_items.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.tbl_items.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tbl_items.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tbl_items.customContextMenuRequested.connect(self._show_item_context_menu)
        self.tbl_items.setMinimumHeight(180)
        self.tbl_items.itemChanged.connect(self._on_table_item_changed)
        # Ambalaj Türü sütunu (6) için QComboBox delegate
        self._packaging_delegate = PackagingTypeDelegate(self.tbl_items)
        self.tbl_items.setItemDelegateForColumn(6, self._packaging_delegate) 
        # Birim sütunu (9) için QComboBox delegate
        self._unit_delegate = UnitDelegate(self.tbl_items)
        self.tbl_items.setItemDelegateForColumn(9, self._unit_delegate)
        # Tek tıkla combo aç
        self.tbl_items.cellClicked.connect(self._on_cell_clicked)
        items_layout.addWidget(self.tbl_items)

        btn_layout = QHBoxLayout()

        btn_delete_item = QPushButton("Secileni Sil")
        btn_delete_item.setObjectName("danger")
        btn_delete_item.clicked.connect(self._delete_selected_item)
        btn_layout.addWidget(btn_delete_item)

        btn_move_up = QPushButton("Yukari")
        btn_move_up.clicked.connect(self._move_item_up)
        btn_layout.addWidget(btn_move_up)

        btn_move_down = QPushButton("Asagi")
        btn_move_down.clicked.connect(self._move_item_down)
        btn_layout.addWidget(btn_move_down)

        btn_layout.addStretch()

        btn_clear_items = QPushButton("Tumunu Temizle")
        btn_clear_items.setObjectName("danger")
        btn_clear_items.clicked.connect(self._clear_all_items)
        btn_layout.addWidget(btn_clear_items)

        items_layout.addLayout(btn_layout)
        layout.addWidget(items_group)

        # Notlar
        notes_group = QGroupBox("Notlar")
        notes_layout = QVBoxLayout(notes_group)

        self.txt_notes = QTextEdit()
        self.txt_notes.setMaximumHeight(80)
        self.txt_notes.setPlaceholderText("Evrakla ilgili notlar...")
        notes_layout.addWidget(self.txt_notes)

        layout.addWidget(notes_group)

        # Alt Butonlar
        bottom_layout = QHBoxLayout()

        btn_validate = QPushButton("DOGRULA")
        btn_validate.setObjectName("primary")
        btn_validate.setMinimumHeight(40)
        btn_validate.clicked.connect(self._validate_and_update)
        bottom_layout.addWidget(btn_validate)

        btn_save = QPushButton("KAYDET")
        btn_save.setObjectName("success")
        btn_save.setMinimumHeight(40)
        btn_save.clicked.connect(self.save_shipment)
        bottom_layout.addWidget(btn_save)

        btn_preview = QPushButton("ONIZLE")
        btn_preview.setMinimumHeight(40)
        btn_preview.clicked.connect(self._preview_document)
        bottom_layout.addWidget(btn_preview)

        btn_pdf = QPushButton("PDF")
        btn_pdf.setMinimumHeight(40)
        btn_pdf.clicked.connect(self.export_pdf)
        bottom_layout.addWidget(btn_pdf)

        btn_print = QPushButton("YAZDIR")
        btn_print.setMinimumHeight(40)
        btn_print.clicked.connect(self.print_document)
        bottom_layout.addWidget(btn_print)

        layout.addLayout(bottom_layout)

        self._load_combobox_data()

    def _load_combobox_data(self):
        self.cmb_sender.clear()
        self.cmb_sender.addItem("-- Seciniz --", None)
        for comp in self.db.get_companies("sender"):
            self.cmb_sender.addItem(f"{comp.name} ({comp.city})", comp.id)

        self.cmb_receiver.clear()
        self.cmb_receiver.addItem("-- Seciniz --", None)
        for comp in self.db.get_companies("receiver"):
            self.cmb_receiver.addItem(f"{comp.name} ({comp.city})", comp.id)

        self.cmb_carrier.clear()
        self.cmb_carrier.addItem("-- Seciniz --", None)
        for comp in self.db.get_companies("carrier"):
            self.cmb_carrier.addItem(f"{comp.name} ({comp.city})", comp.id)

        self.cmb_driver.clear()
        self.cmb_driver.addItem("-- Seciniz --", None)
        for drv in self.db.get_drivers():
            self.cmb_driver.addItem(f"{drv.full_name} - {drv.src5_no}", drv.id)

        self.cmb_vehicle.clear()
        self.cmb_vehicle.addItem("-- Seciniz --", None)
        for veh in self.db.get_vehicles():
            self.cmb_vehicle.addItem(f"{veh.plate}", veh.id)

    def _init_new_document(self):
        self.current_shipment_id = None
        self.txt_doc_no.setText(ADREngine.format_document_number())
        self.date_doc.setDate(QDate.currentDate())
        self.lbl_status.setText("TASLAK")
        self.lbl_status.setStyleSheet("color: #F9E2AF; font-weight: bold;")

        self.cmb_sender.setCurrentIndex(0)
        self.cmb_receiver.setCurrentIndex(0)
        self.cmb_carrier.setCurrentIndex(0)
        self.cmb_driver.setCurrentIndex(0)
        self.cmb_vehicle.setCurrentIndex(0)

        self.items.clear()
        self._refresh_items_table()
        self.txt_notes.clear()

        self.undo_stack = []
        self.undo_index = -1
        self._push_undo()
        if self.parent_window:
            self.parent_window._update_adr_panel()

    def _on_sender_changed(self, index):
        if index <= 0:
            self.lbl_sender_tax.setText("-")
            self.lbl_sender_address.setText("-")
            self.lbl_sender_city.setText("-")
            return

        company_id = self.cmb_sender.itemData(index)
        company = self.db.get_company(company_id)
        if company:
            self.lbl_sender_tax.setText(company.tax_number or "-")
            self.lbl_sender_address.setText(company.address or "-")
            self.lbl_sender_city.setText(f"{company.city or ''} / {company.district or ''}")
            self._push_undo()

    def _on_receiver_changed(self, index):
        if index <= 0:
            self.lbl_receiver_tax.setText("-")
            self.lbl_receiver_address.setText("-")
            self.lbl_receiver_city.setText("-")
            return

        company_id = self.cmb_receiver.itemData(index)
        company = self.db.get_company(company_id)
        if company:
            self.lbl_receiver_tax.setText(company.tax_number or "-")
            self.lbl_receiver_address.setText(company.address or "-")
            self.lbl_receiver_city.setText(f"{company.city or ''} / {company.district or ''}")
            self._push_undo()

    def _on_driver_changed(self, index):
        if index <= 0:
            self.lbl_driver_src5.setText("-")
            self.lbl_driver_adr.setText("-")
            self.lbl_driver_adr_expiry.setText("-")
            return

        driver_id = self.cmb_driver.itemData(index)
        driver = self.db.get_driver(driver_id)
        if driver:
            self.lbl_driver_src5.setText(driver.src5_no or "-")
            self.lbl_driver_adr.setText(driver.adr_certificate_no or "-")

            if driver.adr_certificate_expiry:
                try:
                    expiry = datetime.strptime(driver.adr_certificate_expiry, "%Y-%m-%d")
                    days_left = (expiry - datetime.now()).days
                    if days_left < 0:
                        self.lbl_driver_adr_expiry.setText(f"GECERSIZ ({driver.adr_certificate_expiry})")
                        self.lbl_driver_adr_expiry.setStyleSheet("color: #F38BA8; font-weight: bold;")
                    elif days_left < 30:
                        self.lbl_driver_adr_expiry.setText(f"{days_left} gun kaldi ({driver.adr_certificate_expiry})")
                        self.lbl_driver_adr_expiry.setStyleSheet("color: #F9E2AF;")
                    else:
                        self.lbl_driver_adr_expiry.setText(f"{driver.adr_certificate_expiry}")
                        self.lbl_driver_adr_expiry.setStyleSheet("color: #A6E3A1;")
                except:
                    self.lbl_driver_adr_expiry.setText(driver.adr_certificate_expiry)
            else:
                self.lbl_driver_adr_expiry.setText("-")

            self._push_undo()

    def _on_vehicle_changed(self, index):
        if index <= 0:
            self.lbl_vehicle_trailer.setText("-")
            self.lbl_vehicle_adr.setText("-")
            self.lbl_vehicle_inspection.setText("-")
            return

        vehicle_id = self.cmb_vehicle.itemData(index)
        vehicle = self.db.get_vehicle(vehicle_id)
        if vehicle:
            self.lbl_vehicle_trailer.setText(vehicle.trailer_plate or "-")

            if vehicle.adr_compliance_expiry:
                try:
                    expiry = datetime.strptime(vehicle.adr_compliance_expiry, "%Y-%m-%d")
                    days_left = (expiry - datetime.now()).days
                    if days_left < 0:
                        self.lbl_vehicle_adr.setText("GECERSIZ")
                        self.lbl_vehicle_adr.setStyleSheet("color: #F38BA8; font-weight: bold;")
                    elif days_left < 30:
                        self.lbl_vehicle_adr.setText(f"{days_left} gun")
                        self.lbl_vehicle_adr.setStyleSheet("color: #F9E2AF;")
                    else:
                        self.lbl_vehicle_adr.setText("Gecerli")
                        self.lbl_vehicle_adr.setStyleSheet("color: #A6E3A1;")
                except:
                    self.lbl_vehicle_adr.setText(vehicle.adr_compliance_expiry)
            else:
                self.lbl_vehicle_adr.setText("-")

            if vehicle.inspection_expiry:
                try:
                    expiry = datetime.strptime(vehicle.inspection_expiry, "%Y-%m-%d")
                    days_left = (expiry - datetime.now()).days
                    if days_left < 0:
                        self.lbl_vehicle_inspection.setText("GECERSIZ")
                        self.lbl_vehicle_inspection.setStyleSheet("color: #F38BA8; font-weight: bold;")
                    elif days_left < 30:
                        self.lbl_vehicle_inspection.setText(f"{days_left} gun")
                        self.lbl_vehicle_inspection.setStyleSheet("color: #F9E2AF;")
                    else:
                        self.lbl_vehicle_inspection.setText("Gecerli")
                        self.lbl_vehicle_inspection.setStyleSheet("color: #A6E3A1;")
                except:
                    self.lbl_vehicle_inspection.setText(vehicle.inspection_expiry)
            else:
                self.lbl_vehicle_inspection.setText("-")

            self._push_undo()

    def _on_un_search(self, text):
        if len(text) < 2:
            self.cmb_chemical_results.clear()
            return

        text_upper = text.strip().upper()
        self.cmb_chemical_results.clear()

        # 1) Veritabanı araması (DB kimyasalları) — sevkiyat adı taşıma evrakına yazılır
        chemicals = self.db.search_chemicals(text, limit=20)
        for chem in chemicals:
            display = (f"[DB] {chem.proper_shipping_name_tr} "
                       f"(UN{chem.un_number}, Sınıf {chem.class_code}, PG {chem.packing_group})")
            # UserRole: ('db', chem.id)
            self.cmb_chemical_results.addItem(display, ("db", chem.id))

        # 2) Excel ürünleri — kimyasal adına göre ara, sevkiyat adı taşıma evrakına yazılır
        for p in self.excel_products:
            kim_adi = p.get("kimyasal_adi", "")
            sev_adi = p.get("sevkiyat_adi", "")
            un_no   = p.get("un_numarasi", "")
            sinif   = p.get("sinif", "")
            pg      = p.get("paketleme_grubu", "")
            if (text_upper in kim_adi.upper()
                    or text_upper in sev_adi.upper()
                    or text_upper in un_no.upper()):
                display = (f"[Excel] {kim_adi} → {sev_adi} "
                           f"(UN{un_no}, Sınıf {sinif}, PG {pg})")
                # UserRole: ('excel', product dict)
                self.cmb_chemical_results.addItem(display, ("excel", p))

    def _add_item(self):
        if self.cmb_chemical_results.currentIndex() < 0:
            QMessageBox.warning(self, "Hata", "Lutfen bir urun secin!")
            return

        data = self.cmb_chemical_results.currentData()
        if data is None:
            return

        source, payload = data

        if source == "db":
            # Veritabanı kimyasalı — sevkiyat adı DB'den gelir
            chemical = self.db.get_chemical(payload)
            if not chemical:
                return
            item = ShipmentItem(
                chemical_id=payload,
                un_number=chemical.un_number,
                proper_name=chemical.proper_shipping_name_tr,   # sevkiyat adı
                class_code=chemical.class_code,
                packing_group=chemical.packing_group,
                tunnel_code=chemical.tunnel_code,
                segregation_group=chemical.segregation_group,
                transport_category=chemical.transport_category,
                packaging_type="Kutu",
                packaging_count=1,
                net_quantity=0,
                gross_quantity=0,
                unit="kg",
                is_lq=False,
                is_eq=False
            )

        elif source == "excel":
            # Excel ürünü — taşıma evrakına SEVKİYAT ADI (E sütunu) yazılır
            p = payload
            sevkiyat_adi = p.get("sevkiyat_adi", "").strip()
            if not sevkiyat_adi:
                sevkiyat_adi = p.get("kimyasal_adi", "")

            # Tekrar kontrolü (sevkiyat adına göre)
            for existing in self.items:
                if existing.proper_name.strip().upper() == sevkiyat_adi.upper():
                    QMessageBox.warning(
                        self, "Tekrar",
                        f"Bu sevkiyat adı zaten taşıma evrakında mevcut:\n'{sevkiyat_adi}'"
                    )
                    return

            item = ShipmentItem(
                chemical_id=0,
                un_number=p.get("un_numarasi", ""),
                proper_name=sevkiyat_adi,               # sevkiyat adı taşıma evrakına
                class_code=p.get("sinif", ""),
                packing_group=p.get("paketleme_grubu", ""),
                tunnel_code=p.get("tunel_kodu", ""),
                segregation_group="",
                transport_category=p.get("tasima_kategorisi", ""),
                packaging_type="IBC",
                packaging_count=1,
                net_quantity=0,
                gross_quantity=0,
                unit="kg",
                is_lq=False,
                is_eq=False
            )
        else:
            return

        self.items.append(item)
        self._refresh_items_table()
        self._push_undo()

        if self.parent_window:
            self.parent_window._update_adr_panel()

    def add_item_from_excel(self, product: dict):
        """
        ExcelChemicalPage'den gelen ürün sözlüğünü ShipmentItem'a dönüştürüp
        taşıma evrakına ekler.
        Kimyasal adı (B sütunu) seçimde gösterilir; taşıma evrakına
        SEVKİYAT ADI (E sütunu) yazılır. Aynı sevkiyat adı zaten varsa eklenmez.
        """
        sevkiyat_adi = product.get("sevkiyat_adi", "").strip()
        if not sevkiyat_adi:
            sevkiyat_adi = product.get("kimyasal_adi", "").strip()
        if not sevkiyat_adi:
            return

        # Tekrar kontrolü (sevkiyat adına göre)
        for existing in self.items:
            if existing.proper_name.strip().upper() == sevkiyat_adi.upper():
                QMessageBox.warning(
                    self, "Tekrar",
                    f"Bu sevkiyat adı zaten taşıma evrakında mevcut:\n'{sevkiyat_adi}'"
                )
                return

        item = ShipmentItem(
            chemical_id=0,
            un_number=product.get("un_numarasi", ""),
            proper_name=sevkiyat_adi,               # sevkiyat adı taşıma evrakına
            class_code=product.get("sinif", ""),
            packing_group=product.get("paketleme_grubu", ""),
            tunnel_code=product.get("tunel_kodu", ""),
            segregation_group="",
            transport_category=product.get("tasima_kategorisi", ""),
            packaging_type="IBC",
            packaging_count=1,
            net_quantity=0,
            gross_quantity=0,
            unit="kg",
            is_lq=False,
            is_eq=False
        )

        self.items.append(item)
        self._refresh_items_table()
        self._push_undo()

    def _refresh_items_table(self):
        self.tbl_items.blockSignals(True)
        self.tbl_items.setRowCount(len(self.items))

        for i, item in enumerate(self.items):
            self.tbl_items.setItem(i, 0, QTableWidgetItem(str(i + 1)))
            self.tbl_items.setItem(i, 1, QTableWidgetItem(f"UN{item.un_number}"))
            self.tbl_items.setItem(i, 2, QTableWidgetItem(item.proper_name))

            class_item = QTableWidgetItem(item.class_code)
            bg_color, fg_color = ADREngine.get_class_color(item.class_code)
            class_item.setBackground(QColor(bg_color))
            class_item.setForeground(QColor(fg_color))
            class_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.tbl_items.setItem(i, 3, class_item)

            self.tbl_items.setItem(i, 4, QTableWidgetItem(item.packing_group))

            tunnel_val = item.tunnel_code or "-"
            tunnel_item = QTableWidgetItem(tunnel_val)
            tunnel_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if "A" in tunnel_val:
                tunnel_item.setBackground(QColor("#F38BA8"))
                tunnel_item.setForeground(QColor("#1E1E2E"))
            elif "B" in tunnel_val or "C" in tunnel_val:
                tunnel_item.setBackground(QColor("#F9E2AF"))
                tunnel_item.setForeground(QColor("#1E1E2E"))
            else:
                tunnel_item.setBackground(QColor("#A6E3A1"))
                tunnel_item.setForeground(QColor("#1E1E2E"))
            self.tbl_items.setItem(i, 5, tunnel_item)

            self.tbl_items.setItem(i, 6, QTableWidgetItem(item.packaging_type))
            self.tbl_items.setItem(i, 7, QTableWidgetItem(str(item.packaging_count)))
            self.tbl_items.setItem(i, 8, QTableWidgetItem(str(item.net_quantity)))

            # 9 Birim (unit delegate ile seçilir)
            unit_item = QTableWidgetItem(item.unit or "kg")
            unit_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.tbl_items.setItem(i, 9, unit_item)


            lq_item = QTableWidgetItem("v" if item.is_lq else "")
            lq_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if item.is_lq:
                lq_item.setBackground(QColor("#A6E3A1"))
            self.tbl_items.setItem(i, 10, lq_item)

            eq_item = QTableWidgetItem("v" if item.is_eq else "")
            eq_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if item.is_eq:
                eq_item.setBackground(QColor("#89B4FA"))
            self.tbl_items.setItem(i, 11, eq_item)

        self.tbl_items.blockSignals(False)


    def _on_cell_clicked(self, row, col):
        """Ambalaj Türü (6) ve Birim (9) sütunlarına tek tıkla editörü aç."""
        if col in (6, 9):
            self.tbl_items.edit(self.tbl_items.model().index(row, col))
            
    def _on_table_item_changed(self, table_item):
        row = table_item.row()
        col = table_item.column()

        if row < 0 or row >= len(self.items):
            return

        self.tbl_items.blockSignals(True)

        item = self.items[row]
        val = table_item.text().strip()

        try:
            if col == 5:
                item.tunnel_code = val
            elif col == 6:
                item.packaging_type = val
            elif col == 7:
                item.packaging_count = int(val) if val.isdigit() else item.packaging_count
            elif col == 8:
                item.net_quantity = float(val) if val else 0.0
            elif col == 9:
                item.unit = val if val in UNIT_OPTIONS else "kg"
        except (ValueError, TypeError):
            pass

        self.tbl_items.blockSignals(False)

        if self.parent_window:
            self.parent_window._update_adr_panel()

    def _show_item_context_menu(self, position):
        menu = QMenu(self)

        edit_action = menu.addAction("Duzenle")
        edit_action.triggered.connect(self._edit_selected_item)

        copy_action = menu.addAction("Kopyala")
        copy_action.triggered.connect(self._copy_item)

        menu.addSeparator()

        move_up_action   = menu.addAction("Yukari Tasi")
        move_up_action.triggered.connect(self._move_item_up)

        move_down_action = menu.addAction("Asagi Tasi")
        move_down_action.triggered.connect(self._move_item_down)

        menu.addSeparator()

        lq_action = menu.addAction("LQ Olarak Isaretle")
        lq_action.triggered.connect(self._toggle_lq)

        eq_action = menu.addAction("EQ Olarak Isaretle")
        eq_action.triggered.connect(self._toggle_eq)

        menu.addSeparator()

        delete_action = menu.addAction("Sil")
        delete_action.triggered.connect(self._delete_selected_item)

        menu.exec(self.tbl_items.viewport().mapToGlobal(position))

    def _edit_selected_item(self):
        row = self.tbl_items.currentRow()
        if row < 0 or row >= len(self.items):
            return

        item = self.items[row]

        dialog = QDialog(self)
        dialog.setWindowTitle(f"UN{item.un_number} - Duzenle")
        dialog.setMinimumWidth(400)
        layout = QFormLayout(dialog)

        cmb_packaging = QComboBox()
        cmb_packaging.addItems(["IBC", "Varil", "Bidon", "Kutu", "Çuval", "Kompozit Ambalaj", "Tank", "Dökme"])
        cmb_packaging.setCurrentText(item.packaging_type if item.packaging_type else "Kutu")
        layout.addRow("Ambalaj Turu (*):", cmb_packaging)

        spin_count = QSpinBox()
        spin_count.setValue(item.packaging_count)
        spin_count.setMinimum(1)
        layout.addRow("Ambalaj Adet:", spin_count)

        spin_net = QDoubleSpinBox()
        spin_net.setValue(item.net_quantity)
        spin_net.setMaximum(999999)
        spin_net.setDecimals(2)
        
        
        cmb_unit = QComboBox()
        cmb_unit.addItems(UNIT_OPTIONS)
        cmb_unit.setCurrentText(item.unit if item.unit in UNIT_OPTIONS else "kg")
        cmb_unit.setFixedWidth(65)
        net_row = QHBoxLayout()
        net_row.addWidget(spin_net, 1)
        net_row.addWidget(cmb_unit)
        layout.addRow("Net Miktar / Birim:", net_row)



        cmb_unit = QComboBox()
        cmb_unit.addItems(["kg", "lt", "adet", "m3"])
        cmb_unit.setCurrentText(item.unit)
        layout.addRow("Birim:", cmb_unit)

        chk_lq = QCheckBox("Limited Quantity (LQ)")
        chk_lq.setChecked(item.is_lq)
        layout.addRow(chk_lq)

        chk_eq = QCheckBox("Excepted Quantity (EQ)")
        chk_eq.setChecked(item.is_eq)
        layout.addRow(chk_eq)

        txt_notes = QLineEdit(item.notes)
        layout.addRow("Notlar:", txt_notes)

        buttons = QHBoxLayout()
        btn_ok = QPushButton("Kaydet")
        btn_ok.setObjectName("success")
        btn_ok.clicked.connect(dialog.accept)
        buttons.addWidget(btn_ok)

        btn_cancel = QPushButton("Iptal")
        btn_cancel.clicked.connect(dialog.reject)
        buttons.addWidget(btn_cancel)

        layout.addRow(buttons)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            item.packaging_type  = cmb_packaging.currentText()
            item.packaging_count = spin_count.value()
            item.net_quantity    = spin_net.value()
            item.unit            = cmb_unit.currentText()
            item.is_lq           = chk_lq.isChecked()
            item.is_eq           = chk_eq.isChecked()
            item.notes           = txt_notes.text()

            self._refresh_items_table()
            self._push_undo()
            if self.parent_window:
                self.parent_window._update_adr_panel()

    def _copy_item(self):
        row = self.tbl_items.currentRow()
        if row < 0 or row >= len(self.items):
            return

        copied = copy.deepcopy(self.items[row])
        copied.id = None
        self.items.insert(row + 1, copied)
        self._refresh_items_table()
        self._push_undo()
        if self.parent_window:
            self.parent_window._update_adr_panel()

    def _delete_selected_item(self):
        row = self.tbl_items.currentRow()
        if row < 0 or row >= len(self.items):
            return

        reply = QMessageBox.question(self, "Sil",
            f"UN{self.items[row].un_number} silinecek. Emin misiniz?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        if reply == QMessageBox.StandardButton.Yes:
            del self.items[row]
            self._refresh_items_table()
            self._push_undo()
            if self.parent_window:
                self.parent_window._update_adr_panel()

    def _move_item_up(self):
        row = self.tbl_items.currentRow()
        if row > 0:
            self.items[row], self.items[row - 1] = self.items[row - 1], self.items[row]
            self._refresh_items_table()
            self.tbl_items.selectRow(row - 1)
            self._push_undo()
            if self.parent_window:
                self.parent_window._update_adr_panel()

    def _move_item_down(self):
        row = self.tbl_items.currentRow()
        if row < len(self.items) - 1:
            self.items[row], self.items[row + 1] = self.items[row + 1], self.items[row]
            self._refresh_items_table()
            self.tbl_items.selectRow(row + 1)
            self._push_undo()
            if self.parent_window:
                self.parent_window._update_adr_panel()

    def _toggle_lq(self):
        row = self.tbl_items.currentRow()
        if row < 0 or row >= len(self.items):
            return
        self.items[row].is_lq = not self.items[row].is_lq
        if self.items[row].is_lq:
            self.items[row].is_eq = False
        self._refresh_items_table()
        self._push_undo()
        if self.parent_window:
            self.parent_window._update_adr_panel()

    def _toggle_eq(self):
        row = self.tbl_items.currentRow()
        if row < 0 or row >= len(self.items):
            return
        self.items[row].is_eq = not self.items[row].is_eq
        if self.items[row].is_eq:
            self.items[row].is_lq = False
        self._refresh_items_table()
        self._push_undo()
        if self.parent_window:
            self.parent_window._update_adr_panel()

    def _clear_all_items(self):
        if self.items:
            reply = QMessageBox.question(self, "Temizle",
                f"{len(self.items)} urun silinecek. Emin misiniz?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                self.items.clear()
                self._refresh_items_table()
                self._push_undo()
                if self.parent_window:
                    self.parent_window._update_adr_panel()

    def _validate_and_update(self):
        if self.parent_window:
            self.parent_window._validate_shipment()

    def _preview_document(self):
        if self.parent_window:
            self.parent_window._preview_shipment()

    # --- UNDO / REDO ---

    def _push_undo(self):
        state = {
            'items':    copy.deepcopy(self.items),
            'sender':   self.cmb_sender.currentIndex(),
            'receiver': self.cmb_receiver.currentIndex(),
            'carrier':  self.cmb_carrier.currentIndex(),
            'driver':   self.cmb_driver.currentIndex(),
            'vehicle':  self.cmb_vehicle.currentIndex(),
            'notes':    self.txt_notes.toPlainText()
        }

        self.undo_stack = self.undo_stack[:self.undo_index + 1]
        self.undo_stack.append(state)

        if len(self.undo_stack) > 50:
            self.undo_stack.pop(0)
        else:
            self.undo_index += 1

    def undo(self):
        if self.undo_index > 0:
            self.undo_index -= 1
            self._restore_state(self.undo_stack[self.undo_index])

    def redo(self):
        if self.undo_index < len(self.undo_stack) - 1:
            self.undo_index += 1
            self._restore_state(self.undo_stack[self.undo_index])

    def _restore_state(self, state):
        self.items = copy.deepcopy(state['items'])
        self.cmb_sender.setCurrentIndex(state['sender'])
        self.cmb_receiver.setCurrentIndex(state['receiver'])
        self.cmb_carrier.setCurrentIndex(state['carrier'])
        self.cmb_driver.setCurrentIndex(state['driver'])
        self.cmb_vehicle.setCurrentIndex(state['vehicle'])
        self.txt_notes.setPlainText(state['notes'])
        self._refresh_items_table()
        if self.parent_window:
            self.parent_window._update_adr_panel()

    # --- GETTERS ---

    def get_items(self) -> List[ShipmentItem]:
        return self.items

    def get_selected_sender(self) -> Optional[Company]:
        idx = self.cmb_sender.currentIndex()
        if idx > 0:
            return self.db.get_company(self.cmb_sender.itemData(idx))
        return None

    def get_selected_receiver(self) -> Optional[Company]:
        idx = self.cmb_receiver.currentIndex()
        if idx > 0:
            return self.db.get_company(self.cmb_receiver.itemData(idx))
        return None

    def get_selected_driver(self) -> Optional[Driver]:
        idx = self.cmb_driver.currentIndex()
        if idx > 0:
            return self.db.get_driver(self.cmb_driver.itemData(idx))
        return None

    def get_selected_vehicle(self) -> Optional[Vehicle]:
        idx = self.cmb_vehicle.currentIndex()
        if idx > 0:
            return self.db.get_vehicle(self.cmb_vehicle.itemData(idx))
        return None

    def get_document_info(self) -> Dict[str, str]:
        return {
            'document_no': self.txt_doc_no.text(),
            'date':        self.date_doc.date().toString("dd.MM.yyyy")
        }

    # --- SAVE / LOAD ---

    def save_shipment(self):
        try:
            return self._save_shipment_internal()
        except Exception as e:
            import traceback
            err = traceback.format_exc()
            QMessageBox.critical(self, "Kaydetme Hatasi",
                f"Beklenmeyen hata:\n{str(e)}\n\nDetay:\n{err}")
            return False

    def _save_shipment_internal(self):
        if not self.items:
            QMessageBox.warning(self, "Hata", "En az bir urun eklenmeli!")
            return False

        sender   = self.get_selected_sender()
        receiver = self.get_selected_receiver()

        if not sender or not receiver:
            QMessageBox.warning(self, "Hata", "Gonderici ve alici firma secilmeli!")
            return False

        driver  = self.get_selected_driver()
        vehicle = self.get_selected_vehicle()

        self._sync_items_from_table()

        report = ADREngine.generate_adr_report(self.items, driver, vehicle)

        shipment = Shipment(
            id=self.current_shipment_id,
            document_no=self.txt_doc_no.text(),
            document_date=self.date_doc.date().toString("yyyy-MM-dd"),
            status="Taslak",
            sender_id=sender.id if sender else 0,
            receiver_id=receiver.id if receiver else 0,
            carrier_id=self.cmb_carrier.currentData() or 0,
            driver_id=self.cmb_driver.currentData() or 0,
            vehicle_id=self.cmb_vehicle.currentData() or 0,
            total_points=report.total_points,
            orange_plate_required=report.orange_plate_required,
            written_instructions_required=report.written_instructions_required,
            driver_adr_required=report.driver_adr_required,
            tunnel_restriction_code=report.tunnel_code,
            exemption_type=report.exemption_type,
            is_validated=False,
            validation_errors="",
            notes=self.txt_notes.toPlainText()
        )

        if self.current_shipment_id:
            self.db.update_shipment(shipment)
        else:
            self.current_shipment_id = self.db.add_shipment(shipment)
            shipment.id = self.current_shipment_id

        self.db.delete_shipment_items(shipment.id)
        for item in self.items:
            item.shipment_id = shipment.id
            self.db.add_shipment_item(item)

        self.lbl_status.setText("KAYDEDILDI")
        self.lbl_status.setStyleSheet("color: #A6E3A1; font-weight: bold;")

        if self.parent_window:
            self.parent_window.statusbar.showMessage(
                f"Evrak kaydedildi: {shipment.document_no}", 5000)

        return True

    def load_shipment(self, shipment_id: int):
        shipment = self.db.get_shipment(shipment_id)
        if not shipment:
            return

        self.current_shipment_id = shipment.id
        self.txt_doc_no.setText(shipment.document_no)

        try:
            date = QDate.fromString(shipment.document_date, "yyyy-MM-dd")
            self.date_doc.setDate(date)
        except:
            self.date_doc.setDate(QDate.currentDate())

        self.lbl_status.setText(shipment.status.upper())
        self.lbl_status.setStyleSheet("color: #A6E3A1; font-weight: bold;")

        self._set_combobox_by_id(self.cmb_sender,   shipment.sender_id)
        self._set_combobox_by_id(self.cmb_receiver, shipment.receiver_id)
        self._set_combobox_by_id(self.cmb_carrier,  shipment.carrier_id)
        self._set_combobox_by_id(self.cmb_driver,   shipment.driver_id)
        self._set_combobox_by_id(self.cmb_vehicle,  shipment.vehicle_id)

        self.items = self.db.get_shipment_items(shipment.id)
        self._refresh_items_table()
        self.txt_notes.setPlainText(shipment.notes or "")

        self._push_undo()

        if self.parent_window:
            self.parent_window._update_adr_panel()

    def _set_combobox_by_id(self, combo: QComboBox, id_value: int):
        for i in range(combo.count()):
            if combo.itemData(i) == id_value:
                combo.setCurrentIndex(i)
                return
        combo.setCurrentIndex(0)

    def clear_all(self):
        self._init_new_document()

    # --- EXPORT ---

    def _sync_items_from_table(self):
        for row in range(min(self.tbl_items.rowCount(), len(self.items))):
            item = self.items[row]

            def cell(c):
                w = self.tbl_items.item(row, c)
                return w.text().strip() if w else ""

            item.tunnel_code    = cell(5)
            item.packaging_type = cell(6)

            try:
                item.packaging_count = int(cell(7)) if cell(7) else item.packaging_count
            except ValueError:
                pass
            try:
                item.net_quantity  = float(cell(8)) if cell(8) else 0.0
            except ValueError:
                pass
            if cell(9) in UNIT_OPTIONS:
                item.unit = cell(9)

    def export_pdf(self):
        """Yazdır önizlemesiyle birebir aynı PDF çıktısı oluşturur."""
        self._sync_items_from_table()
        from datetime import datetime as _dt
        default_name = f"ADR_{_dt.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        path, _ = QFileDialog.getSaveFileName(self, "PDF Kaydet",
            default_name, "PDF (*.pdf)")
        if not path:
            return
        try:
            html = self._build_print_html()
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(path)
            page_layout = QPageLayout(
                QPageSize(QPageSize.PageSizeId.A4),
                QPageLayout.Orientation.Portrait,
                QMarginsF(15, 15, 15, 15),
                QPageLayout.Unit.Millimeter
            )
            printer.setPageLayout(page_layout)
            doc = QTextDocument()
            doc.setHtml(html)
            doc.setPageSize(printer.pageRect(QPrinter.Unit.Point).size())
            doc.print(printer)
            QMessageBox.information(self, "Başarılı", f"PDF oluşturuldu:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"PDF oluşturma hatası:\n{str(e)}")



    def _generate_pdf(self, path: str):
        """
        PDF üretici — HTML tabanlı (QTextDocument + QPrinter).
        Eski reportlab implementasyonu kaldırıldı; bu yöntem her ortamda çalışır.
        """
        self._sync_items_from_table()
        html = self._build_print_html()
        try:
            from PyQt6.QtPrintSupport import QPrinter
            from PyQt6.QtGui import QTextDocument
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(path)
            printer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
            doc = QTextDocument()
            doc.setHtml(html)
            doc.setPageSize(printer.pageRect(QPrinter.Unit.Point).size())
            doc.print(printer)
        except Exception as e:
            QMessageBox.critical(self, "PDF Hatası", f"PDF oluşturma hatası:\n{str(e)}")
            return
        # Eski reportlab kodu kaldırıldı — artık burada yok.
        # Yerini alan HTML tabanlı yöntem üstteki try bloğunda.
        return

    def _generate_pdf_LEGACY_UNUSED(self, path: str):
        """
        [KULLANILMIYOR] — reportlab tabanlı eski implementasyon.
        Referans olarak saklandı; çağrılmaz.
        """
        self._sync_items_from_table()

        if not REPORTLAB_AVAILABLE:
            QMessageBox.warning(self, "Eksik Modül", "reportlab modülü yüklü değil.")
            return

        doc = SimpleDocTemplate(path, pagesize=A4,
            rightMargin=15*mm, leftMargin=15*mm,
            topMargin=15*mm, bottomMargin=15*mm)

        styles   = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle(
            'ADRTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=HexColor("#1E3A5F"),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph("ADR TEHLIKELI MADDE TASIMA EVRAKI", title_style))
        elements.append(Spacer(1, 10))

        info_data = [
            ["Evrak No:", self.txt_doc_no.text(), "Tarih:",
             self.date_doc.date().toString("dd.MM.yyyy")]
        ]
        info_table = Table(info_data, colWidths=[30*mm, 60*mm, 20*mm, 50*mm])
        info_table.setStyle(TableStyle([
            ('FONTNAME',      (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, -1), 10),
            ('ALIGN',         (0, 0), (0, -1), 'LEFT'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 10))

        sender   = self.get_selected_sender()
        receiver = self.get_selected_receiver()

        company_data = [
            ["GONDERICI:", "ALICI:"],
            [sender.name    if sender   else "", receiver.name    if receiver else ""],
            [sender.address if sender   else "", receiver.address if receiver else ""],
            [f"{sender.city   if sender   else ''} / {sender.district   if sender   else ''}",
             f"{receiver.city if receiver else ''} / {receiver.district if receiver else ''}"],
        ]
        company_table = Table(company_data, colWidths=[80*mm, 80*mm])
        company_table.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0), HexColor("#1E3A5F")),
            ('TEXTCOLOR',     (0, 0), (-1, 0), white),
            ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, 0), 11),
            ('ALIGN',         (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
            ('GRID',          (0, 0), (-1, -1), 0.5, HexColor("#CCCCCC")),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING',    (0, 0), (-1, -1), 8),
            ('LEFTPADDING',   (0, 0), (-1, -1), 6),
        ]))
        elements.append(company_table)
        elements.append(Spacer(1, 10))

        if self.items:
            item_headers = ["UN No", "Teknik Isim", "Sinif", "PG", "Tunel kodu",
                            "Ambalaj turu", "Adet", "Net", "Birim"]
            item_data = [item_headers]

            for item in self.items:
                row = [
                    f"UN{item.un_number}", item.proper_name, item.class_code,
                    item.packing_group, item.tunnel_code or "-",
                    item.packaging_type,
                    str(item.packaging_count), str(item.net_quantity), item.unit
                ]
                item_data.append(row)

            item_table = Table(item_data,
                colWidths=[16*mm, 50*mm, 11*mm, 9*mm, 12*mm, 22*mm, 11*mm, 14*mm, 12*mm])
            item_table.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0), (-1, 0), HexColor("#555555")),
                ('TEXTCOLOR',     (0, 0), (-1, 0), white),
                ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0), (-1, 0), 9),
                ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE',      (0, 1), (-1, -1), 9),
                ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN',         (1, 0), (1, -1), 'LEFT'),
                ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID',          (0, 0), (-1, -1), 0.5, HexColor("#CCCCCC")),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING',    (0, 0), (-1, -1), 6),
                ('LEFTPADDING',   (0, 0), (-1, -1), 4),
                ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
            ]))
            elements.append(item_table)
            elements.append(Spacer(1, 10))

        driver  = self.get_selected_driver()
        vehicle = self.get_selected_vehicle()
        report  = ADREngine.generate_adr_report(self.items, driver, vehicle)

        # SRC5 Şöför Muafiyeti Değerlendirmesi
        driver_local = self.get_selected_driver()
        src5_status = "—"
        src5_color  = HexColor("#555555")
        if driver_local:
            has_src5 = bool(driver_local.src5_no and driver_local.src5_no.strip())
            if has_src5:
                # SRC5 geçerlilik süresi kontrol
                try:
                    from datetime import datetime, timedelta
                    exp = datetime.strptime(driver_local.src5_expiry, "%Y-%m-%d")
                    if exp < datetime.now():
                        src5_status = f"GEÇERSİZ — Süre dolmuş ({driver_local.src5_expiry})"
                        src5_color  = HexColor("#CC0000")
                    elif exp < datetime.now() + timedelta(days=60):
                        src5_status = f"UYARI — Yakında bitiyor ({driver_local.src5_expiry})"
                        src5_color  = HexColor("#CC7700")
                    else:
                        # SRC5 1.1.3.1 muafiyet koşulunu değerlendir
                        all_lq_or_eq = all(i.is_lq or i.is_eq for i in self.items) if self.items else False
                        below_1136   = report.total_points <= MAX_1136_POINTS
                        if all_lq_or_eq:
                            src5_status = f"GEÇERLİ ({driver_local.src5_no}) — Tüm kalemler LQ/EQ: Şöför ADR muafiyeti uygulanabilir (3.4 / 3.5)"
                        elif below_1136 and not report.orange_plate_required:
                            src5_status = f"GEÇERLİ ({driver_local.src5_no}) — 1.1.3.6 muafiyeti kapsamında şöför ADR belgesi gerekmez"
                        else:
                            src5_status = f"GEÇERLİ ({driver_local.src5_no}) — Tam ADR uygulaması: şöför ADR sertifikası zorunlu"
                        src5_color = HexColor("#1a7a1a")
                except Exception:
                    src5_status = f"GEÇERLİ — SRC5: {driver_local.src5_no}"
                    src5_color  = HexColor("#1a7a1a")
            else:
                src5_status = "SRC5 belgesi girilmemiş"
                src5_color  = HexColor("#CC0000")
        else:
            src5_status = "Sürücü seçilmemiş"

        adr_data = [
            ["ADR KONTROL ÖZETİ", ""],
            ["Turuncu Plaka:",    "ZORUNLU ⚠" if report.orange_plate_required else "Gerekmez ✓"],
            ["Muafiyet:",          report.exemption_type],
            ["Yazılı Talimat:",   "ZORUNLU" if report.written_instructions_required else "Gerekmez"],
            ["SRC5 / Şöför Muafiyeti:", src5_status],
        ]
        adr_table = Table(adr_data, colWidths=[50*mm, 110*mm])
        adr_table.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0), HexColor("#228B22")),
            ('TEXTCOLOR',     (0, 0), (-1, 0), white),
            ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('SPAN',          (0, 0), (-1, 0)),
            ('FONTNAME',      (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, -1), 10),
            ('ALIGN',         (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID',          (0, 0), (-1, -1), 0.5, HexColor("#CCCCCC")),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING',    (0, 0), (-1, -1), 8),
            ('LEFTPADDING',   (0, 0), (-1, -1), 6),
            # SRC5 satırı (son satır) — dinamik renk
            ('TEXTCOLOR',     (1, 4), (1, 4), src5_color),
            ('FONTNAME',      (1, 4), (1, 4), 'Helvetica-Bold'),
        ]))
        elements.append(adr_table)

        if self.txt_notes.toPlainText():
            elements.append(Spacer(1, 10))
            elements.append(Paragraph("NOTLAR:", styles['Heading3']))
            elements.append(Paragraph(self.txt_notes.toPlainText(), styles['Normal']))

        doc.build(elements)

    def export_excel(self):
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "Hata", "openpyxl modulu eksik!\nKurulum: pip install openpyxl")
            return

        path, _ = QFileDialog.getSaveFileName(self, "Excel Kaydet",
            f"ADR_{self.txt_doc_no.text()}.xlsx", "Excel (*.xlsx)")
        if not path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "ADR Evrak"

            ws['A1'] = "ADR TEHLIKELI MADDE TASIMA EVRAKI"
            ws['A1'].font = Font(size=16, bold=True, color="1E3A5F")
            ws.merge_cells('A1:H1')

            ws['A3'] = "Evrak No:"
            ws['A3'].font = Font(bold=True)
            ws['B3'] = self.txt_doc_no.text()
            ws['D3'] = "Tarih:"
            ws['D3'].font = Font(bold=True)
            ws['E3'] = self.date_doc.date().toString("dd.MM.yyyy")

            sender   = self.get_selected_sender()
            receiver = self.get_selected_receiver()

            ws['A5'] = f"GONDERICI: {sender.name if sender else ''}"
            ws['A5'].font = Font(bold=True, color="FFFFFF")
            ws['A5'].fill = PatternFill(start_color="1E3A5F", fill_type="solid")
            ws.merge_cells('A5:H5')

            ws['A6'] = f"ALICI: {receiver.name if receiver else ''}"
            ws['A6'].font = Font(bold=True, color="FFFFFF")
            ws['A6'].fill = PatternFill(start_color="2D5F8B", fill_type="solid")
            ws.merge_cells('A6:H6')

            headers = ["UN No", "Teknik Isim", "Sinif", "PG", "Ambalaj turu",
                       "Adet", "Net Miktar", "Birim", "Tunel", "LQ/EQ"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=8, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="555555", fill_type="solid")

            for i, item in enumerate(self.items, 9):
                ws.cell(row=i, column=1, value=f"UN{item.un_number}")
                ws.cell(row=i, column=2, value=item.proper_name)
                ws.cell(row=i, column=3, value=item.class_code)
                ws.cell(row=i, column=4, value=item.packing_group)
                ws.cell(row=i, column=5, value=item.packaging_type)
                ws.cell(row=i, column=6, value=item.packaging_count)
                ws.cell(row=i, column=7, value=item.net_quantity)
                ws.cell(row=i, column=8, value=item.unit)
                ws.cell(row=i, column=9, value=item.tunnel_code or "-")
                flags = []
                if item.is_lq: flags.append("LQ")
                if item.is_eq: flags.append("EQ")
                ws.cell(row=i, column=10, value="/".join(flags) or "-")

            # ---- [v4.4] ADR kontrol özeti (PDF ile aynı bilgi seti) ----
            driver  = self.get_selected_driver()
            vehicle = self.get_selected_vehicle()
            packaging_types = [item.packaging_type for item in self.items]
            report  = ADREngine.generate_adr_report(self.items, driver, vehicle, packaging_types)

            r = 10 + len(self.items)
            ws.cell(row=r, column=1, value="ADR KONTROL ÖZETİ").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=r, column=1).fill = PatternFill(start_color="1E3A5F", fill_type="solid")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
            summary = [
                ("1.1.3.6 Puanı", f"{report.total_points:.0f} / {MAX_1136_POINTS}"),
                ("Turuncu Plaka", "GEREKLİ" if report.orange_plate_required else "GEREKMEZ"),
                ("Yazılı Talimat", "ZORUNLU" if report.written_instructions_required else "GEREKMEZ"),
                ("Tünel Kısıtı", getattr(report, "tunnel_code", "") or "-"),
                ("Muafiyet", report.exemption_type or "-"),
                ("Sürücü", (driver.full_name if driver else "-")),
                ("Araç", (getattr(vehicle, "plate", "") if vehicle else "-")),
                ("Evrak Durumu", self.lbl_status.text() or "TASLAK"),
            ]
            for off, (k, v) in enumerate(summary, start=1):
                ws.cell(row=r+off, column=1, value=k).font = Font(bold=True)
                ws.cell(row=r+off, column=2, value=v)

            # Kritik hatalar Excel'de de görünsün
            if getattr(report, "errors", None):
                er = r + len(summary) + 2
                ws.cell(row=er, column=1, value="KRİTİK UYARILAR").font = Font(bold=True, color="FFFFFF")
                ws.cell(row=er, column=1).fill = PatternFill(start_color="B91C1C", fill_type="solid")
                ws.merge_cells(start_row=er, start_column=1, end_row=er, end_column=10)
                for off, (_, msg) in enumerate(report.errors, start=1):
                    ws.cell(row=er+off, column=1, value=msg).font = Font(color="B91C1C")
                    ws.merge_cells(start_row=er+off, start_column=1, end_row=er+off, end_column=10)

            # Onaylanmamış evrak uyarısı
            if "ONAYLANDI" not in (self.lbl_status.text() or "").upper():
                ws['A2'] = "TASLAK - Onaylanmamış evrak, resmi kullanım için geçerli değildir"
                ws['A2'].font = Font(bold=True, color="B91C1C")
                ws.merge_cells('A2:J2')

            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 6
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 8
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 8
            ws.column_dimensions['I'].width = 8
            ws.column_dimensions['J'].width = 8

            wb.save(path)
            QMessageBox.information(self, "Basarili", f"Excel olusturuldu:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel olusturma hatasi:\n{str(e)}")

    # =========================================================================
    # YAZDIR - PROFESYONEL HTML ONIZLEME  (v4.1 - yeniden yazildi)
    # =========================================================================

    def print_document(self):
        """Profesyonel önizleme penceresini açar.
        Yazdırma sadece net miktar girilmemişse engellenir.
        TC=4 / LQ / EQ → sınırsız taşıma → yazdırılır.
        pts > 1000 uyarısı → ADR kontrol merkezinde gösterilir, yazdırmayı engellemez."""
        self._sync_items_from_table()

        aktif_items = [i for i in self.items if (i.net_quantity or 0) > 0]
        if not aktif_items:
            QMessageBox.warning(self, "Yazdırma Engellendi",
                "Evrakta hiç ürün yok veya tüm ürünlerin net miktarı 0.\n"
                "Lütfen ürün ekleyip miktarları girin.")
            return

        html = self._build_print_html()
        dlg = PrintPreviewDialog(self, html, title="ADR Taşıma Evrakı — Önizleme / Yazdır")
        dlg.exec()

    @staticmethod
    def _build_letterhead_watermark_b64(logo_b64: str, is_approved: bool,
                                        w: int = 794, h: int = 1120) -> str:
        """Antetli kagit gorunumu icin tek bir arka plan PNG'i uretir:
          - Firma logosu, sayfa ortasinda cok soluk (%8 opaklik) filigran.
          - Onaylanmamis evrakta (TASLAK/DOGRULAMA HATALI) kirmizi, 30 derece
            egik "TASLAK" yazisi ayni goruntuye bindirilir.
        Qt'nin zengin metin motoru CSS opacity/transform desteklemedigi icin
        (test edildi: yalnizca <table>/<td> uzerinde background-image calisir,
        opacity/rotate calismaz) efekt onceden Pillow ile goruntuye islenir.
        Logo yoksa ve evrak onayliysa None doner (arka plan eklenmez)."""
        if not logo_b64 and is_approved:
            return ""
        try:
            from PIL import Image, ImageDraw, ImageFont
            import base64 as _b64, io as _io
        except ImportError:
            return ""  # Pillow kurulu degilse filigransiz devam et

        canvas = Image.new("RGBA", (w, h), (0, 0, 0, 0))

        if logo_b64:
            try:
                raw = _b64.b64decode(logo_b64)
                logo = Image.open(_io.BytesIO(raw)).convert("RGBA")
                target_w = int(w * 0.55)
                ratio = target_w / logo.width if logo.width else 1
                target_h = max(1, int(logo.height * ratio))
                logo = logo.resize((target_w, target_h), Image.LANCZOS)
                r, g, b, a = logo.split()
                a = a.point(lambda px: int(px * 0.08))  # %8 opaklik
                logo.putalpha(a)
                pos = ((w - target_w) // 2, (h - target_h) // 2)
                canvas.alpha_composite(logo, pos)
            except Exception:
                logging.getLogger(__name__).warning(
                    "Antet logosu islenemedi", exc_info=True)

        if not is_approved:
            txt_layer = Image.new("RGBA", (w, h), (0, 0, 0, 0))
            d = ImageDraw.Draw(txt_layer)
            try:
                font = ImageFont.truetype(
                    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 90)
            except Exception:
                font = ImageFont.load_default()
            text = "TASLAK"
            bbox = d.textbbox((0, 0), text, font=font)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
            d.text(((w - tw) / 2 - bbox[0], (h - th) / 2 - bbox[1]),
                   text, font=font, fill=(200, 30, 30, 70))
            txt_layer = txt_layer.rotate(30, expand=False, resample=Image.BICUBIC)
            canvas.alpha_composite(txt_layer)

        buf = _io.BytesIO()
        canvas.save(buf, format="PNG")
        return _b64.b64encode(buf.getvalue()).decode()

    def _build_print_html(self) -> str:
        """
        Profesyonel A4 tek sayfa ADR taşıma evrakı.
        - Gönderici ve Alıcı: ayrı çerçeveli kutucuklar (yan yana)
        - İmza/Kaşe: Gönderici + Sürücü (2 kutu, yalnızca Ad Soyad)
        - ADR uyumluluk özeti compact tek satır şerit
        - Ürün tablosu compact — A4'e sığacak şekilde
        - vehicle.carrier_name hatası düzeltildi (hasattr güvenli)
        - Gözü yormayan beyaz/gri ton, minimal çizgi
        """
        import html as _h
        from datetime import timedelta
        import qrcode  # <--- BURADA OLMALI
        import io      # <--- BURADA OLMALI
        import base64
        
        
        def esc(v):
            return _h.escape(str(v)) if v else ""

        sender   = self.get_selected_sender()
        receiver = self.get_selected_receiver()
        driver   = self.get_selected_driver()
        vehicle  = self.get_selected_vehicle()
        packaging_types = [item.packaging_type for item in self.items]
        report   = ADREngine.generate_adr_report(self.items, driver, vehicle, packaging_types)
        doc_info = self.get_document_info()
        
       


        logo_b64 = self.db.get_company_logo_b64()
        show_qr = self.db.get_setting("doc_show_qr") == "1" # QR ayarını çektik

        # [v4.2] Antetli kagit filigrani: firma logosu (soluk) + onaylanmamis
        # evrakta capraz TASLAK yazisi. Tek PNG olarak onceden islenir.
        _status_text = self.lbl_status.text().upper() if hasattr(self, "lbl_status") else ""
        is_approved = "ONAYLANDI" in _status_text
        letterhead_b64 = self._build_letterhead_watermark_b64(logo_b64, is_approved)

        logo_html = ""
        if logo_b64:
            logo_html = f'<img src="data:image/png;base64,{logo_b64}" style="max-width:100px; max-height:45px; object-fit:contain;">'

        qr_html = ""
        if show_qr:
            # ── vCard kartvizit verisi oluştur ──────────────────────────────
            co_name  = self.db.get_setting("doc_company_name")  or ""
            co_addr  = self.db.get_setting("doc_company_address") or ""
            co_phone = self.db.get_setting("doc_company_phone") or ""
            co_email = self.db.get_setting("doc_company_email") or ""
            co_web   = self.db.get_setting("doc_company_website") or ""

            vcard_lines = [
                "BEGIN:VCARD",
                "VERSION:3.0",
                f"FN:{co_name}",
                f"ORG:{co_name}",
            ]
            if co_addr:
                vcard_lines.append(f"ADR;TYPE=WORK:;;{co_addr};;;;TR")
            # Birden fazla numara varsa (örn. "0850 515 0000 - 0543 271 63 77") her birini ekle
            phone_parts = [p.strip() for p in co_phone.replace("–", "-").split("-") if p.strip()]
            for ph in phone_parts:
                vcard_lines.append(f"TEL;TYPE=WORK,VOICE:{ph}")
            if co_email:
                vcard_lines.append(f"EMAIL;TYPE=WORK:{co_email}")
            if co_web:
                vcard_lines.append(f"URL:{co_web}")
            vcard_lines.append("NOTE:ADR Tehlikeli Madde Danışmanlık")
            vcard_lines.append("END:VCARD")
            vcard_data = "\r\n".join(vcard_lines)

            qr = qrcode.QRCode(
                version=None,
                error_correction=qrcode.constants.ERROR_CORRECT_M,
                box_size=4,
                border=1,
            )
            qr.add_data(vcard_data)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")
            buffer = io.BytesIO()
            qr_img.save(buffer, format="PNG")
            qr_b64 = base64.b64encode(buffer.getvalue()).decode()
            qr_html = (
                f'<img src="data:image/png;base64,{qr_b64}" '
                f'width="55" height="55" '
                f'style="width:55px;height:55px;display:block;" '
                f'title="Firma Kartviziti">'
            )
            
        logo_html = ""

        if logo_b64:
            # max-width ve max-height ile logonun alanı aşması engellenir
            logo_html = f'''
            <img src="data:image/png;base64,{logo_b64}" 
                 style="max-width:60px; max-height:20px; width:auto; height:auto; object-fit:contain;">
            '''
            logo_html = f"""
            <img src="data:image/png;base64,{logo_b64}"
                style="
                    max-width:100px;
                    max-height:45px;
                    object-fit:contain;
                ">
            """

        # ── Renk sabitleri (minimal, gözü yormayan palet) ─────────────────
        NAVY      = "#1C3557"
        NAVY_LITE = "#EEF3FA"
        STEEL     = "#4A6FA5"
        RULE      = "#C8D6E8"
        TXT_MAIN  = "#1A1A2E"
        TXT_MUTED = "#5A6478"
        TXT_LITE  = "#8A94A6"
        ROW_ALT   = "#F7F9FC"
        GREEN     = "#0A6B2E"
        GREEN_BG  = "#E8F5EE"
        RED       = "#B91C1C"
        RED_BG    = "#FEF2F2"
        AMBER     = "#92400E"
        AMBER_BG  = "#FFFBEB"
        WARN_BG   = "#FEF3C7"

        # ── Muafiyet / puan şeridi hesaplamaları ─────────────────────────
        aktif_html = [i for i in self.items if (i.net_quantity or 0) > 0]
        tc4_only = aktif_html and all(
            str(getattr(i, "transport_category", "")).strip().split()[0:1] == ["4"]
            or i.is_lq or i.is_eq
            for i in aktif_html
        )
        has_exemption = (report.exemption_type != ExemptionType.NONE.value) or tc4_only

        # Puan şeridi gösterimi: TC=4/LQ/EQ → "Sınırsız"
        # pts > 1000 → şerit tamamen gizlenir (has_exemption=False)
        if tc4_only:
            pts_display = "Sınırsız"
            pts_sub     = "ADR 3.4 — Miktar Muafiyeti Sınırsız"
            bar_pct     = 0
            bar_color   = GREEN
        elif report.total_points > MAX_1136_POINTS:
            # 1000 puan aşıldı → şeridi gizle
            has_exemption = False
            pts_display = ""
            pts_sub     = ""
            bar_pct     = 0
            bar_color   = TXT_MUTED
        elif report.total_points > 0:
            pts_display = f"{report.total_points:.0f}"
            pts_sub     = f"{report.total_points:.0f} / {MAX_1136_POINTS} puan"
            bar_pct     = min(100, int(report.total_points / MAX_1136_POINTS * 100))
            bar_color   = RED if report.orange_plate_required else GREEN
        else:
            pts_display = "—"
            pts_sub     = "Taşıma Kategorisi girilmemiş"
            bar_pct     = 0
            bar_color   = TXT_MUTED

        # ── Emniyet planı gerekli mi? ──────────────────────────────────────
        try:
            _sp_items = [i for i in self.items if (i.net_quantity or 0) > 0]
            if not _sp_items:
                security_plan_required = False
                security_plan_reasons  = []
            else:
                _sp_pts, _, _ = ADREngine.calculate_1136_points(_sp_items)
                _sp_result = SecurityPlanEngine.check(_sp_items, total_1136_points=_sp_pts)
                security_plan_required = _sp_result.get("required", False)
                security_plan_reasons  = _sp_result.get("reasons", [])
        except Exception:
            security_plan_required = False
            security_plan_reasons  = []

        # ── ADR özet değerleri ────────────────────────────────────────────
        op_color = RED    if report.orange_plate_required else GREEN
        op_bg    = RED_BG if report.orange_plate_required else GREEN_BG
        op_text  = "ZORUNLU ⚠" if report.orange_plate_required else "GEREKMEZ ✓"

        wi_color = AMBER    if report.written_instructions_required else GREEN
        wi_bg    = AMBER_BG if report.written_instructions_required else GREEN_BG
        wi_text  = "ZORUNLU" if report.written_instructions_required else "GEREKMEZ ✓"

        ac_color = AMBER    if report.driver_adr_required else GREEN
        ac_bg    = AMBER_BG if report.driver_adr_required else GREEN_BG
        ac_text  = "ZORUNLU" if report.driver_adr_required else "GEREKMEZ ✓"

        bar_pct   = min(100, int(report.total_points / MAX_1136_POINTS * 100))
        bar_color = RED if report.orange_plate_required else GREEN

        # ── SRC5 durumu ───────────────────────────────────────────────────
        src5_text  = "Sürücü seçilmemiş"
        src5_color = TXT_MUTED
        if driver:
            has_src5 = bool(getattr(driver, 'src5_no', None) and driver.src5_no.strip())
            if has_src5:
                try:
                    exp = datetime.strptime(driver.src5_expiry, "%Y-%m-%d")
                    if exp < datetime.now():
                        src5_text  = f"GEÇERSİZ — Süre Dolmuş ({driver.src5_expiry})"
                        src5_color = RED
                    elif exp < datetime.now() + timedelta(days=60):
                        src5_text  = f"UYARI — {(exp - datetime.now()).days} gün kaldı ({driver.src5_expiry})"
                        src5_color = AMBER
                    else:
                        all_lq_eq  = all(i.is_lq or i.is_eq for i in self.items) if self.items else False
                        below_1136 = report.total_points <= MAX_1136_POINTS
                        if all_lq_eq:
                            src5_text = f"GEÇERLİ — LQ/EQ muafiyeti (ADR 3.4/3.5)"
                        elif below_1136 and not report.orange_plate_required:
                            src5_text = f"GEÇERLİ — 1.1.3.6 muafiyeti kapsamında ADR belgesi gerekmez"
                        else:
                            src5_text = f"GEÇERLİ — Tam ADR: Sürücü ADR sertifikası zorunlu"
                        src5_color = GREEN
                except Exception:
                    src5_text  = f"GEÇERLİ — SRC5: {getattr(driver, 'src5_no', '')}"
                    src5_color = GREEN
            else:
                src5_text  = "SRC5 belgesi girilmemiş"
                src5_color = RED

        # ── Ürün satırları (compact) ──────────────────────────────────────
        item_rows = ""
        for idx, item in enumerate(self.items, 1):
            bg = "#FFFFFF" if idx % 2 == 1 else ROW_ALT
            class_bg, class_fg = ADREngine.get_class_color(item.class_code)
            badges = ""
            if item.is_lq:
                badges += ('<span style="border:1px solid #166534;color:#166534;padding:1px 4px;'
                           'border-radius:2px;font-size:6.5pt;margin-left:3px;">LQ</span>')
            if item.is_eq:
                badges += ('<span style="border:1px solid #1e3a8a;color:#1e3a8a;padding:1px 4px;'
                           'border-radius:2px;font-size:6.5pt;margin-left:2px;">EQ</span>')
            item_rows += f"""
            <tr style="background:{bg};">
              <td style="text-align:center;color:{TXT_LITE};font-size:7.5pt;width:3%;">{idx}</td>
              <td style="color:{NAVY};white-space:nowrap;width:8%;font-size:8pt;">UN&nbsp;{esc(item.un_number)}</td>
              <td style="word-break:break-word;width:28%;font-size:8pt;">{esc(item.proper_name)}{badges}</td>
              <td style="text-align:center;width:6%;">
                <span style="border:1px solid {class_fg};color:{class_fg};padding:1px 5px;
                  border-radius:2px;font-size:7.5pt;">{esc(item.class_code)}</span>
              </td>
              <td style="text-align:center;width:4%;font-size:8pt;">{esc(item.packing_group) or "—"}</td>
              <td style="text-align:center;width:5%;font-size:8pt;color:{NAVY};">{esc(item.tunnel_code) or "—"}</td>
              <td style="font-size:7.5pt;width:14%;color:{TXT_MUTED};">{esc(item.packaging_type) or "—"}</td>
              <td style="text-align:center;width:5%;font-size:8pt;">{item.packaging_count}</td>
              <td style="text-align:right;white-space:nowrap;width:9%;font-size:8pt;">{item.net_quantity}&nbsp;{esc(item.unit)}</td>
            </tr>"""

        if not item_rows:
            item_rows = (f'<tr><td colspan="9" style="text-align:center;color:{TXT_LITE};'
                         f'padding:14px;font-size:8.5pt;">Ürün eklenmemiş</td></tr>')

        # DÜZELTME: GERÇEK motor sonucu burada (self.db erişimi olan bu
        # noktada) hesaplanıp report.compatibility_errors'a yerleştiriliyor
        # — YAZDIRILAN belge artık gerçek bir ADR referansı gösteriyor.
        report.compatibility_errors = _gercek_karisik_yukleme_kontrolu(self.db, self.items)

        # ── Uyumsuzluk uyarısı (varsa) ────────────────────────────────────
        compat_html = ""
        if report.compatibility_errors:
            errs = "".join(f'<li style="margin:2px 0;font-size:8pt;">{esc(e)}</li>'
                           for e in report.compatibility_errors)
            compat_html = f"""
            <div style="margin:6px 0 0;padding:6px 10px;
              border-left:3px solid {RED};border-radius:0 3px 3px 0;">
              <span style="color:{RED};font-size:7.5pt;">⚠ UYUMSUZLUK UYARILARI</span>
              <ul style="margin:3px 0 0 14px;padding:0;color:{RED};">{errs}</ul>
            </div>"""

        # ── Notlar ────────────────────────────────────────────────────────
        notes_html = ""
        notes_text = self.txt_notes.toPlainText().strip()
        if notes_text:
            notes_html = f"""
            <div style="margin-top:8px;padding:6px 10px;
              border-left:3px solid #D97706;border-radius:0 3px 3px 0;page-break-inside:avoid;">
              <span style="color:#92400E;font-size:7.5pt;">NOT:</span>
              <span style="color:{TXT_MUTED};font-size:8pt;"> {esc(notes_text)}</span>
            </div>"""

        # ── İmza / Kaşe — 2 kutu: Gönderici + Sürücü ────────────────────
        sender_name  = esc(sender.name)  if sender  else "—"
        driver_name  = esc(driver.full_name) if driver else "—"

        signature_section = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:8px 0;
          margin-top:5px;page-break-inside:avoid;">
          <tr>
            <!-- GÖNDERİCİ İMZA KUTUSU -->
            <td style="width:50%;vertical-align:top;">
              <div style="border:1px solid {RULE};border-radius:4px;overflow:hidden;">
                <div style="border-bottom:1px solid {RULE};padding:4px 8px;
                  font-size:6.5pt;letter-spacing:0.8px;color:{NAVY};">
                  GÖNDERİCİ / YÜKLETEN
                </div>
                <div style="padding:8px 10px;">
                  <div style="font-size:7pt;color:{TXT_LITE};margin-bottom:1px;">Firma:</div>
                  <div style="font-size:8.5pt;color:{TXT_MAIN};margin-bottom:10px;">
                    {sender_name}
                  </div>
                  <div style="font-size:7pt;color:{TXT_LITE};margin-bottom:2px;">Ad Soyad:</div>
                  <div style="border-bottom:1px solid {RULE};min-height:22px;margin-bottom:10px;"></div>
                  <div style="font-size:7pt;color:{TXT_LITE};margin-bottom:2px;">İmza / Kaşe:</div>
                  <div style="border:1px solid {RULE};min-height:80px;border-radius:2px;margin-bottom:10px;"></div>
                  <div style="font-size:6pt;color:{TXT_LITE};line-height:1.4;border-top:1px solid {RULE};padding-top:4px;">
                    Tehlikeli maddelerin sınıflandırılması, paketlenmesi ve etiketlenmesinin
                    ADR hükümlerine uygun olduğunu beyan ederim. (ADR 5.4.1.1.1/f)
                  </div>
                </div>
              </div>
            </td>
            <!-- SÜRÜCÜ İMZA KUTUSU -->
            <td style="width:50%;vertical-align:top;">
              <div style="border:1px solid {RULE};border-radius:4px;overflow:hidden;">
                <div style="border-bottom:1px solid {RULE};padding:4px 8px;
                  font-size:6.5pt;letter-spacing:0.8px;color:{NAVY};">
                  SÜRÜCÜ
                </div>
                <div style="padding:8px 10px;">
                  <div style="font-size:7pt;color:{TXT_LITE};margin-bottom:1px;">Sürücü:</div>
                  <div style="font-size:8.5pt;color:{TXT_MAIN};margin-bottom:10px;">
                    {driver_name}
                  </div>
                  <div style="font-size:7pt;color:{TXT_LITE};margin-bottom:2px;">Ad Soyad:</div>
                  <div style="border-bottom:1px solid {RULE};min-height:22px;margin-bottom:10px;"></div>
                  <div style="font-size:7pt;color:{TXT_LITE};margin-bottom:2px;">İmza:</div>
                  <div style="border:1px solid {RULE};min-height:80px;border-radius:2px;margin-bottom:10px;"></div>
                  <div style="font-size:6pt;color:{TXT_LITE};line-height:1.4;border-top:1px solid {RULE};padding-top:4px;">
                    Yükü teslim aldığımı ve taşımanın ADR hükümlerine uygun olarak
                    gerçekleştirileceğini kabul ederim.
                  </div>
                </div>
              </div>
            </td>
          </tr>
        </table>"""

        # ── HTML belgesi ──────────────────────────────────────────────────
        html = f"""<!DOCTYPE html>
<html lang="tr">
<head>
<meta charset="UTF-8">
<style>
  @page {{ size: A4 portrait; margin: 8mm 10mm; }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: 'Segoe UI', Arial, Helvetica, sans-serif;
    font-size: 7.5pt;
    color: {TXT_MAIN};
    background: #ffffff;
    line-height: 1.3;
  }}
  table {{ border-collapse: collapse; width: 100%; }}
  .rule {{ border-top: 1px solid {RULE}; margin: 5px 0; }}
  .sec-head {{
    font-size: 6.5pt;
    font-weight: normal;
    letter-spacing: 1.2px;
    text-transform: uppercase;
    color: {NAVY};
    border-bottom: 1.5px solid {NAVY};
    padding-bottom: 2px;
    margin: 5px 0 4px;
  }}
  .items-table th {{
    background: #ffffff;
    color: {NAVY};
    border-bottom: 1.5px solid {NAVY};
    font-size: 6.5pt;
    text-align: center;
    padding: 3px 4px;
    white-space: nowrap;
    font-weight: normal;
    letter-spacing: 0.3px;
  }}
  .items-table td {{
    border-bottom: 1px solid {RULE};
    padding: 2px 4px;
    vertical-align: middle;
    font-weight: normal;
  }}
  .badge {{
    display: inline-block;
    padding: 1px 5px;
    border-radius: 3px;
    font-weight: normal;
    font-size: 7pt;
  }}
  .no-break {{ page-break-inside: avoid; }}
  strong {{ font-weight: normal; }}
  b {{ font-weight: normal; }}
  i, em {{ font-style: normal; }}
</style>
</head>
<body>
{'<table width="100%" style="border-collapse:separate;background-image:url(data:image/png;base64,' + letterhead_b64 + ');background-repeat:no-repeat;background-position:center top;"><tr><td style="padding:0;">' if letterhead_b64 else ''}

<!-- ══════════════════════════ BAŞLIK ══════════════════════════════════ -->
<table style="margin-bottom:5px;border-bottom:2px solid {NAVY};padding-bottom:5px; width:100%; table-layout:fixed;">
  <tr>
    <td style="width:20%; text-align:left; vertical-align:middle; padding:5px;">
        <div style="max-width:80px; overflow:hidden;">
            {logo_html}
        </div>
    </td>
    
    <td style="width:60%; text-align:center; padding-bottom:2px; vertical-align:middle;">
      <div style="font-size:12pt;color:{NAVY};letter-spacing:0.8px;">ADR TEHLİKELİ MADDE TAŞIMA EVRAKI</div>
      <div style="font-size:7pt;color:{TXT_MUTED};margin-top:2px;">Karayolu İle Tehlikeli Madde Taşıma Belgesi &nbsp;·&nbsp; ADR 5.4.1</div>
      <div style="font-size:7pt;color:{TXT_MUTED};margin-top:2px;">Tarih: {esc(doc_info.get('date',''))} &nbsp;|&nbsp; Evrak No: {esc(doc_info.get('document_no',''))}</div>
    </td>
    
    <td style="width:20%; text-align:right; vertical-align:middle; padding:5px;">
        <div style="display:inline-block; max-width:60px;">
            {qr_html}
        </div>
    </td>
  </tr>
</table>

<!-- ══════════════════════ GÖNDERİCİ / ALICI ══════════════════════════ -->
<div class="no-break">
<div class="sec-head">Gönderici ve Alıcı Bilgileri</div>
<table style="border-collapse:separate;border-spacing:6px 0;width:100%;">
  <tr>
    <!-- GÖNDERİCİ -->
    <td style="width:49%;vertical-align:top;border:1px solid {RULE};border-radius:4px;padding:0;overflow:hidden;">
      <div style="border-bottom:1px solid {RULE};padding:4px 8px;font-size:6.5pt;letter-spacing:0.8px;color:{NAVY};">
        GÖNDERİCİ
      </div>
      <div style="padding:6px 8px;">
        <div style="font-size:8.5pt;color:{TXT_MAIN};">
          {esc(sender.name) if sender else '—'}
        </div>
        <div style="font-size:7pt;color:{TXT_MUTED};margin-top:2px;">
          {esc(sender.address) if sender else ''}
          {(' &nbsp;·&nbsp; ' + esc(sender.city)) if sender and sender.city else ''}
          {(' / ' + esc(sender.district)) if sender and getattr(sender,'district','') else ''}
        </div>
        {('<div style="font-size:7pt;color:' + TXT_MUTED + ';margin-top:2px;">Tel: ' + esc(getattr(sender,'phone','')) + '</div>') if sender and getattr(sender,'phone','') else ''}
      </div>
    </td>
    <!-- ALICI -->
    <td style="width:49%;vertical-align:top;border:1px solid {RULE};border-radius:4px;padding:0;overflow:hidden;">
      <div style="border-bottom:1px solid {RULE};padding:4px 8px;font-size:6.5pt;letter-spacing:0.8px;color:{NAVY};">
        ALICI
      </div>
      <div style="padding:6px 8px;">
        <div style="font-size:8.5pt;color:{TXT_MAIN};">
          {esc(receiver.name) if receiver else '—'}
        </div>
        <div style="font-size:7pt;color:{TXT_MUTED};margin-top:2px;">
          {esc(receiver.address) if receiver else ''}
          {(' &nbsp;·&nbsp; ' + esc(receiver.city)) if receiver and receiver.city else ''}
          {(' / ' + esc(receiver.district)) if receiver and getattr(receiver,'district','') else ''}
        </div>
        {('<div style="font-size:7pt;color:' + TXT_MUTED + ';margin-top:2px;">Tel: ' + esc(getattr(receiver,'phone','')) + '</div>') if receiver and getattr(receiver,'phone','') else ''}
      </div>
    </td>
  </tr>
</table>
</div>

<!-- ══════════════════════ ARAÇ VE SÜRÜCÜ ══════════════════════════════ -->
<div class="no-break">
<div class="sec-head">Araç ve Sürücü Bilgileri</div>
<table style="border:1px solid {RULE};">
  <tr>
    <td style="padding:4px 8px;font-size:6.5pt;color:{NAVY};border-right:1px solid {RULE};width:10%;white-space:nowrap;">SÜRÜCÜ</td>
    <td style="padding:4px 8px;width:40%;font-size:8pt;border-right:1px solid {RULE};">
      {esc(driver.full_name) if driver else '—'}
      {(' &nbsp;·&nbsp; SRC5: ' + esc(driver.src5_no)) if driver and getattr(driver,'src5_no','') else ''}
    </td>
    <td style="padding:4px 8px;font-size:6.5pt;color:{NAVY};border-right:1px solid {RULE};width:10%;white-space:nowrap;">ARAÇ</td>
    <td style="padding:4px 8px;width:40%;font-size:8pt;">
      {esc(vehicle.plate) if vehicle else '—'}
    </td>
  </tr>
</table>
</div>

<!-- ══════════════════════ TAŞINAN MADDELER ════════════════════════════ -->
<div class="sec-head">Taşınan Tehlikeli Maddeler</div>
<table class="items-table">
  <thead>
    <tr>
      <th style="width:3%;">#</th>
      <th style="width:8%;">UN No</th>
      <th style="text-align:left;width:28%;">Uygun Sevkiyat Adı</th>
      <th style="width:6%;">Sınıf</th>
      <th style="width:4%;">PG</th>
      <th style="width:5%;">Tünel</th>
      <th style="width:13%;">Ambalaj Türü</th>
      <th style="width:5%;">Adet</th>
      <th style="width:13%;">Net Mik.</th>
    </tr>
  </thead>
  <tbody>
    {item_rows}
  </tbody>
</table>
{compat_html}

<!-- ══════════════════════ ADR UYUMLULUK ŞERİDİ (sadece muafiyet varsa) ═══════ -->
{"" if not has_exemption else f"""
<div class='no-break' style='margin-top:5px;'>
<table style='border:1px solid {RULE};border-radius:3px;overflow:hidden;'>
  <tr>
    <td style='padding:4px 8px;border-right:1px solid {RULE};width:22%;vertical-align:top;'>
      <div style='font-size:6pt;color:{NAVY};letter-spacing:0.5px;'>1.1.3.6 MİKTAR MUAFİYETİ</div>
      <div style='font-size:10pt;color:{bar_color};line-height:1.1;margin-top:1px;'>
        {pts_display}
      </div>
      <div style='font-size:6pt;color:{TXT_MUTED};margin-top:1px;'>{pts_sub}</div>
      <div style='background:#E4EAF2;border-radius:3px;height:4px;margin-top:3px;'>
        <div style='background:{bar_color};width:{bar_pct}%;height:4px;border-radius:3px;'></div>
      </div>
    </td>
    <td style='padding:4px 8px;border-right:1px solid {RULE};width:19%;vertical-align:top;'>
      <div style='font-size:6pt;color:{TXT_MUTED};'>TURUNCU PLAKA</div>
      <span class='badge' style='border:1px solid {op_color};color:{op_color};margin-top:2px;
        display:inline-block;font-size:7pt;'>{op_text}</span>
    </td>
    <td style='padding:4px 8px;border-right:1px solid {RULE};width:19%;vertical-align:top;'>
      <div style='font-size:6pt;color:{TXT_MUTED};'>YAZILI TALİMAT</div>
      <span class='badge' style='border:1px solid {wi_color};color:{wi_color};margin-top:2px;
        display:inline-block;font-size:7pt;'>{wi_text}</span>
    </td>
    <td style='padding:4px 8px;border-right:1px solid {RULE};width:19%;vertical-align:top;'>
      <div style='font-size:6pt;color:{TXT_MUTED};'>ADR SERTİFİKA</div>
      <span class='badge' style='border:1px solid {ac_color};color:{ac_color};margin-top:2px;
        display:inline-block;font-size:7pt;'>{ac_text}</span>
    </td>
    <td style='padding:4px 8px;width:21%;vertical-align:top;'>
      <div style='font-size:6pt;color:{TXT_MUTED};'>SRC5 / MUAFİYET</div>
      <div style='font-size:7pt;color:{src5_color};margin-top:2px;line-height:1.3;'>{esc(src5_text)}</div>
    </td>
  </tr>
  <tr>
    <td colspan='5' style='padding:3px 8px;border-top:1px solid {RULE};font-size:7pt;color:{TXT_MUTED};'>
      Muafiyet: {esc(report.exemption_type)}
    </td>
  </tr>
</table>
</div>
"""}

{notes_html}

<!-- ══════════════════════ EMNİYET PLANI UYARISI (koşullu) ═══════════ -->
{"" if not security_plan_required else f'''
<div class='no-break' style='margin-top:6px;padding:8px 12px;
  border-left:4px solid #B91C1C;border-radius:0 4px 4px 0;'>
  <div style='font-size:8pt;color:#B91C1C;margin-bottom:4px;'>
    ADR 1.10.3 — EMNİYET PLANI ZORUNLU
  </div>
  <div style='font-size:7.5pt;color:#7F1D1D;'>''' + ' &nbsp;·&nbsp; '.join(security_plan_reasons[:3]) + '''
  </div>
</div><br>
'''}

<!-- ══════════════════════ İMZA / KAŞE ════════════════════════════════ -->
<div class="sec-head" style="margin-top:6px;">İmza ve Kaşe</div>
{signature_section}

<!-- ══════════════════════ ALT BİLGİ ══════════════════════════════════ -->
<div style="margin-top:5px;padding:4px 10px;border:1px solid {RULE};
  border-radius:3px;font-size:6pt;color:{TXT_MUTED};text-align:center;">
  <br>
  Bu belge ADR Yönetmeliği Madde 5.4.1 kapsamında düzenlenmiştir.
  <br>
  <span style="font-size:5.5pt;color:{TXT_LITE};">
    ADR Transport Pro {APP_VERSION} &nbsp;|&nbsp; Oluşturma: {datetime.now().strftime('%d.%m.%Y %H:%M')}
  </span>
</div>

{'</td></tr></table>' if letterhead_b64 else ''}
</body>
</html>"""
        return html

    # --- DIALOGS ---

    def _add_new_sender(self):
        self._show_company_dialog("sender")

    def _add_new_receiver(self):
        self._show_company_dialog("receiver")

    def _add_new_carrier(self):
        self._show_company_dialog("carrier")

    def _show_company_dialog(self, company_type: str):
        type_names = {"sender": "Gönderici", "receiver": "Alıcı", "carrier": "Taşıyıcı"}
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Yeni {type_names.get(company_type, '')} Ekle")
        dialog.setMinimumWidth(450)
        layout = QFormLayout(dialog)

        txt_name       = QLineEdit()
        txt_address    = QLineEdit()
        txt_city       = QLineEdit()
        txt_district   = QLineEdit()
        txt_phone      = QLineEdit()
        txt_email      = QLineEdit()
        txt_contact    = QLineEdit()
        chk_favorite   = QCheckBox("Sık Kullanılan")

        layout.addRow("Firma Adı *:", txt_name)
        layout.addRow("Adres:",       txt_address)
        layout.addRow("Şehir:",       txt_city)
        layout.addRow("İlçe:",        txt_district)
        layout.addRow("Telefon:",     txt_phone)
        layout.addRow("E-posta:",     txt_email)
        layout.addRow("Yetkili:",     txt_contact)
        layout.addRow(chk_favorite)

        buttons = QHBoxLayout()
        btn_ok     = QPushButton("Kaydet")
        btn_ok.setObjectName("success")
        btn_ok.clicked.connect(dialog.accept)
        btn_cancel = QPushButton("Iptal")
        btn_cancel.clicked.connect(dialog.reject)
        buttons.addWidget(btn_ok)
        buttons.addWidget(btn_cancel)
        layout.addRow(buttons)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            if not txt_name.text().strip():
                QMessageBox.warning(self, "Hata", "Firma adi zorunlu!")
                return

            company = Company(
                type=company_type,
                name=txt_name.text().strip(),
                address=txt_address.text(),
                city=txt_city.text(),
                district=txt_district.text(),
                phone=txt_phone.text(),
                email=txt_email.text(),
                contact_person=txt_contact.text(),
                is_favorite=chk_favorite.isChecked()
            )

            self.db.add_company(company)
            self._load_combobox_data()

            combo = (self.cmb_sender   if company_type == "sender" else
                     self.cmb_receiver if company_type == "receiver" else
                     self.cmb_carrier)
            self._set_combobox_by_text(combo, company.name)

    def _set_combobox_by_text(self, combo: QComboBox, text: str):
        for i in range(combo.count()):
            if text in combo.itemText(i):
                combo.setCurrentIndex(i)
                return

    def _add_new_driver(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Yeni Surucu Ekle")
        dialog.setMinimumWidth(400)
        layout = QFormLayout(dialog)

        txt_name    = QLineEdit()
        txt_tc      = QLineEdit()
        txt_phone   = QLineEdit()
        txt_adr_no  = QLineEdit()
        date_adr    = QDateEdit()
        date_adr.setCalendarPopup(True)
        date_adr.setDate(QDate.currentDate().addYears(5))
        txt_src5    = QLineEdit()
        date_src5   = QDateEdit()
        date_src5.setCalendarPopup(True)
        date_src5.setDate(QDate.currentDate().addYears(5))
        txt_license = QLineEdit()
        date_license = QDateEdit()
        date_license.setCalendarPopup(True)
        date_license.setDate(QDate.currentDate().addYears(10))

        layout.addRow("Ad Soyad *:",       txt_name)
        layout.addRow("TC Kimlik No:",     txt_tc)
        layout.addRow("Telefon:",          txt_phone)
        layout.addRow("SRC5 No:",          txt_src5)
        layout.addRow("SRC5 Geçerlilik:", date_src5)
        layout.addRow("Ehliyet Sınıfı:",  txt_license)
        layout.addRow("Ehliyet Geçerlilik:", date_license)

        buttons = QHBoxLayout()
        btn_ok     = QPushButton("Kaydet")
        btn_ok.setObjectName("success")
        btn_ok.clicked.connect(dialog.accept)
        btn_cancel = QPushButton("Iptal")
        btn_cancel.clicked.connect(dialog.reject)
        buttons.addWidget(btn_ok)
        buttons.addWidget(btn_cancel)
        layout.addRow(buttons)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            if not txt_name.text().strip():
                QMessageBox.warning(self, "Hata", "Ad soyad zorunlu!")
                return

            driver = Driver(
                full_name=txt_name.text().strip(),
                tc_no=txt_tc.text(),
                phone=txt_phone.text(),
                adr_certificate_no=txt_adr_no.text(),
                adr_certificate_expiry=date_adr.date().toString("yyyy-MM-dd"),
                src5_no=txt_src5.text(),
                src5_expiry=date_src5.date().toString("yyyy-MM-dd"),
                license_class=txt_license.text(),
                license_expiry=date_license.date().toString("yyyy-MM-dd")
            )

            self.db.add_driver(driver)
            self._load_combobox_data()

    def _add_new_vehicle(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Yeni Arac Ekle")
        dialog.setMinimumWidth(400)
        layout = QFormLayout(dialog)

        txt_plate    = QLineEdit()
        txt_trailer  = QLineEdit()
        txt_adr_cert = QLineEdit()
        date_adr     = QDateEdit()
        date_adr.setCalendarPopup(True)
        date_adr.setDate(QDate.currentDate().addYears(1))
        date_inspection = QDateEdit()
        date_inspection.setCalendarPopup(True)
        date_inspection.setDate(QDate.currentDate())
        date_inspection_exp = QDateEdit()
        date_inspection_exp.setCalendarPopup(True)
        date_inspection_exp.setDate(QDate.currentDate().addYears(1))
        txt_tank      = QLineEdit()
        cmb_type      = QComboBox()
        cmb_type.addItems(["Tir", "Kamyon", "Tanker", "Van", "Kamyonet"])
        spin_capacity = QDoubleSpinBox()
        spin_capacity.setMaximum(999999)

        layout.addRow("Plaka *:",                txt_plate)
        layout.addRow("Dorse Plaka:",            txt_trailer)
        layout.addRow("ADR Uygunluk Belge No:",  txt_adr_cert)
        layout.addRow("ADR Gecerlilik:",         date_adr)
        layout.addRow("Muayene Tarihi:",         date_inspection)
        layout.addRow("Muayene Gecerlilik:",     date_inspection_exp)
        layout.addRow("Tank Bilgisi:",           txt_tank)
        layout.addRow("Arac Tipi:",              cmb_type)
        layout.addRow("Max Kapasite (kg):",      spin_capacity)

        buttons = QHBoxLayout()
        btn_ok     = QPushButton("Kaydet")
        btn_ok.setObjectName("success")
        btn_ok.clicked.connect(dialog.accept)
        btn_cancel = QPushButton("Iptal")
        btn_cancel.clicked.connect(dialog.reject)
        buttons.addWidget(btn_ok)
        buttons.addWidget(btn_cancel)
        layout.addRow(buttons)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            if not txt_plate.text().strip():
                QMessageBox.warning(self, "Hata", "Plaka zorunlu!")
                return

            vehicle = Vehicle(
                plate=txt_plate.text().strip().upper(),
                trailer_plate=txt_trailer.text().upper(),
                adr_compliance_cert_no=txt_adr_cert.text(),
                adr_compliance_expiry=date_adr.date().toString("yyyy-MM-dd"),
                inspection_date=date_inspection.date().toString("yyyy-MM-dd"),
                inspection_expiry=date_inspection_exp.date().toString("yyyy-MM-dd"),
                tank_info=txt_tank.text(),
                vehicle_type=cmb_type.currentText(),
                max_capacity=spin_capacity.value()
            )

            self.db.add_vehicle(vehicle)
            self._load_combobox_data()


# =============================================================================
# FİRMA YÖNETİMİ SAYFASI  (Gönderici / Alıcı / Taşıyıcı)
# =============================================================================

class CompanyEditDialog(QDialog):
    """Firma ekle / düzenle diyalogu."""

    TYPE_LABELS = {"sender": "Gönderici", "receiver": "Alıcı", "carrier": "Taşıyıcı"}

    def __init__(self, parent=None, company: "Company | None" = None):
        super().__init__(parent)
        self._company = company
        self.setWindowTitle("Firma Ekle" if company is None else "Firma Düzenle")
        self.setMinimumWidth(480)
        self._build_ui()
        if company:
            self._load(company)

    def _build_ui(self):
        layout = QFormLayout(self)
        layout.setSpacing(8)

        self.cmb_type = QComboBox()
        for key, label in self.TYPE_LABELS.items():
            self.cmb_type.addItem(label, key)
        layout.addRow("Tür *:", self.cmb_type)

        self.txt_name     = QLineEdit(); layout.addRow("Firma Adı *:", self.txt_name)
        self.txt_address  = QLineEdit(); layout.addRow("Adres:", self.txt_address)
        self.txt_city     = QLineEdit(); layout.addRow("Şehir:", self.txt_city)
        self.txt_district = QLineEdit(); layout.addRow("İlçe:", self.txt_district)
        self.txt_phone    = QLineEdit(); layout.addRow("Telefon:", self.txt_phone)
        self.txt_email    = QLineEdit(); layout.addRow("E-posta:", self.txt_email)
        self.txt_contact  = QLineEdit(); layout.addRow("İletişim Kişisi:", self.txt_contact)
        self.chk_fav      = QCheckBox("Sık Kullanılan")
        layout.addRow("", self.chk_fav)

        btn_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(self._accept)
        btn_box.rejected.connect(self.reject)
        layout.addRow(btn_box)

    def _load(self, c: "Company"):
        idx = self.cmb_type.findData(c.type)
        if idx >= 0:
            self.cmb_type.setCurrentIndex(idx)
        self.txt_name.setText(c.name)
        self.txt_address.setText(c.address)
        self.txt_city.setText(c.city)
        self.txt_district.setText(c.district)
        self.txt_phone.setText(c.phone)
        self.txt_email.setText(c.email)
        self.txt_contact.setText(c.contact_person)
        self.chk_fav.setChecked(c.is_favorite)

    def _accept(self):
        if not self.txt_name.text().strip():
            QMessageBox.warning(self, "Eksik Alan", "Firma adı boş bırakılamaz.")
            return
        self.accept()

    def get_company(self) -> "Company":
        c = self._company or Company()
        c.type           = self.cmb_type.currentData()
        c.name           = self.txt_name.text().strip()
        c.address        = self.txt_address.text().strip()
        c.city           = self.txt_city.text().strip()
        c.district       = self.txt_district.text().strip()
        c.phone          = self.txt_phone.text().strip()
        c.email          = self.txt_email.text().strip()
        c.contact_person = self.txt_contact.text().strip()
        c.is_favorite    = self.chk_fav.isChecked()
        return c


class CompanyManagementPage(QWidget):
    """Firma listele / ekle / düzenle / sil."""

    _TYPE_LABELS = {"": "Tümü", "sender": "Gönderici",
                    "receiver": "Alıcı", "carrier": "Taşıyıcı"}

    def __init__(self, db: DatabaseManager, parent=None):
        super().__init__(parent)
        self.db = db
        self._companies: list[Company] = []
        self._build_ui()
        self._load()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(10)

        # ── Başlık ──────────────────────────────────────────────────────────
        title = QLabel("Firma Yönetimi")
        title.setStyleSheet("font-size: 14pt; font-weight: 700; color: #89B4FA;")
        root.addWidget(title)

        # ── Araç çubuğu ─────────────────────────────────────────────────────
        bar = QHBoxLayout()
        self.txt_search = QLineEdit()
        self.txt_search.setPlaceholderText("Ara: firma adı, vergi no, şehir…")
        self.txt_search.textChanged.connect(self._load)
        bar.addWidget(self.txt_search, 1)

        self.cmb_type_filter = QComboBox()
        for key, label in self._TYPE_LABELS.items():
            self.cmb_type_filter.addItem(label, key)
        self.cmb_type_filter.currentIndexChanged.connect(self._load)
        bar.addWidget(self.cmb_type_filter)

        btn_add  = QPushButton("➕ Yeni Firma"); btn_add.clicked.connect(self._add)
        btn_edit = QPushButton("✏ Düzenle");   btn_edit.clicked.connect(self._edit)
        btn_del  = QPushButton("🗑 Sil");       btn_del.clicked.connect(self._delete)
        for btn in (btn_add, btn_edit, btn_del):
            bar.addWidget(btn)
        root.addLayout(bar)

        # ── Tablo ───────────────────────────────────────────────────────────
        self.table = QTableWidget(0, 7)
        self.table.setHorizontalHeaderLabels(
            ["ID", "Tür", "Firma Adı", "Vergi No", "Şehir", "Telefon", "İletişim"])
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSortingEnabled(True)
        self.table.doubleClicked.connect(self._edit)
        self.table.setColumnWidth(0, 45)


        # ← YENİ: tablo tek başına değil, splitter'ın SOL tarafına giriyor
        splitter = QSplitter(Qt.Orientation.Horizontal)
        self.table.selectionModel().selectionChanged.connect(self._on_selection_changed)
        splitter.addWidget(self.table)

        # ← YENİ: SAĞ tarafa detay paneli
        detail = QWidget()
        detail.setMinimumWidth(250)
        detail.setMaximumWidth(310)
        dl = QVBoxLayout(detail)
        dl.setContentsMargins(8, 0, 0, 0)

        dtitle = QLabel("⬡  FİRMA DETAYI")
        dtitle.setStyleSheet("font-size:9pt;font-weight:700;letter-spacing:2px;color:#585B70;padding-bottom:6px;border-bottom:2px solid #313244;background:transparent;")
        dtitle.setAlignment(Qt.AlignmentFlag.AlignLeft)
        dl.addWidget(dtitle)

        self._d_badge = QLabel("")
        self._d_badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._d_badge.setMinimumHeight(40)
        self._d_badge.setStyleSheet(
            "font-size:12pt;font-weight:700;border-radius:6px;"
            "background:#45475A;color:#CDD6F4;padding:4px;")
        dl.addWidget(self._d_badge)

        grp = QGroupBox("Bilgiler")
        gf = QFormLayout(grp)
        gf.setSpacing(5)
        self._d_name    = QLabel("—"); self._d_name.setWordWrap(True)
        self._d_name.setStyleSheet("font-weight:700;font-size:11pt;color:#89B4FA;")
        self._d_city    = QLabel("—")
        self._d_phone   = QLabel("—")
        self._d_email   = QLabel("—"); self._d_email.setWordWrap(True)
        self._d_contact = QLabel("—")
        self._d_address = QLabel("—"); self._d_address.setWordWrap(True)
        gf.addRow("Ad:",        self._d_name)
        gf.addRow("Şehir:",     self._d_city)
        gf.addRow("Tel:",       self._d_phone)
        gf.addRow("E-posta:",   self._d_email)
        gf.addRow("İletişim:",  self._d_contact)
        gf.addRow("Adres:",     self._d_address)
        dl.addWidget(grp)
        dl.addStretch()

        splitter.addWidget(detail)
        splitter.setSizes([680, 280])
        root.addWidget(splitter, 1)     # ← splitter'ı root'a ekle (tablo+detay birlikte)

        self.lbl_count = QLabel("")
        root.addWidget(self.lbl_count)

    def _on_selection_changed(self):
        c = self._selected_company()
        if not c:
            self._d_badge.setText("—")
            for lbl in (self._d_name,self._d_city,self._d_phone,
                        self._d_email,self._d_contact,self._d_address):
                lbl.setText("—")
            return
        colors={"sender":"#A6E3A1","receiver":"#89B4FA","carrier":"#F9E2AF"}
        self._d_badge.setText(self._TYPE_LABELS.get(c.type,c.type))
        self._d_badge.setStyleSheet(
            f"font-size:12pt;font-weight:700;border-radius:6px;"
            f"background:#45475A;color:{colors.get(c.type,'#CDD6F4')};padding:4px;")
        self._d_name.setText(c.name)
        self._d_city.setText(c.city or "—")
        self._d_phone.setText(c.phone or "—")
        self._d_email.setText(getattr(c,"email","") or "—")
        self._d_contact.setText(c.contact_person or "—")
        self._d_address.setText(getattr(c,"address","") or "—")

    def _load(self):
        search  = self.txt_search.text().strip() or None
        c_type  = self.cmb_type_filter.currentData() or None
        self._companies = self.db.get_companies(company_type=c_type, search=search)

        self.table.blockSignals(True)
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(self._companies))
        TYPE_COLORS = {"sender": "#A6E3A1", "receiver": "#89B4FA", "carrier": "#F9E2AF"}
        for row, c in enumerate(self._companies):
            self.table.setItem(row, 0, QTableWidgetItem(str(c.id)))
            type_item = QTableWidgetItem(self._TYPE_LABELS.get(c.type, c.type))
            type_item.setForeground(QColor(TYPE_COLORS.get(c.type, "#CDD6F4")))
            self.table.setItem(row, 1, type_item)
            name_item = QTableWidgetItem(c.name)
            if c.is_favorite:
                name_item.setText(f"⭐ {c.name}")
            self.table.setItem(row, 2, name_item)
            self.table.setItem(row, 3, QTableWidgetItem(c.tax_number))
            self.table.setItem(row, 4, QTableWidgetItem(c.city))
            self.table.setItem(row, 5, QTableWidgetItem(c.phone))
            self.table.setItem(row, 6, QTableWidgetItem(c.contact_person))
        self.table.setSortingEnabled(True)
        self.table.blockSignals(False)
        self.lbl_count.setText(f"{len(self._companies)} firma")

    def _selected_company(self) -> "Company | None":
        row = self.table.currentRow()
        if row < 0:
            return None
        cid = int(self.table.item(row, 0).text())
        return next((c for c in self._companies if c.id == cid), None)

    def _add(self):
        dlg = CompanyEditDialog(self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.db.add_company(dlg.get_company())
            self._load()

    def _edit(self):
        c = self._selected_company()
        if not c:
            QMessageBox.information(self, "Seçim Yok", "Lütfen düzenlenecek firmayı seçin.")
            return
        dlg = CompanyEditDialog(self, c)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.db.update_company(dlg.get_company())
            self._load()

    def _delete(self):
        c = self._selected_company()
        if not c:
            QMessageBox.information(self, "Seçim Yok", "Lütfen silinecek firmayı seçin.")
            return
        ans = QMessageBox.question(
            self, "Silme Onayı",
            f"'{c.name}' silinsin mi?\nBu işlem geri alınamaz.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if ans == QMessageBox.StandardButton.Yes:
            self.db.delete_company(c.id)
            self._load()


# =============================================================================
# SÜRÜCÜ YÖNETİMİ SAYFASI
# =============================================================================

class DriverEditDialog(QDialog):
    """Sürücü ekle / düzenle diyalogu."""

    def __init__(self, parent=None, driver: "Driver | None" = None):
        super().__init__(parent)
        self._driver = driver
        self.setWindowTitle("Sürücü Ekle" if driver is None else "Sürücü Düzenle")
        self.setMinimumWidth(480)
        self._build_ui()
        if driver:
            self._load(driver)

    def _build_ui(self):
        layout = QFormLayout(self)
        layout.setSpacing(8)

        self.txt_name         = QLineEdit(); layout.addRow("Ad Soyad *:", self.txt_name)
        self.txt_tc           = QLineEdit(); layout.addRow("TC Kimlik No:", self.txt_tc)
        self.txt_phone        = QLineEdit(); layout.addRow("Telefon:", self.txt_phone)
        self.txt_src5_no      = QLineEdit(); layout.addRow("SRC5 No:", self.txt_src5_no)
        self.dt_src5_expiry   = QDateEdit()
        self.dt_src5_expiry.setCalendarPopup(True)
        self.dt_src5_expiry.setDate(QDate.currentDate())
        layout.addRow("SRC5 Bitiş:", self.dt_src5_expiry)
        self.txt_license_cls  = QLineEdit(); layout.addRow("Ehliyet Sınıfı:", self.txt_license_cls)
        self.dt_license_exp   = QDateEdit()
        self.dt_license_exp.setCalendarPopup(True)
        self.dt_license_exp.setDate(QDate.currentDate())
        layout.addRow("Ehliyet Bitiş:", self.dt_license_exp)
        self.chk_active       = QCheckBox("Aktif")
        self.chk_active.setChecked(True)
        layout.addRow("", self.chk_active)

        btn_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(self._accept)
        btn_box.rejected.connect(self.reject)
        layout.addRow(btn_box)

    def _load(self, d: "Driver"):
        self.txt_name.setText(d.full_name)
        self.txt_tc.setText(d.tc_no)
        self.txt_phone.setText(d.phone)
        self.txt_src5_no.setText(d.src5_no)
        if d.src5_expiry:
            self.dt_src5_expiry.setDate(QDate.fromString(d.src5_expiry, "yyyy-MM-dd"))
        self.txt_license_cls.setText(d.license_class)
        if d.license_expiry:
            self.dt_license_exp.setDate(QDate.fromString(d.license_expiry, "yyyy-MM-dd"))
        self.chk_active.setChecked(d.is_active)

    def _accept(self):
        if not self.txt_name.text().strip():
            QMessageBox.warning(self, "Eksik Alan", "Ad Soyad boş bırakılamaz.")
            return
        tc = self.txt_tc.text().strip()
        if tc and (not tc.isdigit() or len(tc) != 11):
            QMessageBox.warning(self, "Hatalı TC", "TC Kimlik No 11 rakamdan oluşmalıdır.")
            return
        self.accept()

    def get_driver(self) -> "Driver":
        d = self._driver or Driver()
        d.full_name    = self.txt_name.text().strip()
        d.tc_no        = self.txt_tc.text().strip()
        d.phone        = self.txt_phone.text().strip()
        d.src5_no      = self.txt_src5_no.text().strip()
        d.src5_expiry  = self.dt_src5_expiry.date().toString("yyyy-MM-dd")
        d.license_class = self.txt_license_cls.text().strip()
        d.license_expiry = self.dt_license_exp.date().toString("yyyy-MM-dd")
        d.is_active    = self.chk_active.isChecked()
        return d


# ---------------------------------------------------------------------------
# ORTAK YARDIMCI — Geçerlilik tarihi geri sayım hücresi
# ---------------------------------------------------------------------------

def _make_countdown_item(date_str: str) -> QTableWidgetItem:
    """
    Verilen 'yyyy-MM-dd' tarihinden bugüne kalan günü hesaplar.
    Metin + yeşil→kırmızı renk gradyanı uygulayarak bir QTableWidgetItem döner.

    Renk eşiği:
        > 90 gün  → koyu yeşil
        60–90 gün → sarı-yeşil
        30–60 gün → sarı/amber
        15–30 gün → turuncu
        1–14 gün  → kırmızı
        0 gün     → SON GÜN (koyu kırmızı)
        < 0 gün   → SÜRESI DOLDU (koyu kırmızı)
    """
    if not date_str or date_str.strip() == "":
        item = QTableWidgetItem("—")
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        return item

    today = QDate.currentDate()
    exp   = QDate.fromString(date_str.strip(), "yyyy-MM-dd")
    if not exp.isValid():
        item = QTableWidgetItem("—")
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        return item

    days = today.daysTo(exp)   # pozitif = kalan, negatif = geçmiş

    # ── Metin ────────────────────────────────────────────────────────────
    if days < 0:
        label = f"SÜRESI DOLDU\n({-days} gün önce)"
    elif days == 0:
        label = "⚠ SON GÜN!"
    elif days == 1:
        label = "⚠ 1 gün kaldı"
    else:
        label = f"{days} gün kaldı"

    # ── Renk gradyanı: yeşil(39,174,96) → kırmızı(192,57,43) ────────────
    THRESHOLD = 90   # bu günün üzerinde → tam yeşil
    if days > THRESHOLD:
        fg = QColor(39, 174, 96)
        bg = QColor(228, 252, 238)
    elif days <= 0:
        fg = QColor(176, 28, 28)
        bg = QColor(254, 226, 226)
    else:
        t  = 1.0 - (days / THRESHOLD)   # 0.0=yeşil → 1.0=kırmızı
        r  = int(39  + (192 - 39)  * t)
        g  = int(174 + (57  - 174) * t)
        b  = int(96  + (43  - 96)  * t)
        fg = QColor(r, g, b)
        # Açık arka plan
        br = int(228 + (254 - 228) * t)
        bg_g = max(200, int(252 - 26 * t))
        bg_b = max(200, int(238 - 38 * t))
        bg = QColor(br, bg_g, bg_b)

    item = QTableWidgetItem(label)
    item.setForeground(fg)
    item.setBackground(bg)
    item.setFont(QFont("Segoe UI", 8, QFont.Weight.Bold if days <= 14 else QFont.Weight.Normal))
    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
    # Sayısal sıralama için userData'ya days değerini yaz
    item.setData(Qt.ItemDataRole.UserRole, days)
    return item


class DriverManagementPage(QWidget):
    """Sürücü listele / ekle / düzenle / sil."""

    def __init__(self, db: DatabaseManager, parent=None):
        super().__init__(parent)
        self.db = db
        self._drivers: list[Driver] = []
        self._build_ui()
        self._load()
        # Her dakika geri sayımı güncelle
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._load)
        self._timer.start(60_000)

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(10)

        title = QLabel("Sürücü Yönetimi")
        title.setStyleSheet("font-size: 14pt; font-weight: 700; color: #89B4FA;")
        root.addWidget(title)

        bar = QHBoxLayout()
        self.txt_search = QLineEdit()
        self.txt_search.setPlaceholderText("Ara: ad soyad, TC, SRC5 no…")
        self.txt_search.textChanged.connect(self._load)
        bar.addWidget(self.txt_search, 1)

        self.chk_all = QCheckBox("Pasif sürücüleri de göster")
        self.chk_all.stateChanged.connect(self._load)
        bar.addWidget(self.chk_all)

        btn_add  = QPushButton("➕ Yeni Sürücü"); btn_add.clicked.connect(self._add)
        btn_edit = QPushButton("✏ Düzenle");    btn_edit.clicked.connect(self._edit)
        btn_del  = QPushButton("🗑 Sil");        btn_del.clicked.connect(self._delete)
        for btn in (btn_add, btn_edit, btn_del):
            bar.addWidget(btn)
        root.addLayout(bar)

        self.table = QTableWidget(0, 7)
        self.table.setHorizontalHeaderLabels(
            ["ID", "Ad Soyad", "TC No", "Telefon",
             "SRC5 No", "SRC5 Bitiş", "Kalan"])
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSortingEnabled(True)
        self.table.doubleClicked.connect(self._edit)
        self.table.setColumnWidth(0, 45)
        self.table.verticalHeader().setDefaultSectionSize(36)

        # ── Detay paneli ──────────────────────────────────────────────────────
        detail = QWidget()
        detail.setMinimumWidth(240)
        detail.setMaximumWidth(320)
        detail.setStyleSheet("background:#1E1E2E;border-left:1px solid #313244;")
        dl = QVBoxLayout(detail)
        dl.setContentsMargins(12, 16, 12, 16)
        dl.setSpacing(8)

        title_d = QLabel("⬡  SÜRÜCÜ DETAYI")
        title_d.setStyleSheet("font-size:9pt;font-weight:700;letter-spacing:2px;color:#585B70;padding-bottom:6px;border-bottom:2px solid #313244;background:transparent;")
        title_d.setAlignment(Qt.AlignmentFlag.AlignLeft)
        dl.addWidget(title_d)

        self._dd_name = QLabel("—")
        self._dd_name.setStyleSheet("font-size:13pt;font-weight:700;color:#CDD6F4;")
        self._dd_name.setWordWrap(True)
        dl.addWidget(self._dd_name)

        def _drow(label):
            lbl = QLabel(label)
            lbl.setStyleSheet("font-size:8pt;color:#585B70;")
            dl.addWidget(lbl)
            val = QLabel("—")
            val.setStyleSheet("font-size:10pt;color:#CDD6F4;")
            val.setWordWrap(True)
            dl.addWidget(val)
            return val

        self._dd_tc        = _drow("TC No")
        self._dd_phone     = _drow("Telefon")
        self._dd_src5no    = _drow("SRC5 No")
        self._dd_src5exp   = _drow("SRC5 Bitiş")
        self._dd_src5kalan = _drow("SRC5 Kalan Gün")
        self._dd_status    = _drow("Durum")
        dl.addStretch()

        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.addWidget(self.table)
        splitter.addWidget(detail)
        splitter.setSizes([700, 260])
        root.addWidget(splitter, 1)

        self.table.selectionModel().selectionChanged.connect(self._on_selection_changed)

        # Son güncelleme etiketi
        self.lbl_count = QLabel("")
        root.addWidget(self.lbl_count)

    def _on_selection_changed(self):
        d = self._selected_driver()
        if not d:
            self._dd_name.setText("—")
            for lbl in (self._dd_tc, self._dd_phone, self._dd_src5no,
                        self._dd_src5exp, self._dd_src5kalan, self._dd_status):
                lbl.setText("—")
                lbl.setStyleSheet("font-size:10pt;color:#CDD6F4;")
            return

        def _kalan_stil(exp_str, lbl_tarih, lbl_kalan):
            """Tarihi ve kalan günü renkli göster."""
            if not exp_str:
                lbl_tarih.setText("—"); lbl_kalan.setText("—")
                lbl_tarih.setStyleSheet("font-size:10pt;color:#585B70;")
                lbl_kalan.setStyleSheet("font-size:10pt;color:#585B70;")
                return
            from datetime import date
            try:
                exp  = date.fromisoformat(exp_str)
                kalan = (exp - date.today()).days
            except ValueError:
                lbl_tarih.setText(exp_str); lbl_kalan.setText("—"); return

            lbl_tarih.setText(exp_str)
            if kalan <= 0:
                txt   = f"⛔ SÜRESİ DOLDU ({-kalan} gün önce)" if kalan < 0 else "⚠ SON GÜN!"
                stil  = "font-size:10pt;color:#F38BA8;font-weight:700;"
            elif kalan <= 15:
                txt  = f"🔴 {kalan} gün kaldı"
                stil = "font-size:10pt;color:#F38BA8;font-weight:700;"
            elif kalan <= 30:
                txt  = f"🟠 {kalan} gün kaldı"
                stil = "font-size:10pt;color:#FAB387;font-weight:700;"
            elif kalan <= 90:
                txt  = f"🟡 {kalan} gün kaldı"
                stil = "font-size:10pt;color:#F9E2AF;"
            else:
                txt  = f"🟢 {kalan} gün kaldı"
                stil = "font-size:10pt;color:#A6E3A1;"
            lbl_kalan.setText(txt)
            lbl_kalan.setStyleSheet(stil)
            lbl_tarih.setStyleSheet("font-size:10pt;color:#CDD6F4;")

        self._dd_name.setText(d.full_name)
        self._dd_tc.setText(d.tc_no or "—")
        self._dd_phone.setText(d.phone or "—")
        self._dd_src5no.setText(d.src5_no or "—")
        _kalan_stil(d.src5_expiry, self._dd_src5exp, self._dd_src5kalan)
        self._dd_status.setText("Aktif ✓" if d.is_active else "Pasif ✗")
        self._dd_status.setStyleSheet(
            "font-size:10pt;color:#A6E3A1;" if d.is_active else "font-size:10pt;color:#585B70;")

    def _load(self):
        search      = self.txt_search.text().strip() or None
        active_only = not self.chk_all.isChecked()
        self._drivers = self.db.get_drivers(search=search, active_only=active_only)

        today = QDate.currentDate()

        self.table.blockSignals(True)
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(self._drivers))
        for row, d in enumerate(self._drivers):
            self.table.setItem(row, 0, QTableWidgetItem(str(d.id)))

            name_item = QTableWidgetItem(d.full_name)
            if not d.is_active:
                name_item.setForeground(QColor("#585B70"))
            self.table.setItem(row, 1, name_item)

            self.table.setItem(row, 2, QTableWidgetItem(d.tc_no))
            self.table.setItem(row, 3, QTableWidgetItem(d.phone))
            self.table.setItem(row, 4, QTableWidgetItem(d.src5_no))

            # SRC5 bitiş — renk + tarih
            src5_date_item = QTableWidgetItem(d.src5_expiry)
            if d.src5_expiry:
                exp  = QDate.fromString(d.src5_expiry, "yyyy-MM-dd")
                days = today.daysTo(exp)
                if days <= 0:
                    src5_date_item.setForeground(QColor("#B01C1C"))
                elif days <= 30:
                    src5_date_item.setForeground(QColor("#C05621"))
                elif days <= 90:
                    src5_date_item.setForeground(QColor("#B7791F"))
            self.table.setItem(row, 5, src5_date_item)

            # Geri sayım hücresi (renkli gradyan)
            self.table.setItem(row, 6, _make_countdown_item(d.src5_expiry))

        self.table.setSortingEnabled(True)
        self.table.blockSignals(False)
        from datetime import datetime as _dt
        self.lbl_count.setText(
            f"{len(self._drivers)} sürücü  |  son güncelleme: {_dt.now().strftime('%H:%M:%S')}"
        )

    def _selected_driver(self) -> "Driver | None":
        row = self.table.currentRow()
        if row < 0:
            return None
        did = int(self.table.item(row, 0).text())
        return next((d for d in self._drivers if d.id == did), None)

    def _add(self):
        dlg = DriverEditDialog(self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.db.add_driver(dlg.get_driver())
            self._load()

    def _edit(self):
        d = self._selected_driver()
        if not d:
            QMessageBox.information(self, "Seçim Yok", "Lütfen düzenlenecek sürücüyü seçin.")
            return
        dlg = DriverEditDialog(self, d)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.db.update_driver(dlg.get_driver())
            self._load()

    def _delete(self):
        d = self._selected_driver()
        if not d:
            QMessageBox.information(self, "Seçim Yok", "Lütfen silinecek sürücüyü seçin.")
            return
        ans = QMessageBox.question(
            self, "Silme Onayı",
            f"'{d.full_name}' silinsin mi?\nBu işlem geri alınamaz.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if ans == QMessageBox.StandardButton.Yes:
            self.db.execute_delete("DELETE FROM drivers WHERE id=?", (d.id,))
            self._load()


# =============================================================================
# ARAÇ YÖNETİMİ SAYFASI
# =============================================================================

class VehicleEditDialog(QDialog):
    """Araç ekle / düzenle diyalogu."""

    def __init__(self, parent=None, vehicle: "Vehicle | None" = None):
        super().__init__(parent)
        self._vehicle = vehicle
        self.setWindowTitle("Araç Ekle" if vehicle is None else "Araç Düzenle")
        self.setMinimumWidth(480)
        self._build_ui()
        if vehicle:
            self._load(vehicle)

    def _build_ui(self):
        layout = QFormLayout(self)
        layout.setSpacing(8)

        self.txt_plate       = QLineEdit(); layout.addRow("Plaka *:", self.txt_plate)
        self.txt_trailer     = QLineEdit(); layout.addRow("Dorse Plakası:", self.txt_trailer)
        self.txt_adr_cert    = QLineEdit(); layout.addRow("ADR Uyg. Sertifika No:", self.txt_adr_cert)
        self.dt_adr_exp      = QDateEdit()
        self.dt_adr_exp.setCalendarPopup(True)
        self.dt_adr_exp.setDate(QDate.currentDate())
        layout.addRow("ADR Sertifika Bitiş:", self.dt_adr_exp)
        self.dt_insp_date    = QDateEdit()
        self.dt_insp_date.setCalendarPopup(True)
        self.dt_insp_date.setDate(QDate.currentDate())
        layout.addRow("Son Muayene Tarihi:", self.dt_insp_date)
        self.dt_insp_exp     = QDateEdit()
        self.dt_insp_exp.setCalendarPopup(True)
        self.dt_insp_exp.setDate(QDate.currentDate())
        layout.addRow("Muayene Bitiş:", self.dt_insp_exp)
        self.txt_tank        = QLineEdit(); layout.addRow("Tank Bilgisi:", self.txt_tank)
        self.cmb_type        = QComboBox()
        for t in ["", "Tenteli", "Kapalı Kasa", "Tanker", "Konteyner", "Flatbed", "Diğer"]:
            self.cmb_type.addItem(t)
        layout.addRow("Araç Tipi:", self.cmb_type)
        self.spn_capacity    = QDoubleSpinBox()
        self.spn_capacity.setRange(0, 100000)
        self.spn_capacity.setSuffix(" kg")
        layout.addRow("Maks. Kapasite:", self.spn_capacity)
        self.chk_active      = QCheckBox("Aktif")
        self.chk_active.setChecked(True)
        layout.addRow("", self.chk_active)

        btn_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(self._accept)
        btn_box.rejected.connect(self.reject)
        layout.addRow(btn_box)

    def _load(self, v: "Vehicle"):
        self.txt_plate.setText(v.plate)
        self.txt_trailer.setText(v.trailer_plate)
        self.txt_adr_cert.setText(v.adr_compliance_cert_no)
        if v.adr_compliance_expiry:
            self.dt_adr_exp.setDate(QDate.fromString(v.adr_compliance_expiry, "yyyy-MM-dd"))
        if v.inspection_date:
            self.dt_insp_date.setDate(QDate.fromString(v.inspection_date, "yyyy-MM-dd"))
        if v.inspection_expiry:
            self.dt_insp_exp.setDate(QDate.fromString(v.inspection_expiry, "yyyy-MM-dd"))
        self.txt_tank.setText(v.tank_info)
        idx = self.cmb_type.findText(v.vehicle_type)
        if idx >= 0:
            self.cmb_type.setCurrentIndex(idx)
        self.spn_capacity.setValue(v.max_capacity)
        self.chk_active.setChecked(v.is_active)

    def _accept(self):
        if not self.txt_plate.text().strip():
            QMessageBox.warning(self, "Eksik Alan", "Plaka boş bırakılamaz.")
            return
        self.accept()

    def get_vehicle(self) -> "Vehicle":
        v = self._vehicle or Vehicle()
        v.plate                  = self.txt_plate.text().strip().upper()
        v.trailer_plate          = self.txt_trailer.text().strip().upper()
        v.adr_compliance_cert_no = self.txt_adr_cert.text().strip()
        v.adr_compliance_expiry  = self.dt_adr_exp.date().toString("yyyy-MM-dd")
        v.inspection_date        = self.dt_insp_date.date().toString("yyyy-MM-dd")
        v.inspection_expiry      = self.dt_insp_exp.date().toString("yyyy-MM-dd")
        v.tank_info              = self.txt_tank.text().strip()
        v.vehicle_type           = self.cmb_type.currentText()
        v.max_capacity           = self.spn_capacity.value()
        v.is_active              = self.chk_active.isChecked()
        return v


class VehicleManagementPage(QWidget):
    """Araç listele / ekle / düzenle / sil."""

    def __init__(self, db: DatabaseManager, parent=None):
        super().__init__(parent)
        self.db = db
        self._vehicles: list[Vehicle] = []
        self._build_ui()
        self._load()
        # Her dakika geri sayımı güncelle
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._load)
        self._timer.start(60_000)

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(10)

        title = QLabel("Araç Yönetimi")
        title.setStyleSheet("font-size: 14pt; font-weight: 700; color: #89B4FA;")
        root.addWidget(title)

        bar = QHBoxLayout()
        self.txt_search = QLineEdit()
        self.txt_search.setPlaceholderText("Ara: plaka, dorse plakası…")
        self.txt_search.textChanged.connect(self._load)
        bar.addWidget(self.txt_search, 1)

        self.chk_all = QCheckBox("Pasif araçları da göster")
        self.chk_all.stateChanged.connect(self._load)
        bar.addWidget(self.chk_all)

        btn_add  = QPushButton("➕ Yeni Araç");   btn_add.clicked.connect(self._add)
        btn_edit = QPushButton("✏ Düzenle");      btn_edit.clicked.connect(self._edit)
        btn_del  = QPushButton("🗑 Sil");          btn_del.clicked.connect(self._delete)
        for btn in (btn_add, btn_edit, btn_del):
            bar.addWidget(btn)
        root.addLayout(bar)

        self.table = QTableWidget(0, 10)
        self.table.setHorizontalHeaderLabels([
            "ID", "Plaka", "Dorse", "Araç Tipi",
            "ADR Sertifika No", "ADR Bitiş", "ADR Kalan",
            "Muayene Bitiş", "Mua. Kalan",
            "Kapasite"
        ])
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSortingEnabled(True)
        self.table.doubleClicked.connect(self._edit)
        self.table.setColumnWidth(0, 45)
        self.table.verticalHeader().setDefaultSectionSize(36)   # çift satır yüksekliği
        root.addWidget(self.table, 1)

        # ── Detay paneli ──────────────────────────────────────────────────────
        detail_v = QWidget()
        detail_v.setMinimumWidth(240)
        detail_v.setMaximumWidth(320)
        detail_v.setStyleSheet("background:#1E1E2E;border-left:1px solid #313244;")
        dv = QVBoxLayout(detail_v)
        dv.setContentsMargins(12, 10, 12, 12)
        dv.setSpacing(4)

        title_v = QLabel("⬡  ARAÇ DETAYI")
        title_v.setStyleSheet(
            "font-size:9pt;font-weight:700;letter-spacing:2px;color:#CDD6F4;"
            "padding-bottom:5px;border-bottom:2px solid #89B4FA;background:transparent;")
        title_v.setAlignment(Qt.AlignmentFlag.AlignLeft)
        dv.addWidget(title_v)

        # Plaka — daha küçük, başlığın hemen altında
        self._dv_plate = QLabel("—")
        self._dv_plate.setStyleSheet(
            "font-size:12pt;font-weight:700;color:#89B4FA;"
            "font-family:'Courier New';padding:4px 0 4px 0;")
        dv.addWidget(self._dv_plate)

        # Ayırıcı çizgi
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color:#313244;background:#313244;max-height:1px;")
        dv.addWidget(sep)

        # Etiket + değer aynı satırda — QFormLayout
        form = QFormLayout()
        form.setSpacing(5)
        form.setContentsMargins(0, 6, 0, 0)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        LBL_STYLE = "font-size:7.5pt;color:#585B70;"
        VAL_STYLE  = "font-size:8pt;color:#CDD6F4;font-weight:500;"

        def _frow(label_text):
            lbl = QLabel(label_text + ":")
            lbl.setStyleSheet(LBL_STYLE)
            val = QLabel("—")
            val.setStyleSheet(VAL_STYLE)
            val.setWordWrap(True)
            form.addRow(lbl, val)
            return val

        self._dv_trailer    = _frow("Dorse")
        self._dv_type       = _frow("Araç Tipi")
        self._dv_capacity   = _frow("Kapasite")
        self._dv_adr_cert   = _frow("ADR Sertifika")
        self._dv_adr_exp    = _frow("ADR Bitiş")
        self._dv_adr_kalan  = _frow("ADR Kalan")
        self._dv_insp_exp   = _frow("Muayene Bitiş")
        self._dv_insp_kalan = _frow("Mua. Kalan")
        self._dv_status     = _frow("Durum")

        dv.addLayout(form)
        dv.addStretch()

        # Tabloyu splitter içine al
        table_container = self.table
        splitter_v = QSplitter(Qt.Orientation.Horizontal)
        splitter_v.addWidget(table_container)
        splitter_v.addWidget(detail_v)
        splitter_v.setSizes([700, 260])
        root.addWidget(splitter_v, 1)

        self.table.selectionModel().selectionChanged.connect(self._on_selection_changed)

        self.lbl_count = QLabel("")
        root.addWidget(self.lbl_count)

    def _on_selection_changed(self):
        v = self._selected_vehicle()
        if not v:
            self._dv_plate.setText("—")
            for lbl in (self._dv_trailer, self._dv_type, self._dv_capacity,
                        self._dv_adr_cert, self._dv_adr_exp, self._dv_adr_kalan,
                        self._dv_insp_exp, self._dv_insp_kalan, self._dv_status):
                lbl.setText("—")
                lbl.setStyleSheet("font-size:10pt;color:#CDD6F4;")
            return

        def _kalan_stil(exp_str, lbl_tarih, lbl_kalan):
            if not exp_str:
                lbl_tarih.setText("—"); lbl_kalan.setText("—")
                lbl_tarih.setStyleSheet("font-size:10pt;color:#585B70;")
                lbl_kalan.setStyleSheet("font-size:10pt;color:#585B70;")
                return
            from datetime import date
            try:
                kalan = (date.fromisoformat(exp_str) - date.today()).days
            except ValueError:
                lbl_tarih.setText(exp_str); lbl_kalan.setText("—"); return
            lbl_tarih.setText(exp_str)
            if kalan <= 0:
                txt  = f"⛔ SÜRESİ DOLDU ({-kalan} gün önce)" if kalan < 0 else "⚠ SON GÜN!"
                stil = "font-size:10pt;color:#F38BA8;font-weight:700;"
            elif kalan <= 15:
                txt  = f"🔴 {kalan} gün kaldı"
                stil = "font-size:10pt;color:#F38BA8;font-weight:700;"
            elif kalan <= 30:
                txt  = f"🟠 {kalan} gün kaldı"
                stil = "font-size:10pt;color:#FAB387;font-weight:700;"
            elif kalan <= 90:
                txt  = f"🟡 {kalan} gün kaldı"
                stil = "font-size:10pt;color:#F9E2AF;"
            else:
                txt  = f"🟢 {kalan} gün kaldı"
                stil = "font-size:10pt;color:#A6E3A1;"
            lbl_kalan.setText(txt)
            lbl_kalan.setStyleSheet(stil)
            lbl_tarih.setStyleSheet("font-size:10pt;color:#CDD6F4;")

        self._dv_plate.setText(v.plate)
        self._dv_trailer.setText(v.trailer_plate or "—")
        self._dv_type.setText(v.vehicle_type or "—")
        self._dv_capacity.setText(f"{v.max_capacity:,.0f} kg" if v.max_capacity else "—")
        self._dv_adr_cert.setText(v.adr_compliance_cert_no or "—")
        _kalan_stil(v.adr_compliance_expiry, self._dv_adr_exp, self._dv_adr_kalan)
        _kalan_stil(v.inspection_expiry,     self._dv_insp_exp, self._dv_insp_kalan)
        self._dv_status.setText("Aktif ✓" if v.is_active else "Pasif ✗")
        self._dv_status.setStyleSheet(
            "font-size:10pt;color:#A6E3A1;" if v.is_active else "font-size:10pt;color:#585B70;")

    def _load(self):
        search      = self.txt_search.text().strip() or None
        active_only = not self.chk_all.isChecked()
        self._vehicles = self.db.get_vehicles(search=search, active_only=active_only)

        today = QDate.currentDate()

        self.table.blockSignals(True)
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(self._vehicles))
        for row, v in enumerate(self._vehicles):
            # 0 ID
            self.table.setItem(row, 0, QTableWidgetItem(str(v.id)))

            # 1 Plaka
            plate_item = QTableWidgetItem(v.plate)
            plate_item.setFont(QFont("Courier New", 9, QFont.Weight.Bold))
            if not v.is_active:
                plate_item.setForeground(QColor("#585B70"))
            self.table.setItem(row, 1, plate_item)

            # 2 Dorse
            self.table.setItem(row, 2, QTableWidgetItem(v.trailer_plate))

            # 3 Araç Tipi
            self.table.setItem(row, 3, QTableWidgetItem(v.vehicle_type))

            # 4 ADR Sertifika No
            self.table.setItem(row, 4, QTableWidgetItem(v.adr_compliance_cert_no))

            # 5 ADR Bitiş — tarih + renk
            adr_date_item = QTableWidgetItem(v.adr_compliance_expiry)
            if v.adr_compliance_expiry:
                days_adr = today.daysTo(QDate.fromString(v.adr_compliance_expiry, "yyyy-MM-dd"))
                if days_adr <= 0:
                    adr_date_item.setForeground(QColor("#B01C1C"))
                elif days_adr <= 30:
                    adr_date_item.setForeground(QColor("#C05621"))
                elif days_adr <= 90:
                    adr_date_item.setForeground(QColor("#B7791F"))
            self.table.setItem(row, 5, adr_date_item)

            # 6 ADR Kalan — renkli geri sayım
            self.table.setItem(row, 6, _make_countdown_item(v.adr_compliance_expiry))

            # 7 Muayene Bitiş — tarih + renk
            insp_date_item = QTableWidgetItem(v.inspection_expiry)
            if v.inspection_expiry:
                days_insp = today.daysTo(QDate.fromString(v.inspection_expiry, "yyyy-MM-dd"))
                if days_insp <= 0:
                    insp_date_item.setForeground(QColor("#B01C1C"))
                elif days_insp <= 30:
                    insp_date_item.setForeground(QColor("#C05621"))
                elif days_insp <= 90:
                    insp_date_item.setForeground(QColor("#B7791F"))
            self.table.setItem(row, 7, insp_date_item)

            # 8 Muayene Kalan — renkli geri sayım
            self.table.setItem(row, 8, _make_countdown_item(v.inspection_expiry))

            # 9 Kapasite
            cap_item = QTableWidgetItem(
                f"{v.max_capacity:,.0f} kg" if v.max_capacity else "—")
            cap_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row, 9, cap_item)

        self.table.setSortingEnabled(True)
        self.table.blockSignals(False)
        from datetime import datetime as _dt
        self.lbl_count.setText(
            f"{len(self._vehicles)} araç  |  son güncelleme: {_dt.now().strftime('%H:%M:%S')}"
        )

    def _selected_vehicle(self) -> "Vehicle | None":
        row = self.table.currentRow()
        if row < 0:
            return None
        vid = int(self.table.item(row, 0).text())
        return next((v for v in self._vehicles if v.id == vid), None)

    def _add(self):
        dlg = VehicleEditDialog(self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.db.add_vehicle(dlg.get_vehicle())
            self._load()

    def _edit(self):
        v = self._selected_vehicle()
        if not v:
            QMessageBox.information(self, "Seçim Yok", "Lütfen düzenlenecek aracı seçin.")
            return
        dlg = VehicleEditDialog(self, v)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.db.update_vehicle(dlg.get_vehicle())
            self._load()

    def _delete(self):
        v = self._selected_vehicle()
        if not v:
            QMessageBox.information(self, "Seçim Yok", "Lütfen silinecek aracı seçin.")
            return
        ans = QMessageBox.question(
            self, "Silme Onayı",
            f"'{v.plate}' silinsin mi?\nBu işlem geri alınamaz.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if ans == QMessageBox.StandardButton.Yes:
            self.db.execute_delete("DELETE FROM vehicles WHERE id=?", (v.id,))
            self._load()


# =============================================================================
# SEVKIYAT ARSIV SAYFASI
# =============================================================================

class ShipmentArchivePage(QWidget):

    def __init__(self, db: DatabaseManager, parent=None):
        super().__init__(parent)
        self.db = db
        self.parent_window = parent
        self._setup_ui()
        self._load_shipments()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        # Baslik
        title = QLabel("SEVKİYAT ARŞİVİ")
        title.setObjectName("title")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)

        # Filtre grubu
        filter_group = QGroupBox("Filtre ve Arama")
        filter_layout = QHBoxLayout(filter_group)
        filter_layout.setSpacing(8)

        filter_layout.addWidget(QLabel("Tarih:"))
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addMonths(-3))
        self.date_from.dateChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self.date_from)

        filter_layout.addWidget(QLabel("—"))
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        self.date_to.dateChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self.date_to)

        filter_layout.addWidget(QLabel("Durum:"))
        self.cmb_status = QComboBox()
        self.cmb_status.addItems([
            "Tumu", "Taslak", "Onaylandi", "Yazdirildi", "Arsivlendi", "Iptal Edildi"
        ])
        self.cmb_status.currentIndexChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self.cmb_status)

        filter_layout.addWidget(QLabel("Ara:"))
        self.txt_search = QLineEdit()
        self.txt_search.setPlaceholderText("Evrak no, gönderici veya alıcı...")
        self.txt_search.setMinimumWidth(220)
        self.txt_search.textChanged.connect(self._on_search_changed)
        # Enter ile de arama tetiklensin
        self.txt_search.returnPressed.connect(self._load_shipments)
        filter_layout.addWidget(self.txt_search)

        btn_refresh = QPushButton("Yenile")
        btn_refresh.setMinimumHeight(30)
        btn_refresh.clicked.connect(self._load_shipments)
        filter_layout.addWidget(btn_refresh)

        filter_layout.addStretch()
        self.lbl_count = QLabel("0 kayıt")
        self.lbl_count.setObjectName("subtitle")
        filter_layout.addWidget(self.lbl_count)

        layout.addWidget(filter_group)

        # Tablo
        self.tbl_shipments = QTableWidget()
        self.tbl_shipments.setColumnCount(10)
        self.tbl_shipments.setHorizontalHeaderLabels([
            "ID", "Evrak No", "Tarih", "Gönderici", "Alıcı", "Sürücü",
            "Plaka", "Puan", "Turuncu Plaka", "Durum"
        ])
        hdr = self.tbl_shipments.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        self.tbl_shipments.setColumnWidth(1, 160)
        self.tbl_shipments.setColumnWidth(2, 90)
        self.tbl_shipments.setColumnWidth(5, 130)
        self.tbl_shipments.setColumnWidth(6, 90)
        self.tbl_shipments.setColumnWidth(7, 60)
        self.tbl_shipments.setColumnWidth(8, 100)
        self.tbl_shipments.setColumnWidth(9, 90)
        self.tbl_shipments.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tbl_shipments.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tbl_shipments.setAlternatingRowColors(True)
        self.tbl_shipments.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tbl_shipments.customContextMenuRequested.connect(self._show_context_menu)
        self.tbl_shipments.doubleClicked.connect(self._on_double_click)
        self.tbl_shipments.setColumnHidden(0, True)
        # Sütun başlığına tıkla → sırala
        hdr.sectionClicked.connect(self._on_header_clicked)
        self._sort_col = -1
        self._sort_asc = True

        layout.addWidget(self.tbl_shipments)

        # Alt buton satırı
        btn_layout = QHBoxLayout()

        btn_open = QPushButton("Aç  (Çift Tık)")
        btn_open.setObjectName("primary")
        btn_open.setMinimumHeight(34)
        btn_open.clicked.connect(self._open_selected)
        btn_layout.addWidget(btn_open)

        btn_pdf = QPushButton("PDF Oluştur")
        btn_pdf.setMinimumHeight(34)
        btn_pdf.clicked.connect(self._export_selected_pdf)
        btn_layout.addWidget(btn_pdf)

        btn_duplicate = QPushButton("Kopyala")
        btn_duplicate.setMinimumHeight(34)
        btn_duplicate.clicked.connect(self._duplicate_selected)
        btn_layout.addWidget(btn_duplicate)

        btn_delete = QPushButton("Sil")
        btn_delete.setObjectName("danger")
        btn_delete.setMinimumHeight(34)
        btn_delete.clicked.connect(self._delete_selected)
        btn_layout.addWidget(btn_delete)

        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # Arama geciktirici timer (300ms debounce)
        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.timeout.connect(self._load_shipments)

    def _on_filter_changed(self):
        """Filtre degisince aninda yenile."""
        self._load_shipments()

    def _on_search_changed(self, text: str):
        """300ms debounce ile arama — her tus basiminda sorgu atmaz."""
        self._search_timer.start(300)

    def _on_header_clicked(self, col: int):
        """Sutun basligina tikla → sirala."""
        if self._sort_col == col:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col
            self._sort_asc = True
        self.tbl_shipments.sortItems(col, Qt.SortOrder.AscendingOrder if self._sort_asc
                                     else Qt.SortOrder.DescendingOrder)

    def _load_shipments(self):
        """
        Tek JOIN sorgusuyla tum verileri getirir.
        Eski N+1 yaklasim: 500 sevkiyat = 2000 DB sorgusu.
        Yeni yaklasim: 500 sevkiyat = 1 DB sorgusu.
        """
        status_filter = self.cmb_status.currentText()
        if status_filter == "Tumu":
            status_filter = None

        date_from = self.date_from.date().toString("yyyy-MM-dd")
        date_to   = self.date_to.date().toString("yyyy-MM-dd")
        search    = self.txt_search.text().strip() or None

        rows = self.db.get_shipments_with_details(
            status=status_filter,
            date_from=date_from,
            date_to=date_to,
            search=search,
            limit=500
        )

        self.tbl_shipments.blockSignals(True)
        self.tbl_shipments.setRowCount(len(rows))
        
        
        # Dark tema rozet renkleri: (arka plan = koyu ton, yazı = canlı renk)
        STATUS_COLORS = {
            "Taslak":       ("#3A2F00", "#F9E2AF"),
            "Onaylandi":    ("#1A2B1A", "#A6E3A1"),
            "Yazdirildi":   ("#0F1B2D", "#89B4FA"),
            "Arsivlendi":   ("#262633", "#A6ADC8"),
            "Iptal Edildi": ("#2D1014", "#F38BA8"),
        }


        for i, row in enumerate(rows):
            self.tbl_shipments.setItem(i, 0, QTableWidgetItem(str(row["id"])))
            self.tbl_shipments.setItem(i, 1, QTableWidgetItem(row["document_no"] or ""))
            self.tbl_shipments.setItem(i, 2, QTableWidgetItem(row["document_date"] or ""))
            self.tbl_shipments.setItem(i, 3, QTableWidgetItem(row["sender_name"] or "—"))
            self.tbl_shipments.setItem(i, 4, QTableWidgetItem(row["receiver_name"] or "—"))
            self.tbl_shipments.setItem(i, 5, QTableWidgetItem(row["driver_name"] or "—"))
            self.tbl_shipments.setItem(i, 6, QTableWidgetItem(row["vehicle_plate"] or "—"))

            puan_item = QTableWidgetItem(f"{row['total_points']:.0f}")
            puan_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if row["total_points"] > 1000:
                puan_item.setForeground(QColor("#F38BA8"))
            self.tbl_shipments.setItem(i, 7, puan_item)

            if row["orange_plate_required"]:
                op_item = QTableWidgetItem("⚠ ZORUNLU")
                op_item.setBackground(QColor("#2D1014"))
                op_item.setForeground(QColor("#F38BA8"))
            else:
                op_item = QTableWidgetItem("✓ Gerekmiyor")
                op_item.setForeground(QColor("#A6E3A1"))
            op_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.tbl_shipments.setItem(i, 8, op_item)

            status = row["status"] or "Taslak"
            st_item = QTableWidgetItem(status)
            st_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if status in STATUS_COLORS:
                bg, fg = STATUS_COLORS[status]
                st_item.setBackground(QColor(bg))
                st_item.setForeground(QColor(fg))
            self.tbl_shipments.setItem(i, 9, st_item)

        self.tbl_shipments.blockSignals(False)
        self.lbl_count.setText(f"{len(rows)} kayıt")

    def _show_context_menu(self, position):
        menu = QMenu(self)

        open_action      = menu.addAction("Ac")
        open_action.triggered.connect(self._open_selected)

        pdf_action       = menu.addAction("PDF Olustur")
        pdf_action.triggered.connect(self._export_selected_pdf)

        duplicate_action = menu.addAction("Kopyala")
        duplicate_action.triggered.connect(self._duplicate_selected)

        menu.addSeparator()

        delete_action    = menu.addAction("Sil")
        delete_action.triggered.connect(self._delete_selected)

        menu.exec(self.tbl_shipments.viewport().mapToGlobal(position))

    def _get_selected_id(self) -> Optional[int]:
        row = self.tbl_shipments.currentRow()
        if row >= 0:
            id_item = self.tbl_shipments.item(row, 0)
            if id_item:
                return int(id_item.text())
        return None

    def _open_selected(self):
        shipment_id = self._get_selected_id()
        if shipment_id:
            self.load_selected_shipment(shipment_id)

    def load_selected_shipment(self, shipment_id: int = None):
        if shipment_id is None:
            shipment_id = self._get_selected_id()
        if not shipment_id:
            QMessageBox.warning(self, "Hata", "Lutfen bir evrak secin!")
            return

        if self.parent_window:
            self.parent_window._switch_page(0)
            self.parent_window.shipment_page.load_shipment(shipment_id)
            self.parent_window._update_adr_panel()

    def _export_selected_pdf(self):
        shipment_id = self._get_selected_id()
        if not shipment_id:
            return

        shipment = self.db.get_shipment(shipment_id)
        if not shipment:
            return

        from datetime import datetime as _dt
        default_name = f"ADR_{_dt.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        path, _ = QFileDialog.getSaveFileName(self, "PDF Kaydet",
            default_name, "PDF (*.pdf)")
        if path and self.parent_window:
            self.parent_window.shipment_page.load_shipment(shipment_id)
            self.parent_window.shipment_page._generate_pdf(path)
            QMessageBox.information(self, "Basarili", f"PDF olusturuldu:\n{path}")

    def _duplicate_selected(self):
        shipment_id = self._get_selected_id()
        if not shipment_id:
            return

        shipment = self.db.get_shipment(shipment_id)
        if not shipment:
            return

        shipment.id           = None
        shipment.document_no  = ADREngine.format_document_number()
        shipment.status       = "Taslak"
        shipment.document_date = QDate.currentDate().toString("yyyy-MM-dd")

        new_id = self.db.add_shipment(shipment)

        items = self.db.get_shipment_items(shipment_id)
        for item in items:
            item.id          = None
            item.shipment_id = new_id
            self.db.add_shipment_item(item)

        self._load_shipments()
        QMessageBox.information(self, "Basarili",
            f"Evrak kopyalandi. Yeni No: {shipment.document_no}")

    def _delete_selected(self):
        shipment_id = self._get_selected_id()
        if not shipment_id:
            return

        shipment = self.db.get_shipment(shipment_id)
        if not shipment:
            return

        reply = QMessageBox.question(self, "Sil",
            f"Evrak {shipment.document_no} silinecek. Emin misiniz?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        if reply == QMessageBox.StandardButton.Yes:
            self.db.delete_shipment(shipment_id)
            self._load_shipments()
            QMessageBox.information(self, "Basarili", "Evrak silindi.")

    def _on_double_click(self):
        self._open_selected()


# =============================================================================
# DEMO VERI YUKLEYICI
# =============================================================================

def load_demo_data(db: DatabaseManager):
    companies = [
        Company(type="sender",   name="Petrol Ofisi A.S.", tax_number="1234567890",
                tax_office="Istanbul", address="Kartal", city="Istanbul",
                phone="0212 555 0000", email="info@petrolofisi.com.tr",
                contact_person="Ahmet Yilmaz"),
        Company(type="receiver", name="Shell Turkiye", tax_number="0987654321",
                tax_office="Ankara", address="Cankaya", city="Ankara",
                phone="0312 555 0000", email="info@shell.com.tr",
                contact_person="Mehmet Kaya"),
        Company(type="carrier",  name="Lojistik A.S.", tax_number="1122334455",
                tax_office="Izmir", address="Konak", city="Izmir",
                phone="0232 555 0000", email="info@lojistik.com.tr",
                contact_person="Ali Veli"),
    ]
    for c in companies:
        try:
            db.add_company(c)
        except:
            pass

    drivers = [
        Driver(full_name="Ahmet Yilmaz", tc_no="12345678901", phone="0555 111 2233",
               adr_certificate_no="ADR-2024-001", adr_certificate_expiry="2027-06-01",
               src5_no="SRC5-001", src5_expiry="2027-06-01",
               license_class="C+E", license_expiry="2030-01-01"),
        Driver(full_name="Mehmet Kaya",  tc_no="98765432109", phone="0555 444 5566",
               adr_certificate_no="ADR-2024-002", adr_certificate_expiry="2027-12-01",
               src5_no="SRC5-002", src5_expiry="2027-12-01",
               license_class="C+E", license_expiry="2031-01-01"),
    ]
    for d in drivers:
        try:
            db.add_driver(d)
        except:
            pass

    vehicles = [
        Vehicle(plate="34 ABC 123", trailer_plate="34 DEF 456",
                adr_compliance_cert_no="ADR-V-2024-001",
                adr_compliance_expiry="2026-06-01",
                inspection_date="2025-01-15", inspection_expiry="2026-01-15",
                tank_info="Tanker 20.000 LT", vehicle_type="Tanker", max_capacity=20000),
        Vehicle(plate="06 XYZ 789", trailer_plate="",
                adr_compliance_cert_no="ADR-V-2024-002",
                adr_compliance_expiry="2026-12-01",
                inspection_date="2025-03-20", inspection_expiry="2026-03-20",
                tank_info="Van 5.000 kg", vehicle_type="Van", max_capacity=5000),
    ]
    for v in vehicles:
        try:
            db.add_vehicle(v)
        except:
            pass

    chemicals = [
        Chemical(un_number="1203", proper_shipping_name_tr="BENZIN",
                 proper_shipping_name_en="GASOLINE", class_code="3",
                 packing_group="II", tunnel_code="D/E", transport_category="2",
                 segregation_group="Yanici Maddeler", lq_allowed=True, eq_allowed=True),
        Chemical(un_number="1202", proper_shipping_name_tr="GAZ YAGI (DIESEL)",
                 proper_shipping_name_en="GAS OIL", class_code="3",
                 packing_group="III", tunnel_code="E", transport_category="3",
                 segregation_group="Yanici Maddeler", lq_allowed=True, eq_allowed=True),
        Chemical(un_number="1789", proper_shipping_name_tr="HIDROKLORIK ASIT",
                 proper_shipping_name_en="HYDROCHLORIC ACID", class_code="8",
                 packing_group="II", tunnel_code="E", transport_category="2",
                 segregation_group="Asitler", lq_allowed=True, eq_allowed=False),
        Chemical(un_number="1824", proper_shipping_name_tr="SODYUM HIDROKSIT COZELTISI",
                 proper_shipping_name_en="SODIUM HYDROXIDE SOLUTION", class_code="8",
                 packing_group="II", tunnel_code="E", transport_category="2",
                 segregation_group="Bazlar", lq_allowed=True, eq_allowed=False),
        Chemical(un_number="1075", proper_shipping_name_tr="PETROL GAZI (LPG)",
                 proper_shipping_name_en="LIQUEFIED PETROLEUM GAS", class_code="2.1",
                 packing_group="", tunnel_code="C/E", transport_category="2",
                 segregation_group="Yanici Maddeler", lq_allowed=False, eq_allowed=False),
        Chemical(un_number="1942", proper_shipping_name_tr="AMONYUM NITRAT",
                 proper_shipping_name_en="AMMONIUM NITRATE", class_code="5.1",
                 packing_group="III", tunnel_code="D", transport_category="2",
                 segregation_group="Yukseltgenler", lq_allowed=False, eq_allowed=False),
    ]
    for c in chemicals:
        try:
            db.add_chemical(c)
        except:
            pass


# =============================================================================
# ANA UYGULAMA
# =============================================================================
def main():
    # 1. Önce QApplication oluşturulmalıdır (Şarttır)
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon("icon.ico"))
    app.setApplicationName(APP_NAME)
    app.setApplicationVersion(APP_VERSION)
    app.setOrganizationName(APP_ORGANIZATION)

    font = QFont("Segoe UI", 10)
    app.setFont(font)

    # 2. Açılış Ekranı (Splash Screen) Başlatılıyor
    splash = QSplashScreen()
    splash.setStyleSheet("background-color: #1E1E2E;")

    progress = QProgressBar(splash)
    progress.setGeometry(10, 80, 380, 20)
    progress.setRange(0, 100)
    progress.setValue(0)

    splash.show()
    app.processEvents()

    # %20: Veritabanı Başlatılıyor
    progress.setValue(20)
    app.processEvents()
    db = DatabaseManager()

    # %50: Demo Verileri Yükleniyor
    progress.setValue(50)
    app.processEvents()
    load_demo_data(db)

    # %80: Arka plan yüklemeleri bitti, Güvenlik Kontrolleri Başlıyor
    progress.setValue(80)
    app.processEvents()

    # ── GÜVENLİK & LİSANS ENTEGRASYONU ───────────────────────────────────────────
    lic_info = LicenseManager.validate()
    security = SecurityManager(db.conn)

    # Giriş penceresini açarken açılış ekranını geçici olarak gizliyoruz
    splash.hide() 
    login_dlg = LoginDialog(security, lic_info)
    login_dlg.exec()

    # Eğer giriş başarısızsa veya kullanıcı pencereyi kapattıysa programı sonlandır
    if not login_dlg.logged_in:
        sys.exit(0)

    # Giriş başarılıysa açılış ekranını tekrar gösterip yüklemeyi tamamla
    splash.show()
    app.processEvents()

    # Oturum Zaman Aşımı Timer'ı (Her 5 dakikada bir kontrol eder)
    global session_timer  # Python'ın bellekten silmesini (Garbage Collection) önlemek için global yaptık
    session_timer = QTimer()
    def _check_session():
        if security.check_session_timeout():
            QMessageBox.warning(None, "Oturum Süresi Doldu", 
                                "Oturum süreniz doldu.\nProgram kapanıyor.")
            security.logout()
            sys.exit(0)
        security.ping_session()
    session_timer.timeout.connect(_check_session)
    session_timer.start(5 * 60 * 1000) # 5 dakika

    # Lisans Uyarısı (Lisansın bitmesine 30 günden az kaldıysa uyarı verir)
    if lic_info["valid"] and lic_info["days_left"] <= 30:
        QMessageBox.warning(None, "Lisans Uyarısı",
            f"Lisansınızın süresi {lic_info['days_left']} gün içinde dolacak.\n"
            "Sağlayıcınızla iletişime geçin.")
    # ──────────────────────────────────────────────────────────────────────────

    # %100: Ana pencere güvenlik nesnesiyle birlikte oluşturuluyor
    window = ADRTransportPro(security=security)
    window.session_timer = session_timer # Timer'ı garantiye almak için pencereye bağladık

    progress.setValue(100)
    app.processEvents()

    # Açılış ekranını kapat ve ana pencereyi devreye al
    splash.finish(window)

    sys.exit(app.exec())

if __name__ == "__main__":
    main()