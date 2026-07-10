"""Ana program — Excel evrak çıktısı (v4.4) testleri.

Tespit edilen eksik (main.py'de düzeltilmiş ama ana programa hiç
taşınmamıştı): export_excel() yalnızca ürün listesi basıyordu; ADR
kontrol özeti, LQ/EQ bayrakları, tünel kodu ve TASLAK uyarısı yoktu.
"""
import os
import sys
import tempfile
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

import importlib.util
_spec = importlib.util.spec_from_file_location(
    "anaprog", str(ROOT / "adr_transport_pro_2026.py"))
M = importlib.util.module_from_spec(_spec)
sys.modules["anaprog"] = M
_spec.loader.exec_module(M)

from PyQt6.QtWidgets import QApplication, QMessageBox, QFileDialog  # noqa: E402

_app = QApplication.instance() or QApplication([])
for _f in ("information", "warning", "critical"):
    setattr(QMessageBox, _f, staticmethod(lambda *a, **k: QMessageBox.StandardButton.Ok))


@pytest.fixture()
def page(tmp_path):
    db = M.DatabaseManager(str(tmp_path / "t.db"))
    p = M.ShipmentEditorPage(db, parent=None)
    p.items = [M.ShipmentItem(
        un_number="1203", proper_name="BENZİN", class_code="3", packing_group="II",
        net_quantity=100, packaging_count=2, unit="L", packaging_type="Varil",
        transport_category="2", tunnel_code="D/E", is_lq=True)]
    return p


def _sheet_text(path):
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    return "\n".join(str(c.value) for row in ws.iter_rows() for c in row
                     if c.value is not None)


def _export(page, tmp_path, status="TASLAK", name="out.xlsx"):
    out = str(tmp_path / name)
    QFileDialog.getSaveFileName = staticmethod(lambda *a, **k: (out, ""))
    page.lbl_status.setText(status)
    page.export_excel()
    return _sheet_text(out)


class TestExcelAdrSummary:
    def test_summary_block_present(self, page, tmp_path):
        text = _export(page, tmp_path)
        for needle in ("ADR KONTROL ÖZETİ", "1.1.3.6 Puanı", "Turuncu Plaka",
                       "Yazılı Talimat", "Tünel Kısıtı", "Muafiyet", "Evrak Durumu"):
            assert needle in text, f"Excel'de eksik: {needle}"

    def test_item_row_has_tunnel_and_lq_flag(self, page, tmp_path):
        text = _export(page, tmp_path)
        assert "D/E" in text
        assert "LQ" in text

    def test_draft_banner_present_when_unapproved(self, page, tmp_path):
        text = _export(page, tmp_path, status="TASLAK")
        assert "Onaylanmamış evrak" in text

    def test_draft_banner_absent_when_approved(self, page, tmp_path):
        text = _export(page, tmp_path, status="ONAYLANDI")
        assert "Onaylanmamış evrak" not in text
        assert "ONAYLANDI" in text

    def test_critical_errors_surface_in_excel(self, page, tmp_path):
        # Kategori 0 -> KRİTİK hata bekleniyor (LQ olmayan kalemle test edilmeli;
        # LQ işaretli kalemler 1.1.3.6 puan hesabından zaten muaf tutulur)
        page.items[0].is_lq = False
        page.items[0].transport_category = "0"
        text = _export(page, tmp_path)
        assert "KRİTİK UYARILAR" in text or "Kategori 0" in text

    def test_export_does_not_crash_with_no_items(self, tmp_path):
        db = M.DatabaseManager(str(tmp_path / "t2.db"))
        p = M.ShipmentEditorPage(db, parent=None)
        p.items = []
        out = str(tmp_path / "empty.xlsx")
        QFileDialog.getSaveFileName = staticmethod(lambda *a, **k: (out, ""))
        p.export_excel()  # çökmemeli
        assert os.path.exists(out)
