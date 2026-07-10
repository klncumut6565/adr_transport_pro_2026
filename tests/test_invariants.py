"""Değişmezlik (invariant) ve zorlamalı testler.

Bu dosya, tekil senaryolardan öte, motorun HER koşulda korumak zorunda
olduğu matematiksel/mevzuatsal özellikleri sınar. Piyasadaki sistemlerin
çoğu bu tür sistematik doğrulama yapmaz; sıfır hataya yaklaşmanın yolu
tam olarak budur.
"""
import csv
import itertools
import time

import pytest

from tests.conftest import RULE_FILE
from adr_mix_pro.core.checker import MixChecker
from adr_mix_pro.core.rule_engine import SegregationRuleEngine
from adr_mix_pro.core.risk_engine import score_result
from adr_mix_pro.models import ProductRecord
from tests.conftest import make_record


# =========================================================================
# 1) Simetri: check(A,B) == check(B,A) — sıra sonucu ASLA değiştirmemeli
# =========================================================================
class TestSymmetry:
    LABELS = ["1", "1.4S", "2.1", "2.2", "2.3", "3", "4.1", "4.2", "4.3",
              "5.1", "5.2", "6.1", "6.2", "7", "7A", "8", "9"]

    def test_rule_engine_symmetric(self, rule_engine):
        for l1, l2 in itertools.combinations(self.LABELS, 2):
            v12 = rule_engine.check(l1, l2)
            v21 = rule_engine.check(l2, l1)
            assert v12.status == v21.status, f"Asimetri: {l1}/{l2}"

    def test_checker_symmetric_over_real_db(self, db, rule_engine):
        ck = MixChecker(db, rule_engine)
        uns = [r.un_no for r in db.all_records()][:12]
        for u1, u2 in itertools.combinations(uns, 2):
            r12, _ = ck.check_all([u1, u2])
            r21, _ = ck.check_all([u2, u1])
            assert r12[0].status == r21[0].status, f"Asimetri: {u1}/{u2}"
            assert r12[0].risk_score == r21[0].risk_score


# =========================================================================
# 2) Kural dosyası bütünlüğü: eksik hücre, çelişki, kapsam
# =========================================================================
class TestRuleFileIntegrity:
    def test_all_label_pairs_covered(self, rule_engine):
        """Yaygın etiketlerin TÜM ikilileri için tanımlı kural olmalı;
        UNKNOWN, kullanıcının manuel kontrolüne kaldığı için risklidir."""
        labels = ["1", "1.4", "1.5", "1.6", "1.4S", "2.1", "2.2", "2.3",
                  "3", "4.1", "4.2", "4.3", "5.1", "5.2", "6.1", "6.2",
                  "7", "7A", "7B", "7C", "7E", "8", "9"]
        gaps = [
            (a, b)
            for a, b in itertools.combinations_with_replacement(labels, 2)
            if rule_engine.check(a, b).status == "UNKNOWN"
        ]
        assert not gaps, f"Kural boşlukları: {gaps}"

    def test_no_duplicate_conflicting_rows(self):
        seen = {}
        with open(RULE_FILE, encoding="utf-8-sig") as fh:
            for i, row in enumerate(csv.DictReader(fh), start=2):
                key = tuple(sorted((row["LABEL1"].strip(), row["LABEL2"].strip())))
                if key in seen:
                    assert seen[key][1] == row["STATUS"].strip(), (
                        f"Çelişkili kural: {key} satır {seen[key][0]} vs {i}")
                seen[key] = (i, row["STATUS"].strip())

    def test_class1_never_ok_with_others_except_14s(self, rule_engine):
        """Mevzuat değişmezi: 1/1.4/1.5/1.6 (1.4S hariç) hiçbir sınıf-dışı
        etiketle 'OK' OLAMAZ."""
        explosives = ["1", "1.4", "1.5", "1.6"]
        others = ["2.1", "2.2", "2.3", "3", "4.1", "4.2", "4.3",
                  "5.1", "5.2", "6.1", "6.2", "7", "8", "9"]
        for e in explosives:
            for o in others:
                assert rule_engine.check(e, o).status == "NO", f"{e}/{o} OK çıktı!"

    def test_missing_rule_file_clear_error(self, tmp_path):
        from adr_mix_pro.exceptions import ADRError
        with pytest.raises(ADRError):
            SegregationRuleEngine(tmp_path / "yok.csv")


# =========================================================================
# 3) Risk skoru değişmezleri
# =========================================================================
class TestRiskScore:
    def test_ordering(self):
        """YASAK her zaman en yüksek; OK her zaman en düşük olmalı."""
        s_no = score_result("NO")
        s_ex = score_result("EXPLOSIVE_SPECIAL")
        s_un = score_result("UNKNOWN")
        s_fc = score_result("FOOD_CAUTION")
        s_ok = score_result("OK")
        assert s_no > s_ex > s_un >= s_fc > s_ok

    def test_bounds(self):
        for st in ["NO", "EXPLOSIVE_SPECIAL", "UNKNOWN", "FOOD_CAUTION", "OK"]:
            for f in (False, True):
                for t in (False, True):
                    s = score_result(st, has_food_caution=f, has_tunnel_restriction=t)
                    assert 0 <= s <= 100, f"{st} f={f} t={t} -> {s}"

    def test_addons_never_decrease(self):
        for st in ["OK", "UNKNOWN", "FOOD_CAUTION"]:
            base = score_result(st)
            assert score_result(st, has_food_caution=True) >= base
            assert score_result(st, has_tunnel_restriction=True) >= base


# =========================================================================
# 4) Ölçek ve performans: kombinatorik patlamada doğruluk korunmalı
# =========================================================================
class TestScale:
    def test_100_items_4950_pairs(self, rule_engine):
        db_syn = _SynDB(
            [make_record(f"{5000+i}", f"M{i}", "3", ["3"]) for i in range(99)]
            + [make_record("0335", "PATLAYICI", "1", ["1.3G"], "1.3G")]
        )
        ck = MixChecker(db_syn, rule_engine)
        t0 = time.time()
        results, missing = ck.check_all([r.un_no for r in db_syn.all_records()])
        dt = time.time() - t0
        assert len(results) == 4950 and not missing
        assert dt < 5, f"Çok yavaş: {dt:.2f}s"
        forbidden = [r for r in results if r.status == "NO"]
        # patlayıcı diğer 99 maddenin HER BİRİYLE yasak olmalı — biri bile
        # kaçarsa sahada felaket demektir
        assert len(forbidden) == 99

    def test_result_count_formula(self, checker):
        for n, uns in [(2, ["1203", "1830"]), (3, ["1203", "1830", "1170"]),
                       (4, ["1203", "1830", "1170", "1090"])]:
            results, _ = checker.check_all(uns)
            assert len(results) == n * (n - 1) // 2


class _SynDB:
    def __init__(self, records):
        self._r = {r.un_no: r for r in records}

    def try_get_record(self, un):
        from adr_mix_pro.validators import normalize_un
        return self._r.get(normalize_un(un))

    def all_records(self):
        return list(self._r.values())


# =========================================================================
# 5) Türkçe/Unicode dayanıklılığı
# =========================================================================
class TestUnicode:
    def test_turkish_names_survive_roundtrip(self, rule_engine, tmp_path):
        recs = [
            make_record("1001", "ĞÜŞİÖÇ ığüşöç MADDESİ", "3", ["3"]),
            make_record("1002", "İSTANBUL – %50'lik ÇÖZELTİ (≤ 60 °C)", "8", ["8"]),
        ]
        ck = MixChecker(_SynDB(recs), rule_engine)
        results, _ = ck.check_all(["1001", "1002"])
        assert results[0].name1.startswith("ĞÜŞ")

        from adr_mix_pro.reports.excel_export import export_results_to_excel
        out = tmp_path / "türkçe_rapor_ğüş.xlsx"
        export_results_to_excel(results, out)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        text = " ".join(str(c.value) for row in wb.active.iter_rows() for c in row)
        assert "ĞÜŞİÖÇ" in text
