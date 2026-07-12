"""Raporlar — Faz 2d: özet metrikler, sınıf kırılımı, en-çoklar,
aylık dağılım; Excel (çok sayfalı) ve PDF çıktısı.

Masaüstündeki ReportsPage'in web karşılığı; veriler DatabaseManager'ın
mevcut rapor metotlarından gelir (get_statistics, get_class_breakdown,
get_top_senders, get_top_chemicals).
"""
import io
from datetime import datetime

import streamlit as st
from webcore.errors import turkce_hata_metni

from sayfalar._ortak import db, kullanici

st.title("📈 Raporlar")

d = db()
yil_secimi = st.selectbox("Dönem", ["Tümü"] + [str(y) for y in
                          range(datetime.now().year, datetime.now().year - 6, -1)])
yil = None if yil_secimi == "Tümü" else int(yil_secimi)

ist = d.get_statistics()
kirilim = d.get_class_breakdown(year=yil)
gonderenler = d.get_top_senders(limit=10, year=yil)
kimyasallar = d.get_top_chemicals(limit=10, year=yil)
aylik = ist.get("monthly_shipments") or []

c1, c2, c3, c4 = st.columns(4)
c1.metric("Toplam Sevkiyat", ist.get("total_shipments", 0))
c2.metric("Taslak", ist.get("draft_shipments", 0))
c3.metric("Aktif Sürücü", ist.get("active_drivers", 0))
c4.metric("Kimyasal Kaydı", ist.get("total_chemicals", 0))

sol, sag = st.columns(2)
with sol:
    st.subheader("Sınıf Kırılımı" + (f" — {yil}" if yil else ""))
    if kirilim:
        st.dataframe(kirilim, use_container_width=True, hide_index=True)
        st.bar_chart({r["class_code"] or "—": r["toplam_net_kg"] for r in kirilim})
    else:
        st.info("Seçilen dönemde sevkiyat kalemi yok.")
with sag:
    st.subheader("Son 6 Ay Sevkiyat Dağılımı")
    if aylik:
        st.bar_chart({r["month"]: r["count"] if "count" in r else list(r.values())[1]
                      for r in aylik})
    else:
        st.info("Son 6 ayda sevkiyat yok.")

s2l, s2r = st.columns(2)
with s2l:
    st.subheader("En Çok Gönderenler")
    st.dataframe(gonderenler or [], use_container_width=True, hide_index=True)
with s2r:
    st.subheader("En Çok Taşınan Kimyasallar")
    st.dataframe(kimyasallar or [], use_container_width=True, hide_index=True)

st.divider()
st.subheader("📤 Dışa Aktar")


def _excel_raporu() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()

    def sayfa(ad, basliklar, satirlar):
        ws = wb.create_sheet(ad)
        ws.append(basliklar)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in satirlar:
            ws.append([r.get(k, "") for k in r] if isinstance(r, dict) else list(r))
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(40, w + 2)

    ws0 = wb.active
    ws0.title = "Özet"
    ws0.append(["ADR Transport Pro 2026 — Rapor"])
    ws0.append([f"Dönem: {yil_secimi} | Üretim: "
                f"{datetime.now().strftime('%d.%m.%Y %H:%M')} | "
                f"Hazırlayan: {kullanici().get('full_name') or kullanici().get('username','')}"])
    ws0.append([])
    for k, v in ist.items():
        if k != "monthly_shipments":
            ws0.append([k, v])
    ws0["A1"].font = Font(bold=True, size=13)

    if kirilim:
        sayfa("Sınıf Kırılımı", list(kirilim[0].keys()), kirilim)
    if gonderenler:
        sayfa("En Çok Gönderenler", list(gonderenler[0].keys()), gonderenler)
    if kimyasallar:
        sayfa("En Çok Kimyasallar", list(kimyasallar[0].keys()), kimyasallar)
    if aylik:
        sayfa("Aylık Dağılım", list(aylik[0].keys()), aylik)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdf_raporu() -> bytes:
    from webcore.pdf import html_to_pdf_bytes

    def tablo(baslik, rows):
        if not rows:
            return ""
        keys = list(rows[0].keys())
        thead = "".join(f"<th>{k}</th>" for k in keys)
        trs = "".join("<tr>" + "".join(f"<td>{r.get(k,'')}</td>" for k in keys)
                      + "</tr>" for r in rows)
        return (f"<h3>{baslik}</h3><table><tr>{thead}</tr>{trs}</table>")

    html = f"""<html><head><meta charset="utf-8"><style>
      body {{ font-family: sans-serif; font-size: 10pt; }}
      h1 {{ font-size: 15pt; }} h3 {{ margin: 14px 0 4px; }}
      table {{ border-collapse: collapse; width: 100%; }}
      th, td {{ border: 1px solid #999; padding: 3px 6px; text-align: left; }}
      th {{ background: #1F3864; color: #fff; }}
      .meta {{ color: #666; font-size: 9pt; }}
    </style></head><body>
    <h1>ADR Transport Pro 2026 — Dönem Raporu</h1>
    <p class="meta">Dönem: {yil_secimi} · Üretim: {datetime.now().strftime('%d.%m.%Y %H:%M')}
    · Hazırlayan: {kullanici().get('full_name') or kullanici().get('username','')}</p>
    <h3>Özet</h3>
    <table>{"".join(f"<tr><td>{k}</td><td>{v}</td></tr>" for k, v in ist.items() if k != "monthly_shipments")}</table>
    {tablo("Sınıf Kırılımı", kirilim)}
    {tablo("En Çok Gönderenler", gonderenler)}
    {tablo("En Çok Taşınan Kimyasallar", kimyasallar)}
    {tablo("Aylık Dağılım (son 6 ay)", aylik)}
    </body></html>"""
    return html_to_pdf_bytes(html)


ec1, ec2 = st.columns(2)
with ec1:
    try:
        st.download_button("⬇️ Excel raporu (.xlsx)", data=_excel_raporu(),
                           file_name=f"adr_rapor_{yil_secimi}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument"
                                ".spreadsheetml.sheet",
                           use_container_width=True)
    except Exception as exc:
        st.error(f"Excel üretilemedi: {turkce_hata_metni(exc)}")
with ec2:
    try:
        st.download_button("⬇️ PDF raporu", data=_pdf_raporu(),
                           file_name=f"adr_rapor_{yil_secimi}.pdf",
                           mime="application/pdf", use_container_width=True)
    except ImportError:
        st.info("PDF için WeasyPrint gerekli (Cloud'da otomatik kurulur).")
    except Exception as exc:
        st.error(f"PDF üretilemedi: {turkce_hata_metni(exc)}")
