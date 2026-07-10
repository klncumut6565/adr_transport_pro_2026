"""Sonuçların Excel (.xlsx) biçiminde dışa aktarılması."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..constants import (
    STATUS_EXPLOSIVE_SPECIAL,
    STATUS_FOOD_CAUTION,
    STATUS_FORBIDDEN,
    STATUS_LABELS,
    STATUS_OK,
    STATUS_UNKNOWN,
)
from ..exceptions import ExportError
from ..models import PairCheckResult

_FILL_BY_STATUS = {
    STATUS_OK: PatternFill("solid", fgColor="C6EFCE"),
    STATUS_FORBIDDEN: PatternFill("solid", fgColor="FFC7CE"),
    STATUS_UNKNOWN: PatternFill("solid", fgColor="FFEB9C"),
    STATUS_EXPLOSIVE_SPECIAL: PatternFill("solid", fgColor="D9D2E9"),
    STATUS_FOOD_CAUTION: PatternFill("solid", fgColor="FCE4D6"),
}

_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

_COLUMNS = [
    ("UN 1", 10),
    ("Madde 1", 30),
    ("UN 2", 10),
    ("Madde 2", 30),
    ("Durum", 14),
    ("ADR Referansı", 14),
    ("Açıklama", 50),
    ("Risk Puanı", 10),
    ("Notlar", 60),
]


def export_results_to_excel(results: list[PairCheckResult], filepath: str | Path) -> None:
    if not results:
        raise ExportError("Dışa aktarılacak sonuç bulunmuyor.")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "ADR Karışık Yükleme"

        for col_index, (title, width) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_index, value=title)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_index)].width = width

        for row_index, result in enumerate(results, start=2):
            values = [
                result.un1,
                result.name1,
                result.un2,
                result.name2,
                STATUS_LABELS.get(result.status, result.status),
                result.adr_reference,
                result.reason,
                result.risk_score,
                "\n".join(result.notes),
            ]
            fill = _FILL_BY_STATUS.get(result.status)

            for col_index, value in enumerate(values, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=value)
                cell.alignment = Alignment(
                    wrap_text=True, vertical="top", horizontal="left"
                )
                if fill is not None:
                    cell.fill = fill

        ws.freeze_panes = "A2"
        wb.save(filepath)
    except OSError as exc:
        raise ExportError(f"Excel dosyası kaydedilemedi: {exc}") from exc
